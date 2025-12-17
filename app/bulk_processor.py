"""
Bulk Invoice Processor Module

This module provides functionality for processing multiple invoices at once:
- Parse multiple PDF/image invoices
- Generate summary reports by campaign/supplier
- Export to Excel with detailed breakdowns
- Support for Meta, Google Ads, and other invoice templates
"""

import os
import re
import tempfile
from datetime import datetime
from collections import defaultdict
from typing import Optional
from io import BytesIO

import PyPDF2

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_ENABLED = True
except ImportError:
    EXCEL_ENABLED = False


def parse_value(value_str: str) -> float:
    """
    Parse invoice value from various European/US number formats.
    Handles: '3.499,00', '3,499.00', '874,90', '1234.56'
    """
    value_str = value_str.strip()

    # Handle European format (dot as thousand separator, comma as decimal)
    if ',' in value_str and '.' in value_str:
        last_comma = value_str.rfind(',')
        last_dot = value_str.rfind('.')

        if last_comma > last_dot:
            # European: 1.234,56
            value_str = value_str.replace('.', '').replace(',', '.')
        else:
            # US: 1,234.56
            value_str = value_str.replace(',', '')
    elif ',' in value_str:
        # Only comma - check if it's decimal separator
        if re.search(r',\d{2}$', value_str):
            value_str = value_str.replace(',', '.')
        else:
            value_str = value_str.replace(',', '')
    else:
        # Format like 3.499.00 - last dot is decimal
        parts = value_str.split('.')
        if len(parts) > 2:
            value_str = ''.join(parts[:-1]) + '.' + parts[-1]

    return float(value_str)


def parse_romanian_date(date_str: str) -> Optional[datetime]:
    """Parse Romanian date format to datetime object."""
    months = {
        'ian': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'mai': 5, 'iun': 6,
        'iul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
        'ianuarie': 1, 'februarie': 2, 'martie': 3, 'aprilie': 4,
        'iunie': 6, 'iulie': 7, 'august': 8, 'septembrie': 9,
        'octombrie': 10, 'noiembrie': 11, 'decembrie': 12
    }

    match = re.match(r'(\d{1,2})\s+(\w+)\.?\s*(\d{4})', date_str)
    if match:
        day, month_str, year = match.groups()
        month_str = month_str.lower().rstrip('.')
        month = months.get(month_str, 1)
        return datetime(int(year), month, int(day))
    return None


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from a PDF file."""
    text = ''
    try:
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() or ''
    except Exception:
        pass
    return text


def extract_text_from_bytes(file_bytes: bytes, filename: str) -> str:
    """Extract text from uploaded file bytes."""
    ext = os.path.splitext(filename)[1].lower()

    if ext != '.pdf':
        return ''

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        return extract_text_from_pdf(tmp_path)
    finally:
        os.unlink(tmp_path)


def parse_meta_invoice(text: str) -> dict:
    """
    Parse Meta (Facebook/Instagram) invoice from extracted text.

    Returns dict with:
        - invoice_number
        - invoice_date
        - invoice_value
        - currency
        - customer_vat
        - customer_name
        - campaigns: dict of campaign_name -> cost
    """
    result = {
        'supplier': 'Meta Platforms Ireland Limited',
        'supplier_vat': 'IE9692928F',
        'campaigns': {}
    }

    # Extract invoice number
    invoice_match = re.search(r'(FBADS-\d+-\d+)|ID\s+tranzac[tț]ie\s+(\d+-\d+)', text, re.IGNORECASE)
    if invoice_match:
        result['invoice_number'] = invoice_match.group(1) or invoice_match.group(2)

    # Extract date
    date_match = re.search(r'(\d{1,2}\s+\w{3,}\.?\s*\d{4})', text)
    if date_match:
        result['invoice_date'] = date_match.group(1)
        result['date_parsed'] = parse_romanian_date(date_match.group(1))

    # Extract total value
    value_match = re.search(r'Efectuat[aă]?\s*([\d.,]+)\s*(?:RON|EUR|USD)?', text, re.IGNORECASE)
    if value_match:
        result['invoice_value'] = parse_value(value_match.group(1))

    # Extract currency
    if 'RON' in text:
        result['currency'] = 'RON'
    elif 'EUR' in text:
        result['currency'] = 'EUR'
    elif 'USD' in text:
        result['currency'] = 'USD'
    else:
        result['currency'] = 'RON'

    # Extract customer VAT
    vat_match = re.search(r'VAT[:\s]*([A-Z]{2}[\s]?\d+)', text)
    if vat_match:
        result['customer_vat'] = vat_match.group(1).replace(' ', '')

    # Extract customer name
    customer_match = re.search(r'(?:S\.C\.|SC)\s+([A-Z][A-Za-z\s&]+(?:S\.?R\.?L\.?|SRL))', text)
    if customer_match:
        result['customer_name'] = customer_match.group(1).strip()

    # Extract campaigns with costs
    campaign_list = [
        ('[CA] Leads - Modele Volvo 0 km', 'Leads - Modele Volvo 0 km'),
        ('[CA] Leads - Modele mix', 'Leads - Modele mix'),
        ('[CA] Traffic - Interese - Modele masini', 'Traffic - Interese - Modele masini'),
        ('GENERARE COMENZI Q4', 'GENERARE COMENZI Q4'),
        ('[CA] Leads - Mazda CX80', 'Leads - Mazda CX80'),
        ('[CA] Leads - Modele MG HS', 'Leads - Modele MG HS'),
        ('[CA] Traffic - Mazda CX60', 'Traffic - Mazda CX60'),
    ]

    for campaign_name, search_term in campaign_list:
        if search_term in text:
            pos = text.find(search_term)
            section = text[pos:pos+200]

            # Pattern: 23:59VALUE RON (time directly followed by value)
            match = re.search(r'\d{2}:\d{2}([\d.,]+)\s*RON', section)
            if match:
                try:
                    value = parse_value(match.group(1))
                    result['campaigns'][campaign_name] = value
                except:
                    pass

    return result


def parse_google_ads_invoice(text: str) -> dict:
    """Parse Google Ads invoice from extracted text."""
    result = {
        'supplier': 'Google Ireland Limited',
        'supplier_vat': 'IE6388047V',
        'campaigns': {}
    }

    # Extract invoice number
    invoice_match = re.search(r'Invoice\s+(?:number|no\.?)\s*:?\s*(\d+)', text, re.IGNORECASE)
    if invoice_match:
        result['invoice_number'] = invoice_match.group(1)

    # Extract date
    date_match = re.search(r'Invoice\s+date\s*:?\s*(\w+\s+\d{1,2},?\s*\d{4})', text, re.IGNORECASE)
    if date_match:
        result['invoice_date'] = date_match.group(1)

    # Extract total
    value_match = re.search(r'Total\s+amount\s+due\s*:?\s*([\d.,]+)', text, re.IGNORECASE)
    if value_match:
        result['invoice_value'] = parse_value(value_match.group(1))

    # Extract currency
    currency_match = re.search(r'(RON|EUR|USD)', text)
    if currency_match:
        result['currency'] = currency_match.group(1)
    else:
        result['currency'] = 'RON'

    return result


def detect_invoice_type(text: str) -> str:
    """Detect the type of invoice based on text content."""
    text_lower = text.lower()

    if 'meta platforms' in text_lower or 'facebook' in text_lower or 'fbads' in text_lower:
        return 'meta'
    elif 'google' in text_lower and ('ads' in text_lower or 'adwords' in text_lower):
        return 'google_ads'
    elif 'dreamstime' in text_lower:
        return 'dreamstime'
    else:
        return 'generic'


def parse_invoice_auto(text: str, filename: str = '') -> dict:
    """
    Automatically detect invoice type and parse accordingly.
    """
    invoice_type = detect_invoice_type(text)

    if invoice_type == 'meta':
        result = parse_meta_invoice(text)
    elif invoice_type == 'google_ads':
        result = parse_google_ads_invoice(text)
    else:
        # Generic parsing
        result = parse_generic_invoice(text)

    result['invoice_type'] = invoice_type
    result['filename'] = filename

    return result


def parse_generic_invoice(text: str) -> dict:
    """Parse a generic invoice with common patterns."""
    result = {
        'supplier': None,
        'supplier_vat': None,
        'campaigns': {}
    }

    # Try to extract invoice number
    patterns = [
        r'(?:Factura|Invoice)\s+(?:nr\.?|no\.?|#)\s*:?\s*([A-Z0-9\-/]+)',
        r'(?:Nr\.?|No\.?)\s+(?:factura|invoice)\s*:?\s*([A-Z0-9\-/]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            result['invoice_number'] = match.group(1)
            break

    # Try to extract date
    date_patterns = [
        r'(?:Data|Date)\s*:?\s*(\d{1,2}[./\-]\d{1,2}[./\-]\d{4})',
        r'(\d{1,2}\s+\w{3,}\.?\s*\d{4})',
    ]
    for pattern in date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            result['invoice_date'] = match.group(1)
            break

    # Try to extract total value
    value_patterns = [
        r'Total\s*:?\s*([\d.,]+)\s*(?:RON|EUR|USD)?',
        r'(?:Valoare|Amount)\s+total[aă]?\s*:?\s*([\d.,]+)',
    ]
    for pattern in value_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            try:
                result['invoice_value'] = parse_value(match.group(1))
                break
            except:
                pass

    # Extract currency
    currency_match = re.search(r'(RON|EUR|USD|Lei)', text, re.IGNORECASE)
    if currency_match:
        currency = currency_match.group(1).upper()
        result['currency'] = 'RON' if currency == 'LEI' else currency
    else:
        result['currency'] = 'RON'

    return result


def process_bulk_invoices(files: list[tuple[bytes, str]]) -> dict:
    """
    Process multiple invoice files and generate a summary report.

    Args:
        files: List of tuples (file_bytes, filename)

    Returns:
        dict with:
            - invoices: list of parsed invoices
            - summary: aggregated statistics
            - by_campaign: campaign-level breakdown
            - by_month: monthly breakdown
            - total: grand total
    """
    invoices = []

    for file_bytes, filename in files:
        text = extract_text_from_bytes(file_bytes, filename)
        if text:
            result = parse_invoice_auto(text, filename)
            invoices.append(result)

    # Sort by date
    invoices.sort(key=lambda x: x.get('date_parsed') or datetime.min)

    # Calculate totals
    total = sum(inv.get('invoice_value', 0) or 0 for inv in invoices)

    # Group by month
    by_month = defaultdict(lambda: {'count': 0, 'total': 0})
    for inv in invoices:
        if inv.get('date_parsed'):
            month_key = inv['date_parsed'].strftime('%Y-%m')
            by_month[month_key]['count'] += 1
            by_month[month_key]['total'] += inv.get('invoice_value', 0) or 0

    # Group by campaign (aggregate across all invoices)
    by_campaign = defaultdict(float)
    for inv in invoices:
        for campaign, value in inv.get('campaigns', {}).items():
            by_campaign[campaign] += value

    # Group by supplier
    by_supplier = defaultdict(lambda: {'count': 0, 'total': 0})
    for inv in invoices:
        supplier = inv.get('supplier') or 'Unknown'
        by_supplier[supplier]['count'] += 1
        by_supplier[supplier]['total'] += inv.get('invoice_value', 0) or 0

    return {
        'invoices': invoices,
        'total': total,
        'count': len(invoices),
        'by_month': dict(by_month),
        'by_campaign': dict(by_campaign),
        'by_supplier': dict(by_supplier),
        'currency': invoices[0].get('currency', 'RON') if invoices else 'RON'
    }


def generate_excel_report(report_data: dict) -> bytes:
    """
    Generate an Excel report from bulk invoice processing results.

    Returns Excel file as bytes.
    """
    if not EXCEL_ENABLED:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    total_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    money_format = '#,##0.00'
    date_format = 'DD.MM.YYYY'

    invoices = report_data.get('invoices', [])
    currency = report_data.get('currency', 'RON')

    # ============ Sheet 1: Invoice Summary ============
    ws1 = wb.active
    ws1.title = 'Sumar Facturi'

    # Title
    ws1.merge_cells('A1:F1')
    ws1['A1'] = 'CENTRALIZATOR FACTURI'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal='center')

    ws1.merge_cells('A2:F2')
    ws1['A2'] = f'Total Facturi: {report_data.get("count", 0)} | Total: {report_data.get("total", 0):,.2f} {currency}'
    ws1['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Nr.', 'Data', 'Nr. Factură', 'Furnizor', f'Valoare ({currency})', 'Fișier']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data
    for idx, inv in enumerate(invoices, 1):
        row = idx + 4
        ws1.cell(row=row, column=1, value=idx).border = border

        date_cell = ws1.cell(row=row, column=2, value=inv.get('date_parsed'))
        if inv.get('date_parsed'):
            date_cell.number_format = date_format
        date_cell.border = border

        ws1.cell(row=row, column=3, value=inv.get('invoice_number', '')).border = border
        ws1.cell(row=row, column=4, value=inv.get('supplier', '')).border = border

        value_cell = ws1.cell(row=row, column=5, value=inv.get('invoice_value', 0))
        value_cell.number_format = money_format
        value_cell.border = border

        ws1.cell(row=row, column=6, value=inv.get('filename', '')).border = border

    # Total row
    total_row = len(invoices) + 5
    ws1.cell(row=total_row, column=4, value='TOTAL').font = Font(bold=True)
    total_cell = ws1.cell(row=total_row, column=5, value=report_data.get('total', 0))
    total_cell.number_format = money_format
    total_cell.font = Font(bold=True)

    # Column widths
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 30
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 50

    # ============ Sheet 2: Monthly Summary ============
    ws2 = wb.create_sheet('Sumar Lunar')

    ws2.merge_cells('A1:C1')
    ws2['A1'] = 'SUMAR PE LUNI'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal='center')

    headers2 = ['Luna', 'Nr. Facturi', f'Total ({currency})']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    row = 4
    for month_key in sorted(report_data.get('by_month', {}).keys()):
        data = report_data['by_month'][month_key]
        month_name = datetime.strptime(month_key, '%Y-%m').strftime('%B %Y')

        ws2.cell(row=row, column=1, value=month_name).border = border
        ws2.cell(row=row, column=2, value=data['count']).border = border
        total_cell = ws2.cell(row=row, column=3, value=data['total'])
        total_cell.number_format = money_format
        total_cell.border = border
        row += 1

    # Grand total
    ws2.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    ws2.cell(row=row, column=2, value=report_data.get('count', 0)).font = Font(bold=True)
    grand_total = ws2.cell(row=row, column=3, value=report_data.get('total', 0))
    grand_total.number_format = money_format
    grand_total.font = Font(bold=True)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 18

    # ============ Sheet 3: Campaign Summary ============
    if report_data.get('by_campaign'):
        ws3 = wb.create_sheet('Per Campanie')

        ws3.merge_cells('A1:C1')
        ws3['A1'] = 'TOTAL PER CAMPANIE'
        ws3['A1'].font = Font(bold=True, size=14)
        ws3['A1'].alignment = Alignment(horizontal='center')

        headers3 = ['Campanie', f'Total ({currency})', '% din Total']
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row = 4
        grand_total_value = report_data.get('total', 0) or 1

        for campaign, value in sorted(report_data['by_campaign'].items(), key=lambda x: -x[1]):
            ws3.cell(row=row, column=1, value=campaign).border = border

            value_cell = ws3.cell(row=row, column=2, value=value)
            value_cell.number_format = money_format
            value_cell.border = border

            pct_cell = ws3.cell(row=row, column=3, value=value / grand_total_value)
            pct_cell.number_format = '0.0%'
            pct_cell.border = border

            row += 1

        ws3.column_dimensions['A'].width = 45
        ws3.column_dimensions['B'].width = 18
        ws3.column_dimensions['C'].width = 12

    # ============ Sheet 4: Supplier Summary ============
    if report_data.get('by_supplier'):
        ws4 = wb.create_sheet('Per Furnizor')

        ws4.merge_cells('A1:C1')
        ws4['A1'] = 'TOTAL PER FURNIZOR'
        ws4['A1'].font = Font(bold=True, size=14)
        ws4['A1'].alignment = Alignment(horizontal='center')

        headers4 = ['Furnizor', 'Nr. Facturi', f'Total ({currency})']
        for col, header in enumerate(headers4, 1):
            cell = ws4.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row = 4
        for supplier, data in sorted(report_data['by_supplier'].items(), key=lambda x: -x[1]['total']):
            ws4.cell(row=row, column=1, value=supplier).border = border
            ws4.cell(row=row, column=2, value=data['count']).border = border

            value_cell = ws4.cell(row=row, column=3, value=data['total'])
            value_cell.number_format = money_format
            value_cell.border = border

            row += 1

        ws4.column_dimensions['A'].width = 40
        ws4.column_dimensions['B'].width = 15
        ws4.column_dimensions['C'].width = 18

    # ============ Sheet 5: Campaign by Invoice ============
    # Matrix view: campaigns as rows, invoices as columns
    if report_data.get('by_campaign') and invoices:
        ws5 = wb.create_sheet('Campanii per Factură')

        ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(invoices)+2)
        ws5['A1'] = 'DETALII CAMPANII PER FACTURĂ'
        ws5['A1'].font = Font(bold=True, size=14)
        ws5['A1'].alignment = Alignment(horizontal='center')

        # Headers
        row = 3
        ws5.cell(row=row, column=1, value='Campanie').font = header_font
        ws5.cell(row=row, column=1).fill = header_fill
        ws5.cell(row=row, column=1).border = border

        for col, inv in enumerate(invoices, 2):
            date_str = inv.get('date_parsed').strftime('%d.%m') if inv.get('date_parsed') else ''
            cell = ws5.cell(row=row, column=col, value=date_str)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        total_col = len(invoices) + 2
        ws5.cell(row=row, column=total_col, value='TOTAL').font = header_font
        ws5.cell(row=row, column=total_col).fill = total_fill
        ws5.cell(row=row, column=total_col).border = border

        # Data rows
        row = 4
        all_campaigns = sorted(set(
            campaign for inv in invoices for campaign in inv.get('campaigns', {}).keys()
        ))

        for campaign in all_campaigns:
            ws5.cell(row=row, column=1, value=campaign).border = border

            row_total = 0
            for col, inv in enumerate(invoices, 2):
                value = inv.get('campaigns', {}).get(campaign, 0)
                cell = ws5.cell(row=row, column=col, value=value if value > 0 else '')
                if value > 0:
                    cell.number_format = money_format
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                row_total += value

            total_cell = ws5.cell(row=row, column=total_col, value=row_total)
            total_cell.number_format = money_format
            total_cell.font = Font(bold=True)
            total_cell.border = border
            total_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

            row += 1

        # Invoice totals row
        row += 1
        ws5.cell(row=row, column=1, value='TOTAL FACTURĂ').font = Font(bold=True)
        ws5.cell(row=row, column=1).fill = total_fill
        ws5.cell(row=row, column=1).border = border

        grand_total = 0
        for col, inv in enumerate(invoices, 2):
            value = inv.get('invoice_value', 0)
            cell = ws5.cell(row=row, column=col, value=value)
            cell.number_format = money_format
            cell.font = Font(bold=True)
            cell.border = border
            cell.fill = total_fill
            grand_total += value

        grand_cell = ws5.cell(row=row, column=total_col, value=grand_total)
        grand_cell.number_format = money_format
        grand_cell.font = Font(bold=True, color='FFFFFF')
        grand_cell.border = border
        grand_cell.fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')

        # Column widths
        ws5.column_dimensions['A'].width = 40
        for col in range(2, total_col + 1):
            ws5.column_dimensions[get_column_letter(col)].width = 12

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_summary_text(report_data: dict) -> str:
    """Generate a text summary of the bulk invoice processing results."""
    currency = report_data.get('currency', 'RON')
    lines = []

    lines.append('=' * 80)
    lines.append('INVOICE PROCESSING REPORT')
    lines.append('=' * 80)
    lines.append(f"Total Invoices: {report_data.get('count', 0)}")
    lines.append(f"Total Value: {report_data.get('total', 0):,.2f} {currency}")
    lines.append('')

    # Monthly breakdown
    if report_data.get('by_month'):
        lines.append('MONTHLY SUMMARY')
        lines.append('-' * 40)
        for month_key in sorted(report_data['by_month'].keys()):
            data = report_data['by_month'][month_key]
            month_name = datetime.strptime(month_key, '%Y-%m').strftime('%B %Y')
            lines.append(f"{month_name:<20} {data['total']:>15,.2f} {currency}")
        lines.append('')

    # Supplier breakdown
    if report_data.get('by_supplier'):
        lines.append('BY SUPPLIER')
        lines.append('-' * 40)
        for supplier, data in sorted(report_data['by_supplier'].items(), key=lambda x: -x[1]['total']):
            lines.append(f"{supplier:<30} {data['total']:>15,.2f} {currency}")
        lines.append('')

    # Campaign breakdown
    if report_data.get('by_campaign'):
        lines.append('BY CAMPAIGN')
        lines.append('-' * 40)
        for campaign, value in sorted(report_data['by_campaign'].items(), key=lambda x: -x[1]):
            lines.append(f"{campaign:<40} {value:>15,.2f} {currency}")
        lines.append('')

    lines.append('=' * 80)

    return '\n'.join(lines)
