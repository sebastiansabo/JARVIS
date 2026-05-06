"""Bilant Excel Handler — read Balanta uploads, generate output Excel, ANAF export."""

import io
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from .formula_engine import extract_ct_formula, extract_row_formula
from .anaf_parser import _nr_rd_to_anaf


def read_balanta_from_excel(file_bytes):
    """Read Balanta sheet from uploaded Excel.

    Returns:
        pd.DataFrame with columns at indices 0=Cont, 1=SFD, 2=SFC

    Raises:
        ValueError: if Balanta sheet not found
    """
    xlsx = pd.ExcelFile(io.BytesIO(file_bytes))
    if 'Balanta' not in xlsx.sheet_names:
        raise ValueError('Sheet "Balanta" not found in the uploaded file')
    return pd.read_excel(xlsx, sheet_name='Balanta')


def read_balanta_full(file_bytes):
    """Read balanta Excel with full columns (TSD/TSC + SFD/SFC).

    Supports the standard Romanian accounting software export format:
    SA, NRC, RAD1, CLS, E, NUC, SID, SIC, RPD, RPC, RCD, RCC, TSD, TSC, SFD, SFC, ...

    Header is auto-detected (row containing 'Nr. Cont' or starting with account data).
    Account codes are read from column B (index 1).

    Returns:
        (df_balanta, accounts_dict) where:
        - df_balanta: DataFrame with columns [Cont, SFD, SFC] for F10L formula engine
        - accounts_dict: {code: {'tsd': float, 'tsc': float}} for F20 engine

    Raises:
        ValueError: if file cannot be parsed
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Detect header row (contains 'Nr. Cont' or 'NRC')
    data_start = 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        row_strs = [str(v).lower() if v else '' for v in row[:6]]
        if any('nr. cont' in s or s == 'nrc' for s in row_strs):
            data_start = row_idx + 1
            break
        # Also detect if row has 'sal_sa' (header indicator)
        if any('sal_sa' in s for s in row_strs):
            data_start = row_idx + 1
            break

    accounts = {}
    f10l_rows = []

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        cont = str(row[1]).strip() if row[1] else ''
        if not cont or not cont[0].isdigit() or not cont.replace(' ', '').isdigit():
            continue

        tsd = float(row[12]) if len(row) > 12 and row[12] else 0
        tsc = float(row[13]) if len(row) > 13 and row[13] else 0
        sfd = float(row[14]) if len(row) > 14 and row[14] else 0
        sfc = float(row[15]) if len(row) > 15 and row[15] else 0

        accounts[cont] = {'tsd': tsd, 'tsc': tsc}
        f10l_rows.append([cont, sfd, sfc])

    if not accounts:
        raise ValueError('No account data found in the uploaded file')

    df_balanta = pd.DataFrame(f10l_rows, columns=['Cont', 'SFD', 'SFC'])
    return df_balanta, accounts


def read_bilant_sheet_for_import(file_bytes):
    """Read Bilant sheet and auto-extract formulas for template import.

    Returns:
        list of dicts: [{description, nr_rd, formula_ct, formula_rd, row_type, sort_order}, ...]
    """
    xlsx = pd.ExcelFile(io.BytesIO(file_bytes))
    if 'Bilant' not in xlsx.sheet_names:
        raise ValueError('Sheet "Bilant" not found in the uploaded file')

    df = pd.read_excel(xlsx, sheet_name='Bilant')
    rows = []
    for idx, row in df.iterrows():
        desc = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
        nr_rd = str(row.iloc[1]).replace('.0', '') if pd.notna(row.iloc[1]) else ''
        formula_ct = extract_ct_formula(desc)
        formula_rd = extract_row_formula(desc)

        # Determine row type
        row_type = 'data'
        desc_lower = desc.lower()
        if 'total' in desc_lower or formula_rd:
            row_type = 'total'
        elif not nr_rd and not formula_ct and not formula_rd:
            if desc.strip():
                row_type = 'section_header'
            else:
                row_type = 'separator'

        rows.append({
            'description': desc,
            'nr_rd': nr_rd,
            'formula_ct': formula_ct,
            'formula_rd': formula_rd,
            'row_type': row_type,
            'is_bold': row_type in ('total', 'section_header'),
            'indent_level': 0,
            'sort_order': idx,
        })
    return rows


def generate_output_excel(generation, results, metrics):
    """Generate 3-sheet Excel output from persisted generation data.

    Args:
        generation: dict with generation metadata (company_name, period_label, etc.)
        results: list of result dicts (description, nr_rd, value, verification)
        metrics: dict with summary, ratios, structure

    Returns:
        io.BytesIO containing the Excel file
    """
    output = io.BytesIO()

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    section_font = Font(bold=True, size=11)
    total_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    money_format = '#,##0.00'

    from openpyxl import Workbook
    wb = Workbook()

    # ── Sheet 1: Bilant Results ──
    ws1 = wb.active
    ws1.title = 'Bilant'
    headers = ['Nr. rd.', 'Denumirea elementului', 'Sold Final', 'Verificare']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, r in enumerate(results, 2):
        ws1.cell(row=i, column=1, value=r.get('nr_rd', '')).border = border
        desc_cell = ws1.cell(row=i, column=2, value=r.get('description', ''))
        desc_cell.border = border
        indent = r.get('indent_level', 0)
        if indent:
            desc_cell.alignment = Alignment(indent=indent)
        val_cell = ws1.cell(row=i, column=3, value=r.get('value', 0))
        val_cell.number_format = money_format
        val_cell.border = border
        verif_cell = ws1.cell(row=i, column=4, value=r.get('verification', ''))
        verif_cell.border = border
        verif_cell.alignment = Alignment(wrap_text=True)

        if r.get('is_bold') or r.get('row_type') in ('total', 'section_header'):
            for col in range(1, 5):
                ws1.cell(row=i, column=col).font = total_font if r.get('row_type') == 'total' else section_font

    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 55
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 40

    # ── Sheet 2: Dashboard ──
    ws2 = wb.create_sheet('Dashboard')
    row_num = 1

    # Title
    title_cell = ws2.cell(row=row_num, column=1, value=f"Bilant — {generation.get('company_name', '')} — {generation.get('period_label', '')}")
    title_cell.font = Font(bold=True, size=14)
    row_num += 2

    # Summary
    ws2.cell(row=row_num, column=1, value='SUMAR FINANCIAR').font = section_font
    row_num += 1
    summary = metrics.get('summary', {})
    for label, key in [('Total Active', 'total_active'), ('Active Imobilizate', 'active_imobilizate'),
                       ('Active Circulante', 'active_circulante'), ('Capitaluri Proprii', 'capitaluri_proprii'),
                       ('Total Datorii', 'total_datorii')]:
        ws2.cell(row=row_num, column=1, value=label).border = border
        val_cell = ws2.cell(row=row_num, column=2, value=summary.get(key, 0))
        val_cell.number_format = money_format
        val_cell.border = border
        row_num += 1

    row_num += 1
    ws2.cell(row=row_num, column=1, value='INDICATORI FINANCIARI').font = section_font
    row_num += 1
    for h_col, h_val in enumerate(['Indicator', 'Valoare', 'Interpretare'], 1):
        c = ws2.cell(row=row_num, column=h_col, value=h_val)
        c.font = header_font
        c.fill = header_fill
        c.border = border
    row_num += 1

    from .formula_engine import STANDARD_RATIOS
    ratios = metrics.get('ratios', {})
    for key, spec in STANDARD_RATIOS.items():
        raw = ratios.get(key)
        # Handle both plain numbers and {value, label, interpretation} dicts
        val = raw['value'] if isinstance(raw, dict) and 'value' in raw else raw
        ws2.cell(row=row_num, column=1, value=spec['label']).border = border
        ws2.cell(row=row_num, column=2, value=val if val is not None else 'N/A').border = border
        ws2.cell(row=row_num, column=3, value=spec['interpretation']).border = border
        row_num += 1

    row_num += 1
    # Asset structure
    ws2.cell(row=row_num, column=1, value='STRUCTURA ACTIVELOR').font = section_font
    row_num += 1
    for item in metrics.get('structure', {}).get('assets', []):
        ws2.cell(row=row_num, column=1, value=item['name']).border = border
        val_cell = ws2.cell(row=row_num, column=2, value=item['value'])
        val_cell.number_format = money_format
        val_cell.border = border
        ws2.cell(row=row_num, column=3, value=f"{item['percent']}%").border = border
        row_num += 1

    row_num += 1
    ws2.cell(row=row_num, column=1, value='STRUCTURA PASIVELOR').font = section_font
    row_num += 1
    for item in metrics.get('structure', {}).get('liabilities', []):
        ws2.cell(row=row_num, column=1, value=item['name']).border = border
        val_cell = ws2.cell(row=row_num, column=2, value=item['value'])
        val_cell.number_format = money_format
        val_cell.border = border
        ws2.cell(row=row_num, column=3, value=f"{item['percent']}%").border = border
        row_num += 1

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 20

    wb.save(output)
    output.seek(0)
    return output


def generate_anaf_excel(generation, results, row_mapping=None, prior_results=None):
    """Generate ANAF-format Excel with F10 field codes and human-readable sheets.

    Args:
        generation: dict with generation metadata
        results: list of result dicts (nr_rd, description, value, etc.)
        row_mapping: optional dict from generate_row_mapping() {F10_XXXC: {row, col}}
        prior_results: optional dict {nr_rd: value} for C1 (prior year) column

    Returns:
        io.BytesIO containing the Excel file
    """
    from openpyxl import Workbook

    output = io.BytesIO()

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    total_font = Font(bold=True, size=10)
    section_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    money_format = '#,##0.00'

    wb = Workbook()

    # ── Sheet 1: F10 (ANAF field codes) ──
    ws1 = wb.active
    ws1.title = 'F10'
    for col, h in enumerate(['Field Code', 'Value'], 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    # Build nr_rd → value map from results
    val_map = {}
    for r in results:
        nr = r.get('nr_rd')
        if nr:
            val_map[nr] = r.get('value', 0) or 0

    # Generate F10 entries (sorted by field code)
    if row_mapping:
        mapping = row_mapping
    else:
        # Auto-generate mapping from results
        mapping = {}
        for r in results:
            nr = r.get('nr_rd')
            if not nr:
                continue
            anaf_code = _nr_rd_to_anaf(nr)
            padded = anaf_code.zfill(3)
            for c in ('1', '2'):
                mapping[f'F10_{padded}{c}'] = {'row': nr, 'col': f'C{c}'}

    for field_code in sorted(mapping.keys()):
        info = mapping[field_code]
        nr = info['row']
        col_name = info['col']
        if col_name == 'C1':
            value = prior_results.get(nr, 0) if prior_results else 0
        else:
            value = val_map.get(nr, 0)
        ws1.cell(row=row_num, column=1, value=field_code).border = border
        val_cell = ws1.cell(row=row_num, column=2, value=value or 0)
        val_cell.number_format = money_format
        val_cell.border = border
        row_num += 1

    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 20

    # ── Sheet 2: Bilant (human-readable with C1 + C2) ──
    ws2 = wb.create_sheet('Bilant')
    for col, h in enumerate(['Nr. rd.', 'Denumirea elementului', 'Sold C1', 'Sold C2'], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, r in enumerate(results, 2):
        nr_rd = r.get('nr_rd', '') or ''
        desc = r.get('description', '')
        value = r.get('value', 0) or 0
        c1_val = prior_results.get(nr_rd, 0) if prior_results and nr_rd else 0

        ws2.cell(row=i, column=1, value=nr_rd).border = border
        desc_cell = ws2.cell(row=i, column=2, value=desc)
        desc_cell.border = border
        indent = r.get('indent_level', 0)
        if indent:
            desc_cell.alignment = Alignment(indent=indent)

        c1_cell = ws2.cell(row=i, column=3, value=c1_val)
        c1_cell.number_format = money_format
        c1_cell.border = border

        c2_cell = ws2.cell(row=i, column=4, value=value)
        c2_cell.number_format = money_format
        c2_cell.border = border

        if r.get('is_bold') or r.get('row_type') in ('total', 'section_header', 'section'):
            font = total_font if r.get('row_type') == 'total' else section_font
            for col in range(1, 5):
                ws2.cell(row=i, column=col).font = font

    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18

    wb.save(output)
    output.seek(0)
    return output
