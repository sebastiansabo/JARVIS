"""Tests for Identificare and Mijloace_fixe sheet parsing."""
import io
import pytest
import openpyxl


def _make_xlsx_with_identificare(fields: dict) -> bytes:
    """Helper: create an in-memory xlsx with an Identificare sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Balanta'
    ws.append(['Cont', 'SFD', 'SFC'])
    ws.append(['101', '100', '0'])

    ws2 = wb.create_sheet('Identificare')
    ws2.append(['Câmp', 'Valoare'])
    for key, val in fields.items():
        ws2.append([key, val])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_identificare_basic():
    from accounting.bilant.excel_handler import read_identificare_sheet
    data = _make_xlsx_with_identificare({
        'CUI': '50186890',
        'Denumire': 'AUTOWORLD INTERNATIONAL SRL',
        'CAEN': '4781',
        'Tip entitate': 'Microentitate',
    })
    result = read_identificare_sheet(data)
    assert result['cui'] == '50186890'
    assert result['den'] == 'AUTOWORLD INTERNATIONAL SRL'
    assert result['caen'] == '4781'
    assert result['entity_type'] == 'UU'


def test_parse_identificare_missing_sheet():
    from accounting.bilant.excel_handler import read_identificare_sheet
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    result = read_identificare_sheet(buf.getvalue())
    assert result is None


def test_parse_mijloace_fixe_present():
    from accounting.bilant.excel_handler import read_mijloace_fixe_sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mijloace_fixe'
    ws.append(['Cont', 'Sold Inițial', 'Creșteri', 'Reduceri', 'Reduceri dezmembrări'])
    ws.append(['212', 1000, 500, 100, 50])
    buf = io.BytesIO()
    wb.save(buf)
    result = read_mijloace_fixe_sheet(buf.getvalue())
    assert result is not None
    assert '212' in result


def test_parse_mijloace_fixe_absent():
    from accounting.bilant.excel_handler import read_mijloace_fixe_sheet
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    result = read_mijloace_fixe_sheet(buf.getvalue())
    assert result is None
