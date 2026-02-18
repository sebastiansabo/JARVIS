"""
Bulk Invoice Processor Module

This module provides functionality for processing multiple invoices at once:
- Parse multiple PDF/image invoices
- Generate summary reports by campaign/supplier
- Export to Excel with detailed breakdowns
- Support for Meta, Google Ads, and other invoice templates
"""

import os
import re
import tempfile
from datetime import datetime
from collections import defaultdict
from typing import Optional
from io import BytesIO

import PyPDF2
import json

# AI parsing support
try:
    import anthropic
    AI_ENABLED = True
except ImportError:
    AI_ENABLED = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_ENABLED = True
except ImportError:
    EXCEL_ENABLED = False


def parse_value(value_str: str) -> float:
    """
    Parse invoice value from various European/US number formats.
    Handles: '3.499,00', '3,499.00', '874,90', '1234.56'
    """
    value_str = value_str.strip()

    # Handle European format (dot as thousand separator, comma as decimal)
    if ',' in value_str and '.' in value_str:
        last_comma = value_str.rfind(',')
        last_dot = value_str.rfind('.')

        if last_comma > last_dot:
            # European: 1.234,56
            value_str = value_str.replace('.', '').replace(',', '.')
        else:
            # US: 1,234.56
            value_str = value_str.replace(',', '')
    elif ',' in value_str:
        # Only comma - check if it's decimal separator
        if re.search(r',\d{2}$', value_str):
            value_str = value_str.replace(',', '.')
        else:
            value_str = value_str.replace(',', '')
    else:
        # Format like 3.499.00 - last dot is decimal
        parts = value_str.split('.')
        if len(parts) > 2:
            value_str = ''.join(parts[:-1]) + '.' + parts[-1]

    return float(value_str)


def parse_romanian_date(date_str: str) -> Optional[datetime]:
    """Parse Romanian date format to datetime object."""
    months = {
        'ian': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'mai': 5, 'iun': 6,
        'iul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
        'ianuarie': 1, 'februarie': 2, 'martie': 3, 'aprilie': 4,
        'iunie': 6, 'iulie': 7, 'august': 8, 'septembrie': 9,
        'octombrie': 10, 'noiembrie': 11, 'decembrie': 12
    }

    match = re.match(r'(\d{1,2})\s+(\w+)\.?\s*(\d{4})', date_str)
    if match:
        day, month_str, year = match.groups()
        month_str = month_str.lower().rstrip('.')
        month = months.get(month_str, 1)
        return datetime(int(year), month, int(day))
    return None


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from a PDF file."""
    text = ''
    try:
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() or ''
    except Exception:
        pass
    return text


def extract_text_from_bytes(file_bytes: bytes, filename: str) -> str:
    """Extract text from uploaded file bytes."""
    ext = os.path.splitext(filename)[1].lower()

    if ext != '.pdf':
        return ''

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        return extract_text_from_pdf(tmp_path)
    finally:
        os.unlink(tmp_path)


def parse_meta_invoice(text: str) -> dict:
    """
    Parse Meta (Facebook/Instagram) invoice from extracted text.

    Returns dict with:
        - invoice_number
        - invoice_date
        - invoice_value
        - currency
        - customer_vat
        - customer_name
        - items: dict of item_name -> cost (campaigns, line items, products, etc.)
    """
    result = {
        'supplier': 'Meta Platforms Ireland Limited',
        'supplier_vat': 'IE9692928F',
        'items': {}
    }

    # Extract invoice number
    invoice_match = re.search(r'(FBADS-\d+-\d+)|ID\s+tranzac[tț]ie\s+(\d+-\d+)', text, re.IGNORECASE)
    if invoice_match:
        result['invoice_number'] = invoice_match.group(1) or invoice_match.group(2)

    # Extract date
    date_match = re.search(r'(\d{1,2}\s+\w{3,}\.?\s*\d{4})', text)
    if date_match:
        result['invoice_date'] = date_match.group(1)
        result['date_parsed'] = parse_romanian_date(date_match.group(1))

    # Extract total value
    value_match = re.search(r'Efectuat[aă]?\s*([\d.,]+)\s*(?:RON|EUR|USD)?', text, re.IGNORECASE)
    if value_match:
        result['invoice_value'] = parse_value(value_match.group(1))

    # Extract currency
    if 'RON' in text:
        result['currency'] = 'RON'
    elif 'EUR' in text:
        result['currency'] = 'EUR'
    elif 'USD' in text:
        result['currency'] = 'USD'
    else:
        result['currency'] = 'RON'

    # Extract customer VAT
    vat_match = re.search(r'VAT[:\s]*([A-Z]{2}[\s]?\d+)', text)
    if vat_match:
        result['customer_vat'] = vat_match.group(1).replace(' ', '')

    # Extract customer name
    customer_match = re.search(r'(?:S\.C\.|SC)\s+([A-Z][A-Za-z\s&]+(?:S\.?R\.?L\.?|SRL))', text)
    if customer_match:
        result['customer_name'] = customer_match.group(1).strip()

    # Dynamically extract ALL campaigns with their costs
    # Meta invoices structure:
    # Campaign Name (can start with [CA], Postare:, Stoc_, GENERARE, or other prefixes)
    # DATE_RANGE + VALUE RON (concatenated, e.g., "15 nov. 2025, 00:00 - 18 nov. 2025, 23:59306,23 RON")
    # Date range can end with any time (23:59, 16:33, etc.)

    # Generic pattern: Any campaign header followed by date range ending in HH:MM + VALUE RON
    # Campaigns typically start after "Campanii" section and before the footer
    # Campaign names are on their own line, followed by date range + value on the next line

    # Pattern to match campaign header + date range + value
    # Campaign header: any line that doesn't look like a sub-item (no "de Afişări" or similar)
    # The key insight is that campaign totals have format: "HH:MMVALUE RON" (time directly followed by value)

    # First, preprocess text to split concatenated campaign headers
    # Pattern: "...RON[CA] Campaign Name" becomes "...RON\n[CA] Campaign Name"
    # Also handle other common prefixes that might be concatenated
    text = re.sub(r'RON(\[CA\])', r'RON\n\1', text)
    text = re.sub(r'RON(Postare:)', r'RON\n\1', text)
    text = re.sub(r'RON(Stoc_)', r'RON\n\1', text)
    text = re.sub(r'RON(GENERARE)', r'RON\n\1', text)

    # Split text into lines and find item patterns
    lines = text.split('\n')
    i = 0
    while i < len(lines) - 1:
        line = lines[i].strip()
        next_line = lines[i + 1].strip() if i + 1 < len(lines) else ''

        # Check if next line looks like a date range + value pattern
        # Supports multiple formats:
        # - Romanian: "15 nov. 2025, 00:00 - 18 nov. 2025, 23:59306,23 RON"
        # - English: "From 6 Dec 2025, 00:00 to 9 Dec 2025, 23:59RON50.53"
        # Pattern 1: HH:MMVALUE RON (value after time, RON after value)
        # Pattern 2: HH:MMRONVALUE (RON between time and value)
        date_value_match = re.search(
            r'\d{1,2}\s+\w+\.?,?\s*202\d.*\d{2}:\d{2}\s*(?:RON\s*)?([\d.,]+)\s*(?:RON)?',
            next_line
        )

        if date_value_match:
            # This line might be an item name
            item_name = line

            # Skip if it looks like a sub-item or metadata
            skip_keywords = ['de Afişări', 'Afişări', 'Meta Platforms', 'Merrion Road',
                           'Dublin', 'Ireland', 'VAT Reg', 'S.C.', 'CALEA', 'România',
                           'Factura nr', 'Customer to account', 'Powered by',
                           'ID cont', 'Data facturii', 'Modalitate', 'Număr de referinţă',
                           'ID tranzacţie', 'Tip de produs', 'Ţi s-a emis', 'Campanii']

            should_skip = any(kw in item_name for kw in skip_keywords)

            # Also skip if item name is too short or empty
            if not should_skip and len(item_name) > 3:
                try:
                    value = parse_value(date_value_match.group(1))
                    if value > 0:
                        result['items'][item_name] = value
                except (ValueError, TypeError, AttributeError):
                    pass

            i += 2  # Skip both lines
        else:
            i += 1

    return result


def parse_google_ads_invoice(text: str) -> dict:
    """Parse Google Ads invoice from extracted text."""
    result = {
        'supplier': 'Google Ireland Limited',
        'supplier_vat': 'IE6388047V',
        'items': {}
    }

    # Extract invoice number
    invoice_match = re.search(r'Invoice\s+(?:number|no\.?)\s*:?\s*(\d+)', text, re.IGNORECASE)
    if invoice_match:
        result['invoice_number'] = invoice_match.group(1)

    # Extract date
    date_match = re.search(r'Invoice\s+date\s*:?\s*(\w+\s+\d{1,2},?\s*\d{4})', text, re.IGNORECASE)
    if date_match:
        result['invoice_date'] = date_match.group(1)

    # Extract total
    value_match = re.search(r'Total\s+amount\s+due\s*:?\s*([\d.,]+)', text, re.IGNORECASE)
    if value_match:
        result['invoice_value'] = parse_value(value_match.group(1))

    # Extract currency
    currency_match = re.search(r'(RON|EUR|USD)', text)
    if currency_match:
        result['currency'] = currency_match.group(1)
    else:
        result['currency'] = 'RON'

    return result


def parse_tiktok_invoice(text: str) -> dict:
    """
    Parse TikTok advertising invoice from extracted text.

    TikTok invoices have:
    - Invoice number: BDUK20253368656 format
    - Supplier: TikTok Information Technologies UK Limited
    - VAT: GB485763736
    - Date format: "15, September, 2025"
    - Consumption Details table with campaign info

    Returns dict with invoice details and campaign/items breakdown.
    """
    result = {
        'supplier': 'TikTok Information Technologies UK Limited',
        'supplier_vat': 'GB485763736',
        'currency': 'RON',
        'items': {}
    }

    # Extract invoice number - format: BDUK20253368656
    invoice_match = re.search(r'Invoice\s*#\s*([A-Z]{2,}UK?\d+)', text, re.IGNORECASE)
    if invoice_match:
        result['invoice_number'] = invoice_match.group(1)
    else:
        # Try alternate pattern in Consumption Details section
        alt_match = re.search(r'Invoice\s+Number:\s*([A-Z]{2,}UK?\d+)', text, re.IGNORECASE)
        if alt_match:
            result['invoice_number'] = alt_match.group(1)

    # Extract date - format: "15, September, 2025" or "Invoice Date 15, September, 2025"
    date_match = re.search(r'Invoice\s+Date\s+(\d{1,2}),?\s*(\w+),?\s*(\d{4})', text, re.IGNORECASE)
    if date_match:
        day, month, year = date_match.groups()
        result['invoice_date'] = f"{day} {month} {year}"
        result['date_parsed'] = parse_english_date(f"{month} {day}, {year}")

    # Extract customer VAT - format: RO50186814
    vat_match = re.search(r'VAT\s+number\s+(RO\d+)', text, re.IGNORECASE)
    if vat_match:
        result['customer_vat'] = vat_match.group(1)
    else:
        # Try alternate pattern
        alt_vat = re.search(r'(RO\d{6,})', text)
        if alt_vat:
            result['customer_vat'] = alt_vat.group(1)

    # Extract total value - format: "Total 22.00" or "Total in RON 22.00"
    value_patterns = [
        r'Total\s+in\s+RON\s+([\d.,]+)',
        r'Total\s+([\d.,]+)\s*$',
        r'Total\s+([\d.,]+)',
    ]
    for pattern in value_patterns:
        value_match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if value_match:
            result['invoice_value'] = parse_value(value_match.group(1))
            break

    # Extract currency from text
    currency_match = re.search(r'Amount\s+in\s+(RON|EUR|USD)', text, re.IGNORECASE)
    if currency_match:
        result['currency'] = currency_match.group(1).upper()

    # Extract campaign/consumption details from page 2
    # Format: Campaign ID | Campaign Name | ... | Total Consumption in RON
    # Example: 1843317419438178 | Traffic 20250915112154 | ... | 22.00
    campaign_pattern = re.search(
        r'(\d{15,})\s+([A-Za-z_\s]+\d*)\s+(?:RO|EU|US)?\s*(?:\d{4}-\d{2}-\d{2}\s*~\s*\d{4}-\d{2}-\d{2})?\s*([\d.,]+)',
        text
    )
    if campaign_pattern:
        campaign_id = campaign_pattern.group(1)
        campaign_name = campaign_pattern.group(2).strip()
        campaign_value = parse_value(campaign_pattern.group(3))
        if campaign_name and campaign_value:
            result['items'][f"{campaign_name} ({campaign_id})"] = campaign_value

    # Also support 'campaigns' alias for frontend compatibility
    result['campaigns'] = result['items']

    return result


def parse_anthropic_invoice(text: str) -> dict:
    """
    Parse Anthropic (Claude API) invoice from extracted text.

    Anthropic invoices are typically Stripe invoices with:
    - Invoice number: KCSFWF6E-0001 format
    - Supplier: Anthropic, PBC
    - Items: Credit purchases or API usage
    - Currency: USD

    Returns dict with invoice details and items breakdown.
    """
    result = {
        'supplier': 'Anthropic, PBC',
        'supplier_vat': '',  # US company, no VAT
        'currency': 'USD',
        'items': {}
    }

    # Extract invoice number - format: KCSFWF6E-0001 or similar alphanumeric
    invoice_match = re.search(r'Invoice\s+number\s+([A-Z0-9]+-\d+)', text, re.IGNORECASE)
    if invoice_match:
        result['invoice_number'] = invoice_match.group(1)
    else:
        # Try alternate pattern
        alt_match = re.search(r'#([A-Z0-9]+-\d+)', text)
        if alt_match:
            result['invoice_number'] = alt_match.group(1)

    # Extract date - format: "December 4, 2025" or "Date of issue December 4, 2025"
    date_patterns = [
        r'Date\s+of\s+issue\s+(\w+\s+\d{1,2},?\s*\d{4})',
        r'Payment\s+date\s+(\d{1,2}\s+\w+\s+\d{4})',
        r'(\w+\s+\d{1,2},?\s*\d{4})'
    ]
    for pattern in date_patterns:
        date_match = re.search(pattern, text, re.IGNORECASE)
        if date_match:
            result['invoice_date'] = date_match.group(1)
            result['date_parsed'] = parse_english_date(date_match.group(1))
            break

    # Extract total amount - format: "$50.00" or "US$50.00" or "Amount due $50.00"
    value_patterns = [
        r'Amount\s+due\s+\$?([\d,]+\.?\d*)\s*USD?',
        r'Total\s+\$?([\d,]+\.?\d*)',
        r'US?\$\s*([\d,]+\.\d{2})',
        r'\$([\d,]+\.\d{2})\s*USD'
    ]
    for pattern in value_patterns:
        value_match = re.search(pattern, text, re.IGNORECASE)
        if value_match:
            result['invoice_value'] = parse_value(value_match.group(1))
            break

    # Extract customer VAT (for EU customers)
    vat_match = re.search(r'(?:RO\s*VAT|VAT)\s+(RO\d+)', text, re.IGNORECASE)
    if vat_match:
        result['customer_vat'] = vat_match.group(1)

    # Extract line items - format: "Description Qty Unit price Tax Amount"
    # Look for items like "One-time credit purchase 1 $50.00 0% $50.00"
    item_patterns = [
        r'(One-time credit purchase|API usage|Credit purchase|Usage credits?)\s+\d+\s+\$?([\d,]+\.?\d*)',
        r'(Claude[\s\w]*(?:API|usage)?)\s+\$?([\d,]+\.?\d*)',
    ]
    for pattern in item_patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            item_name = match.group(1).strip()
            try:
                value = parse_value(match.group(2))
                result['items'][item_name] = value
            except ValueError:
                pass

    # If no items found, create a default item from total
    if not result['items'] and result.get('invoice_value'):
        result['items']['API Credits'] = result['invoice_value']

    # Add campaigns alias for compatibility
    result['campaigns'] = result['items']

    return result


def parse_english_date(date_str: str) -> Optional[datetime]:
    """Parse English date formats to datetime object."""
    if not date_str:
        return None

    # Month name mapping
    months = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4,
        'may': 5, 'june': 6, 'july': 7, 'august': 8,
        'september': 9, 'october': 10, 'november': 11, 'december': 12,
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
        'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9,
        'oct': 10, 'nov': 11, 'dec': 12
    }

    # Try "December 4, 2025" format
    match = re.match(r'(\w+)\s+(\d{1,2}),?\s*(\d{4})', date_str)
    if match:
        month_str, day, year = match.groups()
        month = months.get(month_str.lower())
        if month:
            return datetime(int(year), month, int(day))

    # Try "4 December 2025" format
    match = re.match(r'(\d{1,2})\s+(\w+)\s+(\d{4})', date_str)
    if match:
        day, month_str, year = match.groups()
        month = months.get(month_str.lower())
        if month:
            return datetime(int(year), month, int(day))

    return None


def detect_invoice_type(text: str) -> str:
    """Detect the type of invoice based on text content."""
    text_lower = text.lower()

    if 'meta platforms' in text_lower or 'facebook' in text_lower or 'fbads' in text_lower:
        return 'meta'
    elif 'google' in text_lower and ('ads' in text_lower or 'adwords' in text_lower):
        return 'google_ads'
    elif 'anthropic' in text_lower or 'claude' in text_lower:
        return 'anthropic'
    elif 'dreamstime' in text_lower:
        return 'dreamstime'
    elif 'tiktok' in text_lower:
        return 'tiktok'
    elif 'ro efactura' in text_lower or 'efactura' in text_lower or 'nr. factura' in text_lower:
        return 'efactura'
    else:
        return 'generic'


def parse_efactura_invoice(text: str) -> dict:
    """Parse Romanian eFactura format invoices."""
    result = {
        'supplier': None,
        'supplier_vat': None,
        'items': {},
        'invoice_value': None,
        'net_value': None,
        'currency': 'RON'
    }

    # Extract invoice number - format: "XXXX 1234 Nr. factura" or at start
    inv_match = re.search(r'([A-Z]{2,}[\s]?\d+)\s+Nr\.?\s*factura', text)
    if inv_match:
        result['invoice_number'] = inv_match.group(1).strip()

    # Extract date - format: "Data emitere 2025-12-04" or "Data emitere" followed by date
    date_match = re.search(r'Data\s+emitere\s+(\d{4}-\d{2}-\d{2})', text)
    if date_match:
        result['invoice_date'] = date_match.group(1)

    # Extract supplier name - after VANZATOR section
    supplier_match = re.search(r'VANZATOR\s*\n?\s*([A-Z][A-Z\s\.]+(?:S\.?R\.?L\.?|S\.?A\.?))', text)
    if supplier_match:
        result['supplier'] = supplier_match.group(1).strip()
    else:
        # Try simpler pattern
        supplier_match = re.search(r'([A-Z][A-Z\s\.]+(?:S\.?R\.?L\.?|S\.?A\.?))\s+Nume\s*\n?\s*Nr\. inregistrare', text)
        if supplier_match:
            result['supplier'] = supplier_match.group(1).strip()

    # Extract supplier VAT - RO followed by digits
    vat_match = re.search(r'Identificatorul TVA\s*\n?\s*(RO\d+)', text)
    if vat_match:
        result['supplier_vat'] = vat_match.group(1)
    else:
        vat_match = re.search(r'(RO\d{6,})', text)
        if vat_match:
            result['supplier_vat'] = vat_match.group(1)

    # Extract customer info
    customer_match = re.search(r'CUMPARATOR\s*\n?\s*([A-Z][A-Z\s\.]+(?:S\.?R\.?L\.?|S\.?A\.?))', text)
    if customer_match:
        result['customer_name'] = customer_match.group(1).strip()

    customer_vat = re.search(r'Identificator\s*\n?\s*(RO\d+)', text)
    if customer_vat:
        result['customer_vat'] = customer_vat.group(1)

    # Extract total values - look for TOTAL PLATA or VALOARE TOTALA cu TVA
    total_match = re.search(r'TOTAL PLATA\s*\n?\s*([\d.,]+)', text)
    if total_match:
        result['invoice_value'] = parse_value(total_match.group(1))
    else:
        total_match = re.search(r'VALOARE TOTALA cu\s*\n?\s*TVA\s*\n?\s*([\d.,]+)', text)
        if total_match:
            result['invoice_value'] = parse_value(total_match.group(1))

    # Extract net value - look for pattern with net and gross values
    # Format: "535.00 535.00 647.35" followed by "647.35 TOTAL PLATA..."
    # The first value is the net value
    net_match = re.search(r'([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*\n\s*([\d.,]+)\s*TOTAL PLATA', text)
    if net_match:
        # First value is net, last value is gross
        result['net_value'] = parse_value(net_match.group(1))
    else:
        # Try to extract from "Baza de calcul" (VAT base = net value)
        net_match = re.search(r'Baza de calcul\s+Valoare TVA.*?\n\s*([\d.,]+)\s+([\d.,]+)', text, re.DOTALL)
        if net_match:
            result['net_value'] = parse_value(net_match.group(1))

    # Extract line items from eFactura format
    # Format: "21.00 Item description LineNo RON Qty UnitQty NetValue H87 UnitPrice"
    # Example: "21.00 Display plastic prezentare A6 portrait - Mazda 1 RON 1 25.0000 205.00 H87 8.2000"
    # Items in eFactura are NET values - we convert to GROSS by applying VAT rate

    lines = text.split('\n')
    for line in lines:
        # Pattern: VAT% + item description + line# + RON + qty + unit_qty + net_value + H87 + unit_price
        # The item description can contain letters, numbers, spaces, and special chars
        match = re.match(
            r'^\s*(\d+\.\d+)\s+'  # VAT rate (e.g., 21.00)
            r'(.+?)\s+'  # Item description (greedy but minimal)
            r'(\d+)\s+'  # Line number
            r'RON\s+'  # Currency
            r'(\d+)\s+'  # Quantity
            r'([\d.]+)\s+'  # Unit quantity
            r'([\d.,]+)\s+'  # Net value
            r'H87',  # Unit code
            line
        )
        if match:
            vat_rate = float(match.group(1))  # e.g., 21.00
            item_name = match.group(2).strip()
            net_value = parse_value(match.group(6))
            if item_name and net_value > 0:
                # Convert NET to GROSS by applying VAT rate
                gross_value = net_value * (1 + vat_rate / 100)
                result['items'][item_name] = round(gross_value, 2)

    return result


def parse_invoice_auto(text: str, filename: str = '') -> dict:
    """
    Automatically detect invoice type and parse accordingly.
    Uses AI as fallback when regex parsing doesn't extract key data.
    """
    invoice_type = detect_invoice_type(text)

    if invoice_type == 'meta':
        result = parse_meta_invoice(text)
    elif invoice_type == 'google_ads':
        result = parse_google_ads_invoice(text)
    elif invoice_type == 'tiktok':
        result = parse_tiktok_invoice(text)
    elif invoice_type == 'efactura':
        result = parse_efactura_invoice(text)
    else:
        # Generic parsing
        result = parse_generic_invoice(text)

    result['invoice_type'] = invoice_type
    result['filename'] = filename

    # Check if regex parsing was successful - use AI fallback if not
    needs_ai = (
        not result.get('invoice_number') or
        not result.get('invoice_value') or
        (not result.get('items') and invoice_type not in ['meta', 'google_ads', 'tiktok', 'efactura'])
    )

    if needs_ai and AI_ENABLED:
        print(f"Using AI fallback for {filename}")
        ai_result = parse_invoice_with_ai(text)

        # Merge AI results with regex results (AI fills in gaps)
        if ai_result:
            # Fill in missing fields from AI
            for field in ['invoice_number', 'invoice_date', 'invoice_value', 'currency',
                          'supplier', 'supplier_vat', 'customer_name', 'customer_vat']:
                if not result.get(field) and ai_result.get(field):
                    result[field] = ai_result[field]

            # Add items from AI if regex didn't find any
            if not result.get('items') and ai_result.get('items'):
                result['items'] = ai_result['items']

            result['ai_assisted'] = True

    return result


def parse_generic_invoice(text: str) -> dict:
    """Parse a generic invoice with common patterns."""
    result = {
        'supplier': None,
        'supplier_vat': None,
        'items': {}
    }

    # Try to extract invoice number
    patterns = [
        r'(?:Factura|Invoice)\s+(?:nr\.?|no\.?|#)\s*:?\s*([A-Z0-9\-/]+)',
        r'(?:Nr\.?|No\.?)\s+(?:factura|invoice)\s*:?\s*([A-Z0-9\-/]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            result['invoice_number'] = match.group(1)
            break

    # Try to extract date
    date_patterns = [
        r'(?:Data|Date)\s*:?\s*(\d{1,2}[./\-]\d{1,2}[./\-]\d{4})',
        r'(\d{1,2}\s+\w{3,}\.?\s*\d{4})',
    ]
    for pattern in date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            result['invoice_date'] = match.group(1)
            break

    # Try to extract total value
    value_patterns = [
        r'Total\s*:?\s*([\d.,]+)\s*(?:RON|EUR|USD)?',
        r'(?:Valoare|Amount)\s+total[aă]?\s*:?\s*([\d.,]+)',
    ]
    for pattern in value_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            try:
                result['invoice_value'] = parse_value(match.group(1))
                break
            except (ValueError, TypeError, AttributeError):
                pass

    # Extract currency
    currency_match = re.search(r'(RON|EUR|USD|Lei)', text, re.IGNORECASE)
    if currency_match:
        currency = currency_match.group(1).upper()
        result['currency'] = 'RON' if currency == 'LEI' else currency
    else:
        result['currency'] = 'RON'

    return result


def parse_invoice_with_ai(text: str, api_key: str = None) -> dict:
    """
    Use AI to parse invoice text and extract structured data including line items.
    This is used as a fallback when regex parsing fails or for unknown invoice formats.

    Returns dict with:
        - invoice_number, invoice_date, invoice_value, currency
        - supplier, supplier_vat, customer_vat, customer_name
        - items: dict of item_name -> value (line items from the invoice)
    """
    if not AI_ENABLED:
        return {'items': {}}

    if not api_key:
        api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return {'items': {}}

    try:
        client = anthropic.Anthropic(api_key=api_key)

        prompt = f"""Extract invoice data from this text. Return a JSON object with:
- invoice_number: string (invoice/factura number)
- invoice_date: string (date in format YYYY-MM-DD)
- invoice_value: number (total amount)
- currency: string (RON, EUR, USD)
- supplier: string (supplier/vendor name)
- supplier_vat: string (supplier VAT/CUI number)
- customer_name: string (customer/client name)
- customer_vat: string (customer VAT/CUI number)
- items: object mapping item/product/service name to its value (number)

For items, extract ALL line items from the invoice - these could be:
- Products or services with their prices
- Campaign names with costs (for advertising invoices)
- Any other billable line items

If a field cannot be found, use null. For items, include ALL items found on the invoice.

IMPORTANT: Return ONLY valid JSON, no explanations.

Invoice text:
{text[:8000]}"""  # Limit text to avoid token limits

        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}]
        )

        response_text = response.content[0].text.strip()

        # Clean up markdown if present
        if response_text.startswith('```'):
            response_text = response_text.split('```')[1]
            if response_text.startswith('json'):
                response_text = response_text[4:]
            response_text = response_text.strip()

        result = json.loads(response_text)

        # Ensure items dict exists
        if 'items' not in result or not isinstance(result['items'], dict):
            result['items'] = {}

        return result

    except Exception as e:
        print(f"AI invoice parsing error: {e}")
        return {'items': {}}


def process_bulk_invoices(files: list[tuple[bytes, str]]) -> dict:
    """
    Process multiple invoice files and generate a summary report.

    Args:
        files: List of tuples (file_bytes, filename)

    Returns:
        dict with:
            - invoices: list of parsed invoices
            - summary: aggregated statistics
            - by_item: item-level breakdown (campaigns, products, services, etc.)
            - by_month: monthly breakdown
            - total: grand total
    """
    invoices = []

    for file_bytes, filename in files:
        text = extract_text_from_bytes(file_bytes, filename)
        if text:
            result = parse_invoice_auto(text, filename)
            invoices.append(result)

    # Sort by date
    invoices.sort(key=lambda x: x.get('date_parsed') or datetime.min)

    # Calculate totals
    total = sum(inv.get('invoice_value', 0) or 0 for inv in invoices)

    # Group by month
    by_month = defaultdict(lambda: {'count': 0, 'total': 0})
    for inv in invoices:
        if inv.get('date_parsed'):
            month_key = inv['date_parsed'].strftime('%Y-%m')
            by_month[month_key]['count'] += 1
            by_month[month_key]['total'] += inv.get('invoice_value', 0) or 0

    # Group by item (aggregate across all invoices)
    by_item = defaultdict(float)
    for inv in invoices:
        for item, value in inv.get('items', {}).items():
            by_item[item] += value

    # Group by supplier
    by_supplier = defaultdict(lambda: {'count': 0, 'total': 0})
    for inv in invoices:
        supplier = inv.get('supplier') or 'Unknown'
        by_supplier[supplier]['count'] += 1
        by_supplier[supplier]['total'] += inv.get('invoice_value', 0) or 0

    # Add 'campaigns' alias to each invoice for frontend compatibility
    for inv in invoices:
        inv['campaigns'] = inv.get('items', {})

    return {
        'invoices': invoices,
        'total': total,
        'count': len(invoices),
        'by_month': dict(by_month),
        'by_item': dict(by_item),
        'by_campaign': dict(by_item),  # Alias for frontend compatibility
        'by_supplier': dict(by_supplier),
        'currency': invoices[0].get('currency', 'RON') if invoices else 'RON'
    }


def generate_excel_report(report_data: dict) -> bytes:
    """
    Generate an Excel report from bulk invoice processing results.

    Returns Excel file as bytes.
    """
    if not EXCEL_ENABLED:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    total_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    money_format = '#,##0.00'
    date_format = 'DD.MM.YYYY'

    invoices = report_data.get('invoices', [])
    currency = report_data.get('currency', 'RON')

    # ============ Sheet 1: Invoice Summary ============
    ws1 = wb.active
    ws1.title = 'Sumar Facturi'

    # Title
    ws1.merge_cells('A1:F1')
    ws1['A1'] = 'CENTRALIZATOR FACTURI'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal='center')

    ws1.merge_cells('A2:F2')
    ws1['A2'] = f'Total Facturi: {report_data.get("count", 0)} | Total: {report_data.get("total", 0):,.2f} {currency}'
    ws1['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Nr.', 'Data', 'Nr. Factură', 'Furnizor', f'Valoare ({currency})', 'Fișier']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data
    for idx, inv in enumerate(invoices, 1):
        row = idx + 4
        ws1.cell(row=row, column=1, value=idx).border = border

        date_cell = ws1.cell(row=row, column=2, value=inv.get('date_parsed'))
        if inv.get('date_parsed'):
            date_cell.number_format = date_format
        date_cell.border = border

        ws1.cell(row=row, column=3, value=inv.get('invoice_number', '')).border = border
        ws1.cell(row=row, column=4, value=inv.get('supplier', '')).border = border

        value_cell = ws1.cell(row=row, column=5, value=inv.get('invoice_value', 0))
        value_cell.number_format = money_format
        value_cell.border = border

        ws1.cell(row=row, column=6, value=inv.get('filename', '')).border = border

    # Total row
    total_row = len(invoices) + 5
    ws1.cell(row=total_row, column=4, value='TOTAL').font = Font(bold=True)
    total_cell = ws1.cell(row=total_row, column=5, value=report_data.get('total', 0))
    total_cell.number_format = money_format
    total_cell.font = Font(bold=True)

    # Column widths
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 30
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 50

    # ============ Sheet 2: Monthly Summary ============
    ws2 = wb.create_sheet('Sumar Lunar')

    ws2.merge_cells('A1:C1')
    ws2['A1'] = 'SUMAR PE LUNI'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal='center')

    headers2 = ['Luna', 'Nr. Facturi', f'Total ({currency})']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    row = 4
    for month_key in sorted(report_data.get('by_month', {}).keys()):
        data = report_data['by_month'][month_key]
        month_name = datetime.strptime(month_key, '%Y-%m').strftime('%B %Y')

        ws2.cell(row=row, column=1, value=month_name).border = border
        ws2.cell(row=row, column=2, value=data['count']).border = border
        total_cell = ws2.cell(row=row, column=3, value=data['total'])
        total_cell.number_format = money_format
        total_cell.border = border
        row += 1

    # Grand total
    ws2.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    ws2.cell(row=row, column=2, value=report_data.get('count', 0)).font = Font(bold=True)
    grand_total = ws2.cell(row=row, column=3, value=report_data.get('total', 0))
    grand_total.number_format = money_format
    grand_total.font = Font(bold=True)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 18

    # ============ Sheet 3: Item Summary ============
    if report_data.get('by_item'):
        ws3 = wb.create_sheet('Per Pozitie')

        ws3.merge_cells('A1:C1')
        ws3['A1'] = 'TOTAL PER POZITIE'
        ws3['A1'].font = Font(bold=True, size=14)
        ws3['A1'].alignment = Alignment(horizontal='center')

        headers3 = ['Pozitie', f'Total ({currency})', '% din Total']
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row = 4
        grand_total_value = report_data.get('total', 0) or 1

        for item_name, value in sorted(report_data['by_item'].items(), key=lambda x: -x[1]):
            ws3.cell(row=row, column=1, value=item_name).border = border

            value_cell = ws3.cell(row=row, column=2, value=value)
            value_cell.number_format = money_format
            value_cell.border = border

            pct_cell = ws3.cell(row=row, column=3, value=value / grand_total_value)
            pct_cell.number_format = '0.0%'
            pct_cell.border = border

            row += 1

        ws3.column_dimensions['A'].width = 45
        ws3.column_dimensions['B'].width = 18
        ws3.column_dimensions['C'].width = 12

    # ============ Sheet 4: Supplier Summary ============
    if report_data.get('by_supplier'):
        ws4 = wb.create_sheet('Per Furnizor')

        ws4.merge_cells('A1:C1')
        ws4['A1'] = 'TOTAL PER FURNIZOR'
        ws4['A1'].font = Font(bold=True, size=14)
        ws4['A1'].alignment = Alignment(horizontal='center')

        headers4 = ['Furnizor', 'Nr. Facturi', f'Total ({currency})']
        for col, header in enumerate(headers4, 1):
            cell = ws4.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row = 4
        for supplier, data in sorted(report_data['by_supplier'].items(), key=lambda x: -x[1]['total']):
            ws4.cell(row=row, column=1, value=supplier).border = border
            ws4.cell(row=row, column=2, value=data['count']).border = border

            value_cell = ws4.cell(row=row, column=3, value=data['total'])
            value_cell.number_format = money_format
            value_cell.border = border

            row += 1

        ws4.column_dimensions['A'].width = 40
        ws4.column_dimensions['B'].width = 15
        ws4.column_dimensions['C'].width = 18

    # ============ Sheet 5: Item by Invoice ============
    # Matrix view: items as rows, invoices as columns
    if report_data.get('by_item') and invoices:
        ws5 = wb.create_sheet('Pozitii per Factură')

        ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(invoices)+2)
        ws5['A1'] = 'DETALII POZITII PER FACTURĂ'
        ws5['A1'].font = Font(bold=True, size=14)
        ws5['A1'].alignment = Alignment(horizontal='center')

        # Headers
        row = 3
        ws5.cell(row=row, column=1, value='Pozitie').font = header_font
        ws5.cell(row=row, column=1).fill = header_fill
        ws5.cell(row=row, column=1).border = border

        for col, inv in enumerate(invoices, 2):
            date_str = inv.get('date_parsed').strftime('%d.%m') if inv.get('date_parsed') else ''
            cell = ws5.cell(row=row, column=col, value=date_str)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        total_col = len(invoices) + 2
        ws5.cell(row=row, column=total_col, value='TOTAL').font = header_font
        ws5.cell(row=row, column=total_col).fill = total_fill
        ws5.cell(row=row, column=total_col).border = border

        # Data rows
        row = 4
        all_items = sorted(set(
            item_name for inv in invoices for item_name in inv.get('items', {}).keys()
        ))

        for item_name in all_items:
            ws5.cell(row=row, column=1, value=item_name).border = border

            row_total = 0
            for col, inv in enumerate(invoices, 2):
                value = inv.get('items', {}).get(item_name, 0)
                cell = ws5.cell(row=row, column=col, value=value if value > 0 else '')
                if value > 0:
                    cell.number_format = money_format
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                row_total += value

            total_cell = ws5.cell(row=row, column=total_col, value=row_total)
            total_cell.number_format = money_format
            total_cell.font = Font(bold=True)
            total_cell.border = border
            total_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

            row += 1

        # Invoice totals row
        row += 1
        ws5.cell(row=row, column=1, value='TOTAL FACTURĂ').font = Font(bold=True)
        ws5.cell(row=row, column=1).fill = total_fill
        ws5.cell(row=row, column=1).border = border

        grand_total = 0
        for col, inv in enumerate(invoices, 2):
            value = inv.get('invoice_value', 0)
            cell = ws5.cell(row=row, column=col, value=value)
            cell.number_format = money_format
            cell.font = Font(bold=True)
            cell.border = border
            cell.fill = total_fill
            grand_total += value

        grand_cell = ws5.cell(row=row, column=total_col, value=grand_total)
        grand_cell.number_format = money_format
        grand_cell.font = Font(bold=True, color='FFFFFF')
        grand_cell.border = border
        grand_cell.fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')

        # Column widths
        ws5.column_dimensions['A'].width = 40
        for col in range(2, total_col + 1):
            ws5.column_dimensions[get_column_letter(col)].width = 12

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_summary_text(report_data: dict) -> str:
    """Generate a text summary of the bulk invoice processing results."""
    currency = report_data.get('currency', 'RON')
    lines = []

    lines.append('=' * 80)
    lines.append('INVOICE PROCESSING REPORT')
    lines.append('=' * 80)
    lines.append(f"Total Invoices: {report_data.get('count', 0)}")
    lines.append(f"Total Value: {report_data.get('total', 0):,.2f} {currency}")
    lines.append('')

    # Monthly breakdown
    if report_data.get('by_month'):
        lines.append('MONTHLY SUMMARY')
        lines.append('-' * 40)
        for month_key in sorted(report_data['by_month'].keys()):
            data = report_data['by_month'][month_key]
            month_name = datetime.strptime(month_key, '%Y-%m').strftime('%B %Y')
            lines.append(f"{month_name:<20} {data['total']:>15,.2f} {currency}")
        lines.append('')

    # Supplier breakdown
    if report_data.get('by_supplier'):
        lines.append('BY SUPPLIER')
        lines.append('-' * 40)
        for supplier, data in sorted(report_data['by_supplier'].items(), key=lambda x: -x[1]['total']):
            lines.append(f"{supplier:<30} {data['total']:>15,.2f} {currency}")
        lines.append('')

    # Item breakdown
    if report_data.get('by_item'):
        lines.append('BY ITEM')
        lines.append('-' * 40)
        for item_name, value in sorted(report_data['by_item'].items(), key=lambda x: -x[1]):
            lines.append(f"{item_name:<40} {value:>15,.2f} {currency}")
        lines.append('')

    lines.append('=' * 80)

    return '\n'.join(lines)
