"""Marja report xlsx export — styled openpyxl output."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style definitions ──
NAVY = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
DARK_BLUE = PatternFill(start_color='2C3E6B', end_color='2C3E6B', fill_type='solid')
DARKER_BLUE = PatternFill(start_color='1A2744', end_color='1A2744', fill_type='solid')
GREEN = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
BROWN = PatternFill(start_color='5D4037', end_color='5D4037', fill_type='solid')
LIGHT_GREY = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

WHITE_BOLD = Font(name='Arial', size=9, bold=True, color='FFFFFF')
WHITE_NORMAL = Font(name='Arial', size=9, color='FFFFFF')
BLACK_NORMAL = Font(name='Arial', size=9)
RED_NORMAL = Font(name='Arial', size=9, color='FF0000')
HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')

THIN_BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC'),
)

MONTH_NAMES_RO = [
    '', 'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
    'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie',
]

# Section header colors
SECTION_FILLS = {
    'VW PKW INTERN (retail)': DARK_BLUE,
    'VW PKW INTERN (flote)': DARKER_BLUE,
    'Bonus & Discount': BROWN,
    'MARJA FINALĂ': NAVY,
    'VW PKW EXTERN': GREEN,
}


def export_marja_xlsx(report, period_year, period_month):
    """Generate styled Marja report xlsx.

    Args:
        report: dict from compute_marja_report()
        period_year: int
        period_month: int (1-12)

    Returns:
        bytes — xlsx file content
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Raport Marjă'

    eur_rate = report['eur_rate']
    month_name = MONTH_NAMES_RO[period_month]

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 10

    # ── Row 1: Title header ──
    ws.merge_cells('A1:E1')
    cell = ws.cell(row=1, column=1,
                   value=f'RAPORT MARJĂ VÂNZĂRI — {month_name} {period_year}  |  Curs: {eur_rate} LEI/EUR')
    cell.font = HEADER_FONT
    cell.fill = NAVY
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # ── Row 2: Column headers ──
    headers = ['Indicator', 'Valoare (LEI)', 'Valoare (EUR)', 'Conturi', 'KST']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = WHITE_BOLD
        cell.fill = NAVY
        cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 22

    # Freeze panes at row 3
    ws.freeze_panes = 'A3'

    # ── Data rows ──
    row_num = 3

    for section in report['sections']:
        section_name = section['section']

        # Section header
        ws.merge_cells(f'A{row_num}:E{row_num}')
        cell = ws.cell(row=row_num, column=1, value=section_name)
        cell.font = WHITE_BOLD

        # Match section fill color
        fill = LIGHT_GREY
        for key, f in SECTION_FILLS.items():
            if key in section_name:
                fill = f
                cell.font = WHITE_BOLD
                break
        cell.fill = fill
        ws.row_dimensions[row_num].height = 22
        row_num += 1

        # Section rows
        for line in section['rows']:
            is_marja_finala = 'MARJA FINALĂ' in line['label']
            is_negative = line['lei'] < 0

            # Label
            cell_a = ws.cell(row=row_num, column=1, value=f'  {line["label"]}')
            # LEI
            cell_b = ws.cell(row=row_num, column=2, value=float(line['lei']))
            cell_b.number_format = '#,##0.00'
            cell_b.alignment = Alignment(horizontal='right')
            # EUR
            cell_c = ws.cell(row=row_num, column=3, value=float(line['eur']))
            cell_c.number_format = '#,##0.00'
            cell_c.alignment = Alignment(horizontal='right')
            # Accounts
            accts = ', '.join(str(a) for a in line['accounts']) if line['accounts'] else ''
            ws.cell(row=row_num, column=4, value=accts).font = BLACK_NORMAL
            # KST
            ws.cell(row=row_num, column=5, value=line['kst']).font = BLACK_NORMAL

            if is_marja_finala:
                for col in range(1, 6):
                    c = ws.cell(row=row_num, column=col)
                    c.fill = NAVY
                    c.font = WHITE_BOLD
                cell_a.value = line['label']  # no indent for marja finala
                ws.row_dimensions[row_num].height = 24
            elif is_negative:
                cell_b.font = RED_NORMAL
                cell_c.font = RED_NORMAL
                cell_a.font = BLACK_NORMAL
            else:
                cell_a.font = BLACK_NORMAL
                cell_b.font = BLACK_NORMAL
                cell_c.font = BLACK_NORMAL

            # Thin bottom border
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).border = THIN_BORDER

            row_num += 1

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
