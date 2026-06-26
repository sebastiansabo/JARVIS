"""Generate EuroFib 'Note contabile' import xlsx from config + OrderLines.

CRITICAL: Columns AH (kurs_per) and AI (kurs_fix) MUST stay blank.
Putting a date in kurs_per causes EuroFib's SQL Server import to throw:
  'Arithmetic overflow error converting int to data type numeric'
"""
import io
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment

from ..config import JobConfig
from ..models import OrderLine

ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"
TEMPLATE_PATH = ASSETS_DIR / "eurofib_template.xlsx"

# Clean cell styles (template has yellow highlights / red font that confuse review)
_DEFAULT_FONT = Font(name="Arial", size=10, color="000000")
_DEFAULT_FILL = PatternFill(fill_type=None)
_DEFAULT_BORDER = Border()
_DEFAULT_ALIGN = Alignment(horizontal="general", vertical="bottom")
_DATE_FMT = "DD.MM.YYYY"


def _brand_short(model: str, brand_map: dict[str, str]) -> str:
    """Resolve model name to short brand text via config brand_map."""
    m = model.lower()
    for key, value in brand_map.items():
        if key.lower() in m:
            return value
    return "vw"


class EurofibXlsxRenderer:
    """Fills the EuroFib template with debit/credit row pairs."""

    def __init__(self, cfg: JobConfig):
        self.cfg = cfg

    def _load_template(self):
        wb = load_workbook(str(TEMPLATE_PATH))
        ws = wb.active
        # Clear any existing data rows below header (row 2 is header)
        for r in range(3, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=col)
                cell.value = None
                cell.font = _DEFAULT_FONT
                cell.fill = _DEFAULT_FILL
                cell.border = _DEFAULT_BORDER
                cell.alignment = _DEFAULT_ALIGN
                cell.number_format = "General"
        return wb, ws

    def _write_debit(self, ws, r: int, inv_no: int, line: OrderLine, text: str):
        cfg = self.cfg
        buch_date = datetime(
            cfg.invoice.date.year, cfg.invoice.date.month, cfg.invoice.date.day
        )
        kurs_date = datetime(
            cfg.fx.kurs_date.year, cfg.fx.kurs_date.month, cfg.fx.kurs_date.day
        )
        row_kurs = line.kurs if line.kurs is not None else cfg.fx.kurs
        fw_amount = line.advance * line.qty  # negative for storno

        # For STORNO, swap s/h and konto/gegenkonto to produce the reversal entry
        konto = cfg.eurofib.konto_credit if cfg.eurofib.is_storno else cfg.eurofib.konto_debit
        sh = "h" if cfg.eurofib.is_storno else "s"
        gegenkonto = cfg.eurofib.konto_debit if cfg.eurofib.is_storno else cfg.eurofib.konto_credit

        ws.cell(row=r, column=1, value="x")                      # A marker
        ws.cell(row=r, column=2, value=cfg.eurofib.klient)        # B klient
        ws.cell(row=r, column=3, value=konto)                     # C konto
        ws.cell(row=r, column=4, value=sh)                        # D soll_haben
        ws.cell(row=r, column=5, value=buch_date)                 # E buchdatum
        ws.cell(row=r, column=6, value=cfg.eurofib.belegart)      # F belegart
        ws.cell(row=r, column=7, value=buch_date)                 # G belegdatum
        ws.cell(row=r, column=8, value=inv_no)                    # H belegnummer
        ws.cell(row=r, column=9, value=f"=M{r}*AG{r}")           # I betrag (formula)
        # J steuercode: blank for debit
        # K steuerbetrag: blank for debit
        ws.cell(row=r, column=12, value=cfg.fx.currency)          # L fwcd
        ws.cell(row=r, column=13, value=fw_amount)                # M fwbetrag
        # N fw_steuercode: blank for debit
        ws.cell(row=r, column=16, value=int(line.konto_credit_override) if line.konto_credit_override else gegenkonto)  # P gegenkonto
        ws.cell(row=r, column=17, value=text)                     # Q text
        ws.cell(row=r, column=18, value="B")                      # R brutto_netto
        ws.cell(row=r, column=25, value=line.comanda)             # Y extbeleg
        ws.cell(row=r, column=32, value=kurs_date)                # AF kursdatum
        ws.cell(row=r, column=33, value=row_kurs)                 # AG kurs
        # AH (col 34) kurs_per: MUST stay blank — see module docstring
        # AI (col 35) kurs_fix: MUST stay blank
        if line.kostenstelle:
            ws.cell(row=r, column=36, value=line.kostenstelle)    # AJ kostenstelle

    def _write_credit(self, ws, r: int, inv_no: int, line: OrderLine,
                      text: str, debit_row: int):
        cfg = self.cfg
        buch_date = datetime(
            cfg.invoice.date.year, cfg.invoice.date.month, cfg.invoice.date.day
        )
        kurs_date = datetime(
            cfg.fx.kurs_date.year, cfg.fx.kurs_date.month, cfg.fx.kurs_date.day
        )
        row_kurs = line.kurs if line.kurs is not None else cfg.fx.kurs
        fw_amount = line.advance * line.qty  # negative for storno

        # For STORNO, swap s/h: credit row uses konto_debit and "s"
        credit_konto = cfg.eurofib.konto_debit if cfg.eurofib.is_storno else cfg.eurofib.konto_credit
        credit_sh = "s" if cfg.eurofib.is_storno else "h"

        # A marker: blank for credit
        ws.cell(row=r, column=2, value=cfg.eurofib.klient)        # B klient
        ws.cell(row=r, column=3, value=int(line.konto_credit_override) if line.konto_credit_override else credit_konto)  # C konto
        ws.cell(row=r, column=4, value=credit_sh)                 # D soll_haben
        ws.cell(row=r, column=5, value=buch_date)                 # E buchdatum
        ws.cell(row=r, column=6, value=cfg.eurofib.belegart)      # F belegart
        ws.cell(row=r, column=7, value=buch_date)                 # G belegdatum
        ws.cell(row=r, column=8, value=inv_no)                    # H belegnummer
        ws.cell(row=r, column=9, value=f"=I{debit_row}")          # I betrag (ref debit)
        ws.cell(row=r, column=10, value=cfg.eurofib.steuercode)   # J steuercode
        ws.cell(row=r, column=11, value=0)                        # K steuerbetrag
        ws.cell(row=r, column=12, value=cfg.fx.currency)          # L fwcd
        ws.cell(row=r, column=13, value=fw_amount)                # M fwbetrag
        ws.cell(row=r, column=14, value=cfg.eurofib.fw_steuercode) # N fw_steuercode
        # P gegenkonto: blank for credit
        ws.cell(row=r, column=17, value=text)                     # Q text
        ws.cell(row=r, column=18, value="N")                      # R brutto_netto
        ws.cell(row=r, column=25, value=line.comanda)             # Y extbeleg
        ws.cell(row=r, column=32, value=kurs_date)                # AF kursdatum
        ws.cell(row=r, column=33, value=row_kurs)                 # AG kurs
        # AH (col 34) kurs_per: MUST stay blank
        # AI (col 35) kurs_fix: MUST stay blank
        if line.kostenstelle:
            ws.cell(row=r, column=36, value=line.kostenstelle)    # AJ kostenstelle

    def _apply_date_format(self, ws, r: int):
        """Set DD.MM.YYYY format on date columns."""
        for col in (5, 7, 32):  # E, G, AF
            ws.cell(row=r, column=col).number_format = _DATE_FMT

    def _resolve_inv_no(self, idx: int, line: OrderLine) -> int:
        """Per-row invoice number override, or fall back to sequential."""
        return line.start_no if line.start_no is not None else self.cfg.invoice.start_no + idx

    def render(self, lines: list[OrderLine], out_path: Path) -> Path:
        """Generate EuroFib xlsx to disk."""
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb, ws = self._load_template()

        for i, line in enumerate(lines):
            inv_no = self._resolve_inv_no(i, line)
            brand = _brand_short(line.model, self.cfg.eurofib.brand_map)
            text = self.cfg.eurofib.text_template.format(
                model=line.model, comanda=line.comanda,
                brand_short=brand,  # keep for backwards compat
            )
            debit_row = 3 + 2 * i
            credit_row = debit_row + 1

            self._write_debit(ws, debit_row, inv_no, line, text)
            self._apply_date_format(ws, debit_row)

            self._write_credit(ws, credit_row, inv_no, line, text, debit_row)
            self._apply_date_format(ws, credit_row)

        wb.save(str(out_path))
        wb.close()
        return out_path

    def render_to_bytes(self, lines: list[OrderLine]) -> bytes:
        """Generate EuroFib xlsx and return as bytes (for web download)."""
        wb, ws = self._load_template()

        for i, line in enumerate(lines):
            inv_no = self._resolve_inv_no(i, line)
            brand = _brand_short(line.model, self.cfg.eurofib.brand_map)
            text = self.cfg.eurofib.text_template.format(
                model=line.model, comanda=line.comanda,
                brand_short=brand,  # keep for backwards compat
            )
            debit_row = 3 + 2 * i
            credit_row = debit_row + 1

            self._write_debit(ws, debit_row, inv_no, line, text)
            self._apply_date_format(ws, debit_row)

            self._write_credit(ws, credit_row, inv_no, line, text, debit_row)
            self._apply_date_format(ws, credit_row)

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        return buf.getvalue()
