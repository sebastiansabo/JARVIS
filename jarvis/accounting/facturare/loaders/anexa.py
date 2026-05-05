"""Read Anexa xlsx → list[OrderLine] + optional metadata."""
import io
from datetime import date, datetime
import openpyxl
from ..models import OrderLine


# Metadata key aliases → canonical field names
_META_KEYS: dict[str, str] = {
    "client": "customer_name", "customer": "customer_name", "client name": "customer_name",
    "client address": "customer_address", "address": "customer_address", "adresa": "customer_address",
    "client vat": "customer_vat", "vat client": "customer_vat", "cui client": "customer_vat",
    "invoice date": "invoice_date", "date": "invoice_date", "data": "invoice_date", "data factura": "invoice_date",
    "kurs": "kurs", "curs": "kurs",
    "start no": "start_no", "start": "start_no", "numar start": "start_no", "nr start": "start_no",
    "contract ref": "contract_ref", "contract": "contract_ref",
    "anexa ref": "anexa_ref", "anexa": "anexa_ref",
    "intocmit de": "intocmit_de", "intocmit": "intocmit_de",
    "job id": "job_id", "job": "job_id",
}

# Data header keywords that signal end of metadata block
_DATA_HEADERS = {"comanda", "nr. comanda", "nr.comanda", "model", "vin"}


def _parse_metadata(ws) -> tuple[dict[str, str], int]:
    """Extract key-value metadata from top rows.

    Scans rows starting from row 1. Column A = label, column B = value.
    After metadata ends, skips empty rows to find the actual data header row.

    Returns:
        (metadata_dict, data_start_row) — 1-indexed row where data headers begin.
    """
    meta: dict[str, str] = {}
    meta_done = False

    # Phase 1: read metadata key-value pairs
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        cell_a = str(row[0]).strip() if row[0] is not None else ""
        cell_a_lower = cell_a.lower()

        if not meta_done:
            # If cell A looks like a data header, metadata is done
            if cell_a_lower in _DATA_HEADERS:
                return meta, row_idx

            # Empty row = end of metadata block
            if not cell_a:
                meta_done = True
                continue

            # Try to match as metadata key
            if cell_a_lower in _META_KEYS and len(row) > 1 and row[1] is not None:
                value = row[1]
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%Y-%m-%d")
                meta[_META_KEYS[cell_a_lower]] = str(value).strip()
        else:
            # Phase 2: skip empty rows, find data header
            if cell_a_lower in _DATA_HEADERS:
                return meta, row_idx
            # Non-empty row that's not a known header — treat as data header anyway
            if cell_a:
                return meta, row_idx

    # Fallback: data starts at row 3 (legacy format)
    return meta, 3


def _detect_format(ws, header_row: int) -> str:
    """Detect Anexa format from a specific header row.

    Returns 'proforma' if headers match the simple VIN+Rest layout,
    otherwise 'standard' for the full invoice layout.
    """
    for row in ws.iter_rows(min_row=header_row, max_row=header_row + 1, values_only=True):
        headers = [str(c).strip().lower() if c else "" for c in row]
        joined = " ".join(headers)
        if "vin" in headers and ("rest" in joined or "plata" in joined):
            return "proforma"
    return "standard"


def _parse_standard(ws, data_start: int) -> tuple[list[OrderLine], list[str]]:
    """Parse full Anexa format: A=comanda B=model C=culoare D=list E=sell F=disc G=advance H=rest I=vin."""
    lines: list[OrderLine] = []
    errors: list[str] = []

    # Skip header row(s) — data starts 2 rows after header_start
    first_data_row = data_start + 2

    for row_idx, row in enumerate(ws.iter_rows(min_row=first_data_row, values_only=True), start=first_data_row):
        comanda = row[0]  # A
        if comanda is None:
            continue

        model = row[1]    # B
        culoare = row[2]  # C
        advance = row[6] if len(row) > 6 else None  # G

        missing = []
        if model is None:
            missing.append("B (model)")
        if culoare is None:
            missing.append("C (culoare)")
        if advance is None:
            missing.append("G (advance)")
        if missing:
            errors.append(f"Row {row_idx}: missing {', '.join(missing)}")
            continue

        lines.append(OrderLine(
            comanda=int(comanda),
            model=str(model).strip(),
            culoare=str(culoare).strip(),
            list_price=float(row[3]) if row[3] is not None else None,
            selling_price=float(row[4]) if row[4] is not None else None,
            advance=float(advance),
            rest=float(row[7]) if len(row) > 7 and row[7] is not None else None,
            vin=str(row[8]).strip() if len(row) > 8 and row[8] is not None else None,
        ))

    return lines, errors


def _parse_proforma(ws, data_start: int) -> tuple[list[OrderLine], list[str]]:
    """Parse proforma Anexa format.

    A=comanda B=model C=culoare D=VIN E=rest_de_plata
    F=contract_ref G=anexa_ref H=invoice_date I=qty (optional per-row overrides)
    """
    lines: list[OrderLine] = []
    errors: list[str] = []

    # Skip header row — data starts 1 row after header
    first_data_row = data_start + 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=first_data_row, values_only=True), start=first_data_row):
        comanda = row[0]  # A
        if comanda is None:
            continue

        model = row[1]      # B
        culoare = row[2]    # C
        vin = row[3] if len(row) > 3 else None        # D
        rest_amount = row[4] if len(row) > 4 else None  # E — "Rest de plata"

        # Only rest_de_plata (amount) is strictly required
        if rest_amount is None:
            errors.append(f"Row {row_idx}: missing E (rest de plata)")
            continue

        # Optional per-row metadata (F-H)
        contract_ref = str(row[5]).strip() if len(row) > 5 and row[5] is not None else None
        anexa_ref = str(row[6]).strip() if len(row) > 6 and row[6] is not None else None
        invoice_date_val = row[7] if len(row) > 7 and row[7] is not None else None

        # Handle date objects
        inv_date_str = None
        if invoice_date_val is not None:
            if isinstance(invoice_date_val, (date, datetime)):
                inv_date_str = invoice_date_val.strftime("%Y-%m-%d")
            else:
                inv_date_str = str(invoice_date_val).strip()

        # Optional qty (column I)
        qty_val = row[8] if len(row) > 8 and row[8] is not None else 1
        try:
            qty = int(qty_val)
        except (ValueError, TypeError):
            qty = 1

        lines.append(OrderLine(
            comanda=int(comanda),
            model=str(model).strip() if model is not None else "",
            culoare=str(culoare).strip() if culoare is not None else "",
            list_price=None,
            selling_price=None,
            advance=float(rest_amount),
            rest=None,
            vin=str(vin).strip() if vin is not None else None,
            contract_ref=contract_ref,
            anexa_ref=anexa_ref,
            invoice_date=inv_date_str,
            qty=qty,
        ))

    return lines, errors


def parse_anexa_metadata(source, sheet_name: str = "Sheet1") -> dict[str, str]:
    """Extract metadata from an Anexa xlsx.

    Reads both the header block (client info) and per-row fields from
    the first data row (contract_ref, anexa_ref, start_no, invoice_date).

    Returns dict with keys like customer_name, customer_address, invoice_date, kurs, etc.
    """
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(str(source), data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb[sheet_name]
    meta, data_start = _parse_metadata(ws)
    fmt = _detect_format(ws, data_start)

    # For proforma, also grab per-row fields from the first data row
    if fmt == "proforma":
        first_data_row = data_start + 1
        for row in ws.iter_rows(min_row=first_data_row, max_row=first_data_row, values_only=True):
            if row[0] is None:
                continue
            if len(row) > 5 and row[5] is not None:
                meta.setdefault("contract_ref", str(row[5]).strip())
            if len(row) > 6 and row[6] is not None:
                meta.setdefault("anexa_ref", str(row[6]).strip())
            if len(row) > 7 and row[7] is not None:
                val = row[7]
                if isinstance(val, (date, datetime)):
                    meta.setdefault("invoice_date", val.strftime("%Y-%m-%d"))
                else:
                    meta.setdefault("invoice_date", str(val).strip())

    wb.close()
    return meta


def load_anexa(source, sheet_name: str = "Sheet1") -> list[OrderLine]:
    """Parse an Anexa xlsx file into OrderLine objects.

    Auto-detects format and skips metadata header block if present.

    Args:
        source: file path (str/Path) or bytes
        sheet_name: worksheet name to read

    Returns:
        list of OrderLine

    Raises:
        ValueError: on missing columns or invalid data
    """
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(str(source), data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Sheet "{sheet_name}" not found. Available: {wb.sheetnames}')

    ws = wb[sheet_name]
    _, data_start = _parse_metadata(ws)
    fmt = _detect_format(ws, data_start)

    if fmt == "proforma":
        lines, errors = _parse_proforma(ws, data_start)
    else:
        lines, errors = _parse_standard(ws, data_start)

    wb.close()

    if errors:
        raise ValueError(
            f"Anexa validation failed ({len(errors)} errors):\n"
            + "\n".join(errors)
        )

    if not lines:
        raise ValueError("No data rows found in Anexa (all comanda values are empty)")

    return lines
