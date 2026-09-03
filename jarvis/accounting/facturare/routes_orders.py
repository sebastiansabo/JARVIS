"""Comenzi API routes — Contract → Anexa → Invoice lifecycle.

Endpoints:
  Contracts:  GET/POST /facturare/api/contracts
  Anexas:     GET/POST /facturare/api/contracts/<id>/anexas
  Lifecycle:  POST /facturare/api/invoices/proforma|invoice|storno|final
  Detail:     GET  /facturare/api/anexas/<id>
  Delete:     DELETE /facturare/api/invoices/<id>
  Users:      GET  /facturare/api/users?q=
  PDF:        GET  /facturare/api/invoices/<id>/pdf
"""
import logging
import time
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from flask import jsonify, request
from flask_login import login_required, current_user

from . import facturare_bp
from .models import InvoiceTypeEnum
from .repositories.invoice_storage_repository import InvoiceStorageRepository
from .services.invoice_state_machine import InvoiceStateMachine, InvoiceStateMachineError
from .schemas import (
    ContractCreateRequest, AnexaCreateRequest, AnexaLineCreate,
    ProformaCreateRequest, InvoiceCreateRequest,
    StornoCreateRequest, FinalCreateRequest,
)
from core.utils.api_helpers import error_response, handle_api_errors
from core.roles.repositories.permission_repository import PermissionRepository

logger = logging.getLogger("jarvis.facturare.orders")
_repo = InvoiceStorageRepository()
_sm = InvoiceStateMachine(_repo)
_perm_repo = PermissionRepository()

TYPE_LABELS = {"PROFORMA": "Proforma", "INVOICE": "Factura", "STORNO": "Storno", "FINAL": "Final"}


WHOLE_EUR = Decimal("1")
CENTS = Decimal("0.01")


def _quant_for(inv_row) -> Decimal:
    """Per-car rounding granularity for an invoice row: cents when the invoice was
    issued in "zecimale" mode (`round_decimals` true), whole EUR otherwise."""
    return CENTS if (inv_row or {}).get("round_decimals") else WHOLE_EUR


def _round_half_up(value, quant: Decimal = WHOLE_EUR):
    """Round `value` to `quant` (a whole EUR by default) with halves going up
    (1014.50 -> 1015).

    Python's built-in round() uses banker's rounding (round-half-to-even), so a
    5% advance landing exactly on X.50 with an even X rounds DOWN (1014.50 ->
    1014). Per-car advance amounts must round half up per the accounting spec.

    In "zecimale" mode callers pass `quant=CENTS` so each car keeps its two
    decimals (19997 * 10% = 1999.70 instead of 2000). Whole-EUR rounding returns
    an int (unchanged); cent rounding returns a float.
    """
    q = Decimal(str(value)).quantize(quant, rounding=ROUND_HALF_UP)
    return int(q) if quant == WHOLE_EUR else float(q)


def _snap_pct(total_amount, total_selling) -> float:
    """Effective advance percentage, snapped to the nearest whole % when within
    0.5% of one.

    The stored proforma/invoice total is itself a rounded figure (e.g. a 5%
    advance of 8128.20 is stored as 8128.00), so re-deriving each car by slicing
    that total — total_amount * selling / total_selling — loses the exact .50 and
    rounds a 20290 car down to 1014. Snapping the ratio back to a clean 0.05 lets
    each car be computed as selling * pct (1014.50 -> 1015), matching the per-line
    coverage/document-list computation. Non-whole percentages fall through to the
    raw ratio unchanged.
    """
    if not total_selling:
        return 0.0
    raw = total_amount / total_selling
    snapped = round(raw * 100) / 100
    # Never snap a positive advance down to 0% — a real total (e.g. a 0,0999%
    # advance on a 100-car anexa) must keep its raw fraction, else every car
    # renders 0,00 €. Only snap when the clean percent is itself non-zero.
    if snapped != 0 and abs(raw - snapped) < 0.005:
        return snapped
    return raw


_PCT_CLOSE_TOL = 0.005  # how near cumulative fraction must be to 1.0 to "close" a car


def _car_slice_eur(selling, this_fraction, prior_fractions, quant: Decimal = WHOLE_EUR):
    """Amount one invoice books for a single car (cumulative rounding).

    Every non-closing slice rounds its own `selling * fraction` to `quant`. The
    invoice that *closes* the car's coverage — cumulative fraction reaches 1.0 —
    instead books the residual `round(selling) - sum(prior slices)`, so the
    successive slices reconcile to the car's price exactly.

    Without this, a 5% advance and a 95% remainder that both land on X.50 each
    round up and overshoot by 1 EUR: 2288 + 43463 = 45751 instead of
    2288 + 43462 = 45750 (CTR-945 / comanda 352848). Because prior slices are
    themselves rounded to the same `quant`, `sum(prior slices)` telescopes to the
    running total actually invoiced, so the residual reverses exactly what was
    billed. `quant=CENTS` keeps two decimals (zecimale mode); the default whole
    EUR is unchanged.
    """
    if abs(sum(prior_fractions) + this_fraction - 1.0) < _PCT_CLOSE_TOL:
        prior_booked = sum(_round_half_up(selling * f, quant) for f in prior_fractions)
        return _round_half_up(selling, quant) - prior_booked
    return _round_half_up(selling * this_fraction, quant)


def _order_key(inv):
    """Chronological ordering of an anexa's invoices (issued date, then sequence)."""
    return (str(inv.get("issued_date") or ""), inv.get("sequence_number") or 0, inv.get("id") or 0)


def _prior_car_fractions(invoices, this_inv, line_id, price_by_line, all_line_ids):
    """Snapped fractions already booked on `line_id` by same-type invoices issued
    *before* `this_inv` — the cumulative context `_car_slice_eur` needs to know
    whether `this_inv` closes the car and, if so, what to subtract.

    "Same type" (PROFORMA accumulates with PROFORMA, INVOICE/advance with INVOICE)
    keeps the proforma and factura-avans tracks independent, matching how the
    coverage view accumulates line_proforma_eur vs line_invoiced_eur separately.
    """
    this_key = _order_key(this_inv)
    fracs = []
    for inv in sorted(invoices, key=_order_key):
        if inv.get("id") == this_inv.get("id"):
            continue
        if inv.get("invoice_type") != this_inv.get("invoice_type"):
            continue
        if _order_key(inv) >= this_key:
            continue
        raw = inv.get("line_ids")
        if isinstance(raw, str):
            import json as _json
            raw = _json.loads(raw)
        covered = raw or list(all_line_ids)
        if line_id not in covered:
            continue
        covered_total = sum(price_by_line.get(lid, 0) for lid in covered) or 1
        fracs.append(_snap_pct(float(inv["total_amount_eur"]), covered_total))
    return fracs


def _per_car_advance_eur(inv_total, selling, covered_selling, split_mode, n_lines,
                         prior_fractions=None, quant: Decimal = WHOLE_EUR):
    """Whole-EUR per-car slice of an advance/invoice total.

    Single source of truth for the advance, storno and final exports so all three
    book the SAME rounded EUR per car and reconcile to zero. Proportional splits
    snap the ratio to a clean percent then round half up (matching the invoice
    PDF); equal splits divide and round. Slicing the stored total with 2-decimal
    precision instead (the old storno path) echoed fractional-EUR residues that
    never cleared and mismatched the whole-EUR invoice — e.g. a 5% advance stored
    as 1403.89 must reverse as 1404, not 1403.89.

    `prior_fractions` (the snapped fractions of same-track invoices booked earlier
    on this car) enables the cumulative-rounding residual for the closing slice, so
    a 95% remainder reverses as 43462, not 43463. Omitted (None) keeps the legacy
    independent-round behaviour for callers without sibling context.
    """
    if split_mode == "proportional" and covered_selling:
        return _car_slice_eur(selling, _snap_pct(inv_total, covered_selling),
                              prior_fractions or [], quant)
    return _round_half_up(inv_total / max(n_lines, 1), quant)


def _coverage_share(inv_total, selling, covered_total, covered_n, split_mode,
                    prior_fractions, quant: Decimal = WHOLE_EUR):
    """Per-car EUR a single invoice books for one covered car, for the coverage /
    document-items views.

    Mirrors the PDF generator: an *equal* split divides the total evenly across
    the covered cars (a fixed 1999 €/car proforma of 199900 shows 1999, not the
    proportional snap that rounds 9.9965% up to 10% -> 1999.70/2000). Proportional
    splits slice by price with the snapped percent + closing residual. Without the
    equal branch these views disagreed with the actual proforma PDF.
    """
    if split_mode == "equal":
        return _round_half_up(inv_total / max(covered_n, 1), quant)
    return _car_slice_eur(selling, _snap_pct(inv_total, covered_total),
                          prior_fractions or [], quant)

# ── Document items cache (in-memory, per doc_type key) ──────────
_doc_items_cache: dict[str, tuple[float, list]] = {}  # key -> (timestamp, items)
_DOC_ITEMS_TTL = 60  # seconds

def _invalidate_doc_items_cache():
    _doc_items_cache.clear()


def _check_perm(action: str) -> bool:
    if not current_user or not current_user.is_authenticated:
        return False
    role_id = getattr(current_user, "role_id", None)
    if not role_id:
        return getattr(current_user, "can_access_accounting", False)
    perm = _perm_repo.check_permission_v2(role_id, "facturare", "records", action)
    if perm is not None:
        return perm
    perm = _perm_repo.check_permission_v2(role_id, "invoices", "records", action)
    if perm is not None:
        return perm
    return getattr(current_user, "can_access_accounting", False)


def _inv_to_dict(row):
    import json as _json
    raw_line_ids = row.get("line_ids")
    if isinstance(raw_line_ids, str):
        raw_line_ids = _json.loads(raw_line_ids)
    return {
        "id": row["id"], "anexa_id": row["anexa_id"],
        "invoice_type": row["invoice_type"], "invoice_state": row["invoice_state"],
        "sequence_number": row.get("sequence_number", 1),
        "invoice_number": row.get("invoice_number"),
        "issued_date": str(row["issued_date"]) if row.get("issued_date") else None,
        "total_amount_eur": float(row["total_amount_eur"]),
        "total_amount_ron": float(row["total_amount_ron"]),
        "kurs_applied": float(row["kurs_applied"]) if row.get("kurs_applied") else None,
        "currency": row["currency"],
        "intocmit_de": row.get("intocmit_de"),
        "notes": row.get("notes"),
        "line_ids": raw_line_ids,
        "doc_mode": row.get("doc_mode", "per_car"),
        "split_mode": row.get("split_mode", "equal"),
        "round_decimals": bool(row.get("round_decimals")),
        "created_at": str(row["created_at"]) if row.get("created_at") else None,
    }


def _line_to_dict(row):
    return {
        "id": row["id"], "line_number": row["line_number"],
        "nr_comanda": row.get("nr_comanda"), "vin": row.get("vin"),
        "model": row["model"], "culoare": row.get("culoare"),
        "list_price_eur": float(row["list_price_eur"]),
        "selling_price_eur": float(row["selling_price_eur"]),
        "qty": row["qty"],
    }


def _resolve_doc_no(docnum_map: dict, line_id, fallback):
    """Per-car document number: the stored value from facturare_document_numbers
    (keyed by car line_id) when present, else the pre-backfill positional
    derivation (base_no [+ idx]) passed in as `fallback`.

    Shared by the anexa-detail display, the document-items list, and the PDF
    generator so all UI/PDF consumers agree with the numbers actually
    allocated/persisted at issue time (see get_document_number_map).
    """
    return docnum_map.get(line_id, fallback)


# ── Contracts ────────────────────────────────────────────────────

@facturare_bp.route("/facturare/api/contracts")
@login_required
@handle_api_errors
def api_list_contracts():
    if not _check_perm("view"):
        return error_response("Permission denied", 403)
    rows = _repo.list_contracts()
    contracts = []
    for r in rows:
        contracts.append({
            "id": r["id"], "contract_ref": r["contract_ref"],
            "supplier_id": r["supplier_id"], "customer_id": r["customer_id"],
            "supplier_name": r.get("supplier_name", ""),
            "customer_name": r.get("customer_name", ""),
            "contract_date": str(r["contract_date"]) if r.get("contract_date") else None,
            "responsible": r.get("responsible"),
            "anexa_count": r.get("anexa_count", 0),
            "archived_anexa_count": r.get("archived_anexa_count", 0),
            "total_value": float(r.get("total_value", 0) or 0),
            "invoiced_total": float(r.get("invoiced_total", 0) or 0),
            "archived": bool(r.get("archived")),
            "archive_after": str(r["archive_after"]) if r.get("archive_after") else None,
            "notes": r.get("notes"),
            "created_at": str(r["created_at"]) if r.get("created_at") else None,
        })
    return jsonify({"contracts": contracts})


@facturare_bp.route("/facturare/api/contracts", methods=["POST"])
@login_required
@handle_api_errors
def api_create_contract():
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    try:
        req = ContractCreateRequest(**data)
    except Exception as e:
        return error_response(str(e))
    row = _repo.create_contract(
        contract_ref=req.contract_ref, supplier_id=req.supplier_id,
        customer_id=req.customer_id, contract_date=req.contract_date,
        responsible=req.responsible, notes=req.notes, created_by=current_user.id,
    )
    return jsonify({"success": True, "contract": {"id": row["id"], "contract_ref": row["contract_ref"]}}), 201


@facturare_bp.route("/facturare/api/contracts/<int:contract_id>", methods=["PATCH"])
@login_required
@handle_api_errors
def api_update_contract(contract_id):
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    contract = _repo.get_contract_by_id(contract_id)
    if not contract:
        return error_response("Contract not found", 404)
    data = request.get_json(force=True)
    allowed = {"contract_ref", "contract_date", "responsible", "notes"}
    sets = ", ".join(f"{k} = %s" for k in data if k in allowed)
    vals = [data[k] for k in data if k in allowed]
    if not sets:
        return error_response("No valid fields to update")
    vals.append(contract_id)
    _repo.execute(f"UPDATE facturare_contracts SET {sets}, updated_at = now() WHERE id = %s", tuple(vals))
    return jsonify({"success": True})


@facturare_bp.route("/facturare/api/contracts/<int:contract_id>", methods=["DELETE"])
@login_required
@handle_api_errors
def api_delete_contract(contract_id):
    if not _check_perm("delete"):
        return error_response("Permission denied", 403)
    anexas = _repo.list_anexas_by_contract(contract_id)
    if anexas:
        return error_response("Cannot delete contract with existing anexas. Delete anexas first.", 409)
    _repo.delete_contract(contract_id)
    return jsonify({"success": True})


# ── Anexas ───────────────────────────────────────────────────────

@facturare_bp.route("/facturare/api/contracts/<int:contract_id>/anexas")
@login_required
@handle_api_errors
def api_list_anexas(contract_id):
    if not _check_perm("view"):
        return error_response("Permission denied", 403)
    contract = _repo.get_contract_by_id(contract_id)
    if not contract:
        return error_response("Contract not found", 404)

    anexas = _repo.list_anexas_by_contract(contract_id)
    result = []
    for a in anexas:
        lines = _repo.get_lines_by_anexa(a["id"])
        invoices = _repo.get_invoices_by_anexa(a["id"])
        total_value = sum(float(l["selling_price_eur"]) for l in lines)

        types = list({inv["invoice_type"] for inv in invoices})
        if "FINAL" in types:
            stage = "COMPLETE"
        elif "STORNO" in types:
            stage = "STORNO"
        elif "INVOICE" in types:
            stage = "INVOICE"
        elif "PROFORMA" in types:
            stage = "PROFORMA"
        else:
            stage = "NEW"

        status = a.get("status") or "NEW"

        proformas_total = sum(float(inv["total_amount_eur"]) for inv in invoices if inv["invoice_type"] == "PROFORMA")
        invoiced_total = sum(float(inv["total_amount_eur"]) for inv in invoices if inv["invoice_type"] == "INVOICE")
        pct_proforma = round((proformas_total / total_value * 100), 1) if total_value else 0
        pct_invoiced = round((invoiced_total / total_value * 100), 1) if total_value else 0

        # Per-line coverage stats
        import json as _json
        all_line_ids = {l["id"] for l in lines}
        proforma_line_ids = set()
        invoiced_line_ids = set()
        for inv in invoices:
            raw_lids = inv.get("line_ids")
            if isinstance(raw_lids, str):
                raw_lids = _json.loads(raw_lids)
            covered = set(raw_lids) if raw_lids else all_line_ids
            if inv["invoice_type"] == "PROFORMA":
                proforma_line_ids |= covered
            elif inv["invoice_type"] == "INVOICE":
                invoiced_line_ids |= covered

        result.append({
            "id": a["id"], "anexa_number": a["anexa_number"],
            "notes": a.get("notes"),
            "line_count": len(lines), "total_value": total_value,
            "proformas_total": proformas_total, "invoiced_total": invoiced_total,
            "pct_proforma": pct_proforma, "pct_invoiced": pct_invoiced,
            "invoice_count": len(invoices), "stage": stage, "status": status, "types": types,
            "archived": bool(a.get("archived")),
            "archive_after": str(a["archive_after"]) if a.get("archive_after") else None,
            "lines_with_proforma": len(proforma_line_ids | invoiced_line_ids),
            "lines_invoiced": len(invoiced_line_ids),
            "created_at": str(a["created_at"]) if a.get("created_at") else None,
        })
    return jsonify({"anexas": result, "contract": {
        "id": contract["id"], "contract_ref": contract["contract_ref"],
        "supplier_id": contract["supplier_id"], "customer_id": contract["customer_id"],
    }})


ANEXA_STATUSES = ['NEW', 'IN_PROGRESS', 'PAID', 'PROCESSED']

@facturare_bp.route("/facturare/api/anexas/<int:anexa_id>/status", methods=["PATCH"])
@login_required
@handle_api_errors
def api_update_anexa_status(anexa_id):
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    status = data.get("status", "").upper()
    if status not in ANEXA_STATUSES:
        return error_response(f"Invalid status. Must be one of: {', '.join(ANEXA_STATUSES)}")
    _repo.execute("UPDATE facturare_anexas SET status = %s, updated_at = now() WHERE id = %s", (status, anexa_id))
    return jsonify({"success": True, "status": status})


@facturare_bp.route("/facturare/api/anexas/<int:anexa_id>/archive", methods=["PATCH"])
@login_required
@handle_api_errors
def api_toggle_archive(anexa_id):
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    archived = bool(data.get("archived", True))
    if archived:
        _repo.archive_anexa_now(anexa_id)
    else:
        _repo.unarchive_anexa(anexa_id)
    _invalidate_doc_items_cache()
    return jsonify({"success": True, "archived": archived})


@facturare_bp.route("/facturare/api/contracts/<int:contract_id>/archive", methods=["PATCH"])
@login_required
@handle_api_errors
def api_toggle_contract_archive(contract_id):
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    archived = bool(data.get("archived", True))
    if archived:
        _repo.archive_contract_now(contract_id)
    else:
        _repo.unarchive_contract(contract_id)
    _invalidate_doc_items_cache()
    return jsonify({"success": True, "archived": archived})


@facturare_bp.route("/facturare/api/contracts/<int:contract_id>/anexas", methods=["POST"])
@login_required
@handle_api_errors
def api_create_anexa(contract_id):
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    contract = _repo.get_contract_by_id(contract_id)
    if not contract:
        return error_response("Contract not found", 404)

    data = request.get_json(force=True)
    data["contract_id"] = contract_id
    try:
        req = AnexaCreateRequest(**data)
    except Exception as e:
        return error_response(str(e))

    # Check uniqueness
    existing = _repo.get_anexa_by_contract_and_number(contract_id, req.anexa_number)
    if existing:
        return error_response(f"Anexa {req.anexa_number} already exists for this contract", 409)

    anexa_row = _repo.create_anexa(
        contract_id=contract_id, anexa_number=req.anexa_number,
        notes=req.notes, created_by=current_user.id,
    )

    for idx, line in enumerate(req.lines, start=1):
        _repo.create_anexa_line(
            anexa_id=anexa_row["id"], line_number=idx,
            model=line.model, list_price_eur=line.list_price_eur,
            selling_price_eur=line.selling_price_eur, qty=line.qty,
            nr_comanda=line.nr_comanda, vin=line.vin, culoare=line.culoare,
        )

    return jsonify({"success": True, "anexa": {"id": anexa_row["id"], "anexa_number": anexa_row["anexa_number"]}}), 201


# ── Anexa import from Excel ───────────────────────────────────────

@facturare_bp.route("/facturare/api/contracts/<int:contract_id>/anexas/import", methods=["POST"])
@login_required
@handle_api_errors
def api_import_anexa(contract_id):
    """Create an Anexa by importing vehicles from an Anexa XLSX file."""
    from .loaders.anexa import load_anexa

    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    contract = _repo.get_contract_by_id(contract_id)
    if not contract:
        return error_response("Contract not found", 404)

    anexa_file = request.files.get("anexa")
    if not anexa_file:
        return error_response("File required")

    anexa_number = request.form.get("anexa_number")
    if not anexa_number:
        return error_response("anexa_number required")

    notes = request.form.get("notes") or None

    # Check uniqueness
    existing = _repo.get_anexa_by_contract_and_number(contract_id, int(anexa_number))
    if existing:
        return error_response(f"Anexa {anexa_number} already exists for this contract", 409)

    import openpyxl, io as _io
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(anexa_file.read()), data_only=True)
        ws = wb.active
        # Find header row
        headers = []
        header_row = 1
        for r in range(1, min(ws.max_row + 1, 10)):
            row_vals = [str(c.value or "").strip().lower() for c in ws[r]]
            if any(h in " ".join(row_vals) for h in ["comanda", "model", "nr."]):
                headers = row_vals
                header_row = r
                break
        if not headers:
            return error_response("Could not find header row with Nr. Comanda / Model columns")

        # Map columns
        col_map = {}
        for i, h in enumerate(headers):
            if "comanda" in h or "nr." in h: col_map["comanda"] = i
            elif "model" in h: col_map["model"] = i
            elif "cul" in h or "color" in h: col_map["culoare"] = i
            elif "vin" in h: col_map["vin"] = i
            elif ("pret" in h and "lista" in h) or ("list" in h and "price" in h): col_map["list_price"] = i
            elif ("pret" in h and "vanz" in h) or ("sell" in h and "price" in h) or h == "pret vanzare": col_map["selling_price"] = i

        if "model" not in col_map:
            return error_response("Missing 'Model' column in header")

        parsed_lines = []
        for r in range(header_row + 1, ws.max_row + 1):
            row = [c.value for c in ws[r]]
            # Pad row to avoid index errors
            while len(row) < len(headers):
                row.append(None)
            model_val = row[col_map["model"]] if "model" in col_map else None
            model = str(model_val or "").strip()
            if not model:
                continue

            def _get(key):
                if key not in col_map: return None
                v = row[col_map[key]]
                return str(v).strip() if v is not None else None

            def _getf(key):
                if key not in col_map: return 0
                v = row[col_map[key]]
                try: return float(v) if v else 0
                except (ValueError, TypeError): return 0

            parsed_lines.append({
                "nr_comanda": _get("comanda"),
                "model": model,
                "culoare": _get("culoare"),
                "vin": _get("vin") or None,
                "list_price": _getf("list_price"),
                "selling_price": _getf("selling_price"),
            })
        wb.close()
    except Exception as e:
        logger.exception("Anexa import parse error")
        return error_response(f"Parse error: {e}")

    if not parsed_lines:
        return error_response("No vehicles found in file")

    anexa_row = _repo.create_anexa(
        contract_id=contract_id, anexa_number=int(anexa_number),
        notes=notes, created_by=current_user.id,
    )

    for idx, pl in enumerate(parsed_lines, start=1):
        _repo.create_anexa_line(
            anexa_id=anexa_row["id"], line_number=idx,
            model=pl["model"],
            list_price_eur=Decimal(str(pl["list_price"])),
            selling_price_eur=Decimal(str(pl["selling_price"] or pl["list_price"])),
            qty=1,
            nr_comanda=pl["nr_comanda"],
            vin=pl["vin"], culoare=pl["culoare"],
        )

    return jsonify({
        "success": True,
        "anexa": {"id": anexa_row["id"], "anexa_number": anexa_row["anexa_number"]},
        "lines_imported": len(parsed_lines),
    }), 201


# ── Delete anexa (only if no invoices) ───────────────────────────

@facturare_bp.route("/facturare/api/anexas/<int:anexa_id>", methods=["DELETE"])
@login_required
@handle_api_errors
def api_delete_anexa(anexa_id):
    """Delete an anexa. Only allowed if no invoices have been issued."""
    if not _check_perm("delete"):
        return error_response("Permission denied", 403)
    anexa = _repo.get_anexa_by_id(anexa_id)
    if not anexa:
        return error_response("Anexa not found", 404)
    invoices = _repo.get_invoices_by_anexa(anexa_id)
    if invoices:
        return error_response("Cannot delete anexa — invoices already exist. Delete all invoices first.", 409)
    _repo.delete_anexa(anexa_id)
    return jsonify({"success": True})


# ── Anexa line CRUD ──────────────────────────────────────────────

@facturare_bp.route("/facturare/api/anexa-lines", methods=["POST"])
@login_required
@handle_api_errors
def api_create_anexa_line():
    """Add a vehicle to an existing anexa (only if no invoices yet)."""
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    anexa_id = data.get("anexa_id")
    if not anexa_id:
        return error_response("anexa_id required")

    # Check no invoices exist
    invoices = _repo.get_invoices_by_anexa(int(anexa_id))
    if invoices:
        return error_response("Cannot add vehicles after invoices have been issued", 409)

    row = _repo.create_anexa_line(
        anexa_id=int(anexa_id),
        line_number=int(data.get("line_number", 1)),
        model=data.get("model", ""),
        list_price_eur=Decimal(str(data.get("list_price_eur", 0))),
        selling_price_eur=Decimal(str(data.get("selling_price_eur", 0))),
        qty=1,
        nr_comanda=data.get("nr_comanda"),
        vin=data.get("vin"),
        culoare=data.get("culoare"),
    )
    return jsonify({"success": True, "line": _line_to_dict(row)}), 201


@facturare_bp.route("/facturare/api/anexa-lines/<int:line_id>", methods=["DELETE"])
@login_required
@handle_api_errors
def api_delete_anexa_line(line_id):
    """Delete a vehicle from an anexa (only if no invoices yet)."""
    if not _check_perm("delete"):
        return error_response("Permission denied", 403)

    # Find the line's anexa and check for invoices
    line_row = _repo.query_one("SELECT * FROM facturare_anexa_lines WHERE id = %s", (line_id,))
    if not line_row:
        return error_response("Line not found", 404)

    invoices = _repo.get_invoices_by_anexa(line_row["anexa_id"])
    if invoices:
        return error_response("Cannot remove vehicles after invoices have been issued", 409)

    _repo.execute("DELETE FROM facturare_anexa_lines WHERE id = %s", (line_id,))
    return jsonify({"success": True})


# ── Anexa line update (VIN, etc.) ────────────────────────────────

@facturare_bp.route("/facturare/api/anexa-lines/<int:line_id>", methods=["PATCH"])
@login_required
@handle_api_errors
def api_update_anexa_line(line_id):
    """Update an anexa line (e.g., add VIN)."""
    if not _check_perm("edit"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    if not data:
        return error_response("Request body required")

    allowed_fields = {"vin", "nr_comanda", "culoare", "model", "selling_price_eur", "list_price_eur"}
    updates = {k: v for k, v in data.items() if k in allowed_fields}
    if not updates:
        return error_response("No valid fields to update")

    row = _repo.update_anexa_line(line_id, **updates)
    if not row:
        return error_response("Line not found", 404)
    return jsonify({"success": True, "line": _line_to_dict(row)})


# ── Anexa detail ─────────────────────────────────────────────────

@facturare_bp.route("/facturare/api/anexas/<int:anexa_id>")
@login_required
@handle_api_errors
def api_get_anexa_detail(anexa_id):
    if not _check_perm("view"):
        return error_response("Permission denied", 403)

    anexa = _repo.get_anexa_by_id(anexa_id)
    if not anexa:
        return error_response("Anexa not found", 404)

    contract = _repo.get_contract_by_id(anexa["contract_id"])
    raw_lines = _repo.get_lines_by_anexa(anexa_id)
    lines = [_line_to_dict(l) for l in raw_lines]
    # Keep line_number order (raw_lines is ORDER BY line_number) so the per-car
    # display number below (base_no + idx) matches the PDF renderer, which numbers
    # cars in line_number order. A set() here scrambled the order and produced
    # UI invoice numbers that disagreed with the PDF sent to the client.
    all_line_ids = [l["id"] for l in lines]
    invoices = [_inv_to_dict(inv) for inv in _repo.get_invoices_by_anexa(anexa_id)]

    # Enrich storno invoices with the IDs of invoices they reverse
    storno_ids = [inv["id"] for inv in invoices if inv["invoice_type"] == "STORNO"]
    if storno_ids:
        ph = ",".join(["%s"] * len(storno_ids))
        links = _repo.query_all(
            f"SELECT source_invoice_id, target_invoice_id FROM facturare_invoice_links "
            f"WHERE target_invoice_id IN ({ph}) AND link_type = 'REVERSES'",
            tuple(storno_ids))
        storno_reversal_map = {}  # storno_id -> [reversed_invoice_id, ...]
        for lnk in links:
            storno_reversal_map.setdefault(lnk["target_invoice_id"], []).append(lnk["source_invoice_id"])
        for inv in invoices:
            if inv["invoice_type"] == "STORNO":
                inv["reversed_invoice_ids"] = storno_reversal_map.get(inv["id"], [])

    next_actions = _sm.get_next_actions(anexa_id)
    unpaired_raw = _sm.get_unpaired_proformas(anexa_id)
    unpaired = []
    for r in unpaired_raw:
        import json as _json
        raw_lids = r.get("line_ids")
        if isinstance(raw_lids, str):
            raw_lids = _json.loads(raw_lids)
        unpaired.append({
            "sequence_number": r["sequence_number"],
            "total_amount_eur": float(r["total_amount_eur"]),
            "invoice_number": r.get("invoice_number"),
            "line_ids": raw_lids,
        })

    # Get names
    sup = _repo.query_one("SELECT company FROM companies WHERE id = %s", (contract["supplier_id"],))
    cust = _repo.query_one("SELECT display_name FROM crm_clients WHERE id = %s", (contract["customer_id"],))

    # Compute per-line invoicing coverage + per-line EUR amounts
    # line_ids=null means ALL lines are covered by that invoice
    line_prices = {l["id"]: l["selling_price_eur"] for l in lines}
    line_coverage = {}  # line_id -> list of {invoice_id, invoice_type, sequence_number}
    line_proforma_eur = {l["id"]: 0.0 for l in lines}
    line_invoiced_eur = {l["id"]: 0.0 for l in lines}
    for inv in invoices:
        covered = inv.get("line_ids") or list(all_line_ids)  # null = all
        doc_mode = inv.get("doc_mode", "per_car")
        base_no = inv.get("invoice_number")
        # Stored per-car document numbers for this invoice (facturare_document_numbers).
        docnum = _repo.get_document_number_map(inv["id"])
        # Per-car share follows the invoice's split_mode (equal = total/N, matching
        # the PDF; proportional = price-sliced with the snapped percent, never 0).
        # Cents vs whole EUR follows the invoice's own round_decimals.
        covered_total = sum(line_prices.get(lid, 0) for lid in covered)
        inv_total = float(inv["total_amount_eur"])
        split_mode = inv.get("split_mode", "equal")
        quant = _quant_for(inv)
        for idx, lid in enumerate(covered):
            if lid not in line_coverage:
                line_coverage[lid] = []
            # Closing slice books the residual so advance + rest reconcile to the
            # car's price (2288 + 43462 = 45750, not …43463 = 45751).
            priors = _prior_car_fractions(invoices, inv, lid, line_prices, all_line_ids)
            share = _coverage_share(inv_total, line_prices.get(lid, 0), covered_total,
                                    len(covered), split_mode, priors, quant)
            share_ron = 0
            # Per-vehicle document number: prefer the stored number; fall back
            # to the pre-backfill derivation (matches PDF renderer logic: start_no + idx).
            fallback_no = base_no + idx if base_no is not None and doc_mode != 'single_doc' and len(covered) > 1 else base_no
            display_no = _resolve_doc_no(docnum, lid, fallback_no)
            line_coverage[lid].append({
                "invoice_id": inv["id"],
                "invoice_type": inv["invoice_type"],
                "sequence_number": inv.get("sequence_number", 1),
                "amount_eur": round(share, 2),
                "amount_ron": round(share_ron, 2),
                "invoice_number": display_no,
                "kurs_applied": float(inv["kurs_applied"]) if inv.get("kurs_applied") else None,
                "issued_date": str(inv["issued_date"]) if inv.get("issued_date") else None,
            })
            if inv["invoice_type"] == "PROFORMA":
                line_proforma_eur[lid] = line_proforma_eur.get(lid, 0) + share
            elif inv["invoice_type"] == "INVOICE":
                line_invoiced_eur[lid] = line_invoiced_eur.get(lid, 0) + share
            elif inv["invoice_type"] == "STORNO":
                # Storno reverses: subtract from both proforma and invoiced
                line_proforma_eur[lid] = line_proforma_eur.get(lid, 0) - abs(share)
                line_invoiced_eur[lid] = line_invoiced_eur.get(lid, 0) - abs(share)

    # Enrich lines with coverage info + per-line amounts
    for line in lines:
        cov = line_coverage.get(line["id"], [])
        has_invoice = any(c["invoice_type"] == "INVOICE" for c in cov)
        has_proforma = any(c["invoice_type"] == "PROFORMA" for c in cov)
        if has_invoice:
            line["status"] = "INVOICED"
        elif has_proforma:
            line["status"] = "PROFORMA"
        else:
            line["status"] = "NONE"
        line["covered_by"] = cov
        line["proforma_eur"] = round(max(line_proforma_eur.get(line["id"], 0), 0), 2)
        line["invoiced_eur"] = round(max(line_invoiced_eur.get(line["id"], 0), 0), 2)

    # Compute remaining proforma capacity (stornos free up capacity)
    anexa_total = sum(float(l["selling_price_eur"]) for l in raw_lines)
    proformas_total = sum(inv["total_amount_eur"] for inv in invoices if inv["invoice_type"] == "PROFORMA")
    storno_freed = sum(abs(inv["total_amount_eur"]) for inv in invoices if inv["invoice_type"] == "STORNO")
    remaining_eur = anexa_total - float(proformas_total) + float(storno_freed)

    # Line-level stats
    lines_with_proforma = sum(1 for l in lines if l["status"] in ("PROFORMA", "INVOICED"))
    lines_invoiced = sum(1 for l in lines if l["status"] == "INVOICED")

    return jsonify({
        "anexa_id": anexa_id,
        "anexa_number": anexa["anexa_number"],
        "contract_ref": contract["contract_ref"],
        "supplier_name": sup["company"] if sup else "",
        "customer_name": cust["display_name"] if cust else "",
        "lines": lines,
        "invoices": invoices,
        "next_actions": next_actions,
        "unpaired_proformas": unpaired,
        "anexa_total_eur": anexa_total,
        "proformas_total_eur": float(proformas_total),
        "remaining_eur": remaining_eur,
        "lines_with_proforma": lines_with_proforma,
        "lines_invoiced": lines_invoiced,
        "lines_total": len(lines),
    })


# ── Invoice lifecycle endpoints ──────────────────────────────────

@facturare_bp.route("/facturare/api/invoices/proforma", methods=["POST"])
@login_required
@handle_api_errors
def api_issue_proforma():
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    try:
        req = ProformaCreateRequest(**data)
    except Exception as e:
        return error_response(str(e))
    try:
        inv = _sm.issue_proforma(
            anexa_id=req.anexa_id, amount_eur=req.amount_eur,
            split_mode=req.split_mode,
            invoice_number=req.invoice_number, issued_date=req.issued_date,
            intocmit_de=req.intocmit_de, notes=req.notes,
            line_ids=req.line_ids,
            created_by_user_id=current_user.id,
            doc_mode=req.doc_mode,
            round_decimals=req.round_decimals,
        )
    except InvoiceStateMachineError as e:
        return error_response(str(e), 409)
    _invalidate_doc_items_cache()
    return jsonify({"success": True, "invoice": _inv_to_dict(_repo.get_invoice_by_id(inv.id))}), 201


@facturare_bp.route("/facturare/api/invoices/invoice", methods=["POST"])
@login_required
@handle_api_errors
def api_issue_invoice():
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    try:
        req = InvoiceCreateRequest(**data)
    except Exception as e:
        return error_response(str(e))
    try:
        inv = _sm.issue_invoice(
            anexa_id=req.anexa_id, sequence_number=req.sequence_number,
            invoice_number=req.invoice_number, issued_date=req.issued_date,
            intocmit_de=req.intocmit_de, notes=req.notes,
            created_by_user_id=current_user.id,
            doc_mode=req.doc_mode,
            manual_kurs=req.kurs,
            round_decimals=req.round_decimals,
        )
    except InvoiceStateMachineError as e:
        return error_response(str(e), 409)
    _invalidate_doc_items_cache()
    return jsonify({"success": True, "invoice": _inv_to_dict(_repo.get_invoice_by_id(inv.id))}), 201


@facturare_bp.route("/facturare/api/invoices/storno", methods=["POST"])
@login_required
@handle_api_errors
def api_issue_storno():
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    try:
        req = StornoCreateRequest(**data)
    except Exception as e:
        return error_response(str(e))
    try:
        inv = _sm.issue_storno(
            anexa_id=req.anexa_id, invoice_number=req.invoice_number,
            issued_date=req.issued_date, intocmit_de=req.intocmit_de,
            notes=req.notes, line_ids=req.line_ids,
            target_invoice_ids=req.target_invoice_ids,
            created_by_user_id=current_user.id,
            manual_kurs=req.kurs,
            round_decimals=req.round_decimals,
        )
    except InvoiceStateMachineError as e:
        return error_response(str(e), 409)
    _invalidate_doc_items_cache()
    return jsonify({"success": True, "invoice": _inv_to_dict(_repo.get_invoice_by_id(inv.id))}), 201


@facturare_bp.route("/facturare/api/invoices/final", methods=["POST"])
@login_required
@handle_api_errors
def api_issue_final():
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    try:
        req = FinalCreateRequest(**data)
    except Exception as e:
        return error_response(str(e))
    try:
        inv = _sm.issue_final(
            anexa_id=req.anexa_id, invoice_number=req.invoice_number,
            issued_date=req.issued_date, intocmit_de=req.intocmit_de,
            notes=req.notes, line_ids=req.line_ids,
            created_by_user_id=current_user.id,
            manual_kurs=req.kurs,
            round_decimals=req.round_decimals,
        )
    except InvoiceStateMachineError as e:
        return error_response(str(e), 409)
    _invalidate_doc_items_cache()
    return jsonify({"success": True, "invoice": _inv_to_dict(_repo.get_invoice_by_id(inv.id))}), 201


PAYMENT_STATUSES = ['UNPAID', 'PAID', 'PARTIAL']

@facturare_bp.route("/facturare/api/invoices/<int:invoice_id>/payment-status", methods=["PATCH"])
@login_required
@handle_api_errors
def api_update_payment_status(invoice_id):
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    status = data.get("payment_status", "").upper()
    if status not in PAYMENT_STATUSES:
        return error_response(f"Invalid status. Must be one of: {', '.join(PAYMENT_STATUSES)}")
    _repo.execute("UPDATE facturare_invoices SET payment_status = %s, updated_at = now() WHERE id = %s", (status, invoice_id))
    _invalidate_doc_items_cache()
    return jsonify({"success": True, "payment_status": status})


# ── Delete invoice (last only) ───────────────────────────────────

@facturare_bp.route("/facturare/api/invoices/<int:invoice_id>", methods=["DELETE"])
@login_required
@handle_api_errors
def api_delete_invoice(invoice_id):
    if not _check_perm("delete"):
        return error_response("Permission denied", 403)
    inv_row = _repo.get_invoice_by_id(invoice_id)
    if not inv_row:
        return error_response("Invoice not found", 404)

    all_invs = _repo.get_invoices_by_anexa(inv_row["anexa_id"])
    if not all_invs:
        return error_response("No invoices found", 404)

    last = all_invs[-1]
    if last["id"] != invoice_id:
        return error_response(
            f'Only the last document can be deleted. Delete "{TYPE_LABELS.get(last["invoice_type"], last["invoice_type"])} '
            f'#{last.get("sequence_number", 1)}" first.', 409)

    _repo.delete_invoice(invoice_id)
    _invalidate_doc_items_cache()
    return jsonify({"success": True})


# ── Anexa status export (Excel) ──────────────────────────────────

@facturare_bp.route("/facturare/api/anexas/<int:anexa_id>/status-export.xlsx")
@login_required
@handle_api_errors
def api_anexa_status_export(anexa_id):
    """Export per-car invoicing status for an anexa as styled XLSX."""
    if not _check_perm("view"):
        return error_response("Permission denied", 403)

    import io as _io
    import json as _json
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    anexa = _repo.get_anexa_by_id(anexa_id)
    if not anexa:
        return error_response("Anexa not found", 404)
    contract = _repo.get_contract_by_id(anexa["contract_id"])
    lines = _repo.get_lines_by_anexa(anexa_id)
    invoices = _repo.get_invoices_by_anexa(anexa_id)

    sup = _repo.query_one("SELECT company FROM companies WHERE id = %s", (contract["supplier_id"],))
    cust = _repo.query_one("SELECT display_name FROM crm_clients WHERE id = %s", (contract["customer_id"],))
    supplier_name = sup["company"] if sup else ""
    customer_name = cust["display_name"] if cust else ""

    prices = {l["id"]: float(l["selling_price_eur"]) for l in lines}
    nr_map = {l["id"]: l.get("nr_comanda", "") for l in lines}
    model_map = {l["id"]: l.get("model", "") for l in lines}
    total_selling = sum(prices.values()) or 1

    # Compute per-line amounts
    line_data = {lid: {"prof": 0, "inv": 0, "sto": 0, "fin": 0} for lid in prices}
    for inv in invoices:
        raw = inv.get("line_ids")
        if isinstance(raw, str):
            raw = _json.loads(raw)
        inv_lids = raw if raw else list(prices.keys())
        covered_total = sum(prices.get(l, 0) for l in inv_lids)
        if not covered_total:
            continue
        amt = float(inv["total_amount_eur"])
        quant = _quant_for(inv)
        for lid in inv_lids:
            if lid not in line_data:
                continue
            share = _round_half_up(abs(amt) * (prices[lid] / covered_total), quant)
            if inv["invoice_type"] == "PROFORMA":
                line_data[lid]["prof"] += share
            elif inv["invoice_type"] == "INVOICE":
                line_data[lid]["inv"] += share
            elif inv["invoice_type"] == "STORNO":
                line_data[lid]["sto"] += share
            elif inv["invoice_type"] == "FINAL":
                line_data[lid]["fin"] += share

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Anexa {anexa['anexa_number']}"

    bold = Font(bold=True, size=10)
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    complete_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    storno_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    prof90_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    num_fmt = "#,##0"

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{contract['contract_ref']} | Anexa {anexa['anexa_number']} | {supplier_name} → {customer_name}"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A2:I2")
    ws["A2"] = f"{len(lines)} cars | {int(total_selling):,} EUR"
    ws["A2"].font = Font(size=10, italic=True)

    # Headers
    headers = ["Nr Comanda", "Model", "Selling EUR", "Proforma", "Invoiced", "Storno", "Final", "Remaining", "Stage"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin

    # Data rows
    totals = [0] * 6
    row = 5
    for lid in sorted(prices.keys()):
        d = line_data[lid]
        remain = int(prices[lid]) - d["prof"] + d["sto"]
        if d["fin"] > 0:
            stage = "COMPLETE"
        elif d["sto"] > 0 and d["fin"] == 0:
            stage = "STORNO (awaiting FINAL)"
        elif d["inv"] > 0 and d["inv"] >= int(prices[lid]):
            stage = "FULLY INVOICED"
        elif d["prof"] > 0 and d["prof"] >= int(prices[lid]) - 1:
            stage = "PROFORMA COMPLETE"
        elif d["prof"] > 0:
            stage = "PARTIAL"
        else:
            stage = "NEW"

        vals = [nr_map.get(lid, ""), model_map.get(lid, ""), int(prices[lid]),
                d["prof"] or None, d["inv"] or None, d["sto"] or None, d["fin"] or None, remain, stage]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = thin
            if col >= 3 and col <= 8 and v is not None:
                c.number_format = num_fmt
                c.alignment = Alignment(horizontal="right")
            if stage == "COMPLETE":
                c.fill = complete_fill
            elif stage == "FULLY INVOICED":
                c.fill = storno_fill
            elif stage == "PROFORMA COMPLETE":
                c.fill = prof90_fill

        totals[0] += int(prices[lid])
        totals[1] += d["prof"]
        totals[2] += d["inv"]
        totals[3] += d["sto"]
        totals[4] += d["fin"]
        totals[5] += remain
        row += 1

    # Total row
    total_vals = ["", "TOTAL"] + totals + [f"{len(lines)} cars"]
    for col, v in enumerate(total_vals, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = bold
        c.fill = total_fill
        c.border = thin
        if col >= 3 and col <= 8 and isinstance(v, (int, float)):
            c.number_format = num_fmt
            c.alignment = Alignment(horizontal="right")

    # Column widths
    for i, w in enumerate([12, 38, 14, 12, 12, 12, 12, 12, 22], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Legend
    row += 2
    ws.cell(row=row, column=1, value="Legend:").font = bold
    row += 1
    ws.cell(row=row, column=1, value="Complete").fill = complete_fill
    row += 1
    ws.cell(row=row, column=1, value="Fully Invoiced (awaiting Storno)").fill = storno_fill
    row += 1
    ws.cell(row=row, column=1, value="Proforma Complete (awaiting Invoice)").fill = prof90_fill

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    dl_name = f"{contract['contract_ref']}_Anexa{anexa['anexa_number']}_Status.xlsx"
    from flask import send_file
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=dl_name)


# ── User search ──────────────────────────────────────────────────

@facturare_bp.route("/facturare/api/users")
@login_required
@handle_api_errors
def api_search_users():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify({"users": []})
    rows = _repo.query_all(
        "SELECT id, name FROM users WHERE is_active = TRUE AND name ILIKE %s ORDER BY name LIMIT 10",
        (f"%{q}%",))
    return jsonify({"users": [{"id": r["id"], "name": r["name"]} for r in rows]})


# ── Konto configuration ─────────────────────────────────────────

@facturare_bp.route("/facturare/api/konto-config")
@login_required
@handle_api_errors
def api_get_konto_config():
    if not _check_perm("view"):
        return error_response("Permission denied", 403)
    rows = _repo.get_konto_config()
    return jsonify({"configs": [
        {"supplier_id": r["supplier_id"], "supplier_name": r.get("supplier_name", ""),
         "invoice_type": r["invoice_type"], "konto_debit": r.get("konto_debit") or "",
         "konto_credit": r.get("konto_credit") or "", "centru_gestiune": r.get("centru_gestiune") or "",
         "text_template": r.get("text_template") or ""}
        for r in rows
    ]})


@facturare_bp.route("/facturare/api/konto-config", methods=["PUT"])
@login_required
@handle_api_errors
def api_put_konto_config():
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    items = data.get("items", [])
    for item in items:
        _repo.upsert_konto_config(
            supplier_id=item["supplier_id"], invoice_type=item["invoice_type"],
            konto_debit=item.get("konto_debit") or None, konto_credit=item.get("konto_credit") or None,
            centru_gestiune=item.get("centru_gestiune") or None,
            text_template=item.get("text_template") or None,
            updated_by=current_user.id,
        )
    return jsonify({"success": True, "count": len(items)})


# ── Venituri Rules ───────────────────────────────────────────────

@facturare_bp.route("/facturare/api/venituri-rules")
@login_required
@handle_api_errors
def api_get_venituri_rules():
    if not _check_perm("view"):
        return error_response("Permission denied", 403)
    rows = _repo.get_venituri_rules()
    return jsonify({"rules": [
        {"id": r["id"], "supplier_id": r["supplier_id"], "supplier_name": r.get("supplier_name", ""),
         "comanda_prefix": r["comanda_prefix"], "konto_venituri": r["konto_venituri"],
         "kostenstelle": r["kostenstelle"]}
        for r in rows
    ]})


@facturare_bp.route("/facturare/api/venituri-rules", methods=["PUT"])
@login_required
@handle_api_errors
def api_put_venituri_rules():
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    data = request.get_json(force=True)
    items = data.get("items", [])
    for item in items:
        _repo.upsert_venituri_rule(
            supplier_id=item["supplier_id"], comanda_prefix=item["comanda_prefix"],
            konto_venituri=item["konto_venituri"], kostenstelle=item["kostenstelle"],
            updated_by=current_user.id,
        )
    return jsonify({"success": True, "count": len(items)})


@facturare_bp.route("/facturare/api/venituri-rules/<int:rule_id>", methods=["DELETE"])
@login_required
@handle_api_errors
def api_delete_venituri_rule(rule_id):
    if not _check_perm("add"):
        return error_response("Permission denied", 403)
    _repo.delete_venituri_rule(rule_id)
    return jsonify({"success": True})


@facturare_bp.route("/facturare/api/contracts/<int:contract_id>/accounting-summary")
@login_required
@handle_api_errors
def api_contract_accounting_summary(contract_id):
    """Return all accounting data relevant to a contract for display."""
    if not _check_perm("view"):
        return error_response("Permission denied", 403)

    contract = _repo.get_contract_by_id(contract_id)
    if not contract:
        return error_response("Contract not found", 404)

    # Supplier info (firmennr)
    supplier = _repo.query_one(
        "SELECT id, company, eurofib_klient_id FROM companies WHERE id = %s",
        (contract["supplier_id"],))

    # Customer konto debit
    customer = _repo.query_one(
        "SELECT id, display_name, eurofib_konto_debit FROM crm_clients WHERE id = %s",
        (contract["customer_id"],))

    firmennr = supplier.get("eurofib_klient_id") if supplier else None
    kd_map = customer.get("eurofib_konto_debit") if customer else None
    client_konto_debit = kd_map.get(str(firmennr)) if isinstance(kd_map, dict) and firmennr else None

    # Konto configs per invoice type
    konto_configs = {}
    for inv_type in ('INVOICE', 'STORNO', 'FINAL'):
        row = _repo.query_one(
            "SELECT konto_debit, konto_credit, centru_gestiune, text_template FROM facturare_konto_config WHERE supplier_id = %s AND invoice_type = %s",
            (contract["supplier_id"], inv_type))
        konto_configs[inv_type] = dict(row) if row else None

    # Venituri rules for this supplier
    venituri = _repo.query_all(
        "SELECT comanda_prefix, konto_venituri, kostenstelle FROM facturare_venituri_rules WHERE supplier_id = %s ORDER BY comanda_prefix",
        (contract["supplier_id"],))

    return jsonify({
        "firmennr": firmennr,
        "supplier_name": supplier["company"] if supplier else None,
        "customer_name": customer["display_name"] if customer else None,
        "client_konto_debit": client_konto_debit,
        "konto_configs": konto_configs,
        "venituri_rules": [dict(r) for r in venituri] if venituri else [],
    })


@facturare_bp.route("/facturare/api/contracts/<int:contract_id>/accounting-summary", methods=["PUT"])
@login_required
@handle_api_errors
def api_put_contract_accounting_summary(contract_id):
    """Update accounting data for a contract (firmennr, konto debit, konto configs, centru gestiune)."""
    if not _check_perm("add"):
        return error_response("Permission denied", 403)

    contract = _repo.get_contract_by_id(contract_id)
    if not contract:
        return error_response("Contract not found", 404)

    data = request.get_json(force=True)
    supplier_id = contract["supplier_id"]
    customer_id = contract["customer_id"]

    # Update firmennr (companies.eurofib_klient_id)
    if "firmennr" in data:
        _repo.execute(
            "UPDATE companies SET eurofib_klient_id = %s WHERE id = %s",
            (data["firmennr"] or None, supplier_id))

    # Update client konto debit (crm_clients.eurofib_konto_debit JSONB)
    if "client_konto_debit" in data and data.get("firmennr_key"):
        import json as _json
        _repo.execute(
            "UPDATE crm_clients SET eurofib_konto_debit = COALESCE(eurofib_konto_debit, '{}'::jsonb) || %s WHERE id = %s",
            (_json.dumps({str(data["firmennr_key"]): data["client_konto_debit"]}), customer_id))

    # Update konto configs per invoice type
    for inv_type in ("INVOICE", "STORNO", "FINAL"):
        key = f"konto_{inv_type.lower()}"
        if key in data:
            cfg = data[key]
            _repo.upsert_konto_config(
                supplier_id=supplier_id, invoice_type=inv_type,
                konto_debit=cfg.get("konto_debit") or None,
                konto_credit=cfg.get("konto_credit") or None,
                centru_gestiune=cfg.get("centru_gestiune") or None,
                text_template=cfg.get("text_template") or None,
                updated_by=current_user.id)

    return jsonify({"success": True})


# ── Individual document items (per car) ──────────────────────────

@facturare_bp.route("/facturare/api/document-items")
@login_required
@handle_api_errors
def api_document_items():
    """List individual per-car document items.

    Query params:
        type: PROFORMA or INVOICE (required)

    Returns one row per car per document, with calculated per-car amount.
    """
    if not _check_perm("view"):
        return error_response("Permission denied", 403)

    doc_types_raw = request.args.get("type", "").upper()
    allowed = {"PROFORMA", "INVOICE", "STORNO", "FINAL"}
    doc_types = [t.strip() for t in doc_types_raw.split(",") if t.strip() in allowed]
    if not doc_types:
        return error_response("type must be PROFORMA, INVOICE, STORNO or comma-separated")

    cache_key = ",".join(sorted(doc_types))
    cached = _doc_items_cache.get(cache_key)
    if cached and (time.time() - cached[0]) < _DOC_ITEMS_TTL:
        return jsonify({"items": cached[1], "total": len(cached[1])})

    placeholders = ",".join(["%s"] * len(doc_types))
    rows = _repo.query_all(
        f"""SELECT i.id AS invoice_id, i.invoice_type, i.sequence_number,
                  i.invoice_number, i.issued_date, i.total_amount_eur,
                  i.kurs_applied, i.intocmit_de, i.split_mode, i.notes, i.payment_status,
                  i.line_ids, i.round_decimals,
                  a.id AS anexa_id, a.anexa_number,
                  c.id AS contract_id, c.contract_ref,
                  c.supplier_id, c.customer_id,
                  comp.company AS supplier_name,
                  cl.display_name AS customer_name
           FROM facturare_invoices i
           JOIN facturare_anexas a ON a.id = i.anexa_id
           JOIN facturare_contracts c ON c.id = a.contract_id
           JOIN companies comp ON comp.id = c.supplier_id
           JOIN crm_clients cl ON cl.id = c.customer_id
           WHERE i.invoice_type IN ({placeholders})
             AND i.archived = {'TRUE' if request.args.get('archived') == 'true' else 'FALSE'}
           ORDER BY i.issued_date DESC, i.id DESC""",
        tuple(doc_types),
    )

    # Pre-fetch all lines for all referenced anexas in one query (avoid N+1)
    anexa_ids = list({inv["anexa_id"] for inv in rows})
    lines_cache = {}
    if anexa_ids:
        ph = ",".join(["%s"] * len(anexa_ids))
        all_lines = _repo.query_all(
            f"SELECT * FROM facturare_anexa_lines WHERE anexa_id IN ({ph}) ORDER BY anexa_id, line_number",
            tuple(anexa_ids))
        for l in all_lines:
            lines_cache.setdefault(l["anexa_id"], []).append(l)

    # Index same-type invoices per anexa so each car's closing slice can book the
    # cumulative-rounding residual (rest = selling − advances) instead of an
    # independent round that overshoots by 1 EUR. rows already hold every invoice
    # of the queried type(s), so a same-type sibling of `inv` is always present.
    for r in rows:
        r["id"] = r.get("invoice_id")
    invs_by_anexa_type = {}
    for r in rows:
        invs_by_anexa_type.setdefault((r["anexa_id"], r["invoice_type"]), []).append(r)

    items = []
    for inv in rows:
        import json as _json
        lines = lines_cache.get(inv["anexa_id"], [])
        anexa_price_by_line = {l["id"]: float(l["selling_price_eur"]) for l in lines}
        anexa_all_lids = [l["id"] for l in lines]
        same_type_siblings = invs_by_anexa_type.get((inv["anexa_id"], inv["invoice_type"]), [])
        # Filter to invoice's line_ids if set
        raw_lids = inv.get("line_ids")
        if raw_lids:
            if isinstance(raw_lids, str):
                raw_lids = _json.loads(raw_lids)
            lid_set = set(raw_lids)
            inv_lines = [l for l in lines if l["id"] in lid_set]
        else:
            inv_lines = lines
        total_selling = sum(float(l["selling_price_eur"]) for l in inv_lines) or 1
        total_amount = float(inv["total_amount_eur"])
        split_mode = inv.get("split_mode", "equal")
        start_no = inv.get("invoice_number") or inv["invoice_id"]
        # Stored per-car document numbers for this invoice, fetched once (not
        # per car) — facturare_document_numbers.
        docnum = _repo.get_document_number_map(inv["invoice_id"])
        quant = _quant_for(inv)

        for idx, l in enumerate(inv_lines):
            selling = float(l["selling_price_eur"])
            priors = _prior_car_fractions(
                same_type_siblings, inv, l["id"], anexa_price_by_line, anexa_all_lids)
            car_amount = _coverage_share(total_amount, selling, total_selling,
                                         len(inv_lines), split_mode, priors, quant)
            # Prefer the stored number for this car; fall back to the
            # pre-backfill positional derivation (start_no + idx).
            fallback_no = start_no + idx if start_no else None
            doc_number = _resolve_doc_no(docnum, l["id"], fallback_no)

            items.append({
                "invoice_id": inv["invoice_id"],
                "invoice_type": inv["invoice_type"],
                "sequence_number": inv["sequence_number"],
                "doc_number": doc_number,
                "car_index": idx,
                "issued_date": str(inv["issued_date"]) if inv.get("issued_date") else None,
                "kurs_applied": float(inv["kurs_applied"]) if inv.get("kurs_applied") else None,
                "intocmit_de": inv.get("intocmit_de"),
                "contract_ref": inv["contract_ref"],
                "anexa_number": inv["anexa_number"],
                "supplier_name": inv["supplier_name"],
                "customer_name": inv["customer_name"],
                "nr_comanda": l.get("nr_comanda"),
                "model": l["model"],
                "culoare": l.get("culoare"),
                "vin": l.get("vin"),
                "unit_price": selling,
                "doc_amount": car_amount,
                "notes": inv.get("notes"),
                "payment_status": inv.get("payment_status") or "UNPAID",
            })

    _doc_items_cache[cache_key] = (time.time(), items)
    return jsonify({"items": items, "total": len(items)})


# ── PDF generation ───────────────────────────────────────────────

@facturare_bp.route("/facturare/api/invoices/<int:invoice_id>/pdf")
@login_required
@handle_api_errors
def api_generate_pdf(invoice_id):
    from flask import send_file
    from .generators.proforma_pdf import ProformaPdfRenderer
    from .models import OrderLine
    from datetime import date as date_type
    import io

    if not _check_perm("view"):
        return error_response("Permission denied", 403)

    inv_row = _repo.get_invoice_by_id(invoice_id)
    if not inv_row:
        return error_response("Invoice not found", 404)

    anexa = _repo.get_anexa_by_id(inv_row["anexa_id"])
    contract = _repo.get_contract_by_id(anexa["contract_id"])
    all_lines = _repo.get_lines_by_anexa(anexa["id"])
    # All invoices on this anexa — cumulative context so a closing slice books the
    # residual (rest = selling − advances) instead of an independent round that
    # overshoots by 1 EUR. Keyed by id for the storno reversal lookup below.
    _anexa_invoices = list(_repo.get_invoices_by_anexa(anexa["id"]))
    _anexa_price_by_line = {l["id"]: float(l["selling_price_eur"]) for l in all_lines}
    _anexa_all_lids = [l["id"] for l in all_lines]
    _anexa_inv_by_id = {i["id"]: i for i in _anexa_invoices}
    # Filter to selected lines if line_ids is set on this invoice
    # Preserve line_ids order so PDF page numbering matches the UI
    inv_line_ids = inv_row.get("line_ids")
    if inv_line_ids:
        import json as _json
        if isinstance(inv_line_ids, str):
            inv_line_ids = _json.loads(inv_line_ids)
        _lid_map = {l["id"]: l for l in all_lines}
        lines = [_lid_map[lid] for lid in inv_line_ids if lid in _lid_map]
    else:
        lines = all_lines

    # Stored document numbers per car line_id (populated by the numbering
    # module). Consumers below prefer these over the positional derivation.
    docnum = _repo.get_document_number_map(invoice_id)

    # Build supplier/customer dicts
    sup_row = _repo.query_one(
        "SELECT company, vat, reg_no, iban, bank, swift, street, city, county FROM companies WHERE id = %s",
        (contract["supplier_id"],))
    cust_row = _repo.query_one(
        "SELECT display_name, nr_reg, street, city, country FROM crm_clients WHERE id = %s",
        (contract["customer_id"],))

    supplier = {
        "name": sup_row["company"] if sup_row else "",
        "address_lines": [sup_row.get("street"), f"{sup_row.get('city')}, jud. {sup_row.get('county')}" if sup_row.get("city") else None, "Romania"] if sup_row else [],
        "vat": (sup_row or {}).get("vat", ""), "reg_no": (sup_row or {}).get("reg_no", ""),
        "iban": ((sup_row or {}).get("iban") or "").split("+")[0].strip(),
        "bank": ((sup_row or {}).get("bank") or "").split(";")[0].strip(),
        "swift": (sup_row or {}).get("swift", ""),
    }
    supplier["address_lines"] = [a for a in supplier["address_lines"] if a]

    customer = {
        "name": cust_row["display_name"] if cust_row else "",
        "address_lines": [a for a in [(cust_row or {}).get("street"), (cust_row or {}).get("city"), (cust_row or {}).get("country")] if a],
        "vat": (cust_row or {}).get("nr_reg", ""),
    }

    inv_type_str = inv_row["invoice_type"]

    if inv_type_str == "STORNO":
        # Build storno groups: per car, one line per reversed invoice (10%, 90%, etc.)
        import json as _json
        # Only include invoices actually reversed by this storno (via links table)
        reversed_links = _repo.query_all(
            "SELECT source_invoice_id FROM facturare_invoice_links WHERE target_invoice_id = %s AND link_type = 'REVERSES'",
            (invoice_id,))
        reversed_inv_ids = {r["source_invoice_id"] for r in reversed_links} if reversed_links else None
        all_invoices = _repo.query_all(
            "SELECT * FROM facturare_invoices WHERE anexa_id = %s AND invoice_type = 'INVOICE' ORDER BY created_at",
            (anexa["id"],))
        if reversed_inv_ids:
            all_invoices = [inv for inv in all_invoices if inv["id"] in reversed_inv_ids]
        storno_line_set = set(inv_line_ids) if inv_line_ids else {l["id"] for l in all_lines}
        line_map = {l["id"]: l for l in all_lines}
        # Stored document numbers per reversed invoice, fetched once each (not
        # once per car) — used for the "Ref: Factura Nr." text below.
        reversed_docnum = {inv["id"]: _repo.get_document_number_map(inv["id"]) for inv in all_invoices}

        # Per car: collect reversed invoices and their per-car share
        storno_groups = []  # list of list[OrderLine] — one group per car
        storno_group_line_ids = []  # parallel list: car line_id for each group
        for l in lines:
            lid = l["id"]
            car_items = []
            selling = float(l["selling_price_eur"])
            for inv in all_invoices:
                raw = inv.get("line_ids")
                if isinstance(raw, str):
                    raw = _json.loads(raw)
                raw_list = raw if isinstance(raw, list) else None
                inv_lines = set(raw) if raw else {x["id"] for x in all_lines}
                if lid not in inv_lines:
                    continue
                # Per-car share of this invoice — reverse the residual the advance
                # actually booked (43462, not 43463) so the storno nets to zero.
                inv_total = float(inv["total_amount_eur"])
                inv_selling_sum = sum(float(line_map[x]["selling_price_eur"]) for x in inv_lines if x in line_map) or 1
                _priors = _prior_car_fractions(
                    _anexa_invoices, inv, lid, _anexa_price_by_line, _anexa_all_lids)
                car_share = _car_slice_eur(selling, _snap_pct(inv_total, inv_selling_sum), _priors, _quant_for(inv))
                base_no = inv.get("invoice_number") or inv["id"]
                inv_doc_mode = inv.get("doc_mode", "per_car")
                # Per-vehicle document number for the REVERSED invoice: prefer
                # its stored number for this car; fall back to the
                # pre-backfill derivation (matches PDF renderer logic: start_no + idx).
                if inv_doc_mode != 'single_doc' and raw_list and len(raw_list) > 1 and lid in raw_list:
                    fallback_no = base_no + raw_list.index(lid)
                else:
                    fallback_no = base_no
                inv_no = _resolve_doc_no(reversed_docnum.get(inv["id"], {}), lid, fallback_no)
                inv_date = inv.get("issued_date")
                date_fmt = ""
                if inv_date:
                    ds = str(inv_date)
                    if "-" in ds:
                        p = ds.split("-")
                        date_fmt = f"{p[2]}.{p[1]}.{p[0]}"
                    else:
                        date_fmt = ds
                pct = round(car_share / selling * 100) if selling else 0
                inv_kurs = float(inv["kurs_applied"]) if inv.get("kurs_applied") else None
                car_items.append(OrderLine(
                    comanda=int(l["nr_comanda"]) if l.get("nr_comanda") and str(l["nr_comanda"]).isdigit() else 0,
                    model=l["model"], culoare=l.get("culoare") or "",
                    list_price=float(l["list_price_eur"]), selling_price=selling,
                    advance=-car_share,
                    rest=None, vin=l.get("vin"), qty=l.get("qty", 1),
                    anexa_ref=f"Anexa {anexa['anexa_number']} / Contract {contract['contract_ref']}",
                    storno_description=f"Ref: Factura Nr. {inv_no} / {date_fmt} ({pct}%)",
                    kurs=inv_kurs,
                ))
            if car_items:
                storno_groups.append(car_items)
                storno_group_line_ids.append(lid)

        # Flat order_lines for fallback (single-doc mode etc.)
        order_lines = [item for group in storno_groups for item in group]
    else:
        storno_groups = None
        # Per-car order lines
        total_amount = float(inv_row["total_amount_eur"])
        split_mode = inv_row.get("split_mode", "equal")
        total_selling = sum(float(l["selling_price_eur"]) for l in lines) or 1
        quant = _quant_for(inv_row)

        order_lines = []
        for l in lines:
            selling = float(l["selling_price_eur"])
            if split_mode == "proportional" and total_selling > 0:
                priors = _prior_car_fractions(
                    _anexa_invoices, inv_row, l["id"], _anexa_price_by_line, _anexa_all_lids)
                car_advance = _car_slice_eur(
                    selling, _snap_pct(total_amount, total_selling), priors, quant)
            else:
                car_advance = _round_half_up(total_amount / max(len(lines), 1), quant)

            order_lines.append(OrderLine(
                comanda=int(l["nr_comanda"]) if l.get("nr_comanda") and str(l["nr_comanda"]).isdigit() else 0,
                model=l["model"], culoare=l.get("culoare") or "",
                list_price=float(l["list_price_eur"]), selling_price=selling,
                advance=car_advance,
                rest=selling, vin=l.get("vin"), qty=l.get("qty", 1),
                anexa_ref=f"Anexa {anexa['anexa_number']} / Contract {contract['contract_ref']}",
            ))

    issued_date = inv_row.get("issued_date")
    date_str = str(issued_date) if issued_date else date_type.today().strftime("%Y-%m-%d")
    start_no = inv_row.get("invoice_number") or inv_row["id"]
    cust_name = (cust_row["display_name"] if cust_row else "")
    if inv_type_str == "PROFORMA":
        filename = f"Proforma {cust_name} {start_no}"
    else:
        type_label = {"INVOICE": "advance", "STORNO": "storno", "FINAL": "final"}.get(inv_type_str, "")
        filename = f"Invoice {cust_name} {start_no} {type_label}"

    title_map = {
        "PROFORMA": ["FACTURA PROFORMA", "PROFORMA INVOICE"],
        "INVOICE":  ["FACTURA AVANS", "ADVANCE INVOICE"],
        "STORNO":   ["FACTURA STORNO", "STORNO INVOICE"],
        "FINAL":    ["FACTURA FINALA", "FINAL INVOICE"],
    }
    desc_map = {
        "PROFORMA": "1. ADVANCE PAYMENT",
        "INVOICE":  "1. ADVANCE PAYMENT",
        "STORNO":   "1. STORNO ADVANCE",
        "FINAL":    "",
    }

    base_note = inv_row.get("notes") or ""
    renderer = ProformaPdfRenderer(
        supplier=supplier, customer=customer,
        invoice_date=date_str,
        intocmit_de=inv_row.get("intocmit_de") or "",
        title_lines=title_map.get(inv_type_str, ["FACTURA", "INVOICE"]),
        description_prefix=desc_map.get(inv_type_str, "1."),
        note=base_note,
        kurs_applied=float(inv_row["kurs_applied"]) if inv_row.get("kurs_applied") else None,
    )

    doc_mode = inv_row.get("doc_mode", "per_car")

    # For INVOICE type: compute per-vehicle proforma reference number for notes
    import re as _re
    _linked_proforma_no = None
    if inv_type_str == "INVOICE" and doc_mode != "single_doc":
        linked = _repo.query_one(
            "SELECT i.invoice_number, i.doc_mode FROM facturare_invoice_links l "
            "JOIN facturare_invoices i ON i.id = l.source_invoice_id "
            "WHERE l.target_invoice_id = %s AND l.link_type = 'PRECEDES'",
            (invoice_id,))
        if linked and linked.get("doc_mode", "per_car") != "single_doc":
            _linked_proforma_no = linked.get("invoice_number")

    def _note_for_car(idx):
        if not _linked_proforma_no:
            return base_note
        return _re.sub(r'\(No:\s*\d+\)', f'(No: {_linked_proforma_no + idx})', base_note)
    mode = request.args.get("mode", "merged")
    # Storno: use multipage renderer with per-invoice line items
    if inv_type_str == "STORNO" and storno_groups:
        # Single car storno via ?car=N
        car_idx = request.args.get("car")
        if car_idx is not None and car_idx.isdigit():
            idx = int(car_idx)
            if 0 <= idx < len(storno_groups):
                storno_groups = [storno_groups[idx]]
                storno_group_line_ids = [storno_group_line_ids[idx]]
        # Per-car stored document number for THIS storno invoice; fall back to
        # today's start_no + page_idx derivation (render_storno_multipage's
        # own logic) when the map has no entry for that car.
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as rc
        buf = io.BytesIO()
        c = rc.Canvas(buf, pagesize=A4)
        for page_idx, car_items in enumerate(storno_groups):
            page_no = _resolve_doc_no(docnum, storno_group_line_ids[page_idx], start_no + page_idx)
            renderer._render_storno_page(c, page_no, car_items)
            c.showPage()
        c.save()
        pdf_bytes = buf.getvalue()
        return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf", as_attachment=True, download_name=f"{filename}.pdf")

    # Single-document mode: all cars as line items in one PDF
    if doc_mode == "single_doc":
        # All covered cars share ONE stored number under single_doc mode; any
        # car's map entry is representative. Fall back to today's start_no.
        single_doc_no = _resolve_doc_no(docnum, lines[0]["id"], start_no) if lines else start_no
        pdf_bytes = renderer.render_single_doc_to_bytes(order_lines, single_doc_no)
        return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf", as_attachment=True, download_name=f"{filename}.pdf")

    # Single car PDF via ?car=N
    car_idx = request.args.get("car")
    if car_idx is not None and car_idx.isdigit():
        idx = int(car_idx)
        if 0 <= idx < len(order_lines):
            line = order_lines[idx]
            fallback_no = start_no if doc_mode == 'single_doc' else start_no + idx
            inv_no = _resolve_doc_no(docnum, lines[idx]["id"], fallback_no)
            renderer.note = _note_for_car(idx)
            single_buf = io.BytesIO()
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas as rc
            c = rc.Canvas(single_buf, pagesize=A4)
            renderer.render_one(c, inv_no, line)
            c.showPage()
            c.save()
            return send_file(io.BytesIO(single_buf.getvalue()), mimetype="application/pdf", as_attachment=True,
                             download_name=f"{filename}_{inv_no}.pdf")

    if mode == "individual" and len(order_lines) > 1:
        import zipfile
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for i, line in enumerate(order_lines):
                fallback_no = start_no if doc_mode == 'single_doc' else start_no + i
                inv_no = _resolve_doc_no(docnum, lines[i]["id"], fallback_no)
                renderer.note = _note_for_car(i)
                single_buf = io.BytesIO()
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas as rc
                c = rc.Canvas(single_buf, pagesize=A4)
                renderer.render_one(c, inv_no, line)
                c.showPage()
                c.save()
                fname = f"{filename}_{inv_no}_{line.model.replace(' ', '_')}_{i+1}.pdf"
                zf.writestr(fname, single_buf.getvalue())
        zip_buf.seek(0)
        return send_file(zip_buf, mimetype="application/zip", as_attachment=True, download_name=f"{filename}.zip")
    else:
        if _linked_proforma_no:
            # Manual loop to set per-vehicle note before each page
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas as rc
            buf = io.BytesIO()
            c = rc.Canvas(buf, pagesize=A4)
            for i, line in enumerate(order_lines):
                fallback_no = start_no + (0 if doc_mode == 'single_doc' else i)
                inv_no = _resolve_doc_no(docnum, lines[i]["id"], fallback_no)
                renderer.note = _note_for_car(i)
                renderer.render_one(c, inv_no, line)
                c.showPage()
            c.save()
            pdf_bytes = buf.getvalue()
        else:
            # Manual per-car loop (mirrors render_all_to_bytes) so each page
            # can use the stored document number for its car; falls back to
            # render_all_to_bytes' own start_no [+ i] derivation when unmapped.
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas as rc
            buf = io.BytesIO()
            c = rc.Canvas(buf, pagesize=A4)
            for i, line in enumerate(order_lines):
                fallback_no = start_no + (0 if doc_mode == 'single_doc' else i)
                inv_no = _resolve_doc_no(docnum, lines[i]["id"], fallback_no)
                renderer.render_one(c, inv_no, line)
                c.showPage()
            c.save()
            pdf_bytes = buf.getvalue()
        return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf", as_attachment=True, download_name=f"{filename}.pdf")


# ── EuroFib XLSX export ──────────────────────────────────────────

@facturare_bp.route("/facturare/api/invoices/<int:invoice_id>/eurofib")
@login_required
@handle_api_errors
def api_generate_eurofib(invoice_id):
    """Generate EuroFib import XLSX for a non-proforma invoice."""
    from flask import send_file
    from .generators.eurofib_xlsx import EurofibXlsxRenderer
    from .config import JobConfig, InvoiceConfig, FxConfig, EurofibConfig, ContractConfig, InputConfig, PartyConfig
    from .models import OrderLine
    from datetime import date as date_type
    import io

    if not _check_perm("view"):
        return error_response("Permission denied", 403)

    inv_row = _repo.get_invoice_by_id(invoice_id)
    if not inv_row:
        return error_response("Invoice not found", 404)
    if inv_row["invoice_type"] == "PROFORMA":
        return error_response("EuroFib export not available for proformas", 400)

    anexa = _repo.get_anexa_by_id(inv_row["anexa_id"])
    contract = _repo.get_contract_by_id(anexa["contract_id"])
    all_lines = _repo.get_lines_by_anexa(anexa["id"])
    # Cumulative context so a closing slice books the residual (rest = selling −
    # advances), and a storno reverses exactly that residual — see _car_slice_eur.
    _anexa_invoices = list(_repo.get_invoices_by_anexa(anexa["id"]))
    _anexa_price_by_line = {l["id"]: float(l["selling_price_eur"]) for l in all_lines}
    _anexa_all_lids = [l["id"] for l in all_lines]
    _anexa_inv_by_id = {i["id"]: i for i in _anexa_invoices}
    # Filter to selected lines if line_ids is set
    # Preserve line_ids order so numbering matches the UI
    inv_line_ids = inv_row.get("line_ids")
    if inv_line_ids:
        import json as _json
        if isinstance(inv_line_ids, str):
            inv_line_ids = _json.loads(inv_line_ids)
        _lid_map = {l["id"]: l for l in all_lines}
        lines = [_lid_map[lid] for lid in inv_line_ids if lid in _lid_map]
    else:
        lines = all_lines

    # Get Konto config for this supplier + invoice type
    konto_row = _repo.query_one(
        "SELECT * FROM facturare_konto_config WHERE supplier_id = %s AND invoice_type = %s",
        (contract["supplier_id"], inv_row["invoice_type"]))
    if not konto_row or not konto_row.get("konto_credit"):
        return error_response("Konto config not set for this supplier/type. Go to Settings tab.", 400)

    # Stored document numbers per car line_id (populated by the numbering module).
    docnum = _repo.get_document_number_map(invoice_id)

    # Build per-car order lines
    inv_type_str = inv_row["invoice_type"]
    total_amount = float(inv_row["total_amount_eur"])
    split_mode = inv_row.get("split_mode", "equal")
    total_selling = sum(float(l["selling_price_eur"]) for l in lines) or 1
    start_no = inv_row.get("invoice_number") or inv_row["id"]
    issued_date = inv_row.get("issued_date")
    if not issued_date:
        return error_response("Invoice has no issued date. Please set the issued date before exporting.", 400)
    if isinstance(issued_date, str):
        issued_date = date_type.fromisoformat(issued_date)
    kurs = float(inv_row["kurs_applied"]) if inv_row.get("kurs_applied") else 1.0

    # For storno: build lines per reversed invoice (negative amounts)
    reversed_invoices = []  # populated below for STORNO kurs_date usage
    if inv_type_str == "STORNO":
        # Only reverse the invoices actually linked to this storno (not all advances in the anexa)
        reversed_links = _repo.query_all(
            "SELECT source_invoice_id FROM facturare_invoice_links WHERE target_invoice_id = %s AND link_type = 'REVERSES'",
            (invoice_id,))
        reversed_inv_ids = [r["source_invoice_id"] for r in reversed_links] if reversed_links else []

        if not reversed_inv_ids:
            return error_response("STORNO has no linked reversed invoices", 400)

        reversed_invoices = _repo.query_all(
            "SELECT id, invoice_number, total_amount_eur, split_mode, kurs_applied, issued_date, line_ids FROM facturare_invoices "
            "WHERE id IN ({}) ORDER BY sequence_number".format(",".join(["%s"] * len(reversed_inv_ids))),
            tuple(reversed_inv_ids))

        # Build price map for ALL lines in annexa (needed for proportional calculation)
        all_line_prices = {l["id"]: float(l["selling_price_eur"]) for l in all_lines}
        all_line_id_set = set(l["id"] for l in all_lines)

        order_lines = []
        for ri in reversed_invoices:
            ri_total = float(ri["total_amount_eur"])
            ri_split = ri.get("split_mode") or "equal"
            # Kurs from the original advance invoice
            ri_kurs = float(ri["kurs_applied"]) if ri.get("kurs_applied") else kurs
            # Compute covered_selling from the ADVANCE's lines, not the storno's
            import json as _json2
            ri_raw_lids = ri.get("line_ids")
            if isinstance(ri_raw_lids, str):
                ri_raw_lids = _json2.loads(ri_raw_lids)
            ri_line_ids = set(ri_raw_lids) if ri_raw_lids else all_line_id_set
            covered_selling = sum(all_line_prices.get(lid, 0) for lid in ri_line_ids) or 1

            ri_full = _anexa_inv_by_id.get(ri["id"], ri)
            for car in lines:
                # Only reverse cars this advance actually covered. Without this,
                # a single-car advance fans out across every car of a multi-car
                # storno (prod 9104148: 10 advances × 9 cars = 90 rows instead of
                # 18), over-reversing the client account ~8×.
                if car["id"] not in ri_line_ids:
                    continue
                selling = float(car["selling_price_eur"])
                # Reverse the residual the advance actually booked (43462, not 43463).
                ri_priors = _prior_car_fractions(
                    _anexa_invoices, ri_full, car["id"], _anexa_price_by_line, _anexa_all_lids)
                car_amount = _per_car_advance_eur(
                    ri_total, selling, covered_selling, ri_split, len(ri_line_ids),
                    prior_fractions=ri_priors, quant=_quant_for(ri_full))
                order_lines.append(OrderLine(
                    comanda=int(car["nr_comanda"]) if car.get("nr_comanda") and str(car["nr_comanda"]).isdigit() else 0,
                    model=car.get("model", ""), culoare=car.get("culoare") or "",
                    list_price=float(car["list_price_eur"]), selling_price=selling,
                    advance=-car_amount, rest=None,
                    kurs=ri_kurs,
                    # Stored document number for THIS car. A storno is ONE invoice, so
                    # every reversed-advance row of the same car shares that car's
                    # stored number. Pre-backfill fallback: the storno's own start_no
                    # (else the 2nd row would increment into the FINAL's number).
                    start_no=docnum.get(car["id"], start_no),
                ))
    else:
        # For FINAL: derive each car's kurs from the advances it reverses so the
        # final's RON equals the storno's RON (see _final_blended_kurs).
        from datetime import timedelta as _td
        _final_kurs_date_set = False
        final_kurs_acc = {}
        if inv_type_str == "FINAL":
            final_kurs_acc, last_adv_date = _final_blended_kurs(_repo, inv_row, all_lines)
            if last_adv_date:
                kurs_date = last_adv_date - _td(days=1)
                _final_kurs_date_set = True

        order_lines = []
        for l in lines:
            selling = float(l["selling_price_eur"])
            if split_mode == "proportional" and total_selling > 0:
                _priors = _prior_car_fractions(
                    _anexa_invoices, inv_row, l["id"], _anexa_price_by_line, _anexa_all_lids)
                car_advance = _car_slice_eur(
                    selling, _snap_pct(total_amount, total_selling), _priors, _quant_for(inv_row))
            else:
                car_advance = _round_half_up(total_amount / max(len(lines), 1), _quant_for(inv_row))

            # For FINAL: look up venituri rule per line + blended kurs
            line_kostenstelle = None
            line_konto_credit = None
            line_kurs = None
            if inv_type_str == "FINAL":
                nr_cmd = l.get("nr_comanda") or ""
                rule = _repo.match_venituri_rule(contract["supplier_id"], nr_cmd)
                if rule:
                    line_konto_credit = rule["konto_venituri"]
                    line_kostenstelle = rule["kostenstelle"]
                ron_sum, _eur_sum = final_kurs_acc.get(l["id"], (0.0, 0.0))
                if ron_sum and car_advance:
                    line_kurs = ron_sum / car_advance   # RON ÷ EUR ⇒ betrag = storno RON

            order_lines.append(OrderLine(
                comanda=int(l["nr_comanda"]) if l.get("nr_comanda") and str(l["nr_comanda"]).isdigit() else 0,
                model=l["model"], culoare=l.get("culoare") or "",
                list_price=float(l["list_price_eur"]), selling_price=selling,
                advance=car_advance, rest=selling,
                kurs=line_kurs,
                kostenstelle=line_kostenstelle,
                konto_credit_override=line_konto_credit,
                # Stored document number for THIS car; None falls through to the
                # renderer's default (cfg.invoice.start_no + idx) pre-backfill.
                start_no=docnum.get(l["id"]),
            ))

    # Compute kurs_date (day before issued_date)
    from datetime import timedelta
    if inv_type_str == "STORNO" and reversed_invoices:
        # Use the first reversed invoice's issued_date for the global kurs_date
        first_ri_date = reversed_invoices[0].get("issued_date")
        if first_ri_date and isinstance(first_ri_date, str):
            first_ri_date = date_type.fromisoformat(first_ri_date)
        kurs_date = (first_ri_date - timedelta(days=1)) if first_ri_date else issued_date - timedelta(days=1)
    elif inv_type_str == "FINAL":
        # kurs_date already set above from last advance invoice (if found); fall back to day before issued_date
        if not _final_kurs_date_set:
            kurs_date = issued_date - timedelta(days=1)
    else:
        kurs_date = issued_date - timedelta(days=1)

    # Get supplier firmennr (eurofib_klient_id)
    supplier_row = _repo.query_one(
        "SELECT eurofib_klient_id FROM companies WHERE id = %s",
        (contract["supplier_id"],))
    firmennr = supplier_row.get("eurofib_klient_id") if supplier_row else None
    if not firmennr:
        return error_response("Firmennr (eurofib_klient_id) not configured for this supplier. Check company settings.", 400)

    # Get konto_debit from CRM client (per supplier) — mandatory
    crm_client = _repo.query_one(
        "SELECT eurofib_konto_debit FROM crm_clients WHERE id = %s",
        (contract["customer_id"],))
    crm_kd_map = crm_client.get("eurofib_konto_debit") if crm_client else None
    crm_konto_debit = crm_kd_map.get(str(firmennr)) if isinstance(crm_kd_map, dict) else None
    effective_konto_debit = int(crm_konto_debit) if crm_konto_debit else 0

    default_text_templates = {
        'INVOICE': 'avans {model} {comanda}',
        'STORNO': 'storno avans {model} {comanda}',
        'FINAL': '{model} {comanda}',
    }

    # Build JobConfig for the renderer
    cfg = JobConfig(
        job_id=f"inv-{invoice_id}",
        contract=ContractConfig(ref=contract["contract_ref"], anexa_ref=f"Anexa {anexa['anexa_number']}"),
        input=InputConfig(anexa="n/a"),
        invoice=InvoiceConfig(kind="invoice", start_no=start_no, date=issued_date),
        fx=FxConfig(currency="EUR", kurs=kurs, kurs_date=kurs_date),
        supplier=PartyConfig(name="", address_lines=[]),
        customer=PartyConfig(name="", address_lines=[]),
        eurofib=EurofibConfig(
            klient=firmennr,
            konto_debit=effective_konto_debit,
            konto_credit=int(konto_row["konto_credit"]),
            text_template=konto_row.get("text_template") or default_text_templates.get(inv_type_str, "{model} {comanda}"),
            is_storno=(inv_type_str == "STORNO"),
        ),
    )

    renderer = EurofibXlsxRenderer(cfg)
    xlsx_bytes = renderer.render_to_bytes(order_lines)

    cust_row = _repo.query_one("SELECT display_name FROM crm_clients WHERE id = %s", (contract["customer_id"],))
    cust_name = (cust_row["display_name"] if cust_row else "").replace(" ", "_")
    type_label = {"INVOICE": "advance", "STORNO": "storno", "FINAL": "final"}.get(inv_type_str, "")
    dl_name = f"EuroFib_{cust_name}_{start_no}_{type_label}.xlsx"

    return send_file(io.BytesIO(xlsx_bytes), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=dl_name)


def _final_blended_kurs(repo, inv_row, all_lines):
    """Per-car RON/EUR from the advances reversed by THIS final's matching storno.

    A final must net to zero against the storno it closes. That storno is the
    STORNO on the same anexa whose line_ids equal the final's and whose |EUR|
    equals the final's EUR (a 1:1 pairing in production). We sum ONLY the
    advances that storno reverses — each at its own rate — so the final's RON
    equals the storno's RON. Summing every advance that merely *covers* the car
    would double-count when the car also received a separate advance (e.g. an
    earlier 10% advance closed by a different storno), inflating the rate.

    Falls back to all of the anexa's advances when no matching storno is found.
    Returns ({line_id: [ron_sum, eur_sum]}, last_advance_date).
    """
    import json as _json
    anexa_id = inv_row["anexa_id"]
    final_eur = abs(float(inv_row["total_amount_eur"]))
    fraw = inv_row.get("line_ids")
    if isinstance(fraw, str):
        fraw = _json.loads(fraw)
    prices = {l["id"]: float(l["selling_price_eur"]) for l in all_lines}
    all_ids = set(prices)
    final_lids = set(fraw) if fraw else all_ids
    # Cumulative context: each advance's per-car share must equal the storno's
    # residual reversal (43462, not 43463) so final RON == storno RON.
    anexa_invoices = list(repo.get_invoices_by_anexa(anexa_id))
    anexa_inv_by_id = {i["id"]: i for i in anexa_invoices}

    # The storno this final closes: same cars, same |EUR|, most recent.
    matching_storno_id = None
    stornos = repo.query_all(
        "SELECT id, total_amount_eur, line_ids FROM facturare_invoices "
        "WHERE anexa_id = %s AND invoice_type = 'STORNO' ORDER BY id DESC",
        (anexa_id,))
    for s in stornos or []:
        sraw = s.get("line_ids")
        if isinstance(sraw, str):
            sraw = _json.loads(sraw)
        s_lids = set(sraw) if sraw else all_ids
        if s_lids == final_lids and abs(abs(float(s["total_amount_eur"])) - final_eur) < 0.01:
            matching_storno_id = s["id"]
            break

    # Advances to reverse: those the matching storno reverses (else all — fallback).
    advances = None
    if matching_storno_id is not None:
        links = repo.query_all(
            "SELECT source_invoice_id FROM facturare_invoice_links "
            "WHERE target_invoice_id = %s AND link_type = 'REVERSES'",
            (matching_storno_id,))
        adv_ids = [l["source_invoice_id"] for l in (links or [])]
        if adv_ids:
            ph = ",".join(["%s"] * len(adv_ids))
            advances = repo.query_all(
                "SELECT id, total_amount_eur, split_mode, kurs_applied, issued_date, line_ids, round_decimals "
                "FROM facturare_invoices WHERE id IN ({})".format(ph),
                tuple(adv_ids))
    if advances is None:
        # No matching storno (or it reverses nothing) → fall back to ALL anexa
        # advances, i.e. the old blended-rate behaviour. This can over-count when
        # a car has advances the matching storno doesn't reverse, so log it —
        # this bug was previously only caught by a prod audit.
        logger.warning(
            "final_blended_kurs: no matching storno for final invoice %s "
            "(anexa %s, eur %s) — falling back to all anexa advances",
            inv_row.get("id"), anexa_id, final_eur)
        advances = repo.query_all(
            "SELECT id, total_amount_eur, split_mode, kurs_applied, issued_date, line_ids, round_decimals "
            "FROM facturare_invoices "
            "WHERE anexa_id = %s AND invoice_type = 'INVOICE' ORDER BY sequence_number",
            (anexa_id,))
    acc = {}          # line_id -> [ron_sum, eur_sum]
    last_date = None
    for adv in advances or []:
        adv_kurs = float(adv["kurs_applied"]) if adv.get("kurs_applied") else None
        if not adv_kurs:
            continue
        adv_eur = float(adv["total_amount_eur"])
        adv_split = adv.get("split_mode") or "equal"
        raw = adv.get("line_ids")
        if isinstance(raw, str):
            raw = _json.loads(raw)
        adv_lids = set(raw) if raw else all_ids
        covered = sum(prices.get(x, 0) for x in adv_lids) or 1
        adv_full = anexa_inv_by_id.get(adv.get("id"), adv)
        for lid in adv_lids:
            # Whole-EUR share, identical to the storno's per-car amount for this
            # advance, so the final's RON equals the storno's RON to the cent.
            adv_priors = _prior_car_fractions(
                anexa_invoices, adv_full, lid, prices, list(all_ids))
            share = _per_car_advance_eur(
                adv_eur, prices.get(lid, 0), covered, adv_split, len(adv_lids),
                prior_fractions=adv_priors, quant=_quant_for(adv_full))
            slot = acc.setdefault(lid, [0.0, 0.0])
            slot[0] += share * adv_kurs
            slot[1] += share
        adv_date = adv.get("issued_date")
        if isinstance(adv_date, str):
            from datetime import date as _date
            adv_date = _date.fromisoformat(adv_date)
        if adv_date and (last_date is None or adv_date > last_date):
            last_date = adv_date
    return acc, last_date


def _build_eurofib_batch(inv_row):
    """Build (JobConfig, order_lines) for a single invoice. Returns (cfg, lines) or raises ValueError."""
    from .config import JobConfig, InvoiceConfig, FxConfig, EurofibConfig, ContractConfig, InputConfig, PartyConfig
    from .models import OrderLine
    from datetime import date as date_type, timedelta

    invoice_id = inv_row["id"]
    # Stored document numbers per car line_id (populated by the numbering module).
    docnum = _repo.get_document_number_map(invoice_id)
    anexa = _repo.get_anexa_by_id(inv_row["anexa_id"])
    contract = _repo.get_contract_by_id(anexa["contract_id"])
    all_lines = _repo.get_lines_by_anexa(anexa["id"])
    # Cumulative context so a closing slice books the residual and a storno reverses
    # exactly that residual — see _car_slice_eur.
    _anexa_invoices = list(_repo.get_invoices_by_anexa(anexa["id"]))
    _anexa_price_by_line = {l["id"]: float(l["selling_price_eur"]) for l in all_lines}
    _anexa_all_lids = [l["id"] for l in all_lines]
    _anexa_inv_by_id = {i["id"]: i for i in _anexa_invoices}

    inv_line_ids = inv_row.get("line_ids")
    if inv_line_ids:
        import json as _json
        if isinstance(inv_line_ids, str):
            inv_line_ids = _json.loads(inv_line_ids)
        _lid_map = {l["id"]: l for l in all_lines}
        lines = [_lid_map[lid] for lid in inv_line_ids if lid in _lid_map]
    else:
        lines = all_lines

    konto_row = _repo.query_one(
        "SELECT * FROM facturare_konto_config WHERE supplier_id = %s AND invoice_type = %s",
        (contract["supplier_id"], inv_row["invoice_type"]))
    if not konto_row or not konto_row.get("konto_credit"):
        raise ValueError(f"Konto config not set for invoice {inv_row.get('invoice_number')}")

    inv_type_str = inv_row["invoice_type"]
    total_amount = float(inv_row["total_amount_eur"])
    split_mode = inv_row.get("split_mode", "equal")
    total_selling = sum(float(l["selling_price_eur"]) for l in lines) or 1
    start_no = inv_row.get("invoice_number") or inv_row["id"]
    issued_date = inv_row.get("issued_date")
    if not issued_date:
        raise ValueError(f"Invoice {inv_row.get('invoice_number')} has no issued date")
    if isinstance(issued_date, str):
        issued_date = date_type.fromisoformat(issued_date)
    kurs = float(inv_row["kurs_applied"]) if inv_row.get("kurs_applied") else 1.0

    reversed_invoices = []
    if inv_type_str == "STORNO":
        reversed_links = _repo.query_all(
            "SELECT source_invoice_id FROM facturare_invoice_links WHERE target_invoice_id = %s AND link_type = 'REVERSES'",
            (invoice_id,))
        reversed_inv_ids = [r["source_invoice_id"] for r in reversed_links] if reversed_links else []
        if not reversed_inv_ids:
            raise ValueError(f"STORNO {inv_row.get('invoice_number')} has no linked reversed invoices")
        reversed_invoices = _repo.query_all(
            "SELECT id, invoice_number, total_amount_eur, split_mode, kurs_applied, issued_date, line_ids FROM facturare_invoices "
            "WHERE id IN ({}) ORDER BY sequence_number".format(",".join(["%s"] * len(reversed_inv_ids))),
            tuple(reversed_inv_ids))

        all_line_prices = {l["id"]: float(l["selling_price_eur"]) for l in all_lines}
        all_line_id_set = set(l["id"] for l in all_lines)

        order_lines = []
        for ri in reversed_invoices:
            ri_total = float(ri["total_amount_eur"])
            ri_split = ri.get("split_mode") or "equal"
            ri_kurs = float(ri["kurs_applied"]) if ri.get("kurs_applied") else kurs
            import json as _json2
            ri_raw_lids = ri.get("line_ids")
            if isinstance(ri_raw_lids, str):
                ri_raw_lids = _json2.loads(ri_raw_lids)
            ri_line_ids = set(ri_raw_lids) if ri_raw_lids else all_line_id_set
            covered_selling = sum(all_line_prices.get(lid, 0) for lid in ri_line_ids) or 1

            ri_full = _anexa_inv_by_id.get(ri["id"], ri)
            for car in lines:
                # Only reverse cars this advance actually covered. Without this,
                # a single-car advance fans out across every car of a multi-car
                # storno (prod 9104148: 10 advances × 9 cars = 90 rows instead of
                # 18), over-reversing the client account ~8×.
                if car["id"] not in ri_line_ids:
                    continue
                selling = float(car["selling_price_eur"])
                # Reverse the residual the advance actually booked (43462, not 43463).
                ri_priors = _prior_car_fractions(
                    _anexa_invoices, ri_full, car["id"], _anexa_price_by_line, _anexa_all_lids)
                car_amount = _per_car_advance_eur(
                    ri_total, selling, covered_selling, ri_split, len(ri_line_ids),
                    prior_fractions=ri_priors, quant=_quant_for(ri_full))
                order_lines.append(OrderLine(
                    comanda=int(car["nr_comanda"]) if car.get("nr_comanda") and str(car["nr_comanda"]).isdigit() else 0,
                    model=car.get("model", ""), culoare=car.get("culoare") or "",
                    list_price=float(car["list_price_eur"]), selling_price=selling,
                    advance=-car_amount, rest=None,
                    kurs=ri_kurs,
                    # Stored document number for THIS car. A storno is ONE invoice, so
                    # every reversed-advance row of the same car shares that car's
                    # stored number. Pre-backfill fallback: the storno's own start_no
                    # (else the 2nd row would increment into the FINAL's number).
                    start_no=docnum.get(car["id"], start_no),
                ))
    else:
        _final_kurs_date_set = False
        final_kurs_acc = {}
        if inv_type_str == "FINAL":
            final_kurs_acc, last_adv_date = _final_blended_kurs(_repo, inv_row, all_lines)
            if last_adv_date:
                kurs_date = last_adv_date - timedelta(days=1)
                _final_kurs_date_set = True

        order_lines = []
        for l in lines:
            selling = float(l["selling_price_eur"])
            if split_mode == "proportional" and total_selling > 0:
                _priors = _prior_car_fractions(
                    _anexa_invoices, inv_row, l["id"], _anexa_price_by_line, _anexa_all_lids)
                car_advance = _car_slice_eur(
                    selling, _snap_pct(total_amount, total_selling), _priors, _quant_for(inv_row))
            else:
                car_advance = _round_half_up(total_amount / max(len(lines), 1), _quant_for(inv_row))
            line_kostenstelle = None
            line_konto_credit = None
            line_kurs = None
            if inv_type_str == "FINAL":
                nr_cmd = l.get("nr_comanda") or ""
                rule = _repo.match_venituri_rule(contract["supplier_id"], nr_cmd)
                if rule:
                    line_konto_credit = rule["konto_venituri"]
                    line_kostenstelle = rule["kostenstelle"]
                ron_sum, _eur_sum = final_kurs_acc.get(l["id"], (0.0, 0.0))
                if ron_sum and car_advance:
                    line_kurs = ron_sum / car_advance   # RON ÷ EUR ⇒ betrag = storno RON
            order_lines.append(OrderLine(
                comanda=int(l["nr_comanda"]) if l.get("nr_comanda") and str(l["nr_comanda"]).isdigit() else 0,
                model=l["model"], culoare=l.get("culoare") or "",
                list_price=float(l["list_price_eur"]), selling_price=selling,
                advance=car_advance, rest=selling,
                kurs=line_kurs,
                kostenstelle=line_kostenstelle, konto_credit_override=line_konto_credit,
                # Stored document number for THIS car; None falls through to the
                # renderer's default (cfg.invoice.start_no + idx) pre-backfill.
                start_no=docnum.get(l["id"]),
            ))

    # Compute kurs_date
    if inv_type_str == "STORNO" and reversed_invoices:
        first_ri_date = reversed_invoices[0].get("issued_date")
        if first_ri_date and isinstance(first_ri_date, str):
            first_ri_date = date_type.fromisoformat(first_ri_date)
        kurs_date = (first_ri_date - timedelta(days=1)) if first_ri_date else issued_date - timedelta(days=1)
    elif inv_type_str == "FINAL":
        if not _final_kurs_date_set:
            kurs_date = issued_date - timedelta(days=1)
    else:
        kurs_date = issued_date - timedelta(days=1)

    supplier_row = _repo.query_one(
        "SELECT eurofib_klient_id FROM companies WHERE id = %s",
        (contract["supplier_id"],))
    firmennr = supplier_row.get("eurofib_klient_id") if supplier_row else None
    if not firmennr:
        raise ValueError(f"Firmennr not configured for supplier of invoice {inv_row.get('invoice_number')}")

    # Get konto_debit from CRM client (per supplier) — mandatory
    _crm_client = _repo.query_one(
        "SELECT eurofib_konto_debit FROM crm_clients WHERE id = %s",
        (contract["customer_id"],))
    _crm_kd_map = _crm_client.get("eurofib_konto_debit") if _crm_client else None
    _crm_konto = _crm_kd_map.get(str(firmennr)) if isinstance(_crm_kd_map, dict) else None
    effective_konto_debit = int(_crm_konto) if _crm_konto else 0

    default_text_templates = {
        'INVOICE': 'avans {model} {comanda}',
        'STORNO': 'storno avans {model} {comanda}',
        'FINAL': '{model} {comanda}',
    }

    cfg = JobConfig(
        job_id=f"inv-{invoice_id}",
        contract=ContractConfig(ref=contract["contract_ref"], anexa_ref=f"Anexa {anexa['anexa_number']}"),
        input=InputConfig(anexa="n/a"),
        invoice=InvoiceConfig(kind="invoice", start_no=start_no, date=issued_date),
        fx=FxConfig(currency="EUR", kurs=kurs, kurs_date=kurs_date),
        supplier=PartyConfig(name="", address_lines=[]),
        customer=PartyConfig(name="", address_lines=[]),
        eurofib=EurofibConfig(
            klient=firmennr,
            konto_debit=effective_konto_debit,
            konto_credit=int(konto_row["konto_credit"]),
            text_template=konto_row.get("text_template") or default_text_templates.get(inv_type_str, "{model} {comanda}"),
            is_storno=(inv_type_str == "STORNO"),
        ),
    )

    return cfg, order_lines


@facturare_bp.route("/facturare/api/anexas/<int:anexa_id>/eurofib-daily")
@login_required
@handle_api_errors
def api_generate_eurofib_daily(anexa_id):
    """Generate a single concatenated EuroFib XLSX for all invoices on a given date."""
    from flask import send_file
    from .generators.eurofib_xlsx import EurofibXlsxRenderer
    import io

    if not _check_perm("view"):
        return error_response("Permission denied", 403)

    target_date = request.args.get("date")
    if not target_date:
        return error_response("date parameter required (YYYY-MM-DD)", 400)

    invoices = _repo.query_all(
        "SELECT * FROM facturare_invoices WHERE anexa_id = %s AND issued_date = %s AND invoice_type != 'PROFORMA' ORDER BY id",
        (anexa_id, target_date))
    if not invoices:
        return error_response("No invoices found for this date", 404)

    batches = []
    errors = []
    for inv_row in invoices:
        try:
            cfg, order_lines = _build_eurofib_batch(inv_row)
            batches.append((cfg, order_lines))
        except ValueError as e:
            errors.append(str(e))

    if not batches:
        return error_response("; ".join(errors), 400)

    xlsx_bytes = EurofibXlsxRenderer.render_multi_to_bytes(batches)

    dl_name = f"EuroFib_{target_date}.xlsx"
    return send_file(io.BytesIO(xlsx_bytes), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=dl_name)
