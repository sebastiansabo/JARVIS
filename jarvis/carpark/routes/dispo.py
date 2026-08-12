"""Dispo API routes — pipeline dashboard (summary/KPIs) + guarded lifecycle
transitions (reserve/sell/deliver/remove-from-stock/reopen).

No SQL lives here (per repo convention) — every route delegates to
DispoRepository (read-only aggregates) or DispoService (guarded writes).
Exception mapping (mirrors DispoService's documented convention):
  PermissionError -> 403 (cross-tenant vehicle)
  ValueError      -> 400 (validation / guard failure; str(e) is surfaced
                          verbatim so the frontend can branch on the
                          LOW_MARGIN: / MISSING_PV_LIVRARE: prefixes)
  anything else   -> 500 (logged)
"""
import io
import logging

from flask import request, jsonify, send_file
from flask_login import login_required, current_user

from carpark import carpark_bp
from carpark.repositories.dispo_repository import DispoRepository
from carpark.services.dispo_service import DispoService
from carpark.services.import_service import CentralizatorImporter
from carpark.routes.vehicles import (
    carpark_required, carpark_edit_required, carpark_delete_required,
    _serialize, _verify_vehicle_ownership, _user_company_id, _acting_company_id,
)

logger = logging.getLogger('jarvis.carpark')

_dispo_service = DispoService()
_dispo_repo = DispoRepository()
_importer = CentralizatorImporter()

_MAX_IMPORT_UPLOAD_BYTES = 10 * 1024 * 1024  # 10MB

# Money fields hidden from the Dispo summary/export/KPIs for users without
# carpark.view_finance — kept in one place so the JSON and xlsx paths (and
# kpis()) can never drift from each other.
_FINANCE_ROW_FIELDS = ('acquisition_price', 'total_costs', 'gross_margin',
                        'margin_pct', 'bonus_leasing')
_FINANCE_KPI_FIELDS = ('gross_margin_mtd',)

# Query-string filter keys accepted by /dispo/summary, passed through
# verbatim to DispoRepository.summary()/_build_where().
_SUMMARY_FILTER_KEYS = (
    'stage', 'brand', 'location_id', 'salesperson_user_id', 'source',
    'sale_type', 'date_from', 'date_to', 'sale_date_from', 'sale_date_to',
    'search', 'stock_removed',
)

# Column order for the xlsx export. Built from the summary row shape
# (carpark_vehicles.* plus DispoRepository's computed columns) — not every
# `v.*` column (equipment/optional_packages are large JSON blobs unsuited to
# a spreadsheet), just the fields relevant to a Dispo pipeline export.
_EXPORT_COLUMNS = (
    'id', 'vin', 'nr_stoc', 'brand', 'model', 'variant', 'status',
    'location_id', 'salesperson_user_id', 'source',
    'acquisition_date', 'listing_date', 'sale_date', 'sale_type',
    'delivery_date', 'days_in_stock',
    'current_price', 'sale_price',
    'acquisition_price', 'total_costs', 'gross_margin', 'margin_pct',
    'bonus_leasing',
    'buyer_name', 'stock_removed',
)


def _has_finance_perm() -> bool:
    return bool(getattr(current_user, 'can_view_carpark_finance', False))


def _strip_finance(result):
    """Mutates a DispoRepository.summary() result in place, removing money
    fields from every row and nulling the totals block, for callers without
    carpark.view_finance."""
    for row in result.get('rows') or []:
        for field in _FINANCE_ROW_FIELDS:
            row.pop(field, None)
    result['totals'] = None
    return result


def _build_xlsx(rows, has_finance: bool) -> io.BytesIO:
    """Build an xlsx workbook of the Dispo summary rows, respecting finance
    gating (money columns simply aren't emitted when has_finance is False)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    columns = [c for c in _EXPORT_COLUMNS if has_finance or c not in _FINANCE_ROW_FIELDS]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dispo'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col))

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════
# DISPO — SUMMARY (pipeline dashboard)
# ═══════════════════════════════════════════════

@carpark_bp.route('/dispo/summary', methods=['GET'])
@login_required
@carpark_required
def dispo_summary():
    """Paginated, filtered Dispo pipeline summary. Add `?export=xlsx` to
    download the visible (finance-gated) columns as a spreadsheet instead of
    JSON.

    Query params: stage, brand, location_id, salesperson_user_id, source,
    sale_type, date_from, date_to, sale_date_from, sale_date_to, search,
    stock_removed, page, per_page, sort_by, sort_dir, export
    """
    try:
        page = max(1, int(request.args.get('page', 1)))
        per_page = max(1, min(200, int(request.args.get('per_page', 25))))
    except (ValueError, TypeError):
        page, per_page = 1, 25

    sort_by = request.args.get('sort_by', 'acquisition_date')
    sort_dir = request.args.get('sort_dir', 'desc')

    filters = {}
    for key in _SUMMARY_FILTER_KEYS:
        val = request.args.get(key)
        if val is not None and val != '':
            filters[key] = val

    company_id = _acting_company_id()

    try:
        result = _dispo_repo.summary(company_id, filters, page, per_page, sort_by, sort_dir)
    except Exception as e:
        logger.error(f'Dispo summary query failed: {e}', exc_info=True)
        return jsonify({'error': 'Internal error'}), 500

    has_finance = _has_finance_perm()
    if not has_finance:
        _strip_finance(result)

    serialized = _serialize(result)

    if request.args.get('export') == 'xlsx':
        buf = _build_xlsx(serialized['rows'], has_finance)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name='dispo_export.xlsx',
        )

    return jsonify(serialized)


# ═══════════════════════════════════════════════
# DISPO — KPIs
# ═══════════════════════════════════════════════

@carpark_bp.route('/dispo/kpis', methods=['GET'])
@login_required
@carpark_required
def dispo_kpis():
    """Company-wide Dispo KPI tile values."""
    company_id = _acting_company_id()
    try:
        kpis = _dispo_repo.kpis(company_id)
    except Exception as e:
        logger.error(f'Dispo KPIs query failed: {e}', exc_info=True)
        return jsonify({'error': 'Internal error'}), 500

    if not _has_finance_perm():
        kpis = dict(kpis)
        for field in _FINANCE_KPI_FIELDS:
            kpis.pop(field, None)

    return jsonify(_serialize(kpis))


# ═══════════════════════════════════════════════
# DISPO — GUARDED LIFECYCLE TRANSITIONS
# ═══════════════════════════════════════════════

@carpark_bp.route('/vehicles/<int:vehicle_id>/reserve', methods=['POST'])
@login_required
@carpark_edit_required
def reserve_vehicle(vehicle_id):
    """Reserve a vehicle. Body: {client_name|client_id, reservation_end,
    reservation_start?, client_company?, client_phone?, client_email?,
    deposit_amount?, deposit_paid?, notes?}"""
    _, err = _verify_vehicle_ownership(vehicle_id)
    if err:
        return err
    data = request.get_json(silent=True) or {}
    try:
        result = _dispo_service.reserve(vehicle_id, _user_company_id(), current_user, data)
        return jsonify(_serialize(result))
    except PermissionError as e:
        return jsonify({'error': str(e)}), 403
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f'Reserve failed for vehicle {vehicle_id}: {e}', exc_info=True)
        return jsonify({'error': 'Internal error'}), 500


@carpark_bp.route('/vehicles/<int:vehicle_id>/sell', methods=['POST'])
@login_required
@carpark_edit_required
def sell_vehicle(vehicle_id):
    """Record a sale. Body: {sale_price, sale_type, sale_date,
    buyer_name|buyer_client_id, confirm_low_margin?}"""
    _, err = _verify_vehicle_ownership(vehicle_id)
    if err:
        return err
    data = request.get_json(silent=True) or {}
    try:
        result = _dispo_service.sell(vehicle_id, _user_company_id(), current_user, data)
        return jsonify(_serialize(result))
    except PermissionError as e:
        return jsonify({'error': str(e)}), 403
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f'Sell failed for vehicle {vehicle_id}: {e}', exc_info=True)
        return jsonify({'error': 'Internal error'}), 500


@carpark_bp.route('/vehicles/<int:vehicle_id>/deliver', methods=['POST'])
@login_required
@carpark_edit_required
def deliver_vehicle(vehicle_id):
    """Deliver a vehicle. Hard-blocked (400, MISSING_PV_LIVRARE) without an
    uploaded pv_livrare document. Body: {delivery_date}"""
    _, err = _verify_vehicle_ownership(vehicle_id)
    if err:
        return err
    data = request.get_json(silent=True) or {}
    try:
        result = _dispo_service.deliver(vehicle_id, _user_company_id(), current_user, data)
        return jsonify(_serialize(result))
    except PermissionError as e:
        return jsonify({'error': str(e)}), 403
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f'Deliver failed for vehicle {vehicle_id}: {e}', exc_info=True)
        return jsonify({'error': 'Internal error'}), 500


@carpark_bp.route('/vehicles/<int:vehicle_id>/remove-from-stock', methods=['POST'])
@login_required
@carpark_delete_required
def remove_vehicle_from_stock(vehicle_id):
    """Flag a vehicle out of the pipeline. Requires a terminal status
    (DELIVERED/TRANSFERRED/SCRAPPED/RETURNED)."""
    _, err = _verify_vehicle_ownership(vehicle_id)
    if err:
        return err
    try:
        result = _dispo_service.remove_from_stock(vehicle_id, _user_company_id(), current_user)
        return jsonify(_serialize(result))
    except PermissionError as e:
        return jsonify({'error': str(e)}), 403
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f'Remove-from-stock failed for vehicle {vehicle_id}: {e}', exc_info=True)
        return jsonify({'error': 'Internal error'}), 500


@carpark_bp.route('/vehicles/<int:vehicle_id>/reopen', methods=['POST'])
@login_required
@carpark_delete_required
def reopen_vehicle(vehicle_id):
    """Move a vehicle backward in the pipeline. Body: {reason, target_status}"""
    _, err = _verify_vehicle_ownership(vehicle_id)
    if err:
        return err
    data = request.get_json(silent=True) or {}
    try:
        result = _dispo_service.reopen(
            vehicle_id, _user_company_id(), current_user,
            reason=data.get('reason'), target_status=data.get('target_status'))
        return jsonify(_serialize(result))
    except PermissionError as e:
        return jsonify({'error': str(e)}), 403
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f'Reopen failed for vehicle {vehicle_id}: {e}', exc_info=True)
        return jsonify({'error': 'Internal error'}), 500


# ═══════════════════════════════════════════════
# DISPO — CENTRALIZATOR IMPORTER (Phase 5)
# ═══════════════════════════════════════════════

@carpark_bp.route('/dispo/import', methods=['POST'])
@login_required
@carpark_edit_required
def dispo_import():
    """Import the backoffice centralizator xlsx tracker into
    carpark_vehicles (+ cost/revenue rows). All parsing/validation/DB work
    delegates to CentralizatorImporter — this route only handles the
    multipart upload and the dry-run/commit permission split.

    Multipart form: `file` (.xlsx, required). The write target is ALWAYS
    the authenticated caller's own company (`_user_company_id()`) — a
    client-supplied `company_id` is deliberately NOT honored here, so this
    endpoint can't be used to bulk-write vehicles into another tenant
    (IDOR). The server-side CLI (`scripts/import_centralizator.py`) keeps
    its `--company-id` arg for admin use; the HTTP surface does not.

    Query param `?commit=true` actually writes to the DB (default: dry-run
    validation report only). Committing requires `can_delete_carpark`
    (manager) on top of the `can_edit_carpark` the route already requires,
    since it's a bulk write across the whole pipeline, not a single
    vehicle's fields.
    """
    uploaded = request.files.get('file') if request.files else None
    if not uploaded or not uploaded.filename:
        return jsonify({'error': 'file is required (multipart form field "file", .xlsx)'}), 400
    if not uploaded.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Only .xlsx files are supported'}), 400

    file_bytes = uploaded.read()
    if not file_bytes:
        return jsonify({'error': 'Uploaded file is empty'}), 400
    if len(file_bytes) > _MAX_IMPORT_UPLOAD_BYTES:
        return jsonify({'error': 'File exceeds the 10MB upload limit'}), 400

    # Write target is the authenticated caller's own company — never a
    # value from the request body/params. This is the IDOR guard: the
    # importer's own cross-company VIN check (import_service) is the second
    # layer, but the write target itself is fixed here so a caller can't
    # even aim the import at another tenant.
    company_id = _user_company_id()
    if not company_id:
        return jsonify({'error': 'Your account has no company assigned; cannot import'}), 400

    commit = request.args.get('commit', 'false').strip().lower() == 'true'
    if commit and not getattr(current_user, 'can_delete_carpark', False):
        return jsonify({'error': 'Committing an import requires manager (can_delete_carpark) permission'}), 403

    try:
        report = _importer.run(io.BytesIO(file_bytes), company_id, dry_run=not commit)
        return jsonify(_serialize(report.to_dict()))
    except ValueError as e:
        # ValueError from the importer is a safe, user-facing validation
        # message (e.g. "Could not read xlsx file") — never raw internals.
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f'Centralizator import failed (company_id={company_id}): {e}', exc_info=True)
        return jsonify({'error': 'Internal error'}), 500
