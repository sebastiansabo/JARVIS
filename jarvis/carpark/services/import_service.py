"""Centralizator Importer — migrates the backoffice's Excel sales tracker
(`docs/centralizator AAP.xlsx`, sheet AUDI, 149 data rows) into
carpark_vehicles + cost/revenue rows, so the Dispo workspace starts
populated instead of empty.

One-off migration tool, not a recurring sync: a Flask route (`POST
/dispo/import`) and a CLI (`scripts/import_centralizator.py`) both call
`CentralizatorImporter.run()`. Design:

  parse()  — pure Excel -> ParsedRow[]. Iterates every sheet, header row 1.
             A row with no VIN is `skipped_no_vin`. A VIN longer than 17
             chars is a hard `reject` (carpark_vehicles.vin is
             VARCHAR(17) — silently truncating would corrupt the value, so
             this is the one case where "tolerate with a warning" isn't
             possible). A VIN that's simply *short* (16 chars, the dirty
             AUDI-sheet case) still fits the column and is imported with a
             warning. Vânzător/Achizitor names are resolved against the
             `users` table (exact/case-insensitive, then a fuzzy fallback);
             an unmatched name doesn't fail the row — it's recorded as an
             INFO note and the id is simply left unset.

  run()    — wraps parse() with counting + (if dry_run=False) a commit.
             Dry-run never touches the DB for writes. Commit upserts every
             `ok` row by VIN inside ONE transaction
             (BaseRepository.execute_many): any hard DB error rolls back
             the *entire* batch — this is a data migration, not a
             best-effort per-row import, so there's no partial-write mode.
             cost_type='istoric_import'/'cheltuieli_vanzare' and
             revenue_type='bonus_leasing' rows for a vehicle are deleted
             and re-inserted on every commit, so re-running the same file
             (or a corrected re-export) is idempotent rather than
             duplicating cost/revenue history.

Deliberate exception to the "services orchestrate, repositories own SQL"
convention: `_commit()` issues its INSERT/UPDATE/DELETE statements directly
against a single shared cursor (via VehicleRepository.execute_many, which
is exactly BaseRepository's sanctioned "multiple statements, one
transaction" building block — see its docstring). None of the per-call
repository methods (VehicleRepository.create/update, CostRepository.create,
RevenueRepository.create) can share a transaction across ~150 rows since
each opens and commits its own connection, and re-implementing every repo
around an optional external cursor is out of scope for a one-off importer.
The field whitelist itself (VEHICLE_UPDATABLE_FIELDS) is still the single
source of truth, imported from vehicle_repository.py rather than
duplicated.
"""
import difflib
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional

import openpyxl

from carpark.repositories.vehicle_repository import VehicleRepository, VEHICLE_UPDATABLE_FIELDS
from carpark.repositories.cost_repository import CostRepository
from carpark.repositories.revenue_repository import RevenueRepository
from core.auth.repositories.user_repository import UserRepository

logger = logging.getLogger('jarvis.carpark.import')

# carpark_vehicles.vin is VARCHAR(17) — see migrations/domains/schema_carpark.py.
_VIN_COLUMN_MAX_LEN = 17

# Cost/revenue types this importer owns exclusively. Re-running a commit
# deletes-then-reinserts rows of these types per vehicle (see _commit),
# which is what keeps repeated imports idempotent instead of duplicating
# history every time the source file changes.
_IMPORT_COST_TYPES = ('istoric_import', 'cheltuieli_vanzare')
_IMPORT_REVENUE_TYPES = ('bonus_leasing',)

# Excel header (sheet "AUDI", header row 1) -> canonical field name.
# Kept close to the spec's §2.3/§3 wording; a couple of alternate spellings
# are tolerated in case a re-export renames a column slightly.
_HEADER_ALIASES: Dict[str, tuple] = {
    'brand': ('Marca',),
    'model': ('Model',),
    'impus_raw': ('IMPUS',),
    'scos_raw': ('SCOS DIN EVIDENTA',),
    'vin_raw': ('Serie_sasiu', 'Serie sasiu', 'Serie_sasiu(VIN)', 'VIN'),
    'source': ('Furnizor',),
    'location_text': ('Locatie', 'Locație'),
    'acquisition_date': ('Data achizitie', 'Data achiziție'),
    'intake_pv_date': ('PV intrare',),
    'supplier_payment_date': ('Data plata', 'Data plată'),
    'acquisition_price': ('Pret achizitie', 'Preț achiziție'),
    'listing_date': ('Data promovare',),
    'sale_date': ('Data vanzarii', 'Data vânzării'),
    'total_costs': ('Total costuri',),
    'sale_price': ('Pret vanzare', 'Preț vânzare'),
    'bonus_leasing': ('Bonus leasing',),
    'sale_type': ('Tip vanzare', 'Tip vânzare'),
    'buyer_name': ('Client',),
    'salesperson_name': ('Vanzator', 'Vânzător'),
    'acquisition_manager_name': ('Achizitor',),
    'gw_file_number': ('DOSAR GW',),
    'delivery_date': ('PV livrare',),
    'cheltuieli_vanzare': ('Chelt cu vanzarea', 'Chelt cu vânzarea'),
}


# ── Value parsing helpers ──

def _clean_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _parse_money(value: Any) -> Optional[Decimal]:
    """Robust money parser: passes through numeric Excel cells as-is,
    strips spaces/commas from string cells (e.g. "1 234,56" / "1,234.56")
    before converting. Returns None for blank/unparseable values."""
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(' ', '').replace('\xa0', '').replace(',', '')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _parse_date(value: Any) -> Optional[date]:
    """Excel dates arrive as datetime objects when the cell is
    date-formatted; tolerate plain date objects and a few common string
    formats too (blank/unparseable -> None)."""
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _norm_name(name: str) -> str:
    return ' '.join(name.strip().lower().split())


# ── Result shapes ──

@dataclass
class ParsedRow:
    row_index: int
    sheet_name: str
    vin_raw: Optional[str] = None
    vin: Optional[str] = None
    status: str = 'ok'  # 'ok' | 'skipped_no_vin' | 'reject'
    reject_reason: Optional[str] = None
    inferred_status: Optional[str] = None
    action: Optional[str] = None  # 'create' | 'update', filled in by run()
    warnings: List[str] = field(default_factory=list)
    infos: List[str] = field(default_factory=list)
    unmatched_names: List[str] = field(default_factory=list)
    vehicle_fields: Dict[str, Any] = field(default_factory=dict)
    cost_rows: List[Dict[str, Any]] = field(default_factory=list)
    revenue_rows: List[Dict[str, Any]] = field(default_factory=list)

    def to_detail(self) -> Dict[str, Any]:
        return {
            'sheet': self.sheet_name,
            'row': self.row_index,
            'vin': self.vin or self.vin_raw,
            'status': self.status,
            'action': self.action,
            'inferred_status': self.inferred_status,
            'reject_reason': self.reject_reason,
            'warnings': self.warnings,
            'infos': self.infos,
        }


@dataclass
class ImportReport:
    dry_run: bool
    company_id: Optional[int] = None
    total: int = 0
    ok: int = 0
    warnings: int = 0  # count of `ok` rows carrying >=1 warning
    skipped_no_vin: int = 0
    rejects: int = 0
    committed_vehicles_created: int = 0
    committed_vehicles_updated: int = 0
    committed_cost_rows: int = 0
    committed_revenue_rows: int = 0
    error: Optional[str] = None
    rows: List[Dict[str, Any]] = field(default_factory=list)
    unmatched_names: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            'dry_run': self.dry_run,
            'company_id': self.company_id,
            'total': self.total,
            'ok': self.ok,
            'warnings': self.warnings,
            'skipped_no_vin': self.skipped_no_vin,
            'rejects': self.rejects,
            'committed_vehicles_created': self.committed_vehicles_created,
            'committed_vehicles_updated': self.committed_vehicles_updated,
            'committed_cost_rows': self.committed_cost_rows,
            'committed_revenue_rows': self.committed_revenue_rows,
            'error': self.error,
            'rows': self.rows,
            'unmatched_names': self.unmatched_names,
        }


class CentralizatorImporter:
    """Parses the centralizator xlsx and upserts it into carpark_vehicles."""

    def __init__(self,
                 vehicle_repo: VehicleRepository = None,
                 cost_repo: CostRepository = None,
                 revenue_repo: RevenueRepository = None,
                 user_repo: UserRepository = None):
        self._vehicle_repo = vehicle_repo or VehicleRepository()
        self._cost_repo = cost_repo or CostRepository()
        self._revenue_repo = revenue_repo or RevenueRepository()
        self._user_repo = user_repo or UserRepository()
        self._user_by_name: Optional[Dict[str, int]] = None

    # ── USER RESOLUTION ──

    def _load_users(self) -> Dict[str, int]:
        if self._user_by_name is None:
            users = self._user_repo.get_all() or []
            by_name: Dict[str, int] = {}
            for u in users:
                key = _norm_name(u['name'])
                if key and key not in by_name:  # first occurrence wins on duplicate names
                    by_name[key] = u['id']
            self._user_by_name = by_name
        return self._user_by_name

    def _resolve_user(self, name: str) -> Optional[int]:
        """Exact (case-insensitive/trimmed) match first, then a fuzzy
        fallback (difflib, cutoff=0.84) for near-misses like "Alex Vilcan"
        vs "Alexandru Vilcan". Returns None (never raises) on no match."""
        by_name = self._load_users()
        key = _norm_name(name)
        if key in by_name:
            return by_name[key]
        close = difflib.get_close_matches(key, by_name.keys(), n=1, cutoff=0.84)
        if close:
            return by_name[close[0]]
        return None

    # ── PARSE ──

    def parse(self, file_path_or_stream) -> List[ParsedRow]:
        """Iterate every sheet (header row 1) and return one ParsedRow per
        data row (rows 2..N). Never touches the DB except to resolve
        Vânzător/Achizitor names against the users table."""
        try:
            wb = openpyxl.load_workbook(file_path_or_stream, data_only=True, read_only=True)
        except Exception as e:
            raise ValueError(f'Could not read xlsx file: {e}')

        rows: List[ParsedRow] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
            if header_row is None:
                continue
            headers = [str(c.value).strip() if c.value is not None else '' for c in header_row]
            col_index = {h: i for i, h in enumerate(headers) if h}

            for excel_row_num, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
                values = [c.value for c in row_cells]
                if all(v is None for v in values):
                    continue  # fully blank row (trailing sheet padding) — not counted at all
                rows.append(self._parse_row(sheet_name, excel_row_num, col_index, values))
        return rows

    @staticmethod
    def _get(col_index: Dict[str, int], values: list, field_key: str):
        for candidate in _HEADER_ALIASES[field_key]:
            idx = col_index.get(candidate)
            if idx is not None and idx < len(values):
                return values[idx]
        return None

    def _parse_row(self, sheet_name: str, excel_row_num: int,
                   col_index: Dict[str, int], values: list) -> ParsedRow:
        row = ParsedRow(row_index=excel_row_num, sheet_name=sheet_name)
        get = lambda key: self._get(col_index, values, key)  # noqa: E731

        vin_raw = get('vin_raw')
        row.vin_raw = _clean_str(vin_raw)
        if not row.vin_raw:
            row.status = 'skipped_no_vin'
            return row

        vin = row.vin_raw.upper()
        if len(vin) > _VIN_COLUMN_MAX_LEN:
            row.status = 'reject'
            row.reject_reason = (
                f'VIN "{vin}" is {len(vin)} chars — exceeds the '
                f'{_VIN_COLUMN_MAX_LEN}-char carpark_vehicles.vin column limit'
            )
            return row
        row.vin = vin
        if len(vin) != _VIN_COLUMN_MAX_LEN:
            row.warnings.append(f'VIN "{vin}" is {len(vin)} chars (expected {_VIN_COLUMN_MAX_LEN})')

        brand = _clean_str(get('brand'))
        model = _clean_str(get('model'))
        if not brand or not model:
            row.status = 'reject'
            row.reject_reason = 'Missing brand or model'
            return row

        vehicle_fields: Dict[str, Any] = {'vin': vin, 'brand': brand, 'model': model}

        source = _clean_str(get('source'))
        if source:
            vehicle_fields['source'] = source
        location_text = _clean_str(get('location_text'))
        if location_text:
            vehicle_fields['location_text'] = location_text

        acquisition_date = _parse_date(get('acquisition_date'))
        if acquisition_date:
            vehicle_fields['acquisition_date'] = acquisition_date
        intake_pv_date = _parse_date(get('intake_pv_date'))
        if intake_pv_date:
            vehicle_fields['intake_pv_date'] = intake_pv_date
        supplier_payment_date = _parse_date(get('supplier_payment_date'))
        if supplier_payment_date:
            vehicle_fields['supplier_payment_date'] = supplier_payment_date

        acquisition_price = _parse_money(get('acquisition_price'))
        if acquisition_price is not None:
            vehicle_fields['acquisition_price'] = acquisition_price

        listing_date = _parse_date(get('listing_date'))
        if listing_date:
            vehicle_fields['listing_date'] = listing_date
        sale_date = _parse_date(get('sale_date'))
        if sale_date:
            vehicle_fields['sale_date'] = sale_date
        sale_price = _parse_money(get('sale_price'))
        if sale_price is not None:
            vehicle_fields['sale_price'] = sale_price
        sale_type = _clean_str(get('sale_type'))
        if sale_type:
            vehicle_fields['sale_type'] = sale_type
        buyer_name = _clean_str(get('buyer_name'))
        if buyer_name:
            vehicle_fields['buyer_name'] = buyer_name
        gw_file_number = _clean_str(get('gw_file_number'))
        if gw_file_number:
            vehicle_fields['gw_file_number'] = gw_file_number

        impus_raw = _clean_str(get('impus_raw'))
        if impus_raw:
            upper = impus_raw.upper()
            if upper == 'DA':
                vehicle_fields['is_impus'] = True
            elif 'LIPSA CIV' in upper:
                vehicle_fields['missing_civ'] = True
            else:
                row.warnings.append(f'Unrecognized IMPUS value: {impus_raw!r}')

        scos_raw = _clean_str(get('scos_raw'))
        stock_removed = bool(scos_raw and scos_raw.upper() == 'DA')
        if stock_removed:
            vehicle_fields['stock_removed'] = True

        delivery_date = _parse_date(get('delivery_date'))
        if delivery_date:
            vehicle_fields['delivery_date'] = delivery_date

        if stock_removed:
            vehicle_fields['stock_removed_date'] = delivery_date or sale_date or date.today()

        # ── user resolution (never fails the row) ──
        salesperson_name = _clean_str(get('salesperson_name'))
        if salesperson_name:
            uid = self._resolve_user(salesperson_name)
            if uid:
                vehicle_fields['salesperson_user_id'] = uid
            else:
                row.infos.append(f'Unmatched Vanzator: {salesperson_name!r}')
                row.unmatched_names.append(salesperson_name)

        acquisition_manager_name = _clean_str(get('acquisition_manager_name'))
        if acquisition_manager_name:
            uid = self._resolve_user(acquisition_manager_name)
            if uid:
                vehicle_fields['acquisition_manager_id'] = uid
            else:
                row.infos.append(f'Unmatched Achizitor: {acquisition_manager_name!r}')
                row.unmatched_names.append(acquisition_manager_name)

        # ── status inference ──
        if delivery_date or stock_removed:
            inferred_status = 'DELIVERED'
        elif sale_date:
            inferred_status = 'SOLD'
        elif listing_date:
            inferred_status = 'LISTED'
        else:
            inferred_status = 'READY_FOR_SALE'
        vehicle_fields['status'] = inferred_status
        row.inferred_status = inferred_status

        # sanity: every key we're about to write must be whitelisted (vin
        # is handled separately in _commit, the rest must round-trip
        # through VEHICLE_UPDATABLE_FIELDS like any other vehicle write).
        for key in vehicle_fields:
            if key not in VEHICLE_UPDATABLE_FIELDS and key != 'vin':
                raise AssertionError(
                    f'import_service field {key!r} is not in VEHICLE_UPDATABLE_FIELDS — '
                    f'fix the whitelist in vehicle_repository.py')

        # ── cost / revenue rows ──
        total_costs = _parse_money(get('total_costs'))
        if total_costs is not None and total_costs > 0:
            row.cost_rows.append({
                'cost_type': 'istoric_import',
                'amount': total_costs,
                'description': 'Import centralizator — total costuri',
                'date': acquisition_date or date.today(),
            })

        bonus_leasing = _parse_money(get('bonus_leasing'))
        if bonus_leasing is not None and bonus_leasing > 0:
            row.revenue_rows.append({
                'revenue_type': 'bonus_leasing',
                'amount': bonus_leasing,
                'description': 'Import centralizator — bonus leasing',
                'date': sale_date or acquisition_date or date.today(),
            })

        cheltuieli_vanzare = _parse_money(get('cheltuieli_vanzare'))
        if cheltuieli_vanzare is not None and cheltuieli_vanzare > 0:
            row.cost_rows.append({
                'cost_type': 'cheltuieli_vanzare',
                'amount': cheltuieli_vanzare,
                'description': 'Import centralizator — cheltuieli vanzare',
                'date': sale_date or acquisition_date or date.today(),
            })

        row.vehicle_fields = vehicle_fields
        row.status = 'ok'
        return row

    # ── RUN (parse + count + optional commit) ──

    def run(self, file_path_or_stream, company_id: int, dry_run: bool = True) -> ImportReport:
        parsed_rows = self.parse(file_path_or_stream)
        report = ImportReport(dry_run=dry_run, company_id=company_id, total=len(parsed_rows))

        ok_rows = [r for r in parsed_rows if r.status == 'ok']

        # Annotate create/update per VIN with a single batch query instead
        # of one lookup per row.
        existing_vins: set = set()
        vins = [r.vin for r in ok_rows]
        if vins:
            found = self._vehicle_repo.query_all(
                'SELECT vin FROM carpark_vehicles WHERE vin = ANY(%s)', (vins,))
            existing_vins = {f['vin'] for f in found}

        seen_vins: set = set()
        for r in ok_rows:
            if r.vin in seen_vins:
                r.warnings.append(f'Duplicate VIN within import file: {r.vin} (last occurrence wins)')
            seen_vins.add(r.vin)
            r.action = 'update' if r.vin in existing_vins else 'create'

        for row in parsed_rows:
            if row.status == 'skipped_no_vin':
                report.skipped_no_vin += 1
            elif row.status == 'reject':
                report.rejects += 1
            else:
                report.ok += 1
                if row.warnings:
                    report.warnings += 1
            report.rows.append(row.to_detail())
            for n in row.unmatched_names:
                if n not in report.unmatched_names:
                    report.unmatched_names.append(n)

        if dry_run:
            return report

        if ok_rows:
            try:
                created, updated, cost_ct, rev_ct = self._commit(ok_rows, company_id)
                report.committed_vehicles_created = created
                report.committed_vehicles_updated = updated
                report.committed_cost_rows = cost_ct
                report.committed_revenue_rows = rev_ct
            except Exception as e:
                logger.error(f'Centralizator import commit failed, rolled back: {e}', exc_info=True)
                report.error = str(e)
                raise
        return report

    # ── COMMIT (single transaction) ──

    def _commit(self, ok_rows: List[ParsedRow], company_id: int):
        """Upsert every row inside ONE transaction via
        BaseRepository.execute_many — any exception rolls back the whole
        batch, so a bad row never leaves a half-imported file in the DB."""

        def _work(cursor):
            created = updated = cost_count = revenue_count = 0

            for row in ok_rows:
                vin = row.vin
                cursor.execute('SELECT id, status FROM carpark_vehicles WHERE vin = %s', (vin,))
                existing = cursor.fetchone()

                fields = {k: v for k, v in row.vehicle_fields.items() if k != 'vin'}

                if existing:
                    vehicle_id = existing['id']
                    if fields:
                        set_sql = ', '.join(f'{k} = %s' for k in fields)
                        params = list(fields.values()) + [vehicle_id]
                        cursor.execute(
                            f'UPDATE carpark_vehicles SET {set_sql}, updated_at = CURRENT_TIMESTAMP '
                            f'WHERE id = %s',
                            params,
                        )
                    new_status = fields.get('status')
                    if new_status and new_status != existing['status']:
                        cursor.execute('''
                            INSERT INTO carpark_status_history
                                (vehicle_id, old_status, new_status, notes)
                            VALUES (%s, %s, %s, %s)
                        ''', (vehicle_id, existing['status'], new_status,
                              'Centralizator import (update)'))
                    updated += 1
                else:
                    cols = ['vin', 'company_id', 'category'] + list(fields.keys())
                    vals = [vin, company_id, 'SH'] + list(fields.values())
                    placeholders = ', '.join(['%s'] * len(vals))
                    cursor.execute(
                        f'INSERT INTO carpark_vehicles ({", ".join(cols)}) '
                        f'VALUES ({placeholders}) RETURNING id',
                        vals,
                    )
                    vehicle_id = cursor.fetchone()['id']
                    cursor.execute('''
                        INSERT INTO carpark_status_history
                            (vehicle_id, old_status, new_status, notes)
                        VALUES (%s, %s, %s, %s)
                    ''', (vehicle_id, None, fields.get('status'),
                          'Centralizator import (create)'))
                    created += 1

                # Delete-then-reinsert the import-owned cost/revenue types
                # so re-running the same file is idempotent rather than
                # duplicating history.
                cursor.execute(
                    'DELETE FROM carpark_vehicle_costs WHERE vehicle_id = %s AND cost_type = ANY(%s)',
                    (vehicle_id, list(_IMPORT_COST_TYPES)),
                )
                for c in row.cost_rows:
                    cursor.execute('''
                        INSERT INTO carpark_vehicle_costs (vehicle_id, cost_type, amount, description, date)
                        VALUES (%s, %s, %s, %s, %s)
                    ''', (vehicle_id, c['cost_type'], c['amount'], c.get('description'), c.get('date')))
                    cost_count += 1

                cursor.execute(
                    'DELETE FROM carpark_vehicle_revenues WHERE vehicle_id = %s AND revenue_type = ANY(%s)',
                    (vehicle_id, list(_IMPORT_REVENUE_TYPES)),
                )
                for rvn in row.revenue_rows:
                    cursor.execute('''
                        INSERT INTO carpark_vehicle_revenues (vehicle_id, revenue_type, amount, description, date)
                        VALUES (%s, %s, %s, %s, %s)
                    ''', (vehicle_id, rvn['revenue_type'], rvn['amount'],
                          rvn.get('description'), rvn.get('date')))
                    revenue_count += 1

            return created, updated, cost_count, revenue_count

        return self._vehicle_repo.execute_many(_work)
