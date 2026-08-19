"""Pure builders + workbook generation for the Pontaje export.

No SQL here — callers pass already-fetched rows/maps. Kept pure so the row
logic is unit-testable under the psycopg2-mocked test harness.
"""
import datetime as _dt
from io import BytesIO

HEADERS = ['Date', 'Weekday', 'Name', 'Group', 'Company',
           'Checked In', 'Checked Out', 'Status',
           'Actual In', 'Actual Out', 'Actual Status',
           'Lunch', 'Duration', 'Punches', 'Schedule', 'Sincron', 'Permit']

_WEEKDAYS = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

# Sincron day-codes that mean the employee is on a motivated full-day leave/absence.
# These override a punch-driven "Present" in the official Status column and blank the
# worked Duration — the raw punch still surfaces in the "Actual *" columns so a
# badge-vs-leave discrepancy stays visible. Work codes (OZ/OS/OSW), delegation (DLG)
# and contract-gap (X) are intentionally excluded. Used as the fallback set because
# sincron_activity_codes.is_leave is currently unpopulated in prod.
LEAVE_STATUS_CODES = frozenset({
    'CO', 'CM', 'CMS', 'CES', 'CFS', 'CFP', 'CIC', 'INV', 'ZLC', 'ZLS',
})


def _fmt_time(value):
    """Return 'HH:MM' from a datetime/str, by slice (no tz conversion)."""
    if value is None:
        return ''
    s = str(value)
    # match 'T08:03' or ' 08:03'
    for sep in ('T', ' '):
        i = s.find(sep)
        if i != -1 and len(s) >= i + 6 and s[i + 3] == ':':
            return s[i + 1:i + 6]
    return s[11:16] if len(s) >= 16 else s


def _fmt_hm(total_sec):
    if not total_sec or total_sec <= 0:
        return ''
    total_min = round(total_sec / 60)
    h, m = divmod(total_min, 60)
    return f'{h}:{m:02d}'


def _net_seconds(gross_sec, lunch_min):
    if not gross_sec or gross_sec <= 0:
        return 0
    lunch_sec = (lunch_min or 0) * 60
    return gross_sec - lunch_sec if gross_sec > lunch_sec else gross_sec


def _lunch_cell(lunch_min):
    return '' if lunch_min is None else f'{int(lunch_min)} min'


def _span_seconds(a, b):
    a = _dt.datetime.fromisoformat(a) if isinstance(a, str) else a
    b = _dt.datetime.fromisoformat(b) if isinstance(b, str) else b
    if not a or not b:
        return None
    return (b - a).total_seconds()


def _absent_label(is_weekend, is_holiday, has_permit):
    """Label for a non-worked day. Precedence: Holiday > Weekend > Permit > Absent."""
    if is_holiday:
        return 'Holiday'
    if is_weekend:
        return 'Weekend'
    if has_permit:
        return 'Permit'
    return 'Absent'


def _actual_status(raw_first, total, is_weekend, is_holiday, has_permit):
    """Status from physical punches only — ignores manual adjustments."""
    if not raw_first:
        return _absent_label(is_weekend, is_holiday, has_permit)
    if total == 1:
        return 'Not exited'
    return 'Present'


def _adjusted_status(has_punch, has_adj, single_punch_no_adj, is_weekend, is_holiday,
                     has_permit, leave_code=''):
    """Official status after manual adjustments are applied.

    A Sincron motivated-leave code (CO/CM/CMS/…) is authoritative for the day and
    overrides a punch-driven 'Present': the physical punch still shows in the
    'Actual *' columns, but the official pontaj reflects the leave.
    """
    if leave_code:
        return leave_code
    if single_punch_no_adj:
        return 'Not exited'
    if has_punch or has_adj:
        return 'Present'
    return _absent_label(is_weekend, is_holiday, has_permit)


def _permit_cell(entry):
    """Format a permit entry as e.g. '2h (Connecteam)' or '3h (Connecteam, JARVIS)'."""
    if not entry:
        return ''
    hours = entry.get('hours') or 0
    hs = f'{hours:g}'  # 2 -> '2', 2.5 -> '2.5'
    sources = ', '.join(entry.get('sources') or [])
    return f'{hs}h ({sources})' if sources else f'{hs}h'


def build_rows(punch_rows, sched_map, code_map, holidays=None, permit_map=None,
               excluded_keys=None, leave_codes=None):
    holidays = holidays or set()
    permit_map = permit_map or {}
    excluded_keys = excluded_keys or set()
    leave_codes = leave_codes or set()
    out = []
    for r in punch_rows:
        raw_day = r['day']
        d = raw_day if hasattr(raw_day, 'weekday') else _dt.date.fromisoformat(str(raw_day)[:10])
        is_weekend = d.weekday() >= 5
        is_holiday = d.isoformat() in holidays
        juid = r.get('jarvis_user_id')
        company_id = r.get('company_id')
        sched = sched_map.get((juid, company_id, raw_day)) or {}
        lunch = sched.get('lunch_break_minutes')  # None allowed -> blank
        # Schedule is Sincron-authoritative: no BioStar static fallback. A row with
        # no non-excluded Sincron schedule is flagged instead of showing a generic
        # BioStar default that may be wrong.
        sstart = sched.get('schedule_start')
        send = sched.get('schedule_end')
        if sstart or send:
            schedule_cell = f'{sstart or ""}–{send or ""}'
        elif (juid, company_id) in excluded_keys:
            schedule_cell = 'Exclus din pontaj'
        else:
            schedule_cell = 'Fără orar Sincron'

        adj_first = r.get('adjusted_first_punch')
        adj_last = r.get('adjusted_last_punch')
        has_adj = bool(adj_first or adj_last)
        raw_first = r.get('first_punch')
        raw_last = r.get('last_punch')
        total = r.get('total_punches') or 0
        single_no_adj = total == 1 and not has_adj

        eff_in = adj_first or raw_first
        eff_out = adj_last or raw_last
        checked_out = eff_out if (eff_out and eff_out != eff_in) else None

        # "Checked In/Out" columns show ONLY manual adjustments (blank when none);
        # the raw punch lives in the "Actual In/Out" columns.
        checked_in_disp = _fmt_time(adj_first) if adj_first else ''
        checked_out_disp = _fmt_time(adj_last) if (adj_last and adj_last != adj_first) else ''

        code = code_map.get((juid, company_id, raw_day), '')
        # A motivated full-day leave code is authoritative: it overrides a
        # punch-driven "Present" status and zeroes the worked duration.
        leave_code = code if code in leave_codes else ''
        permit = permit_map.get((juid, d.isoformat()))
        has_permit = bool(permit)
        wd = _WEEKDAYS[d.weekday()]

        # gross seconds: adjusted span if both adjusted, else duration_seconds
        if adj_first and adj_last:
            gross = _span_seconds(adj_first, adj_last)
        else:
            gross = r.get('duration_seconds')
        duration = '' if (single_no_adj or not eff_in or not checked_out or leave_code) \
            else _fmt_hm(_net_seconds(gross, lunch))

        out.append([
            d.isoformat(),
            wd,
            r.get('name') or '',
            r.get('group') or '',
            r.get('company') or '',
            checked_in_disp,
            checked_out_disp,
            _adjusted_status(bool(raw_first), has_adj, single_no_adj, is_weekend, is_holiday, has_permit, leave_code),
            _fmt_time(raw_first) if raw_first else '',
            _fmt_time(raw_last) if (raw_last and raw_last != raw_first) else '',
            _actual_status(raw_first, total, is_weekend, is_holiday, has_permit),
            _lunch_cell(lunch),
            duration,
            str(total),
            schedule_cell,
            code,
            _permit_cell(permit),
        ])
    return out


def build_workbook(rows):
    """rows: list of 15-col lists (no header). Returns xlsx bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pontaje'
    fill = PatternFill(start_color='0F6D63', end_color='0F6D63', fill_type='solid')
    font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal='center')
    for i, r in enumerate(rows, 2):
        for col, val in enumerate(r, 1):
            ws.cell(row=i, column=col, value=val)
    widths = [11, 8, 22, 16, 16, 10, 11, 15, 10, 10, 14, 8, 9, 8, 14, 9, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws.freeze_panes = 'A2'
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_code_map(code_rows):
    """First row per (user, company, day) wins — get_day_codes_for_users returns
    them pre-ordered by priority (leave codes before OZ/OS). Keyed per company so
    a code never bleeds across a multi-contract employee's other-company rows."""
    code_map = {}
    for row in code_rows:
        code_map.setdefault(
            (row['mapped_jarvis_user_id'], row['company_id'], row['day']),
            row['short_code'])
    return code_map


def _months_between(start, end):
    s = _dt.date.fromisoformat(str(start)[:10])
    e = _dt.date.fromisoformat(str(end)[:10])
    out, y, m = [], s.year, s.month
    while (y, m) <= (e.year, e.month):
        out.append((y, m))
        m += 1
        if m > 12:
            m = 1; y += 1
    return out


def _fetch_holidays(start, end):
    """Return a set of ISO 'YYYY-MM-DD' public-holiday dates spanning [start, end].

    Best-effort: if the holiday table/repo is unavailable, return an empty set so
    the export still succeeds (weekend detection is unaffected).
    """
    try:
        from core.utils.holidays_repository import HolidayRepository
    except Exception:
        return set()
    repo = HolidayRepository()
    y0 = int(str(start)[:4])
    y1 = int(str(end)[:4])
    out = set()
    for y in range(y0, y1 + 1):
        for h in repo.get_holidays_for_year(y):
            d = h['date']
            out.add(d.isoformat() if hasattr(d, 'isoformat') else str(d)[:10])
    return out


def _fetch_permits(start, end):
    """Return {(jarvis_user_id, 'YYYY-MM-DD'): {'hours': float, 'sources': [..]}}.

    Merges partial-day leave permits from two sources:
      • Connecteam  — connecteam_form_submissions (Bilet de Invoire webhook)
      • JARVIS      — internal 'bilet-de-invoire' form submissions
    Best-effort: any source that errors is skipped so the export still succeeds.
    """
    from core.base_repository import BaseRepository
    base = BaseRepository()
    permit_map = {}

    def _add(uid, day, hours, source):
        if uid is None or not day:
            return
        key = (uid, str(day)[:10])
        e = permit_map.setdefault(key, {'hours': 0.0, 'sources': []})
        e['hours'] += float(hours or 0)
        if source not in e['sources']:
            e['sources'].append(source)

    try:
        for r in base.query_all('''
            SELECT mapped_jarvis_user_id AS uid, leave_date AS day,
                   COALESCE(SUM(leave_hours), 0) AS hours
            FROM connecteam_form_submissions
            WHERE mapped_jarvis_user_id IS NOT NULL
              AND leave_date BETWEEN %s AND %s
              AND COALESCE(status, 'submitted') NOT IN ('rejected', 'cancelled')
            GROUP BY mapped_jarvis_user_id, leave_date
        ''', (str(start)[:10], str(end)[:10])):
            _add(r['uid'], r['day'], r['hours'], 'Connecteam')
    except Exception:
        pass

    try:
        for r in base.query_all('''
            SELECT fs.respondent_user_id AS uid,
                   (fs.answers->>'f_bi_leave_date') AS day,
                   COALESCE(SUM((fs.answers->>'f_bi_hours')::numeric), 0) AS hours
            FROM form_submissions fs
            JOIN forms f ON f.id = fs.form_id
            WHERE f.slug = 'bilet-de-invoire'
              AND fs.status NOT IN ('cancelled', 'rejected')
              AND fs.respondent_user_id IS NOT NULL
              AND (fs.answers->>'f_bi_leave_date') ~ %s
              AND (fs.answers->>'f_bi_leave_date')::date BETWEEN %s AND %s
            GROUP BY fs.respondent_user_id, (fs.answers->>'f_bi_leave_date')
        ''', (r'^\d{4}-\d{2}-\d{2}$', str(start)[:10], str(end)[:10])):
            _add(r['uid'], r['day'], r['hours'], 'JARVIS')
    except Exception:
        pass

    return permit_map


def resolve_export_ids(allowed, group_ids=None, employee_ids=None):
    """Intersect a requested group/employee filter with the caller's permission scope.

    allowed:      None (see all) | list[int] of permitted jarvis_user_ids | [-1] (deny).
    group_ids:    list[int] | None — jarvis ids belonging to a chosen group.
    employee_ids: list[int] | None — explicitly chosen jarvis ids (takes precedence).

    Returns the id list to hand to generate(), or None when no filter is requested and
    the scope is see-all. An empty list means "filter requested but nothing in scope".
    """
    requested = None
    if employee_ids:
        requested = list(dict.fromkeys(int(x) for x in employee_ids))
    elif group_ids:
        requested = list(dict.fromkeys(int(x) for x in group_ids))

    if requested is None:
        return allowed  # unchanged behaviour: None=all, [ids], [-1]=deny

    if allowed is None:
        return requested  # see-all: honour the request as-is
    allowed_set = set(allowed)
    return [uid for uid in requested if uid in allowed_set]


def generate(start, end, jarvis_user_ids):
    """Fetch + assemble + build workbook. Returns (xlsx_bytes, filename)."""
    from core.connectors.biostar.repositories.biostar_repository import BioStarRepository
    from core.connectors.sincron.repositories.sincron_repository import SincronRepository
    b_repo = BioStarRepository()
    s_repo = SincronRepository()

    punch_rows = b_repo.get_pontaje_rows(start, end, jarvis_user_ids)

    ids = sorted({r['jarvis_user_id'] for r in punch_rows if r.get('jarvis_user_id')})
    sched_map = {}
    code_rows = []
    excluded_keys = set()
    if ids:
        for s in s_repo.get_day_schedules_for_users(ids, start, end):
            sched_map[(s['jarvis_user_id'], s['company_id'], s['day'])] = s
        for (y, m) in _months_between(start, end):
            code_rows.extend(s_repo.get_day_codes_for_users(ids, y, m))
        excluded_keys = {
            (k['jarvis_user_id'], k['company_id'])
            for k in s_repo.get_excluded_contract_keys(ids)
        }
    code_map = _build_code_map(code_rows)

    # Prefer the DB-configured leave flag; fall back to the built-in set because
    # sincron_activity_codes.is_leave is currently unpopulated in prod.
    try:
        leave_codes = set(s_repo.get_leave_codes()) or set(LEAVE_STATUS_CODES)
    except Exception:
        leave_codes = set(LEAVE_STATUS_CODES)

    holidays = _fetch_holidays(start, end)
    permit_map = _fetch_permits(start, end)
    rows = build_rows(punch_rows, sched_map, code_map, holidays, permit_map,
                      excluded_keys, leave_codes)
    xlsx = build_workbook(rows)
    filename = f'pontaje_{start}_{end}.xlsx'
    return xlsx, filename
