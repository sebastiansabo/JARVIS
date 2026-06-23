"""Shared parser utilities — file reading, type coercion, phone normalization."""

import re
import unicodedata
import logging
from decimal import Decimal, InvalidOperation

logger = logging.getLogger('jarvis.crm.parsers.utils')


def read_file(file_path, sheet_name=None):
    """Read Excel (.xlsx/.xls) or CSV file into a pandas DataFrame.

    For Excel files with multiple sheets: if sheet_name is given, read that sheet.
    Otherwise try to find a sheet matching the import type, or fall back to the first sheet.
    """
    import pandas as pd
    try:
        if file_path.endswith('.csv'):
            return pd.read_csv(file_path, dtype=str, keep_default_na=False)
        else:
            kwargs = {'dtype': str, 'keep_default_na': False}
            if sheet_name:
                kwargs['sheet_name'] = sheet_name
            return pd.read_excel(file_path, **kwargs)
    except Exception as e:
        logger.error(f'Failed to read file {file_path}: {e}')
        return None


def safe_str(val):
    """Convert to string, return None for empty/nan."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', 'nan', 'None', 'NaN', 'none', 'null'):
        return None
    return s


def safe_date(val):
    """Parse date string to ISO format or return None."""
    from datetime import datetime
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', 'nan', 'None', 'NaN', 'none', 'null'):
        return None
    # Try common formats
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y',
                '%Y-%m-%dT%H:%M:%S', '%d.%m.%Y %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt).isoformat()
        except ValueError:
            continue
    # Pandas timestamp
    try:
        import pandas as pd
        return pd.Timestamp(s).isoformat()
    except Exception:
        pass
    return None


def safe_decimal(val):
    """Convert to Decimal, return None on failure."""
    if val is None:
        return None
    s = str(val).strip().replace(',', '.')
    if s in ('', 'nan', 'None', 'NaN', 'none', 'null'):
        return None
    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return None


def safe_int(val):
    """Convert to int, return None on failure."""
    if val is None:
        return None
    s = str(val).strip().replace('.0', '')
    if s in ('', 'nan', 'None', 'NaN', 'none', 'null'):
        return None
    try:
        return int(float(s))
    except (ValueError, OverflowError):
        return None


def normalize_phone(raw):
    """Normalize phone to 40XXXXXXXXX format. Returns (normalized, raw) or (None, raw)."""
    if not raw:
        return None, raw
    digits = re.sub(r'\D', '', str(raw))
    if not digits:
        return None, raw
    # Romanian numbers
    if digits.startswith('0') and len(digits) == 10:
        digits = '40' + digits[1:]
    elif digits.startswith('4') and len(digits) == 11:
        pass  # already 40XXXXXXXXX
    elif digits.startswith('40') and len(digits) == 12:
        pass
    elif len(digits) >= 9 and not digits.startswith('40'):
        digits = '40' + digits
    if len(digits) < 10 or len(digits) > 12:
        return None, raw
    return digits, raw


def read_file_chunked(file_path, sheet_name=None, chunk_size=5000):
    """Read Excel/CSV in chunks using streaming to avoid memory spikes.

    For CSV: uses pandas native chunksize.
    For XLSX: uses openpyxl read_only mode (never loads full file into memory).
    For XLS: falls back to pandas full read, then yields in chunks.

    Yields pandas DataFrames of up to chunk_size rows each.
    """
    import pandas as pd

    if file_path.endswith('.csv'):
        try:
            for chunk in pd.read_csv(file_path, dtype=str, keep_default_na=False, chunksize=chunk_size):
                yield chunk
        except Exception as e:
            logger.error(f'Failed to read CSV {file_path}: {e}')
        return

    if file_path.endswith('.xls'):
        # Old format — no openpyxl support, fall back to full read + chunking
        try:
            df = pd.read_excel(file_path, dtype=str, keep_default_na=False)
            for start in range(0, len(df), chunk_size):
                yield df.iloc[start:start + chunk_size]
        except Exception as e:
            logger.error(f'Failed to read XLS {file_path}: {e}')
        return

    # XLSX: stream via openpyxl read_only mode
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)

        ws = None
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        if ws is None:
            wb.close()
            return

        rows_iter = ws.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            wb.close()
            return

        headers = [str(h).strip() if h is not None else f'_col_{i}'
                   for i, h in enumerate(header_row)]

        chunk_rows = []
        for row_values in rows_iter:
            row_dict = {}
            for i, v in enumerate(row_values):
                if i < len(headers):
                    row_dict[headers[i]] = str(v).strip() if v is not None else ''
            chunk_rows.append(row_dict)
            if len(chunk_rows) >= chunk_size:
                yield pd.DataFrame(chunk_rows)
                chunk_rows = []

        if chunk_rows:
            yield pd.DataFrame(chunk_rows)

        wb.close()
    except Exception as e:
        logger.error(f'Failed to stream Excel {file_path}: {e}')


def normalize_columns(df, column_map):
    """Rename DataFrame columns: accept both original Excel headers and DB field names.

    If a column already has a DB field name (value in column_map), keep it as-is.
    If a column matches an Excel header (key in column_map), rename it to the DB field name.
    """
    reverse_map = {v: v for v in column_map.values()}  # identity for DB names
    forward_map = {k: v for k, v in column_map.items()}  # Excel → DB
    rename = {}
    for col in df.columns:
        if col in reverse_map:
            pass  # already a DB field name
        elif col in forward_map:
            rename[col] = forward_map[col]
    if rename:
        df = df.rename(columns=rename)
    return df


def normalize_name(name):
    """Normalize name: strip diacritics, lowercase, collapse whitespace."""
    if not name:
        return ''
    # Remove diacritics
    nfkd = unicodedata.normalize('NFKD', name)
    ascii_str = ''.join(c for c in nfkd if not unicodedata.combining(c))
    # Lowercase and collapse spaces
    return re.sub(r'\s+', ' ', ascii_str.lower().strip())
