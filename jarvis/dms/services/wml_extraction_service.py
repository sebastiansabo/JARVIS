"""WML/Document extraction service — extract text from .docx, .xlsx, .pdf, .txt files."""
import os
import re
import logging
from typing import Optional, List, Dict

logger = logging.getLogger('jarvis.dms.services.wml_extraction')


class WmlExtractionService:
    """Extract text and structure from uploaded document files."""

    # Max file size for extraction (20MB)
    MAX_EXTRACT_SIZE = 20 * 1024 * 1024
    # Max text length to store
    MAX_TEXT_LENGTH = 500_000

    def extract_from_file(self, file_path: str, mime_type: Optional[str] = None) -> Dict:
        """Extract text from a file.

        Returns:
            {
                'raw_text': str,
                'structured_json': dict or None,
                'method': str,
            }
        """
        if not os.path.exists(file_path):
            return {'raw_text': '', 'structured_json': None, 'method': 'error', 'error': 'File not found'}

        size = os.path.getsize(file_path)
        if size > self.MAX_EXTRACT_SIZE:
            return {'raw_text': '', 'structured_json': None, 'method': 'error', 'error': 'File too large'}

        ext = os.path.splitext(file_path)[1].lower()

        if ext == '.docx':
            return self._extract_docx(file_path)
        elif ext == '.xlsx':
            return self._extract_xlsx(file_path)
        elif ext == '.pdf':
            return self._extract_pdf(file_path)
        elif ext in ('.txt', '.csv', '.md', '.log'):
            return self._extract_text(file_path)
        else:
            return {'raw_text': '', 'structured_json': None, 'method': 'unsupported',
                    'error': f'Unsupported file type: {ext}'}

    def _extract_docx(self, file_path: str) -> Dict:
        """Extract from .docx using mammoth (preferred) or python-docx fallback."""
        try:
            import mammoth
            with open(file_path, 'rb') as f:
                result = mammoth.extract_raw_text(f)
            text = result.value[:self.MAX_TEXT_LENGTH]
            # Also try to get structured HTML for headings
            structured = None
            try:
                with open(file_path, 'rb') as f:
                    html_result = mammoth.convert_to_html(f)
                structured = self._parse_html_structure(html_result.value)
            except Exception:
                pass
            return {'raw_text': text, 'structured_json': structured, 'method': 'mammoth'}
        except ImportError:
            pass

        # Fallback: python-docx
        try:
            from docx import Document
            doc = Document(file_path)
            paragraphs = []
            structured = {'headings': [], 'paragraphs': []}
            for para in doc.paragraphs:
                text = para.text.strip()
                if not text:
                    continue
                paragraphs.append(text)
                if para.style and para.style.name and para.style.name.startswith('Heading'):
                    structured['headings'].append({
                        'level': para.style.name,
                        'text': text,
                        'index': len(paragraphs) - 1,
                    })
                else:
                    structured['paragraphs'].append(text)
            return {
                'raw_text': '\n'.join(paragraphs)[:self.MAX_TEXT_LENGTH],
                'structured_json': structured,
                'method': 'python-docx',
            }
        except ImportError:
            pass

        # Last resort: zipfile extraction
        try:
            import zipfile
            from xml.etree import ElementTree
            with zipfile.ZipFile(file_path) as zf:
                xml_content = zf.read('word/document.xml')
            tree = ElementTree.fromstring(xml_content)
            ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            texts = []
            for t in tree.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t'):
                if t.text:
                    texts.append(t.text)
            return {
                'raw_text': ' '.join(texts)[:self.MAX_TEXT_LENGTH],
                'structured_json': None,
                'method': 'zipfile',
            }
        except Exception as e:
            return {'raw_text': '', 'structured_json': None, 'method': 'error', 'error': str(e)}

    def _extract_xlsx(self, file_path: str) -> Dict:
        """Extract from .xlsx using openpyxl."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = {}
            all_text = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else '' for c in row]
                    if any(cells):
                        rows.append(cells)
                        all_text.append(' | '.join(c for c in cells if c))
                sheets[sheet_name] = rows[:100]  # Limit rows per sheet
            wb.close()
            return {
                'raw_text': '\n'.join(all_text)[:self.MAX_TEXT_LENGTH],
                'structured_json': {'sheets': sheets},
                'method': 'openpyxl',
            }
        except ImportError:
            return {'raw_text': '', 'structured_json': None, 'method': 'error',
                    'error': 'openpyxl not installed'}
        except Exception as e:
            return {'raw_text': '', 'structured_json': None, 'method': 'error', 'error': str(e)}

    def _extract_pdf(self, file_path: str) -> Dict:
        """Extract from .pdf using pdfplumber."""
        try:
            import pdfplumber
            all_text = []
            pages = []
            with pdfplumber.open(file_path) as pdf:
                for i, page in enumerate(pdf.pages[:50]):  # Limit to 50 pages
                    text = page.extract_text() or ''
                    all_text.append(text)
                    pages.append({'page': i + 1, 'text': text[:2000]})
            return {
                'raw_text': '\n\n'.join(all_text)[:self.MAX_TEXT_LENGTH],
                'structured_json': {'pages': pages},
                'method': 'pdfplumber',
            }
        except ImportError:
            return {'raw_text': '', 'structured_json': None, 'method': 'error',
                    'error': 'pdfplumber not installed'}
        except Exception as e:
            return {'raw_text': '', 'structured_json': None, 'method': 'error', 'error': str(e)}

    def _extract_text(self, file_path: str) -> Dict:
        """Extract from plain text files."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                text = f.read(self.MAX_TEXT_LENGTH)
            return {'raw_text': text, 'structured_json': None, 'method': 'text'}
        except Exception as e:
            return {'raw_text': '', 'structured_json': None, 'method': 'error', 'error': str(e)}

    def _parse_html_structure(self, html: str) -> Optional[Dict]:
        """Parse mammoth HTML output into heading/paragraph structure."""
        try:
            sections = []
            # Split by heading tags
            parts = re.split(r'(<h[1-6][^>]*>.*?</h[1-6]>)', html, flags=re.DOTALL)
            current_heading = None
            current_content = []
            for part in parts:
                heading_match = re.match(r'<h([1-6])[^>]*>(.*?)</h\1>', part, re.DOTALL)
                if heading_match:
                    if current_heading or current_content:
                        sections.append({
                            'heading': current_heading,
                            'content': self._strip_tags(' '.join(current_content)),
                        })
                    current_heading = self._strip_tags(heading_match.group(2))
                    current_content = []
                else:
                    text = self._strip_tags(part).strip()
                    if text:
                        current_content.append(text)
            if current_heading or current_content:
                sections.append({
                    'heading': current_heading,
                    'content': self._strip_tags(' '.join(current_content)),
                })
            return {'sections': sections} if sections else None
        except Exception:
            return None

    @staticmethod
    def _strip_tags(html: str) -> str:
        """Remove HTML tags."""
        return re.sub(r'<[^>]+>', '', html).strip()

    def chunk_text(self, raw_text: str, structured_json: Optional[Dict] = None,
                   max_chunk_size: int = 1000) -> List[Dict]:
        """Split text into chunks for indexing.

        Tries to split at heading boundaries if structured data available.
        Falls back to paragraph/size-based splitting.
        """
        chunks = []

        # Use structured sections if available
        if structured_json and structured_json.get('sections'):
            for section in structured_json['sections']:
                content = section.get('content', '')
                if not content:
                    continue
                # Split large sections
                if len(content) > max_chunk_size:
                    for sub in self._split_by_size(content, max_chunk_size):
                        chunks.append({
                            'heading': section.get('heading'),
                            'content': sub,
                            'token_count': max(1, len(sub) // 4),
                        })
                else:
                    chunks.append({
                        'heading': section.get('heading'),
                        'content': content,
                        'token_count': max(1, len(content) // 4),
                    })
            if chunks:
                return chunks

        # Fallback: split by paragraphs / size
        if not raw_text:
            return []

        paragraphs = [p.strip() for p in raw_text.split('\n\n') if p.strip()]
        current = []
        current_size = 0

        for para in paragraphs:
            if current_size + len(para) > max_chunk_size and current:
                text = '\n\n'.join(current)
                chunks.append({
                    'heading': None,
                    'content': text,
                    'token_count': max(1, len(text) // 4),
                })
                current = []
                current_size = 0
            current.append(para)
            current_size += len(para)

        if current:
            text = '\n\n'.join(current)
            chunks.append({
                'heading': None,
                'content': text,
                'token_count': max(1, len(text) // 4),
            })

        return chunks

    @staticmethod
    def _split_by_size(text: str, max_size: int) -> List[str]:
        """Split text by sentence boundaries within size limit."""
        parts = []
        sentences = re.split(r'(?<=[.!?])\s+', text)
        current = []
        current_len = 0
        for sent in sentences:
            if current_len + len(sent) > max_size and current:
                parts.append(' '.join(current))
                current = []
                current_len = 0
            current.append(sent)
            current_len += len(sent)
        if current:
            parts.append(' '.join(current))
        return parts
