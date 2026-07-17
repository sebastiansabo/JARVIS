"""Export routes for foi de parcurs: session list (xlsx) + contract PDFs (zip)."""
import io
import re
import zipfile
from datetime import datetime
from flask import Response
from ._shared import foi_parcurs_bp, jsonify, request, login_required, logger, _fp_repo
from .pdf import _ensure_pdf_path

_MAX_ROWS = 100000


def _num(v):
    """Coerce a value to a number for spreadsheet cells, else None."""
    if v in (None, ''):
        return None
    try:
        return float(v) if isinstance(v, str) and '.' in v else int(v)
    except (ValueError, TypeError):
        return v


def _fmt_dt(v):
    """Format an ISO/datetime value as 'YYYY-MM-DD HH:MM', else ''. """
    if not v:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M')
    s = str(v)
    try:
        return datetime.fromisoformat(s.replace('Z', '')).strftime('%Y-%m-%d %H:%M')
    except ValueError:
        return s


def _filters_from_request():
    company_id = request.args.get('company_id', type=int)
    return {
        'company_id': company_id,
        'vin': (request.args.get('vin') or '').strip() or None,
        'route_type': (request.args.get('route_type') or '').strip() or None,
        'date_from': (request.args.get('date_from') or '').strip() or None,
        'date_to': (request.args.get('date_to') or '').strip() or None,
    }


def _fetch_rows(filters):
    rows, _ = _fp_repo.get_contracts(
        company_id=filters['company_id'],
        vin=filters['vin'],
        route_type=filters['route_type'],
        date_from=filters['date_from'],
        date_to=filters['date_to'],
        page=1, per_page=_MAX_ROWS,
        sort_by='created_at', sort_dir='DESC',
    )
    return rows or []


def _period_label(filters):
    a = filters['date_from'] or 'inceput'
    b = filters['date_to'] or 'azi'
    return f'{a}_{b}'


@foi_parcurs_bp.route('/api/foi-parcurs/export/xlsx', methods=['GET'])
@login_required
def api_export_xlsx():
    """Export the filtered session list as an .xlsx spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    filters = _filters_from_request()
    rows = _fetch_rows(filters)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sesiuni'
    headers = [
        'Data plecare', 'Data retur', 'Companie', 'Nr. înmatriculare', 'VIN',
        'Client', 'Consilier', 'Tip', 'Status', 'KM plecare', 'KM sosire', 'Distanță (km)',
    ]
    ws.append(headers)
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='2F3640')
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal='center')

    for r in rows:
        ws.append([
            _fmt_dt(r.get('departure_datetime')),
            _fmt_dt(r.get('return_datetime')),
            r.get('company_name') or '',
            r.get('registration_number') or '',
            r.get('vin') or '',
            r.get('client_name') or '',
            r.get('advisor_name') or '',
            r.get('route_type') or '',
            r.get('status') or '',
            _num(r.get('km_start')),
            _num(r.get('km_end')),
            _num(r.get('distance_km')),
        ])

    widths = [17, 17, 26, 16, 20, 24, 20, 8, 12, 12, 12, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    filename = f'sesiuni_{_period_label(filters)}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


@foi_parcurs_bp.route('/api/foi-parcurs/export/contracts-zip', methods=['GET'])
@login_required
def api_export_contracts_zip():
    """Export the filtered sessions' legal contract PDFs as a .zip (one per file).
    Only sessions with a generated contract (FILLED/COMPLETED) are included."""
    filters = _filters_from_request()
    rows = _fetch_rows(filters)

    buf = io.BytesIO()
    included = 0
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for r in rows:
            if r.get('status') not in ('FILLED', 'COMPLETED'):
                continue
            try:
                pdf_path = _ensure_pdf_path(r, r['id'], 'legal')
                with open(pdf_path, 'rb') as f:
                    pdf_bytes = f.read()
            except Exception:
                logger.warning('Skipping contract %s in zip export (PDF failed)', r.get('id'), exc_info=True)
                continue
            safe = re.sub(r'[^A-Za-z0-9._-]+', '_', str(r.get('contract_id') or r.get('id')))
            zf.writestr(f'foaie-parcurs-{safe}.pdf', pdf_bytes)
            included += 1

    if included == 0:
        return jsonify({'success': False, 'error': 'Niciun contract de exportat în perioada selectată.'}), 404

    filename = f'contracte_{_period_label(filters)}.zip'
    return Response(
        buf.getvalue(),
        mimetype='application/zip',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )
