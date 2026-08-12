"""Monthly *Foaie de Parcurs* generator (per car).

Aggregates one vehicle's driving sessions for a given month into a single
document. PDF output is AI-drafted (prose only) into a locked HTML skeleton,
rendered with Playwright. Excel output is a deterministic tabular log.

See docs/superpowers/specs/2026-07-22-monthly-foaie-parcurs-design.md
"""
import io
import os
import re
import json
import html
import logging
from datetime import datetime

from ..repositories.foi_parcurs_repository import FoiParcursRepository
from ..repositories.vehicle_repository import FPVehicleRepository
from .pdf_service import _build_prestator_intro, _PRESTATOR_FALLBACK

logger = logging.getLogger(__name__)

_MODEL = 'claude-sonnet-4-6'
# The Foaie de Parcurs is structured data → the prose is generated from a
# deterministic template by default. The AI is only an OPTIONAL polish pass,
# enabled per-deploy via FP_ROUTESHEET_AI (off by default: reliable, fast, and
# it can't invent filler like "vehiculul nu a înregistrat curse de tip Comodat").
_AI_ENHANCE = os.environ.get('FP_ROUTESHEET_AI', '').strip().lower() in ('1', 'true', 'yes', 'on')
_MONTHS_RO = [
    '', 'ianuarie', 'februarie', 'martie', 'aprilie', 'mai', 'iunie',
    'iulie', 'august', 'septembrie', 'octombrie', 'noiembrie', 'decembrie',
]

_fp_repo = FoiParcursRepository()
_veh_repo = FPVehicleRepository()


def _period(c: dict):
    """(year, month) for a session — explicit columns, falling back to created_at.

    `get_contracts` returns rows via `dict_from_row`, which serializes `created_at`
    to an ISO *string*, so parse it with `_as_dt` rather than isinstance-checking
    for `datetime` (that check always failed on string rows, dropping every session
    with NULL year/month — i.e. all test-drive sessions — and blanking the sheet)."""
    dt = _as_dt(c.get('created_at'))
    return (c.get('year') or (dt.year if dt else None),
            c.get('month') or (dt.month if dt else None))


def _as_dt(v):
    if isinstance(v, datetime):
        return v
    try:
        return datetime.fromisoformat(str(v).replace('Z', '')) if v else None
    except ValueError:
        return None


def _fmt_date(v) -> str:
    dt = _as_dt(v)
    return dt.strftime('%d.%m.%Y') if dt else (str(v) if v else '')


def _fmt_dt(v) -> str:
    """Date + time, 'dd.mm.yyyy HH:MM' — '' when missing."""
    dt = _as_dt(v)
    return dt.strftime('%d.%m.%Y %H:%M') if dt else ''


def _iso_date(v) -> str:
    """ISO date 'YYYY-MM-DD' (for matching a session against an event interval)."""
    dt = _as_dt(v)
    return dt.strftime('%Y-%m-%d') if dt else ''


def aggregate_month(vin: str, year: int, month: int) -> dict:
    """Collect the locked-facts view of one car's month of driving sessions."""
    rows, _ = _fp_repo.get_contracts(vin=vin, per_page=2000, lean=True)
    sessions = [c for c in rows if _period(c) == (year, month)]
    # chronological by drive date (departure, else created)
    sessions.sort(key=lambda c: str(c.get('departure_datetime') or c.get('created_at') or ''))

    veh = _veh_repo.get_by_vin(vin) or {}
    company_id = sessions[0].get('company_id') if sessions else veh.get('company_id')
    company_name = (sessions[0].get('company_name') if sessions else '') or ''

    prestator = _PRESTATOR_FALLBACK
    if company_id:
        try:
            from core.organization.repositories.company_repository import CompanyRepository
            company_row = CompanyRepository().get(company_id)
            if company_row and (company_row.get('company') or '').strip():
                company_name = company_row.get('company').strip()
                phone = ''
                try:
                    from ..dealer_config import get_dealer_config
                    phone = get_dealer_config(company_name, veh.get('brand') or '').get('phone', '')
                except Exception:
                    logger.warning('dealer_config lookup failed', exc_info=True)
                prestator = _build_prestator_intro(company_row, phone)
        except Exception:
            logger.warning('Prestator lookup failed for company_id=%s', company_id, exc_info=True)

    # Traseu rule inputs from the Settings tab: TD max distance (Route KM Limits)
    # + the manually-configured Comodat routes (Itinerary Routes).
    td_max = 50
    comodat_routes = []
    if company_id:
        try:
            km_cfg = _fp_repo.query_one('SELECT td_km_max FROM fp_km_configs WHERE company_id=%s', (company_id,))
            if km_cfg and km_cfg.get('td_km_max'):
                td_max = int(km_cfg['td_km_max'])
        except Exception:
            logger.warning('td_km_max lookup failed', exc_info=True)
        try:
            rr = _fp_repo.query_all(
                "SELECT itinerary FROM fp_routes WHERE company_id=%s AND route_type='Comodat' ORDER BY id",
                (company_id,))
            comodat_routes = [r['itinerary'] for r in (rr or []) if r.get('itinerary')]
        except Exception:
            logger.warning('comodat routes lookup failed', exc_info=True)

    # Tied promo project per session (already set in the driving-session module).
    project_by_id = {}
    try:
        pr = _fp_repo.query_all(
            'SELECT fp.id, mp.name AS project FROM foi_de_parcurs fp '
            'LEFT JOIN mkt_projects mp ON mp.id = fp.mkt_project_id WHERE fp.vin=%s', (vin,))
        project_by_id = {r['id']: (r.get('project') or '') for r in (pr or [])}
    except Exception:
        logger.warning('mkt_project lookup failed', exc_info=True)

    model_label = ' '.join(x for x in (veh.get('mark'), veh.get('model')) if x) or 'vehicul'
    ci = 0  # rotates through the configured Comodat routes
    trips = []
    for i, c in enumerate(sessions):
        dist = c.get('distance_km') or 0
        is_td = dist <= td_max
        project = project_by_id.get(c.get('id'), '')
        if is_td:
            traseu = f'Test Drive {model_label}'
        elif not project:
            # Comodat with no promo project → deterministic Settings route
            base = 'Deplasare în interes de serviciu'
            if comodat_routes:
                traseu = f'{base} — {comodat_routes[ci % len(comodat_routes)]}'
                ci += 1
            else:
                traseu = base
        else:
            # Comodat tied to a promo project — deterministic fallback; the AI
            # rephrases this into the promovare framing for the PDF.
            traseu = f'Deplasare în interes de serviciu — participare la {project}, în scop de promovare'
        trips.append({
            'id': i,
            'date': _fmt_date(c.get('departure_datetime') or c.get('created_at')),
            'iso': _iso_date(c.get('departure_datetime') or c.get('created_at')),
            'plecare': _fmt_dt(c.get('departure_datetime')),
            # Sosire: the return date/time; if the return isn't recorded, assume a
            # same-day trip and show the departure date (a multi-day return keeps
            # its own different date).
            'sosire': _fmt_dt(c.get('return_datetime')) or _fmt_date(c.get('departure_datetime')),
            'km_start': c.get('km_start') or 0,
            'km_end': c.get('km_end') or 0,
            'distance_km': dist,
            'is_td': is_td,
            'project': project,
            'route_type': c.get('route_type') or '',
            'driver': (c.get('client_name') or c.get('advisor_name') or '').strip(),
            'itinerary': (c.get('itinerary') or '').strip(),
            'traseu': traseu,
            'fuel_consumed': float(c.get('fuel_consumed_liters') or 0),
        })

    total_km = _span_km(trips)
    km_start = min((t['km_start'] for t in trips), default=0)
    km_end = max((t['km_end'] for t in trips), default=0)
    clients = len({t['driver'] for t in trips if t['driver']})
    consum_efectiv = round(sum(t['fuel_consumed'] for t in trips), 2)

    return {
        'company': {'id': company_id, 'name': company_name, 'prestator': prestator},
        'vehicle': {
            'vin': vin,
            'make': veh.get('mark') or '',
            'model': veh.get('model') or '',
            'fuel_type': veh.get('fuel_type') or '',
            'category': veh.get('category') or '',
            'registration_number': veh.get('registration_number') or (sessions[0].get('registration_number') if sessions else '') or '',
        },
        'period': {'year': year, 'month': month, 'label': f'{_MONTHS_RO[month]} {year}' if 1 <= month <= 12 else f'{month}/{year}'},
        'trips': trips,
        'totals': {'km': total_km, 'km_start': km_start, 'km_end': km_end,
                   'sessions': len(trips), 'clients': clients, 'consum_efectiv': consum_efectiv},
    }


# ── Prose: deterministic template (default) with optional AI polish ──
# The route sheet is structured data, so each trip already has a deterministic
# `traseu` (from the Settings rules) and the monthly summary is templated from
# the totals. When FP_ROUTESHEET_AI is on, the AI rephrases Comodat/promo trips
# and may replace the summary. Returns {'trips': {id: text}, 'summary'}.

def _template_summary(data: dict) -> str:
    """Factual monthly summary from the structured totals — describes only the
    activity that happened (never the absence of a trip category)."""
    tot = data['totals']
    td = sum(1 for t in data['trips'] if t['is_td'])
    comodat = tot['sessions'] - td
    parts = []
    if td:
        parts.append(f'{td} test drive')
    if comodat:
        parts.append(f'{comodat} comodat')
    if not parts:
        return ''
    v = data['vehicle']
    return (f"În {data['period']['label']}, {v['make']} {v['model']} a efectuat "
            f"{tot['sessions']} curse ({', '.join(parts)}), total {tot['km']} km.")


def _ai_prose(data: dict) -> dict:
    """Deterministic template prose (each trip's `traseu` + a factual summary).
    The AI is an OPTIONAL polish pass (FP_ROUTESHEET_AI) that rephrases Comodat
    promo trips; on failure or when disabled the template stands. Never raises."""
    template = {'trips': {t['id']: (t['traseu'] or '').strip() for t in data['trips']},
                'summary': _template_summary(data)}
    if not _AI_ENHANCE or not data['trips']:
        return template
    fallback = template
    # Normalize events to {name, start, end}; an event can be tied to a session
    # ONLY if the session's date falls within [start, end].
    events = []
    for e in (data.get('events') or []):
        name = (e.get('name') or '').strip()
        if not name:
            continue
        start = (e.get('start') or e.get('date') or '').strip()
        end = (e.get('end') or e.get('date') or start).strip()
        events.append({'name': name, 'start': start, 'end': end})

    def _eligible_event(iso: str) -> str:
        """Name of an event whose interval contains the session's date, else ''."""
        if not iso:
            return ''
        for ev in events:
            if ev['start'] and ev['start'] <= iso <= (ev['end'] or ev['start']):
                return ev['name']
        return ''

    # Every Comodat session may be tied to its own project or an in-interval event.
    to_fill = [
        {'id': t['id'], 'data': t['date'], 'km_parcursi': session_actual_km(t),
         'sofer': t['driver'], 'proiect_promovare': t['project'],
         'eveniment_eligibil': _eligible_event(t.get('iso', ''))}
        for t in data['trips'] if not t['is_td']
    ]
    # No Comodat/promo activity → nothing to summarize. Skip the AI so it can't
    # write filler like "vehiculul nu a înregistrat curse de tip Comodat".
    if not to_fill:
        return fallback

    system = (
        'Ești asistent pentru foi de parcurs auto (document legal românesc). '
        'Nu inventa și nu modifica numere, date sau kilometraj. '
        'Pentru fiecare cursă de tip Comodat, compune un text scurt de traseu/scop care '
        'leagă deplasarea de participarea la un eveniment/proiect de promovare '
        '(ex: "Deplasare în interes de serviciu — participare la {eveniment}, în scop de promovare"). '
        'Folosește proiect_promovare al cursei dacă există; altfel folosește exact '
        'eveniment_eligibil (deja filtrat pe data cursei). NU lega o cursă de un eveniment '
        'dacă eveniment_eligibil este gol — în acest caz păstrează un scop generic. '
        'Scrie și un rezumat lunar de 1-2 fraze care descrie DOAR activitatea reală; '
        'NU menționa absența unui tip de cursă sau lipsa deplasărilor. '
        'Răspunde STRICT în JSON, fără alt text.'
    )
    prompt = (
        'Vehicul: {make} {model} ({vin}). Perioada: {period}.\n'
        'Curse Comodat (folosește exact aceste id-uri):\n{trips}\n\n'
        'Returnează JSON: {{"trips": {{"<id>": "traseu/scop"}}, "summary": "rezumat lunar"}}'
    ).format(
        make=data['vehicle']['make'], model=data['vehicle']['model'], vin=data['vehicle']['vin'],
        period=data['period']['label'], trips=json.dumps(to_fill, ensure_ascii=False),
    )

    try:
        from ai_agent.services.llm_client import ask
        raw = ask(prompt, system=system, model=_MODEL, max_tokens=1500)
        match = re.search(r'\{.*\}', raw or '', re.DOTALL)
        parsed = json.loads(match.group(0)) if match else {}
        ai_trips = parsed.get('trips', {}) or {}
        out_trips = {}
        for t in data['trips']:
            ai_txt = ai_trips.get(str(t['id'])) or ai_trips.get(t['id'])
            out_trips[t['id']] = (ai_txt or t['traseu'] or '').strip()
        return {'trips': out_trips, 'summary': (parsed.get('summary') or template['summary']).strip()}
    except Exception:
        logger.warning('Route-sheet AI enhance failed; using template prose', exc_info=True)
        return fallback


def session_actual_km(row: dict) -> int:
    """Actual distance a session drove = odometer delta (km_end − km_start).
    0 while the car is still out (km_end == km_start placeholder). This is what
    an export shows — never the advisor's entered estimate (distance_km)."""
    return int(row.get('km_end') or 0) - int(row.get('km_start') or 0)


def _span_km(trips: list) -> int:
    """Total physical distance the car moved across the sheet = odometer span
    (max km_end − min km_start). Authoritative even when individual session
    ranges overlap on bad data; 0 when there are no trips."""
    if not trips:
        return 0
    return (max(int(t.get('km_end') or 0) for t in trips)
            - min(int(t.get('km_start') or 0) for t in trips))


def _rows_with_gaps(trips: list) -> list:
    """Order trips by odometer and interleave gap markers wherever the odometer
    jumps between logged sessions (km moved without a logged drive). Gap dict:
    {'gap': True, 'date', 'km_start', 'km_end', 'distance_km'}; trip:
    {'gap': False, 'trip': <trip>}. The gap's date is the session that revealed it."""
    ordered = sorted(trips, key=lambda t: (t['km_start'], t['km_end']))
    rows = []
    prev_end = None
    for t in ordered:
        if prev_end is not None and t['km_start'] > prev_end:
            rows.append({'gap': True, 'date': t['date'], 'km_start': prev_end,
                         'km_end': t['km_start'], 'distance_km': t['km_start'] - prev_end})
        rows.append({'gap': False, 'trip': t})
        prev_end = max(prev_end or 0, t['km_end'])
    return rows


# ── HTML skeleton (locked numbers) + Playwright render ──

def _fuel_section_html(unit: str, norma, entries: list, km, consum_efectiv=None) -> str:
    """One consumption section — Combustibil (l) or Energie (kWh) — as HTML.
    A hybrid renders both; every entry carries its receipt value (lei)."""
    e = html.escape
    is_e = unit == 'kWh'
    title = 'Energie' if is_e else 'Combustibil'
    qty_h = 'kWh' if is_e else 'Litri'
    op = 'încărcare' if is_e else 'alimentare'
    op_pl = 'încărcări' if is_e else 'alimentări'
    rows = ''.join(
        f'<tr><td>{e(str(a.get("date", "") or "—"))}</td>'
        f'<td>{e(str(a.get("bon", "") or "—"))}</td>'
        f'<td class="n">{float(a.get("liters", 0) or 0):g}</td>'
        f'<td class="n">{float(a.get("lei", 0) or 0):g}</td></tr>'
        for a in entries
    ) or f'<tr><td colspan="4" class="empty">Fără {op_pl} înregistrate.</td></tr>'
    total = round(sum(float(a.get('liters', 0) or 0) for a in entries), 2)
    cost = round(sum(float(a.get('lei', 0) or 0) for a in entries), 2)
    pret = round(cost / total, 2) if total else None
    consum_normat = round(float(norma) * km / 100, 2) if norma else None
    kv = [
        f'<tr><td class="k">Normă consum</td><td>{norma if norma is not None else "—"} {unit}/100 km</td></tr>',
        f'<tr><td class="k">Consum normat</td><td>{consum_normat if consum_normat is not None else "—"} {unit}</td></tr>',
    ]
    if consum_efectiv is not None:
        kv.append(f'<tr><td class="k">Consum efectiv</td><td>{consum_efectiv:g} {unit}</td></tr>')
    kv.append(f'<tr><td class="k">Cost total {title.lower()}</td><td>{cost:g} lei</td></tr>')
    kv.append(f'<tr><td class="k">Preț mediu</td><td>{pret if pret is not None else "—"} lei/{unit}</td></tr>')
    return f"""<div class="fuel-section"><div class="fuel-title">{title}</div>
<div class="fuel">
  <table class="kv">{''.join(kv)}</table>
  <table class="trips alim">
    <thead><tr><th>Data {op}</th><th>Bon fiscal</th><th>{qty_h}</th><th>Valoare (lei)</th></tr></thead>
    <tbody>{rows}
      <tr class="totals"><td>Total</td><td></td><td class="n">{total:g}</td><td class="n">{cost:g}</td></tr>
    </tbody>
  </table>
</div></div>"""


def _xlsx_fuel_section(ws, start_row, unit, norma, entries, km, consum_efectiv, head, fill, bold):
    """Write one Combustibil (l) / Energie (kWh) section to the worksheet; returns
    the next free row so a hybrid can stack both sections."""
    from openpyxl.styles import Alignment
    is_e = unit == 'kWh'
    title = 'Energie' if is_e else 'Combustibil'
    op = 'încărcare' if is_e else 'alimentare'
    consum_normat = round(float(norma) * km / 100, 2) if norma else None
    cost = round(sum(float(a.get('lei', 0) or 0) for a in entries), 2)
    total = round(sum(float(a.get('liters', 0) or 0) for a in entries), 2)
    pret = round(cost / total, 2) if total else None
    row = start_row
    ws.cell(row=row, column=1, value=title).font = bold
    ws.cell(row=row + 1, column=1, value=f'Normă consum ({unit}/100km)'); ws.cell(row=row + 1, column=2, value=float(norma) if norma else None)
    ws.cell(row=row + 2, column=1, value=f'Consum normat ({unit})'); ws.cell(row=row + 2, column=2, value=consum_normat)
    row += 3
    if consum_efectiv is not None:
        ws.cell(row=row, column=1, value=f'Consum efectiv ({unit})'); ws.cell(row=row, column=2, value=consum_efectiv); row += 1
    ws.cell(row=row, column=1, value=f'Cost total {title.lower()} (lei)'); ws.cell(row=row, column=2, value=cost); row += 1
    ws.cell(row=row, column=1, value=f'Preț mediu (lei/{unit})'); ws.cell(row=row, column=2, value=pret); row += 2
    for col, h in enumerate([f'Data {op}', 'Bon fiscal', 'kWh' if is_e else 'Litri', 'Valoare (lei)'], start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = head; cell.fill = fill; cell.alignment = Alignment(horizontal='center')
    row += 1
    for a in entries:
        ws.cell(row=row, column=1, value=str(a.get('date', '') or ''))
        ws.cell(row=row, column=2, value=str(a.get('bon', '') or ''))
        ws.cell(row=row, column=3, value=float(a.get('liters', 0) or 0))
        ws.cell(row=row, column=4, value=float(a.get('lei', 0) or 0))
        row += 1
    ws.cell(row=row, column=1, value='Total').font = bold
    ws.cell(row=row, column=3, value=total).font = bold
    ws.cell(row=row, column=4, value=cost).font = bold
    return row + 2


def _skeleton_html(data: dict, prose: dict) -> str:
    e = html.escape
    v = data['vehicle']
    rows = []
    for r in _rows_with_gaps(data['trips']):
        if r['gap']:
            rows.append(
                '<tr class="gap">'
                f'<td class="c">{e(r["date"])}</td>'
                '<td class="c">—</td>'
                '<td class="route">Gap kilometraj (nejustificat)</td>'
                '<td>—</td>'
                f'<td class="n">{r["km_start"]:,}</td>'
                f'<td class="n">{r["km_end"]:,}</td>'
                f'<td class="n">{r["distance_km"]:,}</td>'
                '</tr>'
            )
            continue
        t = r['trip']
        rows.append(
            '<tr>'
            f'<td class="c">{e(t.get("plecare") or "—")}</td>'
            f'<td class="c">{e(t.get("sosire") or "—")}</td>'
            f'<td class="route">{e(prose["trips"].get(t["id"], "") or t.get("traseu") or "—")}</td>'
            f'<td>{e(t["driver"] or "—")}</td>'
            f'<td class="n">{t["km_start"]:,}</td>'
            f'<td class="n">{t["km_end"]:,}</td>'
            f'<td class="n">{session_actual_km(t):,}</td>'
            '</tr>'
        )
    if not rows:
        rows.append('<tr><td colspan="7" class="empty">Nicio sesiune în această lună.</td></tr>')
    tot = data['totals']
    summary_html = f'<p class="summary">{e(prose["summary"])}</p>' if prose.get('summary') else ''

    # Fuel/charging sections. A car may have a fuel tank (Benzină/Diesel/Hybrid)
    # and/or a battery (Electric/Hybrid) — a hybrid shows BOTH sections. Each
    # alimentare entry carries a unit ('l' by default, 'kWh' for charging).
    fuel = data.get('fuel', {}) or {}
    alim = fuel.get('alimentari') or []
    fuel_entries = [a for a in alim if (a.get('unit') or 'l') != 'kWh']
    energy_entries = [a for a in alim if a.get('unit') == 'kWh']
    ft = v.get('fuel_type') or ''
    uses_tank = ft in ('Benzina', 'Diesel', 'Hybrid')
    uses_batt = ft in ('Electric', 'Hybrid')
    if not uses_tank and not uses_batt:
        uses_tank = True  # unknown fuel_type → treat as a fuel car
    sections = []
    if uses_tank:
        sections.append(_fuel_section_html('l', fuel.get('norma'), fuel_entries, tot['km'], tot.get('consum_efectiv', 0)))
    if uses_batt:
        sections.append(_fuel_section_html('kWh', fuel.get('norma_energie'), energy_entries, tot['km'], None))
    fuel_block = '\n'.join(sections)

    # Signatures — Întocmit = generating user's stored signature (image);
    # Aprobat = blank line (superior signs manually).
    sigs = data.get('signatures', {}) or {}
    intocmit_img = (f'<img class="sig-img" src="{sigs["intocmit"]}" />'
                    if sigs.get('intocmit') else '<div class="sig-ph"></div>')
    intocmit_name = f' — {e(sigs["intocmit_name"])}' if sigs.get('intocmit_name') else ''
    sig_block = (
        '<div class="sign">'
        f'<div class="box">{intocmit_img}<div class="line">Întocmit{intocmit_name}</div></div>'
        '<div class="box"><div class="sig-ph"></div><div class="line">Aprobat</div></div>'
        '</div>'
    )

    return f"""<!doctype html><html lang="ro"><head><meta charset="utf-8"><style>
* {{ box-sizing: border-box; }}
body {{ font-family: 'Helvetica Neue', Arial, sans-serif; color:#1a1a2e; font-size:11px; margin:0; }}
.doc {{ padding:0; }}
h1 {{ font-size:18px; text-align:center; margin:0 0 2px; }}
.sub {{ text-align:center; color:#555; margin:0 0 10px; font-size:11px; }}
.rule {{ border:0; border-top:1.5px solid #1a1a2e; margin:8px 0 12px; }}
.prestator {{ font-size:9px; text-align:justify; color:#333; line-height:1.4; margin-bottom:12px; }}
.meta {{ width:100%; margin-bottom:10px; border-collapse:collapse; }}
.meta td {{ padding:2px 6px; font-size:11px; }}
.meta td.k {{ color:#555; width:120px; }}
table.trips {{ width:100%; border-collapse:collapse; margin-top:6px; }}
table.trips th, table.trips td {{ border:0.5px solid #cfcfd8; padding:4px 6px; font-size:10px; }}
table.trips th {{ background:#1a1a2e; color:#fff; text-align:left; font-weight:600; }}
table.trips td.n {{ text-align:right; white-space:nowrap; }}
table.trips td.c {{ text-align:center; white-space:nowrap; }}
table.trips td.route {{ color:#222; }}
tr.totals td {{ font-weight:700; background:#f2f2f6; }}
tr.gap td {{ background:#fff7ed; font-style:italic; color:#9a6a00; }}
.empty {{ text-align:center; color:#888; }}
.summary {{ margin-top:12px; font-size:10.5px; line-height:1.5; }}
.fuel-section {{ margin-top:12px; }}
.fuel-title {{ font-weight:700; font-size:11px; margin-bottom:2px; color:#1a1a2e; }}
.fuel {{ display:flex; gap:16px; align-items:flex-start; }}
.fuel .kv {{ border-collapse:collapse; }}
.fuel .kv td {{ padding:3px 8px; font-size:10px; border:0.5px solid #cfcfd8; }}
.fuel .kv td.k {{ color:#555; background:#f7f7fa; }}
table.alim {{ flex:1; margin-top:0; }}
.sign {{ margin-top:30px; display:flex; justify-content:space-around; font-size:10px; }}
.sign .box {{ width:42%; text-align:center; }}
.sign .sig-img {{ height:44px; max-width:80%; object-fit:contain; margin-bottom:2px; }}
.sign .sig-ph {{ height:44px; }}
.sign .line {{ border-top:0.5px solid #999; padding-top:4px; color:#555; }}
</style></head><body><div class="doc">
<h1>Foaie de Parcurs</h1>
<p class="sub">{e(v['make'])} {e(v['model'])} • {e(v['vin'])} • {e(data['period']['label'])}</p>
<hr class="rule">
<p class="prestator">{e(data['company']['prestator'])}</p>
<table class="meta">
  <tr><td class="k">Companie</td><td>{e(data['company']['name'] or '—')}</td>
      <td class="k">Nr. înmatriculare</td><td>{e(v['registration_number'] or '—')}</td></tr>
  <tr><td class="k">Vehicul</td><td>{e((v['make'] + ' ' + v['model']).strip() or '—')}</td>
      <td class="k">VIN</td><td>{e(v['vin'])}</td></tr>
  <tr><td class="k">Categorie</td><td colspan="3">{e(v.get('category') or '—')}</td></tr>
</table>
<table class="trips">
  <thead><tr>
    <th>Plecare</th><th>Sosire</th><th>Traseu / Scop</th><th>Șofer</th>
    <th>KM start</th><th>KM end</th><th>KM parcurși</th>
  </tr></thead>
  <tbody>
    {''.join(rows)}
    <tr class="totals"><td colspan="4">Total — {tot['sessions']} sesiuni, {tot['clients']} șoferi</td>
      <td class="n">{tot['km_start']:,}</td><td class="n">{tot['km_end']:,}</td><td class="n">{tot['km']:,}</td></tr>
  </tbody>
</table>
{fuel_block}
{summary_html}
{sig_block}
</div></body></html>"""


def _html_to_pdf_bytes(html_str: str) -> bytes:
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        browser = p.chromium.launch()
        try:
            page = browser.new_page()
            page.set_content(html_str, wait_until='load')
            return page.pdf(
                format='A4', print_background=True,
                margin={'top': '16mm', 'bottom': '16mm', 'left': '16mm', 'right': '16mm'},
            )
        finally:
            browser.close()


# ── Storage: durable per (vin, year, month) record in fp_route_sheets ──

from core.base_repository import BaseRepository  # noqa: E402
_store = BaseRepository()


def _gap_context(vin: str):
    """(company_id, registration_number, tank_liters, td_km_max) used when
    inserting gap-fill rows — shared by redistribute_gap and absorb_gap."""
    veh = _veh_repo.get_by_vin(vin) or {}
    ref = _fp_repo.query_one('SELECT company_id, registration_number FROM foi_de_parcurs WHERE vin=%s LIMIT 1', (vin,)) or {}
    company_id = ref.get('company_id') or veh.get('company_id')
    if not company_id:
        raise ValueError('Company nu a putut fi determinată pentru acest VIN')
    reg = ref.get('registration_number') or veh.get('registration_number') or ''
    tank = veh.get('fuel_tank_capacity_liters') or 50
    td_max = 50
    try:
        km_cfg = _fp_repo.query_one('SELECT td_km_max FROM fp_km_configs WHERE company_id=%s', (company_id,))
        if km_cfg and km_cfg.get('td_km_max'):
            td_max = int(km_cfg['td_km_max'])
    except Exception:
        logger.warning('td_km_max lookup failed', exc_info=True)
    return company_id, reg, tank, td_max


def _insert_gap_fill(vin, year, month, ctx, item, user_name) -> int:
    """Insert one synthetic 'gap-fill' session covering item km_start→km_end.
    item may also carry advisor_name / client_signature / driver_license_* for a
    fully-documented "client extra". Idempotent per km range. Returns 0 or 1."""
    company_id, reg, tank, td_max = ctx
    ks, ke = int(item['km_start']), int(item['km_end'])
    if ke <= ks:
        return 0
    dist = ke - ks
    client = (item.get('client_name') or '').strip()
    date = (item.get('date') or '').strip()
    dep = f'{date} 10:00:00' if date else None
    route_type = 'TD' if dist <= td_max else 'Comodat'
    # Optional "client extra" documentation — advisor (consilier), the signed
    # client signature and the driver-license photo/number/expiry.
    advisor = (item.get('advisor_name') or user_name or 'Redistribuire')
    client_sig = item.get('client_signature') or ''
    dl_photo = item.get('driver_license_photo') or None
    dl_number = (item.get('driver_license_number') or '').strip() or None
    dl_expiry = (item.get('driver_license_expiry') or '').strip() or None
    cid = f'GAPFILL_{re.sub(r"[^A-Za-z0-9]", "", vin)[-6:]}_{year}{month:02d}_{ks}_{ke}'
    _fp_repo.execute(
        '''INSERT INTO foi_de_parcurs
             (contract_id, vin, company_id, year, month, route_type, slot_number,
              km_start, km_end, distance_km, registration_number,
              fuel_tank_capacity_liters, fuel_gauge_start_level, fuel_gauge_end_level,
              fuel_start_liters, fuel_end_liters, fuel_consumed_liters,
              status, advisor_name, client_name, itinerary, departure_datetime, source,
              client_signature, driver_license_photo, driver_license_number, driver_license_expiry)
           VALUES (%s,%s,%s,%s,%s,%s,0,%s,%s,%s,%s,%s,'1','1',0,0,0,
                   'COMPLETED',%s,%s,'',%s,'gap-fill',%s,%s,%s,%s)
           ON CONFLICT (contract_id) DO UPDATE SET
             km_start=EXCLUDED.km_start, km_end=EXCLUDED.km_end,
             distance_km=EXCLUDED.distance_km, client_name=EXCLUDED.client_name,
             departure_datetime=EXCLUDED.departure_datetime, route_type=EXCLUDED.route_type,
             advisor_name=EXCLUDED.advisor_name, client_signature=EXCLUDED.client_signature,
             driver_license_photo=EXCLUDED.driver_license_photo,
             driver_license_number=EXCLUDED.driver_license_number,
             driver_license_expiry=EXCLUDED.driver_license_expiry''',
        (cid, vin, company_id, year, month, route_type, ks, ke, dist, reg, tank,
         advisor, client, dep, client_sig, dl_photo, dl_number, dl_expiry),
    )
    return 1


def redistribute_gap(vin: str, year: int, month: int, items: list, user_name=None) -> int:
    """Insert up to 3 documented "client extra" gap-fill sessions. Each item:
    {date, client_name, km_start, km_end} and OPTIONALLY {advisor_name,
    client_signature, driver_license_photo, driver_license_number,
    driver_license_expiry}. Returns the number inserted. Idempotent per km range."""
    items = [it for it in (items or []) if it][:3]
    if not items:
        return 0
    ctx = _gap_context(vin)
    inserted = 0
    for it in items:
        try:
            inserted += _insert_gap_fill(vin, year, month, ctx, it, user_name)
        except (KeyError, TypeError, ValueError):
            continue
    return inserted


def absorb_gap(vin: str, year: int, month: int, before_id: int, after_id: int,
               before_km: int, after_km: int, middles=None, user_name=None) -> dict:
    """Close an odometer gap by TILING it left→right, leaving no orphan km:
      • `before_km` grows the earlier session's km_end forward,
      • each `middles` entry ({km, client_name?, date?}) becomes a documented
        gap-fill row filling the next contiguous slice (max 3, in order),
      • `after_km` pulls the later session's km_start back.
    before_km + Σmiddles + after_km must equal the gap. distance_km/route_type
    are recomputed on the two bounding rows and an audit note is stamped."""
    before = _fp_repo.query_one(
        'SELECT id, km_start, km_end, company_id FROM foi_de_parcurs WHERE id=%s AND vin=%s',
        (before_id, vin)) or {}
    after = _fp_repo.query_one(
        'SELECT id, km_start, km_end, company_id FROM foi_de_parcurs WHERE id=%s AND vin=%s',
        (after_id, vin)) or {}
    if not before or not after:
        raise ValueError('Sesiunile care mărginesc gap-ul nu au fost găsite')

    b_start, b_end = int(before['km_start']), int(before['km_end'])
    a_start, a_end = int(after['km_start']), int(after['km_end'])
    gap = a_start - b_end
    if gap <= 0:
        raise ValueError('Nu există un gap între aceste sesiuni')

    middles = [m for m in (middles or []) if m][:3]
    try:
        kb, ka = int(before_km), int(after_km)
        mids = [(int(m['km']), (m.get('client_name') or '').strip(), (m.get('date') or '').strip())
                for m in middles]
    except (KeyError, TypeError, ValueError):
        raise ValueError('KM alocați sunt invalizi')
    if kb < 0 or ka < 0 or any(km <= 0 for km, _, _ in mids):
        raise ValueError('KM alocați trebuie să fie pozitivi')
    total = kb + ka + sum(km for km, _, _ in mids)
    if total != gap:
        raise ValueError(f'Suma alocată ({total} km) trebuie să fie egală cu gap-ul ({gap} km)')

    ctx = _gap_context(vin)
    td_max = ctx[3]

    def _extend(row_id, new_start, new_end, added):
        dist = new_end - new_start
        rtype = 'TD' if dist <= td_max else 'Comodat'
        _fp_repo.execute(
            '''UPDATE foi_de_parcurs
                 SET km_start=%s, km_end=%s, distance_km=%s, route_type=%s,
                     general_observation = TRIM(BOTH ' ' FROM
                       COALESCE(general_observation, '') || %s)
               WHERE id=%s''',
            (new_start, new_end, dist, rtype, f' [KM ajustat +{added} km prin redistribuire gap]', row_id),
        )

    cursor = b_end
    if kb > 0:
        _extend(before_id, b_start, b_end + kb, kb)
        cursor = b_end + kb
    inserted = 0
    for km, name, date in mids:
        inserted += _insert_gap_fill(
            vin, year, month, ctx,
            {'client_name': name, 'km_start': cursor, 'km_end': cursor + km, 'date': date}, user_name)
        cursor += km
    if ka > 0:
        _extend(after_id, a_start - ka, a_end, ka)
    return {'before_id': before_id, 'after_id': after_id, 'before_km': kb, 'after_km': ka,
            'middles_inserted': inserted, 'gap': gap}


def retile_gap(vin: str, year: int, month: int, allocations: list, user_name=None) -> dict:
    """Distribute a gap across a window of consecutive EXISTING sessions — no new
    rows. `allocations`: ordered [{id, distance}] giving each session its NEW
    distance. The window is re-tiled contiguously from the first session's
    km_start; the sum of new distances must equal the window span
    (first.km_start → last.km_end), which is exactly the original distances plus
    the gap. Each touched session's km/route_type is updated in place."""
    allocations = [a for a in (allocations or []) if a]
    if len(allocations) < 2:
        raise ValueError('Sunt necesare cel puțin două sesiuni')
    rows = []
    for a in allocations:
        r = _fp_repo.query_one(
            'SELECT id, km_start, km_end, company_id FROM foi_de_parcurs WHERE id=%s AND vin=%s',
            (a.get('id'), vin))
        if not r:
            raise ValueError('O sesiune din fereastră nu a fost găsită')
        try:
            dist = int(a.get('distance'))
        except (TypeError, ValueError):
            raise ValueError('Distanțe invalide')
        if dist < 0:
            raise ValueError('Distanțele trebuie să fie pozitive')
        rows.append([r, dist])

    rows.sort(key=lambda x: (int(x[0]['km_start']), int(x[0]['km_end'])))
    first_start = int(rows[0][0]['km_start'])
    last_end = int(rows[-1][0]['km_end'])
    span = last_end - first_start
    total = sum(d for _, d in rows)
    if total != span:
        raise ValueError(f'Suma distanțelor ({total} km) trebuie să fie egală cu intervalul ferestrei ({span} km)')

    company_id = rows[0][0].get('company_id')
    td_max = 50
    try:
        km_cfg = _fp_repo.query_one('SELECT td_km_max FROM fp_km_configs WHERE company_id=%s', (company_id,))
        if km_cfg and km_cfg.get('td_km_max'):
            td_max = int(km_cfg['td_km_max'])
    except Exception:
        logger.warning('td_km_max lookup failed', exc_info=True)

    cursor = first_start
    updated = 0
    for r, dist in rows:
        orig_dist = int(r['km_end']) - int(r['km_start'])
        new_start, new_end = cursor, cursor + dist
        rtype = 'TD' if dist <= td_max else 'Comodat'
        note = f' [KM ajustat +{dist - orig_dist} km prin redistribuire gap]' if dist != orig_dist else ''
        _fp_repo.execute(
            '''UPDATE foi_de_parcurs SET km_start=%s, km_end=%s, distance_km=%s, route_type=%s,
                 general_observation = TRIM(BOTH ' ' FROM COALESCE(general_observation,'') || %s)
               WHERE id=%s''',
            (new_start, new_end, dist, rtype, note, r['id']))
        cursor = new_end
        updated += 1
    return {'sessions': updated, 'span': span}


def get_stored_pdf(vin: str, year: int, month: int) -> bytes | None:
    """The stored PDF bytes for a sheet, or None if not generated yet."""
    row = _store.query_one(
        'SELECT pdf_bytes FROM fp_route_sheets WHERE vin=%s AND year=%s AND month=%s',
        (vin, year, month),
    )
    if row and row.get('pdf_bytes') is not None:
        return bytes(row['pdf_bytes'])
    return None


def list_stored(company_id: int | None, year: int, month: int) -> list:
    """Metadata for every stored sheet in a period (badge + modal prefill)."""
    return _store.query_all(
        'SELECT vin, session_count, total_km, norma_combustibil, norma_energie, alimentari, evenimente, '
        'generated_by_name, generated_at '
        'FROM fp_route_sheets WHERE (%s IS NULL OR company_id=%s) AND year=%s AND month=%s',
        (company_id, company_id, year, month),
    )


def _save_sheet(data: dict, pdf_bytes: bytes, prose: dict, user_id, user_name) -> None:
    from psycopg2 import Binary
    from psycopg2.extras import Json
    fuel = data.get('fuel', {}) or {}
    _store.execute(
        '''INSERT INTO fp_route_sheets
             (vin, company_id, year, month, pdf_bytes, ai_summary, ai_trips_json,
              norma_combustibil, norma_energie, alimentari, evenimente, session_count, total_km,
              generated_by, generated_by_name, updated_at)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
           ON CONFLICT (vin, year, month) DO UPDATE SET
             pdf_bytes=EXCLUDED.pdf_bytes, ai_summary=EXCLUDED.ai_summary,
             ai_trips_json=EXCLUDED.ai_trips_json, norma_combustibil=EXCLUDED.norma_combustibil,
             norma_energie=EXCLUDED.norma_energie,
             alimentari=EXCLUDED.alimentari, evenimente=EXCLUDED.evenimente,
             session_count=EXCLUDED.session_count,
             total_km=EXCLUDED.total_km, generated_by=EXCLUDED.generated_by,
             generated_by_name=EXCLUDED.generated_by_name, updated_at=NOW()''',
        (data['vehicle']['vin'], data['company'].get('id'),
         data['period']['year'], data['period']['month'], Binary(pdf_bytes),
         prose.get('summary', ''), Json(prose.get('trips', {})),
         fuel.get('norma'), fuel.get('norma_energie'), Json(fuel.get('alimentari') or []),
         Json(data.get('events') or []),
         data['totals']['sessions'], data['totals']['km'], user_id, user_name),
    )


# Only a clean base64 image data-URL may be embedded — the signature is rendered
# inside an <img src="…"> by Playwright (full JS execution), so an unvalidated
# value could break out of the attribute (stored-XSS / SSRF in the renderer).
_SIG_DATA_URL_RE = re.compile(r'^data:image/(?:png|jpe?g|gif|webp);base64,[A-Za-z0-9+/=\s]+$')


def _user_signature(user_id) -> str | None:
    """The 'Întocmit' user's stored signature as a validated base64 image
    data-URL, or None if absent/malformed (rejects anything that isn't a clean
    data:image/*;base64 URL — no HTML/attribute breakout possible)."""
    if not user_id:
        return None
    row = _store.query_one('SELECT signature FROM users WHERE id=%s', (user_id,))
    sig = (row or {}).get('signature')
    if not sig:
        return None
    sig = str(sig).strip()
    if not sig.startswith('data:'):
        sig = f'data:image/png;base64,{sig}'
    return sig if _SIG_DATA_URL_RE.match(sig) else None


def generate_and_store(vin: str, year: int, month: int, user_id=None, user_name=None,
                       regenerate: bool = False, norma=None, norma_energie=None,
                       alimentari=None, events=None) -> bytes:
    """Return the stored PDF (unless `regenerate`), else build it with AI +
    Playwright, persist it to fp_route_sheets, and return the bytes. `norma`
    (l/100km) and `alimentari` (list of {date, bon, liters, lei}) are user-entered;
    when `norma` is omitted it falls back to the car's profile normă.
    `events` (list of {name, date}) are promo events the AI ties Comodat sessions
    to. The generating user's stored signature is embedded in the 'Întocmit' box."""
    if not regenerate:
        cached = get_stored_pdf(vin, year, month)
        if cached is not None:
            return cached
    data = aggregate_month(vin, year, month)
    if norma is None or norma_energie is None:  # fall back to the car-profile norms
        veh = _veh_repo.get_by_vin(vin) or {}
        norma = norma if norma is not None else veh.get('norma_combustibil')
        norma_energie = norma_energie if norma_energie is not None else veh.get('norma_energie')
    data['fuel'] = {'norma': norma, 'norma_energie': norma_energie, 'alimentari': alimentari or []}
    data['events'] = events or []
    data['signatures'] = {'intocmit': _user_signature(user_id), 'intocmit_name': user_name or ''}
    prose = _ai_prose(data)
    pdf_bytes = _html_to_pdf_bytes(_skeleton_html(data, prose))
    _save_sheet(data, pdf_bytes, prose, user_id, user_name)
    return pdf_bytes


def render_xlsx(vin: str, year: int, month: int) -> bytes:
    """Deterministic monthly route-sheet workbook (no AI)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    data = aggregate_month(vin, year, month)
    v = data['vehicle']
    wb = Workbook()
    ws = wb.active
    ws.title = 'Foaie de parcurs'

    bold = Font(bold=True)
    head = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='1A1A2E')

    ws['A1'] = 'Foaie de Parcurs'; ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"{v['make']} {v['model']}".strip(); ws['A3'] = f"VIN: {v['vin']}"
    ws['A4'] = f"Nr. înmatriculare: {v['registration_number'] or '—'}  ·  Categorie: {v.get('category') or '—'}"
    ws['A5'] = f"Companie: {data['company']['name'] or '—'}"
    ws['A6'] = f"Perioada: {data['period']['label']}"

    headers = ['Plecare', 'Sosire', 'Traseu / Scop', 'Șofer', 'KM start', 'KM end', 'KM parcurși']
    hrow = 8
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=hrow, column=col, value=h)
        cell.font = head; cell.fill = fill; cell.alignment = Alignment(horizontal='center')

    r = hrow + 1
    for row in _rows_with_gaps(data['trips']):
        if row['gap']:
            ws.cell(row=r, column=1, value=row['date'])
            ws.cell(row=r, column=3, value='Gap kilometraj (nejustificat)')
            ws.cell(row=r, column=5, value=row['km_start'])
            ws.cell(row=r, column=6, value=row['km_end'])
            ws.cell(row=r, column=7, value=row['distance_km'])
            r += 1
            continue
        t = row['trip']
        ws.cell(row=r, column=1, value=t.get('plecare') or '')
        ws.cell(row=r, column=2, value=t.get('sosire') or '')
        ws.cell(row=r, column=3, value=t.get('traseu') or '')
        ws.cell(row=r, column=4, value=t['driver'] or '')
        ws.cell(row=r, column=5, value=t['km_start'])
        ws.cell(row=r, column=6, value=t['km_end'])
        ws.cell(row=r, column=7, value=session_actual_km(t))
        r += 1

    tot = data['totals']
    ws.cell(row=r, column=1, value='Total').font = bold
    ws.cell(row=r, column=4, value=f"{tot['sessions']} sesiuni").font = bold
    ws.cell(row=r, column=5, value=tot['km_start']).font = bold
    ws.cell(row=r, column=6, value=tot['km_end']).font = bold
    ws.cell(row=r, column=7, value=tot['km']).font = bold

    # Fuel/charging sections — Normă + Alimentări from the stored sheet (norms
    # fall back to the car profile). A hybrid renders both Combustibil + Energie.
    veh = _veh_repo.get_by_vin(vin) or {}
    ft = veh.get('fuel_type') or ''
    uses_tank = ft in ('Benzina', 'Diesel', 'Hybrid')
    uses_batt = ft in ('Electric', 'Hybrid')
    if not uses_tank and not uses_batt:
        uses_tank = True
    stored = _store.query_one(
        'SELECT norma_combustibil, norma_energie, alimentari FROM fp_route_sheets WHERE vin=%s AND year=%s AND month=%s',
        (vin, year, month),
    ) or {}
    norma = stored.get('norma_combustibil') if stored.get('norma_combustibil') is not None else veh.get('norma_combustibil')
    norma_energie = stored.get('norma_energie') if stored.get('norma_energie') is not None else veh.get('norma_energie')
    alimentari = stored.get('alimentari') or []
    fuel_entries = [a for a in alimentari if (a.get('unit') or 'l') != 'kWh']
    energy_entries = [a for a in alimentari if a.get('unit') == 'kWh']

    row = r + 3
    if uses_tank:
        row = _xlsx_fuel_section(ws, row, 'l', norma, fuel_entries, tot['km'], tot.get('consum_efectiv', 0), head, fill, bold)
    if uses_batt:
        row = _xlsx_fuel_section(ws, row, 'kWh', norma_energie, energy_entries, tot['km'], None, head, fill, bold)

    widths = [20, 20, 42, 22, 12, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
