"""Bulk import of driving sessions from an Excel file (tenant-scoped, keyed by
VIN). Unknown VINs auto-create the car; duplicates are skipped. See
docs/superpowers/specs/2026-07-22-session-bulk-import-design.md"""
import io
import re
import logging
from datetime import datetime

from ..repositories.vehicle_repository import FPVehicleRepository
from ..repositories.foi_parcurs_repository import FoiParcursRepository

logger = logging.getLogger(__name__)

_veh_repo = FPVehicleRepository()
_fp_repo = FoiParcursRepository()

SESSION_COLUMNS = [
    'VIN', 'Marcă', 'Model', 'Nr. înmatriculare', 'Combustibil',
    'Capacitate rezervor (L)', 'Brand', 'Plecare', 'Sosire',
    'KM start', 'KM end', 'Șofer',
    'Telefon', 'Email', 'Permis', 'Link Poza Permis', 'Link Semnatura',
]

_DT_FORMATS = (
    '%d.%m.%Y %H:%M', '%d.%m.%Y', '%Y-%m-%d %H:%M', '%Y-%m-%d', '%Y-%m-%dT%H:%M',
    '%b %d, %Y %H:%M', '%b %d, %Y', '%B %d, %Y %H:%M', '%B %d, %Y',  # "Jul 20, 2026 10:22"
)


def parse_dt(v):
    """Parse a cell value to datetime, or None."""
    if isinstance(v, datetime):
        return v
    s = str(v or '').strip()
    if not s:
        return None
    for fmt in _DT_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _int(v):
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return None


def row_error(row: dict, vehicle: dict | None, company_id: int) -> str | None:
    """Validation message for a row, or None when valid."""
    vin = str(row.get('VIN') or '').strip()
    if not vin:
        return 'VIN lipsă'
    if vehicle is not None and vehicle.get('company_id') not in (None, company_id):
        return 'VIN aparține altei companii'
    if vehicle is None and not (str(row.get('Marcă') or '').strip() and str(row.get('Model') or '').strip()):
        return 'Mașină nouă necesită Marcă și Model'
    if parse_dt(row.get('Plecare')) is None:
        return 'Plecare invalidă'
    ks, ke = _int(row.get('KM start')), _int(row.get('KM end'))
    if ks is None or ke is None or ke <= ks:
        return 'KM invalizi (KM end trebuie > KM start)'
    sos = parse_dt(row.get('Sosire'))
    if sos is not None and sos < parse_dt(row.get('Plecare')):
        return 'Sosire înainte de Plecare'
    return None


def build_template_xlsx(company_id: int) -> bytes:
    """Excel template: a 'Sesiuni' sheet (headers + example) and a 'Mașini'
    reference sheet with the company's cars."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sesiuni'
    head = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='1A1A2E')
    for col, name in enumerate(SESSION_COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = head; c.fill = fill
    example = ['WAUZZZ00000000000', 'Audi', 'A4', 'B 123 ABC', 'Diesel', 55, 'Audi',
               '02.07.2026 10:00', '02.07.2026 12:30', 13000, 13025, 'Ion Popescu',
               '0720000000', 'ion@example.com', 'B00123456',
               'https://…/permis.jpg', 'https://…/semnatura.png']
    for col, val in enumerate(example, start=1):
        ws.cell(row=2, column=col, value=val)
    widths = [20, 12, 14, 16, 12, 16, 12, 18, 18, 10, 10, 18, 14, 22, 14, 34, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ref = wb.create_sheet('Mașini')
    for col, name in enumerate(['Marcă', 'Model', 'VIN'], start=1):
        c = ref.cell(row=1, column=col, value=name)
        c.font = head; c.fill = fill
    cars = [v for v in (_veh_repo.get_all(active_only=False) or []) if v.get('company_id') == company_id]
    for r, v in enumerate(cars, start=2):
        ref.cell(row=r, column=1, value=v.get('mark') or '')
        ref.cell(row=r, column=2, value=v.get('model') or '')
        ref.cell(row=r, column=3, value=v.get('vin') or '')
    ref.column_dimensions['A'].width = 14
    ref.column_dimensions['B'].width = 14
    ref.column_dimensions['C'].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _td_max(company_id: int) -> int:
    try:
        r = _fp_repo.query_one('SELECT td_km_max FROM fp_km_configs WHERE company_id=%s', (company_id,))
        if r and r.get('td_km_max'):
            return int(r['td_km_max'])
    except Exception:
        logger.warning('td_km_max lookup failed', exc_info=True)
    return 50


def _create_vehicle(row: dict, company_id: int) -> None:
    vin = str(row['VIN']).strip()
    _fp_repo.execute(
        '''INSERT INTO fp_vehicles
             (vin, mark, model, registration_number, fuel_type,
              fuel_tank_capacity_liters, brand, company_id, is_active)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,TRUE)
           ON CONFLICT (vin) DO NOTHING''',
        (vin, str(row.get('Marcă') or '').strip() or '—',
         str(row.get('Model') or '').strip() or '—',
         str(row.get('Nr. înmatriculare') or '').strip() or None,
         str(row.get('Combustibil') or '').strip() or 'Diesel',
         _int(row.get('Capacitate rezervor (L)')) or 50,
         str(row.get('Brand') or row.get('Marcă') or '').strip() or None,
         company_id),
    )


def import_sessions(company_id: int, file_bytes: bytes, user_name: str | None) -> dict:
    """Parse the Excel, validate + insert sessions (auto-creating cars),
    returning {'inserted', 'skipped', 'cars_created', 'errors':[{row, message}]}."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb['Sesiuni'] if 'Sesiuni' in wb.sheetnames else wb.active

    td_max = _td_max(company_id)
    inserted = skipped = cars_created = 0
    errors = []
    for i, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if raw is None or all(c in (None, '') for c in raw):
            continue
        row = {SESSION_COLUMNS[j]: (raw[j] if j < len(raw) else '') for j in range(len(SESSION_COLUMNS))}
        vin = str(row.get('VIN') or '').strip()
        vehicle = _veh_repo.get_by_vin(vin) if vin else None
        err = row_error(row, vehicle, company_id)
        if err:
            errors.append({'row': i, 'message': err})
            continue
        if vehicle is None:
            _create_vehicle(row, company_id)
            cars_created += 1
            vehicle = {'fuel_tank_capacity_liters': _int(row.get('Capacitate rezervor (L)')) or 50,
                       'registration_number': str(row.get('Nr. înmatriculare') or '').strip()}

        dep = parse_dt(row['Plecare'])
        sos = parse_dt(row.get('Sosire'))
        ks, ke = _int(row['KM start']), _int(row['KM end'])
        dist = ke - ks
        safe_vin = re.sub(r'[^A-Za-z0-9]', '', vin)
        cid = f'IMPORT_{safe_vin}_{dep.strftime("%Y%m%d")}_{ks}_{ke}'
        if _fp_repo.query_one('SELECT 1 AS x FROM foi_de_parcurs WHERE contract_id=%s', (cid,)):
            skipped += 1
            continue
        route_type = 'TD' if dist <= td_max else 'Comodat'
        tank = (vehicle or {}).get('fuel_tank_capacity_liters') or 50
        reg = (vehicle or {}).get('registration_number') or ''
        def _s(col):
            return str(row.get(col) or '').strip() or None
        _fp_repo.execute(
            '''INSERT INTO foi_de_parcurs
                 (contract_id, vin, company_id, year, month, route_type, slot_number,
                  km_start, km_end, distance_km, registration_number,
                  fuel_tank_capacity_liters, fuel_gauge_start_level, fuel_gauge_end_level,
                  fuel_start_liters, fuel_end_liters, fuel_consumed_liters,
                  status, advisor_name, client_name, client_phone, client_email,
                  driver_license_number, driver_license_photo, client_signature,
                  itinerary, departure_datetime, return_datetime, source)
               VALUES (%s,%s,%s,%s,%s,%s,0,%s,%s,%s,%s,%s,'1','1',0,0,0,
                       'COMPLETED',%s,%s,%s,%s,%s,%s,%s,'',%s,%s,'import')
               ON CONFLICT (contract_id) DO NOTHING''',
            (cid, vin, company_id, dep.year, dep.month, route_type, ks, ke, dist, reg, tank,
             (user_name or 'Import'), str(row.get('Șofer') or '').strip(),
             _s('Telefon'), _s('Email'), _s('Permis'), _s('Link Poza Permis'), _s('Link Semnatura'),
             dep.strftime('%Y-%m-%d %H:%M:%S'),
             sos.strftime('%Y-%m-%d %H:%M:%S') if sos else None),
        )
        inserted += 1
    return {'inserted': inserted, 'skipped': skipped, 'cars_created': cars_created, 'errors': errors}
