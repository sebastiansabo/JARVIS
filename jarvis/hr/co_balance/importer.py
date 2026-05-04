"""CO balance Excel importer.

Reads `Raport_concedii_contracte_lunar - <COMPANY>.xlsx` files and upserts
the 5 static input columns (carry prev year, carry 2y, CIM, vechime, reglare)
into `sincron_co_balance`. All other numeric columns in the file (Total,
Expirate, CO 1..N, Sold CO) are ignored on purpose — they're computed or
derived from Sincron timesheets.
"""

import io
import logging
from datetime import date, datetime

import openpyxl

from core.base_repository import BaseRepository
from .repository import CoBalanceRepository

logger = logging.getLogger('jarvis.hr.co_balance.importer')

# Expected column positions (0-indexed). The HR template has been stable.
COL_NUME = 0
COL_PRENUME = 1
COL_CNP = 2
COL_NR_CONTRACT = 3
COL_STATUS = 4
COL_DATA_START = 5
COL_COMPANIE = 6
COL_DEPARTAMENT = 7
COL_CARRY_PREV = 8          # Sold CO an anterior
COL_CARRY_2Y = 9            # Sold CO ramas de acum 2 ani
COL_CIM = 10                # Zile CO cuvenite/an conf CIM
COL_VECHIME = 11            # Zile CO cuvenite/an conf vechime
COL_REGLARE = 12            # Reglare sold CO

# The 5 inputs we import. Any file whose header row doesn't contain all of these
# (normalized) is rejected.
REQUIRED_HEADERS = [
    'nume', 'prenume', 'cnp', 'nr. contract',
    'sold co an anterior',
    'sold co ramas de acum 2 ani',
    'zile co cuvenite/an conf cim',
    'zile co cuvenite/an conf vechime',
    'reglare sold co',
]


def _to_int(v):
    if v is None or v == '':
        return 0
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _norm(s):
    return (s or '').strip().lower()


def parse_co_balance_file(path_or_bytes, year):
    """Parse an xlsx file and return {rows, parse_errors}.

    `rows` is a list of dicts ready for upsert — identity + 5 inputs.
    """
    if isinstance(path_or_bytes, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(path_or_bytes), data_only=True)
    else:
        wb = openpyxl.load_workbook(path_or_bytes, data_only=True)

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return {'rows': [], 'parse_errors': ['File is empty']}

    header_row = all_rows[0]
    headers_norm = [_norm(h) for h in header_row]
    missing = [h for h in REQUIRED_HEADERS if h not in headers_norm]
    if missing:
        return {
            'rows': [],
            'parse_errors': [f'Missing required columns: {", ".join(missing)}'],
        }

    rows, parse_errors = [], []
    for idx, r in enumerate(all_rows[1:], start=2):
        if not r or not r[COL_NUME]:
            continue  # blank row
        try:
            cnp = (r[COL_CNP] or '').strip() if isinstance(r[COL_CNP], str) else r[COL_CNP]
            if cnp is None or str(cnp).strip() == '':
                parse_errors.append(f'Row {idx}: missing CNP for {r[COL_NUME]} {r[COL_PRENUME]}')
                continue
            rows.append({
                'year': year,
                'company_name': (r[COL_COMPANIE] or '').strip() if r[COL_COMPANIE] else '',
                'cnp': str(cnp).strip(),
                'nume': (r[COL_NUME] or '').strip(),
                'prenume': (r[COL_PRENUME] or '').strip() if r[COL_PRENUME] else '',
                'nr_contract': str(r[COL_NR_CONTRACT]).strip() if r[COL_NR_CONTRACT] is not None else None,
                'data_incepere_contract': _to_date(r[COL_DATA_START]),
                'departament': (r[COL_DEPARTAMENT] or '').strip() if r[COL_DEPARTAMENT] else None,
                'carry_prev_year': _to_int(r[COL_CARRY_PREV]),
                'carry_two_years_ago': _to_int(r[COL_CARRY_2Y]),
                'annual_cim': _to_int(r[COL_CIM]),
                'seniority_bonus': _to_int(r[COL_VECHIME]),
                'manual_adjustment': _to_int(r[COL_REGLARE]),
            })
        except Exception as e:  # noqa: BLE001 — report and continue
            parse_errors.append(f'Row {idx}: {e}')

    return {'rows': rows, 'parse_errors': parse_errors}


# ── Employee matching ──

class _UserMatcher:
    """Match CO-balance rows against JARVIS users via Sincron's CNP mapping.

    Sincron is the authoritative source: every employee with a payroll record
    already has `sincron_employees.mapped_jarvis_user_id` resolved via their
    CNP. So for CO-balance we mirror that exact map (keyed on CNP) and avoid
    duplicating a name-fallback here — if an employee isn't in Sincron they
    shouldn't auto-map to a JARVIS user from an HR xlsx alone.
    """

    def __init__(self):
        repo = BaseRepository()
        rows = repo.query_all(
            """
            SELECT DISTINCT ON (cnp_key) cnp_key, mapped_jarvis_user_id
              FROM (
                SELECT regexp_replace(COALESCE(cnp,''), '\\s', '', 'g') AS cnp_key,
                       mapped_jarvis_user_id,
                       last_synced_at
                  FROM sincron_employees
                 WHERE mapped_jarvis_user_id IS NOT NULL
                   AND cnp IS NOT NULL
                   AND cnp <> ''
              ) s
             ORDER BY cnp_key, last_synced_at DESC NULLS LAST
            """
        )
        self._by_cnp = {r['cnp_key']: r['mapped_jarvis_user_id'] for r in rows if r.get('cnp_key')}

    @staticmethod
    def _norm_cnp(cnp):
        if cnp is None:
            return ''
        return ''.join(str(cnp).split())

    def match(self, cnp, nume=None, prenume=None):  # noqa: ARG002 — kept for call-site compat
        key = self._norm_cnp(cnp)
        if not key:
            return None, None
        uid = self._by_cnp.get(key)
        return (uid, 'cnp') if uid else (None, None)


# ── Full pipeline ──

def import_co_balance(file_bytes, filename, year, imported_by):
    """Full import pipeline: parse → match → upsert.

    Returns a summary dict with counts and error messages.
    """
    parsed = parse_co_balance_file(file_bytes, year)
    errors = list(parsed['parse_errors'])
    rows = parsed['rows']
    if not rows:
        return {
            'source_file': filename,
            'year': year,
            'rows_total': 0,
            'rows_matched': 0,
            'rows_unmatched': 0,
            'companies': [],
            'errors': errors or ['No rows found in file'],
        }

    matcher = _UserMatcher()
    repo = CoBalanceRepository()

    matched = unmatched = 0
    companies = set()
    for r in rows:
        user_id, method = matcher.match(r['cnp'], r['nume'], r['prenume'])
        if user_id:
            matched += 1
        else:
            unmatched += 1
        companies.add(r['company_name'])
        try:
            repo.upsert(
                year=r['year'],
                company_name=r['company_name'],
                cnp=r['cnp'],
                nume=r['nume'],
                prenume=r['prenume'],
                nr_contract=r['nr_contract'],
                data_incepere_contract=r['data_incepere_contract'],
                departament=r['departament'],
                carry_prev_year=r['carry_prev_year'],
                carry_two_years_ago=r['carry_two_years_ago'],
                annual_cim=r['annual_cim'],
                seniority_bonus=r['seniority_bonus'],
                manual_adjustment=r['manual_adjustment'],
                mapped_jarvis_user_id=user_id,
                source_file=filename,
                imported_by=imported_by,
            )
        except Exception as e:  # noqa: BLE001
            logger.exception('upsert failed for %s %s', r['nume'], r['prenume'])
            errors.append(f"Upsert failed for {r['nume']} {r['prenume']}: {e}")

    return {
        'source_file': filename,
        'year': year,
        'rows_total': len(rows),
        'rows_matched': matched,
        'rows_unmatched': unmatched,
        'companies': sorted(c for c in companies if c),
        'errors': errors,
    }
