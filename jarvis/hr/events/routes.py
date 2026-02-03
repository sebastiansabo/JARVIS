"""HR Module Routes."""
from functools import wraps
from flask import render_template, request, jsonify, redirect, url_for, flash
from flask_login import login_required, current_user

from . import events_bp

# Import permission checking from main database
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from database import check_permission_v2
from .database import (
    get_all_hr_employees, get_hr_employee, save_hr_employee,
    update_hr_employee, delete_hr_employee, search_hr_employees,
    get_all_hr_events, get_hr_event, save_hr_event, update_hr_event, delete_hr_event, delete_hr_events_bulk,
    get_all_event_bonuses, get_event_bonus, save_event_bonus,
    save_event_bonuses_bulk, update_event_bonus, delete_event_bonus, delete_event_bonuses_bulk,
    delete_event_bonuses_by_employee, delete_event_bonuses_by_event,
    get_event_bonuses_summary, get_bonuses_by_month, get_bonuses_by_employee, get_bonuses_by_event,
    get_all_bonus_types, get_bonus_type, save_bonus_type, update_bonus_type, delete_bonus_type,
    # Company CRUD
    get_all_companies_with_brands, create_company, update_company, delete_company,
    # Company Brands CRUD
    get_all_company_brands, create_company_brand, update_company_brand, delete_company_brand,
    # Department Structure CRUD
    get_all_department_structures, create_department_structure, update_department_structure,
    delete_department_structure, get_name_by_id,
    # Master tables CRUD
    get_all_master_brands, create_master_brand, update_master_brand, delete_master_brand,
    get_all_master_departments, create_master_department, update_master_department, delete_master_department,
    get_all_master_subdepartments, create_master_subdepartment, update_master_subdepartment, delete_master_subdepartment
)

# Import from app root for structure data
import sys
import os
# Go up two levels: events/ -> hr/ -> app/
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from models import get_companies, get_brands_for_company, get_departments_for_company


def hr_required(f):
    """Decorator to require HR access permission (view only)."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not getattr(current_user, 'can_access_hr', False):
            flash('HR access required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


def hr_manager_required(f):
    """Decorator to require HR Manager permission for write operations.
    Uses is_hr_manager flag as fallback if permissions_v2 not configured.
    """
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not getattr(current_user, 'can_access_hr', False):
            flash('HR access required.', 'error')
            return redirect(url_for('index'))
        if not getattr(current_user, 'is_hr_manager', False):
            # For API calls, return JSON error
            if request.path.startswith('/hr/events/api/'):
                return jsonify({'success': False, 'error': 'HR Manager permission required'}), 403
            flash('HR Manager permission required.', 'error')
            return redirect(url_for('hr.events.event_bonuses'))
        return f(*args, **kwargs)
    return decorated


def hr_permission_required(entity: str, action: str):
    """Decorator to check specific HR permission using permissions_v2.

    Args:
        entity: Entity within HR module (events, bonuses, employees, structure)
        action: Action to check (view, add, edit, delete)

    Usage:
        @hr_permission_required('events', 'edit')
        def api_update_event():
            ...
    """
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))

            # Check basic HR access first
            if not getattr(current_user, 'can_access_hr', False):
                if request.path.startswith('/hr/events/api/'):
                    return jsonify({'success': False, 'error': 'HR access required'}), 403
                flash('HR access required.', 'error')
                return redirect(url_for('index'))

            # Check specific permission using permissions_v2
            role_id = getattr(current_user, 'role_id', None)
            if role_id:
                perm = check_permission_v2(role_id, 'hr', entity, action)
                if perm['has_permission']:
                    return f(*args, **kwargs)

            # Fallback to is_hr_manager for write operations
            if action in ('add', 'edit', 'delete') and getattr(current_user, 'is_hr_manager', False):
                return f(*args, **kwargs)

            # Permission denied
            if request.path.startswith('/hr/events/api/'):
                return jsonify({'success': False, 'error': f'Permission denied: hr.{entity}.{action}'}), 403
            flash(f'Permission denied: HR {entity} {action}', 'error')
            return redirect(url_for('hr.events.event_bonuses'))
        return decorated
    return decorator


# Romanian month names for display
MONTH_NAMES = {
    1: 'Ianuarie', 2: 'Februarie', 3: 'Martie', 4: 'Aprilie',
    5: 'Mai', 6: 'Iunie', 7: 'Iulie', 8: 'August',
    9: 'Septembrie', 10: 'Octombrie', 11: 'Noiembrie', 12: 'Decembrie'
}


# ============== Events Routes ==============

@events_bp.route('/')
@events_bp.route('/event-bonuses')
@login_required
@hr_required
def event_bonuses():
    """Events main page."""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    bonuses = get_all_event_bonuses(year=year, month=month)
    events = get_all_hr_events()
    employees = get_all_hr_employees()
    summary = get_event_bonuses_summary(year=year)
    employee_summary = get_bonuses_by_employee(year=year, month=month)
    event_summary = get_bonuses_by_event(year=year, month=month)

    # Get unique years from bonuses for filter
    years = sorted(set(b['year'] for b in bonuses), reverse=True) if bonuses else [2025]

    return render_template('event_bonuses.html',
                           bonuses=bonuses,
                           events=events,
                           employees=employees,
                           summary=summary,
                           employee_summary=employee_summary,
                           event_summary=event_summary,
                           years=years,
                           months=MONTH_NAMES,
                           selected_year=year,
                           selected_month=month,
                           is_hr_manager=current_user.is_hr_manager)


@events_bp.route('/api/event-bonuses', methods=['GET'])
@login_required
@hr_required
def api_get_event_bonuses():
    """API: Get event bonuses with filters."""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    employee_id = request.args.get('employee_id', type=int)
    event_id = request.args.get('event_id', type=int)

    bonuses = get_all_event_bonuses(year=year, month=month,
                                    employee_id=employee_id, event_id=event_id)
    return jsonify(bonuses)


@events_bp.route('/api/event-bonuses', methods=['POST'])
@login_required
@hr_permission_required('bonuses', 'add')
def api_create_event_bonus():
    """API: Create a new event bonus."""
    data = request.get_json()

    bonus_id = save_event_bonus(
        employee_id=data['employee_id'],
        event_id=data['event_id'],
        year=data['year'],
        month=data['month'],
        participation_start=data.get('participation_start'),
        participation_end=data.get('participation_end'),
        bonus_days=data.get('bonus_days'),
        hours_free=data.get('hours_free'),
        bonus_net=data.get('bonus_net'),
        details=data.get('details'),
        allocation_month=data.get('allocation_month'),
        created_by=current_user.id
    )

    return jsonify({'success': True, 'id': bonus_id})


@events_bp.route('/api/event-bonuses/bulk', methods=['POST'])
@login_required
@hr_permission_required('bonuses', 'add')
def api_create_event_bonuses_bulk():
    """API: Bulk create event bonuses."""
    data = request.get_json()
    bonuses = data.get('bonuses', [])

    if not bonuses:
        return jsonify({'success': False, 'error': 'No bonuses provided'}), 400

    created_ids = save_event_bonuses_bulk(bonuses, created_by=current_user.id)
    return jsonify({'success': True, 'ids': created_ids, 'count': len(created_ids)})


@events_bp.route('/api/event-bonuses/<int:bonus_id>', methods=['GET'])
@login_required
@hr_required
def api_get_event_bonus(bonus_id):
    """API: Get a single event bonus."""
    bonus = get_event_bonus(bonus_id)
    if not bonus:
        return jsonify({'error': 'Bonus not found'}), 404
    return jsonify(bonus)


@events_bp.route('/api/event-bonuses/<int:bonus_id>', methods=['PUT'])
@login_required
@hr_permission_required('bonuses', 'edit')
def api_update_event_bonus(bonus_id):
    """API: Update an event bonus."""
    data = request.get_json()

    update_event_bonus(
        bonus_id=bonus_id,
        employee_id=data['employee_id'],
        event_id=data['event_id'],
        year=data['year'],
        month=data['month'],
        participation_start=data.get('participation_start'),
        participation_end=data.get('participation_end'),
        bonus_days=data.get('bonus_days'),
        hours_free=data.get('hours_free'),
        bonus_net=data.get('bonus_net'),
        details=data.get('details'),
        allocation_month=data.get('allocation_month')
    )

    return jsonify({'success': True})


@events_bp.route('/api/event-bonuses/<int:bonus_id>', methods=['DELETE'])
@login_required
@hr_permission_required('bonuses', 'delete')
def api_delete_event_bonus(bonus_id):
    """API: Delete an event bonus."""
    delete_event_bonus(bonus_id)
    return jsonify({'success': True})


@events_bp.route('/api/event-bonuses/bulk-delete', methods=['POST'])
@login_required
@hr_permission_required('bonuses', 'delete')
def api_bulk_delete_event_bonuses():
    """API: Delete multiple event bonuses."""
    data = request.get_json()
    bonus_ids = data.get('ids', [])

    if not bonus_ids:
        return jsonify({'success': False, 'error': 'No IDs provided'}), 400

    # Validate all IDs are integers
    try:
        bonus_ids = [int(id) for id in bonus_ids]
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid ID format'}), 400

    deleted_count = delete_event_bonuses_bulk(bonus_ids)
    return jsonify({'success': True, 'deleted': deleted_count})


@events_bp.route('/api/event-bonuses/bulk-delete-by-employee', methods=['POST'])
@login_required
@hr_permission_required('bonuses', 'delete')
def api_bulk_delete_event_bonuses_by_employee():
    """API: Delete all bonuses for given employee IDs."""
    data = request.get_json()
    employee_ids = data.get('employee_ids', [])

    if not employee_ids:
        return jsonify({'success': False, 'error': 'No employee IDs provided'}), 400

    try:
        employee_ids = [int(id) for id in employee_ids]
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid ID format'}), 400

    deleted_count = delete_event_bonuses_by_employee(employee_ids)
    return jsonify({'success': True, 'deleted': deleted_count})


@events_bp.route('/api/event-bonuses/bulk-delete-by-event', methods=['POST'])
@login_required
@hr_permission_required('bonuses', 'delete')
def api_bulk_delete_event_bonuses_by_event():
    """API: Delete all bonuses for given event/year/month combinations."""
    data = request.get_json()
    selections = data.get('selections', [])

    if not selections:
        return jsonify({'success': False, 'error': 'No selections provided'}), 400

    deleted_count = delete_event_bonuses_by_event(selections)
    return jsonify({'success': True, 'deleted': deleted_count})


# ============== Events Routes ==============

@events_bp.route('/events')
@login_required
@hr_required
def events():
    """Events management page."""
    all_events = get_all_hr_events()
    companies = get_companies()
    return render_template('events.html', events=all_events, companies=companies)


@events_bp.route('/events/new')
@login_required
@hr_required
def add_event():
    """Add new event page."""
    companies = get_companies()
    employees = get_all_hr_employees(active_only=True)
    bonus_types = get_all_bonus_types(active_only=True)
    return render_template('add_event.html',
                           companies=companies,
                           employees=employees,
                           bonus_types=bonus_types,
                           is_hr_manager=current_user.is_hr_manager)


@events_bp.route('/bonuses/new')
@login_required
@hr_required
def add_bonus():
    """Add new bonus page."""
    from datetime import datetime
    events = get_all_hr_events()
    employees = get_all_hr_employees()
    bonus_types = get_all_bonus_types(active_only=True)
    return render_template('add_bonus.html',
                           events=events,
                           employees=employees,
                           bonus_types=bonus_types,
                           months=MONTH_NAMES,
                           current_year=datetime.now().year,
                           current_month=datetime.now().month,
                           is_hr_manager=current_user.is_hr_manager)


@events_bp.route('/api/events', methods=['GET'])
@login_required
@hr_required
def api_get_events():
    """API: Get all events."""
    events = get_all_hr_events()
    return jsonify(events)


@events_bp.route('/api/events', methods=['POST'])
@login_required
@hr_permission_required('events', 'add')
def api_create_event():
    """API: Create a new event."""
    data = request.get_json()

    event_id = save_hr_event(
        name=data['name'],
        start_date=data['start_date'],
        end_date=data['end_date'],
        company=data.get('company'),
        brand=data.get('brand'),
        description=data.get('description'),
        created_by=current_user.id
    )

    return jsonify({'success': True, 'id': event_id})


@events_bp.route('/api/events/<int:event_id>', methods=['GET'])
@login_required
@hr_required
def api_get_event(event_id):
    """API: Get a single event."""
    event = get_hr_event(event_id)
    if not event:
        return jsonify({'error': 'Event not found'}), 404
    return jsonify(event)


@events_bp.route('/api/events/<int:event_id>', methods=['PUT'])
@login_required
@hr_permission_required('events', 'edit')
def api_update_event(event_id):
    """API: Update an event."""
    data = request.get_json()

    update_hr_event(
        event_id=event_id,
        name=data['name'],
        start_date=data['start_date'],
        end_date=data['end_date'],
        company=data.get('company'),
        brand=data.get('brand'),
        description=data.get('description')
    )

    return jsonify({'success': True})


@events_bp.route('/api/events/<int:event_id>', methods=['DELETE'])
@login_required
@hr_permission_required('events', 'delete')
def api_delete_event(event_id):
    """API: Delete an event."""
    delete_hr_event(event_id)
    return jsonify({'success': True})


@events_bp.route('/api/events/bulk-delete', methods=['POST'])
@login_required
@hr_permission_required('events', 'delete')
def api_bulk_delete_events():
    """API: Delete multiple events."""
    data = request.get_json()
    event_ids = data.get('ids', [])

    if not event_ids:
        return jsonify({'success': False, 'error': 'No IDs provided'}), 400

    try:
        event_ids = [int(id) for id in event_ids]
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid ID format'}), 400

    deleted_count = delete_hr_events_bulk(event_ids)
    return jsonify({'success': True, 'deleted': deleted_count})


# ============== Employees API Routes ==============
# Note: Employee management page is now in core Settings (Settings → HR → Employees)
# These API routes are kept for HR Events module to fetch employees for bonuses

@events_bp.route('/api/employees', methods=['GET'])
@login_required
@hr_required
def api_get_employees():
    """API: Get all employees."""
    active_only = request.args.get('active_only', 'true').lower() == 'true'
    employees = get_all_hr_employees(active_only=active_only)
    return jsonify(employees)


@events_bp.route('/api/employees/search', methods=['GET'])
@login_required
@hr_required
def api_search_employees():
    """API: Search employees by name."""
    query = request.args.get('q', '')
    if len(query) < 2:
        return jsonify([])
    employees = search_hr_employees(query)
    return jsonify(employees)


@events_bp.route('/api/employees', methods=['POST'])
@login_required
@hr_permission_required('employees', 'add')
def api_create_employee():
    """API: Create a new employee."""
    data = request.get_json()

    employee_id = save_hr_employee(
        name=data['name'],
        department=data.get('departments'),  # frontend sends 'departments'
        subdepartment=data.get('subdepartment'),
        brand=data.get('brand'),
        company=data.get('company'),
        email=data.get('email'),
        phone=data.get('phone'),
        notify_on_allocation=data.get('notify_on_allocation', True)
    )

    return jsonify({'success': True, 'id': employee_id})


@events_bp.route('/api/employees/<int:employee_id>', methods=['GET'])
@login_required
@hr_required
def api_get_employee(employee_id):
    """API: Get a single employee."""
    employee = get_hr_employee(employee_id)
    if not employee:
        return jsonify({'error': 'Employee not found'}), 404
    return jsonify(employee)


@events_bp.route('/api/employees/<int:employee_id>', methods=['PUT'])
@login_required
@hr_permission_required('employees', 'edit')
def api_update_employee(employee_id):
    """API: Update an employee."""
    data = request.get_json()

    update_hr_employee(
        employee_id=employee_id,
        name=data['name'],
        department=data.get('departments'),  # frontend sends 'departments'
        subdepartment=data.get('subdepartment'),
        brand=data.get('brand'),
        company=data.get('company'),
        email=data.get('email'),
        phone=data.get('phone'),
        notify_on_allocation=data.get('notify_on_allocation', True),
        is_active=data.get('is_active', True)
    )

    return jsonify({'success': True})


@events_bp.route('/api/employees/<int:employee_id>', methods=['DELETE'])
@login_required
@hr_permission_required('employees', 'delete')
def api_delete_employee(employee_id):
    """API: Soft delete an employee."""
    delete_hr_employee(employee_id)
    return jsonify({'success': True})


# ============== Summary/Stats Routes ==============

@events_bp.route('/api/summary', methods=['GET'])
@login_required
@hr_required
def api_get_summary():
    """API: Get summary statistics."""
    year = request.args.get('year', type=int)
    summary = get_event_bonuses_summary(year=year)
    return jsonify(summary)


@events_bp.route('/api/summary/by-month', methods=['GET'])
@login_required
@hr_required
def api_get_by_month():
    """API: Get bonuses grouped by month."""
    year = request.args.get('year', type=int, default=2025)
    data = get_bonuses_by_month(year)
    return jsonify(data)


@events_bp.route('/api/summary/by-employee', methods=['GET'])
@login_required
@hr_required
def api_get_by_employee():
    """API: Get bonuses grouped by employee."""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    data = get_bonuses_by_employee(year=year, month=month)
    return jsonify(data)


# ============== Structure Data Routes ==============

@events_bp.route('/api/structure/companies', methods=['GET'])
@login_required
@hr_required
def api_get_companies():
    """API: Get all companies."""
    companies = get_companies()
    return jsonify(companies)


@events_bp.route('/api/structure/brands/<company>', methods=['GET'])
@login_required
@hr_required
def api_get_brands(company):
    """API: Get brands for a company."""
    brands = get_brands_for_company(company)
    return jsonify(brands)


@events_bp.route('/api/structure/departments/<company>', methods=['GET'])
@login_required
@hr_required
def api_get_departments(company):
    """API: Get departments for a company."""
    departments = get_departments_for_company(company)
    return jsonify(departments)


# ============== Companies CRUD API ==============

@events_bp.route('/api/structure/companies-full', methods=['GET'])
@login_required
@hr_required
def api_get_companies_full():
    """API: Get all companies with full details including brands from company_brands."""
    companies = get_all_companies_with_brands()
    # Format for API response
    for company in companies:
        if company.get('created_at'):
            company['created_at'] = company['created_at'].isoformat() if hasattr(company['created_at'], 'isoformat') else company['created_at']
        brand_list = company.get('brands', [])
        company['brands'] = ', '.join(brand_list) if isinstance(brand_list, list) else brand_list
        company['brands_list'] = [{'brand': b} for b in (brand_list if isinstance(brand_list, list) else [])]
    return jsonify(companies)


@events_bp.route('/api/structure/companies', methods=['POST'])
@login_required
@hr_permission_required('structure', 'edit')
def api_create_company():
    """API: Create a new company."""
    data = request.get_json()
    company_id = create_company(data['company'], data.get('vat'))
    return jsonify({'success': True, 'id': company_id})


@events_bp.route('/api/structure/companies/<int:company_id>', methods=['PUT'])
@login_required
@hr_permission_required('structure', 'edit')
def api_update_company(company_id):
    """API: Update a company."""
    data = request.get_json()
    update_company(company_id, data['company'], data.get('vat'))
    return jsonify({'success': True})


@events_bp.route('/api/structure/companies/<int:company_id>', methods=['DELETE'])
@login_required
@hr_permission_required('structure', 'edit')
def api_delete_company(company_id):
    """API: Delete a company."""
    delete_company(company_id)
    return jsonify({'success': True})


# ============== Company Brands CRUD API ==============

@events_bp.route('/api/structure/company-brands', methods=['GET'])
@login_required
@hr_required
def api_get_company_brands():
    """API: Get all company brands."""
    company_id = request.args.get('company_id', type=int)
    brands = get_all_company_brands(company_id)
    return jsonify(brands)


@events_bp.route('/api/structure/company-brands', methods=['POST'])
@login_required
@hr_permission_required('structure', 'edit')
def api_create_company_brand():
    """API: Create a new company brand."""
    from models import clear_structure_cache
    data = request.get_json()
    try:
        brand_id = create_company_brand(data['company_id'], data['brand'])
        clear_structure_cache()
        return jsonify({'success': True, 'id': brand_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@events_bp.route('/api/structure/company-brands/<int:brand_id>', methods=['PUT'])
@login_required
@hr_permission_required('structure', 'edit')
def api_update_company_brand(brand_id):
    """API: Update a company brand."""
    from models import clear_structure_cache
    data = request.get_json()
    update_company_brand(brand_id, data.get('company_id'), data['brand'], data.get('is_active', True))
    clear_structure_cache()
    return jsonify({'success': True})


@events_bp.route('/api/structure/company-brands/<int:brand_id>', methods=['DELETE'])
@login_required
@hr_permission_required('structure', 'edit')
def api_delete_company_brand(brand_id):
    """API: Delete a company brand."""
    from models import clear_structure_cache
    delete_company_brand(brand_id)
    clear_structure_cache()
    return jsonify({'success': True})


@events_bp.route('/api/structure/companies/<int:company_id>/brands', methods=['GET'])
@login_required
@hr_required
def api_get_brands_for_company(company_id):
    """API: Get brands for a specific company."""
    brands = get_all_company_brands(company_id)
    return jsonify(brands)


# ============== Departments CRUD API ==============

@events_bp.route('/api/structure/departments-full', methods=['GET'])
@login_required
@hr_required
def api_get_departments_full():
    """API: Get all department structure entries with full details via JOINs."""
    structures = get_all_department_structures()
    return jsonify(structures)


@events_bp.route('/api/structure/departments', methods=['POST'])
@login_required
@hr_permission_required('structure', 'edit')
def api_create_department():
    """API: Create a new department structure entry."""
    from models import clear_structure_cache

    data = request.get_json()

    # Look up the text values from master tables (if IDs provided)
    company_name = get_name_by_id('companies', data.get('company_id')) or data.get('company', '')
    brand_name = get_name_by_id('brands', data.get('brand_id')) or data.get('brand', '')
    dept_name = get_name_by_id('departments', data.get('department_id')) or data.get('department', '')
    subdept_name = get_name_by_id('subdepartments', data.get('subdepartment_id')) or data.get('subdepartment', '')

    dept_id = create_department_structure(
        data.get('company_id'),
        data.get('manager', ''),
        company_name, brand_name, dept_name, subdept_name,
        data.get('manager_ids')
    )
    clear_structure_cache()
    return jsonify({'success': True, 'id': dept_id})


@events_bp.route('/api/structure/departments/<int:dept_id>', methods=['PUT'])
@login_required
@hr_permission_required('structure', 'edit')
def api_update_department(dept_id):
    """API: Update a department structure entry."""
    from models import clear_structure_cache

    data = request.get_json()

    # Look up the text values from master tables (if IDs provided)
    company_name = get_name_by_id('companies', data.get('company_id')) or data.get('company', '')
    brand_name = get_name_by_id('brands', data.get('brand_id')) or data.get('brand', '')
    dept_name = get_name_by_id('departments', data.get('department_id')) or data.get('department', '')
    subdept_name = get_name_by_id('subdepartments', data.get('subdepartment_id')) or data.get('subdepartment', '')

    update_department_structure(
        dept_id, data.get('company_id'),
        data.get('manager', ''),
        company_name, brand_name, dept_name, subdept_name,
        data.get('manager_ids')
    )
    clear_structure_cache()
    return jsonify({'success': True})


@events_bp.route('/api/structure/departments/<int:dept_id>', methods=['DELETE'])
@login_required
@hr_permission_required('structure', 'edit')
def api_delete_department(dept_id):
    """API: Delete a department."""
    delete_department_structure(dept_id)
    return jsonify({'success': True})


# ============== Master Tables CRUD API ==============

# --- Brands Master Table ---

@events_bp.route('/api/master/brands', methods=['GET'])
@login_required
@hr_required
def api_get_master_brands():
    """API: Get all brands from master table."""
    brands = get_all_master_brands()
    return jsonify(brands)


@events_bp.route('/api/master/brands', methods=['POST'])
@login_required
@hr_permission_required('structure', 'edit')
def api_create_master_brand():
    """API: Create a new brand in master table."""
    from models import clear_structure_cache
    data = request.get_json()
    try:
        brand_id = create_master_brand(data['name'])
        clear_structure_cache()
        return jsonify({'success': True, 'id': brand_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@events_bp.route('/api/master/brands/<int:brand_id>', methods=['PUT'])
@login_required
@hr_permission_required('structure', 'edit')
def api_update_master_brand(brand_id):
    """API: Update a brand in master table."""
    from models import clear_structure_cache
    data = request.get_json()
    update_master_brand(brand_id, data['name'], data.get('is_active', True))
    clear_structure_cache()
    return jsonify({'success': True})


@events_bp.route('/api/master/brands/<int:brand_id>', methods=['DELETE'])
@login_required
@hr_permission_required('structure', 'edit')
def api_delete_master_brand(brand_id):
    """API: Delete a brand from master table."""
    from models import clear_structure_cache
    delete_master_brand(brand_id)
    clear_structure_cache()
    return jsonify({'success': True})


# --- Departments Master Table ---

@events_bp.route('/api/master/departments', methods=['GET'])
@login_required
@hr_required
def api_get_master_departments():
    """API: Get all departments from master table."""
    departments = get_all_master_departments()
    return jsonify(departments)


@events_bp.route('/api/master/departments', methods=['POST'])
@login_required
@hr_permission_required('structure', 'edit')
def api_create_master_department():
    """API: Create a new department in master table."""
    from models import clear_structure_cache
    data = request.get_json()
    try:
        dept_id = create_master_department(data['name'])
        clear_structure_cache()
        return jsonify({'success': True, 'id': dept_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@events_bp.route('/api/master/departments/<int:dept_id>', methods=['PUT'])
@login_required
@hr_permission_required('structure', 'edit')
def api_update_master_department(dept_id):
    """API: Update a department in master table."""
    from models import clear_structure_cache
    data = request.get_json()
    update_master_department(dept_id, data['name'], data.get('is_active', True))
    clear_structure_cache()
    return jsonify({'success': True})


@events_bp.route('/api/master/departments/<int:dept_id>', methods=['DELETE'])
@login_required
@hr_permission_required('structure', 'edit')
def api_delete_master_department(dept_id):
    """API: Delete a department from master table."""
    from models import clear_structure_cache
    delete_master_department(dept_id)
    clear_structure_cache()
    return jsonify({'success': True})


# --- Subdepartments Master Table ---

@events_bp.route('/api/master/subdepartments', methods=['GET'])
@login_required
@hr_required
def api_get_master_subdepartments():
    """API: Get all subdepartments from master table."""
    subdepartments = get_all_master_subdepartments()
    return jsonify(subdepartments)


@events_bp.route('/api/master/subdepartments', methods=['POST'])
@login_required
@hr_permission_required('structure', 'edit')
def api_create_master_subdepartment():
    """API: Create a new subdepartment in master table."""
    from models import clear_structure_cache
    data = request.get_json()
    try:
        subdept_id = create_master_subdepartment(data['name'])
        clear_structure_cache()
        return jsonify({'success': True, 'id': subdept_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@events_bp.route('/api/master/subdepartments/<int:subdept_id>', methods=['PUT'])
@login_required
@hr_permission_required('structure', 'edit')
def api_update_master_subdepartment(subdept_id):
    """API: Update a subdepartment in master table."""
    from models import clear_structure_cache
    data = request.get_json()
    update_master_subdepartment(subdept_id, data['name'], data.get('is_active', True))
    clear_structure_cache()
    return jsonify({'success': True})


@events_bp.route('/api/master/subdepartments/<int:subdept_id>', methods=['DELETE'])
@login_required
@hr_permission_required('structure', 'edit')
def api_delete_master_subdepartment(subdept_id):
    """API: Delete a subdepartment from master table."""
    from models import clear_structure_cache
    delete_master_subdepartment(subdept_id)
    clear_structure_cache()
    return jsonify({'success': True})


# ============== Bonus Types API Routes ==============

@events_bp.route('/api/bonus-types', methods=['GET'])
@login_required
@hr_required
def api_get_bonus_types():
    """API: Get all bonus types."""
    active_only = request.args.get('active_only', 'true').lower() == 'true'
    bonus_types = get_all_bonus_types(active_only=active_only)
    return jsonify(bonus_types)


@events_bp.route('/api/bonus-types', methods=['POST'])
@login_required
@hr_permission_required('bonuses', 'add')
def api_create_bonus_type():
    """API: Create a new bonus type."""
    data = request.get_json()

    bonus_type_id = save_bonus_type(
        name=data['name'],
        amount=data['amount'],
        days_per_amount=data.get('days_per_amount', 1),
        description=data.get('description')
    )

    return jsonify({'success': True, 'id': bonus_type_id})


@events_bp.route('/api/bonus-types/<int:bonus_type_id>', methods=['GET'])
@login_required
@hr_required
def api_get_bonus_type(bonus_type_id):
    """API: Get a single bonus type."""
    bonus_type = get_bonus_type(bonus_type_id)
    if not bonus_type:
        return jsonify({'error': 'Bonus type not found'}), 404
    return jsonify(bonus_type)


@events_bp.route('/api/bonus-types/<int:bonus_type_id>', methods=['PUT'])
@login_required
@hr_permission_required('bonuses', 'edit')
def api_update_bonus_type(bonus_type_id):
    """API: Update a bonus type."""
    data = request.get_json()

    update_bonus_type(
        bonus_type_id=bonus_type_id,
        name=data['name'],
        amount=data['amount'],
        days_per_amount=data.get('days_per_amount', 1),
        description=data.get('description'),
        is_active=data.get('is_active', True)
    )

    return jsonify({'success': True})


@events_bp.route('/api/bonus-types/<int:bonus_type_id>', methods=['DELETE'])
@login_required
@hr_permission_required('bonuses', 'delete')
def api_delete_bonus_type(bonus_type_id):
    """API: Soft delete a bonus type."""
    delete_bonus_type(bonus_type_id)
    return jsonify({'success': True})


# ============== Export Routes ==============

@events_bp.route('/api/export', methods=['GET'])
@login_required
@hr_permission_required('bonuses', 'export')
def api_export_bonuses():
    """API: Export event bonuses to Excel."""
    from flask import Response
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    bonuses = get_all_event_bonuses(year=year, month=month)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"

    # Headers
    headers = ['An', 'Luna', 'Nume', 'Dep', 'Brand', 'Compania', 'Eveniment',
               'Start event', 'End event', 'Data Start Participare', 'Data End Participare',
               'Zile Bonusabile', 'Ore / Libere', 'Prima (Net)', 'Detalii']

    header_fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, bonus in enumerate(bonuses, 2):
        ws.cell(row=row_idx, column=1, value=bonus['year'])
        ws.cell(row=row_idx, column=2, value=MONTH_NAMES.get(bonus['month'], bonus['month']))
        ws.cell(row=row_idx, column=3, value=bonus['employee_name'])
        ws.cell(row=row_idx, column=4, value=bonus.get('department', ''))
        ws.cell(row=row_idx, column=5, value=bonus.get('brand', ''))
        ws.cell(row=row_idx, column=6, value=bonus.get('company', ''))
        ws.cell(row=row_idx, column=7, value=bonus['event_name'])
        ws.cell(row=row_idx, column=8, value=bonus.get('event_start', ''))
        ws.cell(row=row_idx, column=9, value=bonus.get('event_end', ''))
        ws.cell(row=row_idx, column=10, value=bonus.get('participation_start', ''))
        ws.cell(row=row_idx, column=11, value=bonus.get('participation_end', ''))
        ws.cell(row=row_idx, column=12, value=bonus.get('bonus_days'))
        ws.cell(row=row_idx, column=13, value=bonus.get('hours_free'))
        ws.cell(row=row_idx, column=14, value=bonus.get('bonus_net'))
        ws.cell(row=row_idx, column=15, value=bonus.get('details', ''))

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Generate filename
    filename = f"Events"
    if year:
        filename += f"_{year}"
    if month:
        filename += f"_{MONTH_NAMES.get(month, month)}"
    filename += ".xlsx"

    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )
