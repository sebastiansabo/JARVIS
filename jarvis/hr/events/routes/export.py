from ._shared import *


# ============== Export Routes ==============

@events_bp.route('/api/export', methods=['GET'])
@login_required
@hr_permission_required('bonuses', 'export')
def api_export_bonuses():
    """API: Export event bonuses to Excel."""
    from flask import Response
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    bonuses = get_all_event_bonuses(year=year, month=month)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"

    # Headers
    headers = ['An', 'Luna', 'Nume', 'Dep', 'Brand', 'Compania', 'Eveniment',
               'Start event', 'End event', 'Data Start Participare', 'Data End Participare',
               'Zile Bonusabile', 'Ore / Libere', 'Ore Eveniment', 'Prima (Net)', 'Detalii']

    header_fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, bonus in enumerate(bonuses, 2):
        ws.cell(row=row_idx, column=1, value=bonus['year'])
        ws.cell(row=row_idx, column=2, value=MONTH_NAMES.get(bonus['month'], bonus['month']))
        ws.cell(row=row_idx, column=3, value=bonus['employee_name'])
        ws.cell(row=row_idx, column=4, value=bonus.get('department', ''))
        ws.cell(row=row_idx, column=5, value=bonus.get('brand', ''))
        ws.cell(row=row_idx, column=6, value=bonus.get('company', ''))
        ws.cell(row=row_idx, column=7, value=bonus['event_name'])
        ws.cell(row=row_idx, column=8, value=bonus.get('event_start', ''))
        ws.cell(row=row_idx, column=9, value=bonus.get('event_end', ''))
        ws.cell(row=row_idx, column=10, value=bonus.get('participation_start', ''))
        ws.cell(row=row_idx, column=11, value=bonus.get('participation_end', ''))
        ws.cell(row=row_idx, column=12, value=bonus.get('bonus_days'))
        ws.cell(row=row_idx, column=13, value=bonus.get('hours_free'))
        ws.cell(row=row_idx, column=14, value=bonus.get('event_hours'))
        ws.cell(row=row_idx, column=15, value=bonus.get('bonus_net'))
        ws.cell(row=row_idx, column=16, value=bonus.get('details', ''))

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Generate filename
    filename = f"Events"
    if year:
        filename += f"_{year}"
    if month:
        filename += f"_{MONTH_NAMES.get(month, month)}"
    filename += ".xlsx"

    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )
