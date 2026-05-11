"""Time Bank T0 Excel importer.

Reads an xlsx file with columns: Nume, Prenume, CNP, Ore (hours).
Matches employees via CNP using the same Sincron-based mapping as
co_balance/importer.py.
"""

import io
import logging

import openpyxl

from core.base_repository import BaseRepository
from .service import TimeBankService

logger = logging.getLogger('jarvis.hr.time_bank.importer')

REQUIRED_HEADERS = ['nume', 'prenume', 'cnp', 'ore']


def _norm(s):
    return (s or '').strip().lower()


def _to_float(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


class _UserMatcher:
    """Match rows against JARVIS users via Sincron's CNP mapping.

    Identical to co_balance/importer.py — Sincron is the authoritative
    source for CNP → JARVIS user mapping.
    """

    def __init__(self):
        repo = BaseRepository()
        rows = repo.query_all(
            """
            SELECT DISTINCT ON (cnp_key) cnp_key, mapped_jarvis_user_id
              FROM (
                SELECT regexp_replace(COALESCE(cnp,''), '\\s', '', 'g') AS cnp_key,
                       mapped_jarvis_user_id,
                       last_synced_at
                  FROM sincron_employees
                 WHERE mapped_jarvis_user_id IS NOT NULL
                   AND cnp IS NOT NULL
                   AND cnp <> ''
              ) s
             ORDER BY cnp_key, last_synced_at DESC NULLS LAST
            """
        )
        self._by_cnp = {r['cnp_key']: r['mapped_jarvis_user_id'] for r in rows if r.get('cnp_key')}

    @staticmethod
    def _norm_cnp(cnp):
        if cnp is None:
            return ''
        return ''.join(str(cnp).split())

    def match(self, cnp):
        key = self._norm_cnp(cnp)
        if not key:
            return None
        return self._by_cnp.get(key)


def import_t0_file(file_bytes, filename, imported_by=None):
    """Parse T0 xlsx and insert transactions.

    Returns summary dict with counts and errors.
    """
    if isinstance(file_bytes, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    else:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return {'rows_total': 0, 'rows_matched': 0, 'rows_skipped': 0,
                'rows_unmatched': 0, 'errors': ['File is empty']}

    # Find column positions from header
    header_row = all_rows[0]
    headers_norm = [_norm(h) for h in header_row]
    missing = [h for h in REQUIRED_HEADERS if h not in headers_norm]
    if missing:
        return {'rows_total': 0, 'rows_matched': 0, 'rows_skipped': 0,
                'rows_unmatched': 0,
                'errors': [f'Missing required columns: {", ".join(missing)}']}

    col_nume = headers_norm.index('nume')
    col_prenume = headers_norm.index('prenume')
    col_cnp = headers_norm.index('cnp')
    col_ore = headers_norm.index('ore')

    matcher = _UserMatcher()
    svc = TimeBankService()
    errors = []
    matched = 0
    unmatched = 0
    skipped = 0

    for idx, r in enumerate(all_rows[1:], start=2):
        if not r or not r[col_nume]:
            continue
        try:
            cnp_raw = r[col_cnp]
            cnp = (cnp_raw or '').strip() if isinstance(cnp_raw, str) else cnp_raw
            if cnp is None or str(cnp).strip() == '':
                errors.append(f'Row {idx}: missing CNP for {r[col_nume]} {r[col_prenume]}')
                unmatched += 1
                continue

            cnp_str = str(cnp).strip()
            hours = _to_float(r[col_ore])
            if hours <= 0:
                skipped += 1
                continue

            user_id = matcher.match(cnp_str)
            if not user_id:
                errors.append(f'Row {idx}: no JARVIS user for CNP {cnp_str} ({r[col_nume]} {r[col_prenume]})')
                unmatched += 1
                continue

            svc.set_t0(user_id, hours, created_by=imported_by)
            matched += 1

        except Exception as e:  # noqa: BLE001
            errors.append(f'Row {idx}: {e}')

    total = matched + unmatched + skipped
    return {
        'source_file': filename,
        'rows_total': total,
        'rows_matched': matched,
        'rows_skipped': skipped,
        'rows_unmatched': unmatched,
        'errors': errors,
    }
