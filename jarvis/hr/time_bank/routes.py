"""HTTP endpoints for Time Bank.

Routes mounted under `/hr/time-bank/*` via `time_bank_bp`.
"""

import io
import logging

from flask import request, jsonify, send_file
from flask_login import current_user

from core.utils.api_helpers import api_login_required, safe_error_response
from . import time_bank_bp
from .service import TimeBankService

logger = logging.getLogger('jarvis.hr.time_bank.routes')


def _require_hr():
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'error': 'Authentication required'}), 401
    if not getattr(current_user, 'can_access_hr', False):
        return jsonify({'success': False, 'error': 'HR access required'}), 403
    return None


def _get_managed_ids():
    """Return set of user IDs the current user manages (None = admin sees all)."""
    from hr.events.database import get_managed_employee_ids, is_manager
    if getattr(current_user, 'is_admin', False):
        return None  # admin sees all
    if is_manager(current_user.id):
        return set(get_managed_employee_ids(current_user.id))
    return {current_user.id}  # regular user sees only self


# ── Balances ──

@time_bank_bp.route('/api/time-bank/balances', methods=['GET'])
@api_login_required
def time_bank_balances():
    """All employee balances (scoped by manager hierarchy)."""
    forbidden = _require_hr()
    if forbidden:
        return forbidden
    try:
        svc = TimeBankService()
        show_all = request.args.get('all', '').lower() in ('1', 'true', 'yes')
        all_balances = svc.get_all_balances(include_all_employees=show_all)
        managed = _get_managed_ids()
        if managed is not None:
            all_balances = [b for b in all_balances if b['user_id'] in managed]
        return jsonify({'success': True, 'data': all_balances})
    except Exception as e:  # noqa: BLE001
        return safe_error_response(e)


@time_bank_bp.route('/api/time-bank/balances/<int:user_id>', methods=['GET'])
@api_login_required
def time_bank_balance_single(user_id):
    """Single employee balance."""
    forbidden = _require_hr()
    if forbidden:
        return forbidden
    try:
        managed = _get_managed_ids()
        if managed is not None and user_id not in managed:
            return jsonify({'success': False, 'error': 'Access denied'}), 403
        svc = TimeBankService()
        balance = svc.get_balance(user_id)
        return jsonify({'success': True, 'data': {'user_id': user_id, 'balance': balance}})
    except Exception as e:  # noqa: BLE001
        return safe_error_response(e)


# ── Transactions ──

@time_bank_bp.route('/api/time-bank/transactions', methods=['GET'])
@api_login_required
def time_bank_transactions():
    """All transactions (admin/manager scoped)."""
    forbidden = _require_hr()
    if forbidden:
        return forbidden
    try:
        limit = int(request.args.get('limit', 100))
        offset = int(request.args.get('offset', 0))
        tx_type = request.args.get('tx_type') or None
        filter_user = request.args.get('user_id')
        filter_user = int(filter_user) if filter_user else None

        svc = TimeBankService()
        managed = _get_managed_ids()

        if filter_user:
            if managed is not None and filter_user not in managed:
                return jsonify({'success': False, 'error': 'Access denied'}), 403
            rows = svc.get_transactions(filter_user, limit=limit, offset=offset, tx_type=tx_type)
        else:
            rows = svc.get_all_transactions(limit=limit, offset=offset, tx_type=tx_type)
            if managed is not None:
                rows = [r for r in rows if r.get('jarvis_user_id') in managed]

        total = svc.count_transactions(user_id=filter_user, tx_type=tx_type)
        return jsonify({'success': True, 'data': rows, 'total': total})
    except Exception as e:  # noqa: BLE001
        return safe_error_response(e)


@time_bank_bp.route('/api/time-bank/transactions/<int:user_id>', methods=['GET'])
@api_login_required
def time_bank_user_transactions(user_id):
    """Transactions for a specific employee."""
    forbidden = _require_hr()
    if forbidden:
        return forbidden
    try:
        managed = _get_managed_ids()
        if managed is not None and user_id not in managed:
            return jsonify({'success': False, 'error': 'Access denied'}), 403

        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
        tx_type = request.args.get('tx_type') or None

        svc = TimeBankService()
        rows = svc.get_transactions(user_id, limit=limit, offset=offset, tx_type=tx_type)
        total = svc.count_transactions(user_id=user_id, tx_type=tx_type)
        balance = svc.get_balance(user_id)
        return jsonify({
            'success': True, 'data': rows, 'total': total,
            'balance': balance, 'user_id': user_id,
        })
    except Exception as e:  # noqa: BLE001
        return safe_error_response(e)


# ── Manual Credit/Debit ──

@time_bank_bp.route('/api/time-bank/credit', methods=['POST'])
@api_login_required
def time_bank_credit():
    """Manual credit: {user_id, amount, description}."""
    forbidden = _require_hr()
    if forbidden:
        return forbidden
    try:
        data = request.get_json(silent=True) or {}
        user_id = data.get('user_id')
        amount = data.get('amount')
        description = data.get('description', '')
        if not user_id or not amount:
            return jsonify({'success': False, 'error': 'user_id and amount required'}), 400

        svc = TimeBankService()
        row = svc.credit(
            user_id=int(user_id),
            amount=float(amount),
            tx_type='manual_credit',
            description=description,
            created_by=current_user.id,
        )
        return jsonify({'success': True, 'data': row})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:  # noqa: BLE001
        return safe_error_response(e)


@time_bank_bp.route('/api/time-bank/debit', methods=['POST'])
@api_login_required
def time_bank_debit():
    """Manual debit: {user_id, amount, description}."""
    forbidden = _require_hr()
    if forbidden:
        return forbidden
    try:
        data = request.get_json(silent=True) or {}
        user_id = data.get('user_id')
        amount = data.get('amount')
        description = data.get('description', '')
        if not user_id or not amount:
            return jsonify({'success': False, 'error': 'user_id and amount required'}), 400

        svc = TimeBankService()
        row = svc.debit(
            user_id=int(user_id),
            amount=float(amount),
            tx_type='manual_debit',
            description=description,
            created_by=current_user.id,
        )
        return jsonify({'success': True, 'data': row})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:  # noqa: BLE001
        return safe_error_response(e)


# ── T0 ──

@time_bank_bp.route('/api/time-bank/set-t0', methods=['POST'])
@api_login_required
def time_bank_set_t0():
    """Set T0 for a single employee: {user_id, amount}."""
    forbidden = _require_hr()
    if forbidden:
        return forbidden
    try:
        data = request.get_json(silent=True) or {}
        user_id = data.get('user_id')
        amount = data.get('amount')
        if not user_id or amount is None:
            return jsonify({'success': False, 'error': 'user_id and amount required'}), 400

        svc = TimeBankService()
        row = svc.set_t0(int(user_id), float(amount), created_by=current_user.id)
        return jsonify({'success': True, 'data': row})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:  # noqa: BLE001
        return safe_error_response(e)


@time_bank_bp.route('/api/time-bank/import-t0', methods=['POST'])
@api_login_required
def time_bank_import_t0():
    """Upload T0 Excel file."""
    forbidden = _require_hr()
    if forbidden:
        return forbidden
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Missing file'}), 400
        f = request.files['file']
        if not f or not f.filename:
            return jsonify({'success': False, 'error': 'Empty file'}), 400

        file_bytes = f.read()
        from .importer import import_t0_file
        result = import_t0_file(file_bytes, f.filename, imported_by=current_user.id)
        return jsonify({'success': True, 'data': result})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:  # noqa: BLE001
        return safe_error_response(e)


@time_bank_bp.route('/api/time-bank/t0-template', methods=['GET'])
@api_login_required
def time_bank_t0_template():
    """Download a blank T0 import template (.xlsx)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Time Bank T0'
    headers = ['Nume', 'Prenume', 'CNP', 'Ore']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10
    # Example row
    ws.cell(row=2, column=1, value='Popescu')
    ws.cell(row=2, column=2, value='Ion')
    ws.cell(row=2, column=3, value='1234567890123')
    ws.cell(row=2, column=4, value=8)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='time_bank_t0_template.xlsx',
    )
