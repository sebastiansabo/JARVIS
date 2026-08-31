"""Hermetic tests for the EuroFib XLSX export line-building logic.

Reproduces the production scenario for comanda 352716 (VW Caddy, anexa 21):
a STORNO reversing two advances (2,290 € @ 5.0927 and 20,610 € @ 5.2429)
followed by a FINAL for the full 22,900 €.

Two accounting requirements are asserted:

1. Every row of a single STORNO invoice must carry that storno's own
   invoice number (both lines = 9103805), not a per-row incremented number
   (the 2nd line was wrongly getting 9103806 — the FINAL's number).

2. The FINAL invoice's RON value must equal the summed RON of the reversed
   advances (11,662.28 + 108,056.17 = 119,718.45), so the client ledger nets
   to zero — not 22,900 × (last advance's kurs) = 120,062.41.

The route module opens a DB pool at import, so DATABASE_URL must point at a
reachable Postgres. No query actually runs: the module-level `_repo` is
replaced with a fake that returns canned rows for this scenario.
"""
import io
import os
import sys

import pytest

JARVIS_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if JARVIS_ROOT not in sys.path:
    sys.path.insert(0, JARVIS_ROOT)

# Import-time only: the pool connects but every query is intercepted by FakeRepo.
os.environ.setdefault("DATABASE_URL", "postgresql://localhost/defaultdb")

from openpyxl import load_workbook  # noqa: E402

from accounting.facturare import routes_orders  # noqa: E402
from accounting.facturare.generators.eurofib_xlsx import EurofibXlsxRenderer  # noqa: E402

# ── Column indices in the EuroFib template (1-based) ─────────────
COL_BELEGNUMMER = 8   # H
COL_FWBETRAG = 13     # M
COL_KURS = 33         # AG

# ── Canned scenario data (single car) ────────────────────────────
LINE_840 = {
    "id": 840, "anexa_id": 1, "line_number": 1, "nr_comanda": "352716",
    "model": "Caddy Maxi Cargo 2.0 TDI 75 kW", "culoare": "",
    "list_price_eur": 0, "selling_price_eur": 22900, "qty": 1,
}
ADV_A = {  # first advance — 2,290 € @ 5.0927
    "id": 91, "invoice_number": 9103741, "total_amount_eur": 2290,
    "split_mode": "proportional", "kurs_applied": 5.0927,
    "issued_date": "2026-04-27", "line_ids": [840],
}
ADV_B = {  # second advance — 20,610 € @ 5.2429
    "id": 92, "invoice_number": 9103742, "total_amount_eur": 20610,
    "split_mode": "proportional", "kurs_applied": 5.2429,
    "issued_date": "2026-07-19", "line_ids": [840],
}
STORNO_ROW = {
    "id": 100, "invoice_type": "STORNO", "invoice_number": 9103805, "anexa_id": 1,
    "total_amount_eur": -22900, "split_mode": "equal", "kurs_applied": 5.1893,
    "issued_date": "2026-07-27", "line_ids": [840],
}
FINAL_ROW = {
    "id": 101, "invoice_type": "FINAL", "invoice_number": 9103806, "anexa_id": 1,
    "total_amount_eur": 22900, "split_mode": "proportional", "kurs_applied": 5.2279,
    "issued_date": "2026-07-27", "line_ids": [840],
}

# Stored document numbers per invoice (facturare_document_numbers), keyed by
# invoice id — this is what the export now reads instead of deriving.
DOCNUM_MAP_BY_INVOICE = {
    100: {840: 9103805},  # storno
    101: {840: 9103806},  # final
}

# ── Canned scenario data (multi-car anexa, ids 838/839/840) ──────
LINE_838 = {
    "id": 838, "anexa_id": 2, "line_number": 1, "nr_comanda": "838-1",
    "model": "Model A", "culoare": "", "list_price_eur": 0,
    "selling_price_eur": 10000, "qty": 1,
}
LINE_839 = {
    "id": 839, "anexa_id": 2, "line_number": 2, "nr_comanda": "838-2",
    "model": "Model B", "culoare": "", "list_price_eur": 0,
    "selling_price_eur": 10000, "qty": 1,
}
LINE_840_MULTI = {
    "id": 840, "anexa_id": 2, "line_number": 3, "nr_comanda": "838-3",
    "model": "Model C", "culoare": "", "list_price_eur": 0,
    "selling_price_eur": 10000, "qty": 1,
}
ADVANCE_3CAR = {  # single_doc: all three cars share ONE stored number
    "id": 200, "invoice_type": "INVOICE", "invoice_number": 300, "anexa_id": 2,
    "total_amount_eur": 9126, "split_mode": "equal", "kurs_applied": 5.0,
    "issued_date": "2026-07-27", "line_ids": [838, 839, 840],
}
ADVANCE_2CAR = {  # per_car: each car has its OWN stored number
    "id": 201, "invoice_type": "INVOICE", "invoice_number": 301, "anexa_id": 2,
    "total_amount_eur": 6084, "split_mode": "equal", "kurs_applied": 5.0,
    "issued_date": "2026-07-27", "line_ids": [838, 839],
}


class FakeRepo:
    """Returns canned rows for the 352716 scenario; dispatches raw SQL by substring."""

    def get_anexa_by_id(self, anexa_id):
        return {"id": 1, "contract_id": 1, "anexa_number": 21}

    def get_contract_by_id(self, contract_id):
        return {"id": 1, "contract_ref": "C-1", "supplier_id": 16, "customer_id": 18}

    def get_lines_by_anexa(self, anexa_id):
        return [dict(LINE_840)]

    def get_invoices_by_anexa(self, anexa_id):
        # Cumulative-rounding context: the two advances (INVOICE), the storno and
        # the final on anexa 1. Advances carry invoice_type so _prior_car_fractions
        # can identify same-track siblings.
        return [
            dict(ADV_A, invoice_type="INVOICE"),
            dict(ADV_B, invoice_type="INVOICE"),
            dict(STORNO_ROW),
            dict(FINAL_ROW),
        ]

    def get_document_number_map(self, invoice_id):
        return dict(DOCNUM_MAP_BY_INVOICE.get(invoice_id, {}))

    def match_venituri_rule(self, supplier_id, nr_comanda):
        return None  # FINAL falls back to konto_config credit (707128)

    def query_one(self, sql, params=None):
        if "facturare_konto_config" in sql:
            invoice_type = params[1]
            credit = 419968 if invoice_type == "STORNO" else 707128
            return {"konto_credit": credit, "text_template": None}
        if "FROM companies" in sql:
            return {"eurofib_klient_id": 139}
        if "FROM crm_clients" in sql:
            return {"eurofib_konto_debit": {"139": "41217835"}}
        if "ORDER BY sequence_number DESC" in sql:   # legacy last-advance lookup
            return dict(ADV_B)
        return None

    def query_all(self, sql, params=None):
        if "facturare_invoice_links" in sql:
            return [{"source_invoice_id": 91}, {"source_invoice_id": 92}]
        if "invoice_type = 'STORNO'" in sql:          # matching-storno lookup (final)
            return [{"id": 100, "total_amount_eur": -22900, "line_ids": [840]}]
        if "id IN" in sql:                            # reversed invoices for storno
            return [dict(ADV_A), dict(ADV_B)]
        if "invoice_type = 'INVOICE'" in sql:         # all advances for anexa (fallback)
            return [dict(ADV_A), dict(ADV_B)]
        return []


class MultiCarFakeRepo(FakeRepo):
    """3-car anexa (line ids 838/839/840) used for the multi-car advance tests.

    Only `get_document_number_map` is scenario-specific — the caller supplies
    the stored map (single_doc: all cars share one number; per_car: each car
    has its own).
    """

    def __init__(self, docnum_map):
        self._docnum_map = docnum_map

    def get_anexa_by_id(self, anexa_id):
        return {"id": 2, "contract_id": 1, "anexa_number": 99}

    def get_lines_by_anexa(self, anexa_id):
        return [dict(LINE_838), dict(LINE_839), dict(LINE_840_MULTI)]

    def get_invoices_by_anexa(self, anexa_id):
        return [dict(ADVANCE_3CAR), dict(ADVANCE_2CAR)]

    def get_document_number_map(self, invoice_id):
        return dict(self._docnum_map)


@pytest.fixture(autouse=True)
def fake_repo(monkeypatch):
    monkeypatch.setattr(routes_orders, "_repo", FakeRepo())


def _render(inv_row):
    cfg, lines = routes_orders._build_eurofib_batch(inv_row)
    wb = load_workbook(io.BytesIO(EurofibXlsxRenderer(cfg).render_to_bytes(lines)))
    return wb.active


def _betrag(ws, row):
    """betrag = fwbetrag * kurs (the renderer writes betrag as an Excel formula)."""
    return round(ws.cell(row=row, column=COL_FWBETRAG).value
                 * ws.cell(row=row, column=COL_KURS).value, 2)


def test_storno_lines_share_the_storno_invoice_number():
    """Belegnummer comes from DOCNUM_MAP_BY_INVOICE[100] = {840: 9103805}, not
    derived — every reversed-advance row of car 840 gets that stored number."""
    ws = _render(STORNO_ROW)
    # Two reversed advances → rows 3/4 (line 0) and 5/6 (line 1)
    belegnummers = {ws.cell(row=r, column=COL_BELEGNUMMER).value for r in (3, 4, 5, 6)}
    assert belegnummers == {9103805}, (
        f"all storno rows must carry 9103805, got {sorted(belegnummers)}"
    )
    # Sanity: the two reversed advances are still represented at their own rates
    assert _betrag(ws, 3) == -11662.28
    assert _betrag(ws, 5) == -108056.17


def test_final_ron_equals_summed_storno_ron():
    """Belegnummer comes from DOCNUM_MAP_BY_INVOICE[101] = {840: 9103806}."""
    ws = _render(FINAL_ROW)
    assert ws.cell(row=3, column=COL_BELEGNUMMER).value == 9103806
    assert ws.cell(row=3, column=COL_FWBETRAG).value == 22900
    # 11,662.28 + 108,056.17 = 119,718.45 (net-zero against the storno)
    assert _betrag(ws, 3) == 119718.45


def test_daily_export_storno_plus_final_nets_to_zero():
    """The combined daily sheet (storno batch + final batch) must net to ~0 RON."""
    storno_batch = routes_orders._build_eurofib_batch(STORNO_ROW)
    final_batch = routes_orders._build_eurofib_batch(FINAL_ROW)
    xlsx = EurofibXlsxRenderer.render_multi_to_bytes([storno_batch, final_batch])
    ws = load_workbook(io.BytesIO(xlsx)).active

    # Debit rows only (odd export rows 3,5,7) hold each position once.
    belegnummers = [ws.cell(row=r, column=COL_BELEGNUMMER).value for r in (3, 5, 7)]
    assert belegnummers == [9103805, 9103805, 9103806]

    net = sum(_betrag(ws, r) for r in (3, 5, 7))
    assert abs(net) < 0.01, f"storno + final should net to zero, got {net}"


def test_multi_car_advance_single_doc_shares_stored_number(monkeypatch):
    """single_doc: all three cars were stored under the SAME document number.

    The old `start_no + idx` derivation would have produced 9103042/43/44 for
    the three rows — wrong, since this invoice's per-car numbers were actually
    stored identically (single_doc mode). The export must read the map.
    """
    monkeypatch.setattr(routes_orders, "_repo", MultiCarFakeRepo(
        {838: 9103042, 839: 9103042, 840: 9103042}))
    ws = _render(ADVANCE_3CAR)
    belegnummers = [ws.cell(row=r, column=COL_BELEGNUMMER).value for r in (3, 5, 7)]
    assert belegnummers == [9103042, 9103042, 9103042]


def test_multi_car_advance_per_car_stored_numbers(monkeypatch):
    """per_car: each car was stored with its OWN document number."""
    monkeypatch.setattr(routes_orders, "_repo", MultiCarFakeRepo(
        {838: 9103042, 839: 9103043}))
    ws = _render(ADVANCE_2CAR)
    belegnummers = [ws.cell(row=r, column=COL_BELEGNUMMER).value for r in (3, 5)]
    assert belegnummers == [9103042, 9103043]


# ── Double-advance scenario (prod: Taigo Prime, car 714, anexa 17) ───
# A car that received TWO advances — a 10% advance (@ 5.0924) closed by an
# earlier storno, and a ~100% advance (@ 5.2460) reversed by THIS final's
# storno. The final must mirror ONLY its matching storno (16,113 € @ 5.2460 =
# 84,528.80 RON), not the blend of both advances (which inflated the rate to
# ~5.7552 → 92,734.18, the reported bug).
CAR_714 = {
    "id": 714, "anexa_id": 3, "line_number": 1, "nr_comanda": "714",
    "model": "Taigo Prime 1.0", "culoare": "", "list_price_eur": 0,
    "selling_price_eur": 16113, "qty": 1,
}
ADV_10PCT = {  # 10% advance — NOT reversed by this final's storno
    "id": 920, "total_amount_eur": 1611, "kurs_applied": 5.0924,
    "issued_date": "2026-04-24", "line_ids": [714],
}
ADV_100PCT = {  # ~100% advance — reversed by storno 300
    "id": 921, "total_amount_eur": 16113, "kurs_applied": 5.2460,
    "issued_date": "2026-06-24", "line_ids": [714],
}
STORNO_714 = {
    "id": 300, "invoice_type": "STORNO", "invoice_number": 9103789, "anexa_id": 3,
    "total_amount_eur": -16113, "split_mode": "equal", "kurs_applied": 5.2460,
    "issued_date": "2026-07-27", "line_ids": [714],
}
FINAL_714 = {
    "id": 301, "invoice_type": "FINAL", "invoice_number": 9103790, "anexa_id": 3,
    "total_amount_eur": 16113, "split_mode": "proportional", "kurs_applied": 5.7552,
    "issued_date": "2026-07-27", "line_ids": [714],
}


class DoubleAdvanceFakeRepo(FakeRepo):
    """Car 714: two advances, but the final's storno reverses only the 100% one."""

    def get_anexa_by_id(self, anexa_id):
        return {"id": 3, "contract_id": 1, "anexa_number": 17}

    def get_lines_by_anexa(self, anexa_id):
        return [dict(CAR_714)]

    def get_document_number_map(self, invoice_id):
        return {714: 9103790}

    def query_all(self, sql, params=None):
        if "facturare_invoice_links" in sql:          # storno 300 reverses ONLY the 100%
            return [{"source_invoice_id": 921}]
        if "invoice_type = 'STORNO'" in sql:
            return [dict(STORNO_714)]
        if "id IN" in sql:                            # advances the matching storno reverses
            return [dict(ADV_100PCT)]
        if "invoice_type = 'INVOICE'" in sql:         # fallback: BOTH advances cover the car
            return [dict(ADV_10PCT), dict(ADV_100PCT)]
        return []


def test_final_mirrors_matching_storno_not_all_advances(monkeypatch):
    """The final's RON must equal its matching storno's RON (84,528.80 @ 5.2460),
    NOT the blend of every advance covering the car (which gave ~92,734 @ 5.7552).
    """
    monkeypatch.setattr(routes_orders, "_repo", DoubleAdvanceFakeRepo())
    ws = _render(FINAL_714)
    assert ws.cell(row=3, column=COL_FWBETRAG).value == 16113
    assert round(ws.cell(row=3, column=COL_KURS).value, 4) == 5.2460
    assert _betrag(ws, 3) == 84528.80


# ── Decimal-advance scenario (prod: Audi A3, comanda 152392, anexa 1) ──────
# The 5% and 95% advances were stored with sub-EUR decimals (1403.89 / 26674.10)
# — themselves fractional slices of the car total. The advance export already
# rounds each car to whole EUR (1404 / 26674) via _snap_pct + round-half-up, and
# the STORNO must reverse those SAME whole amounts (matching the invoice PDF).
# The old storno path echoed the stored decimals into fwbetrag, so the EuroFib
# EUR column showed 1403.89 / 26674.10 instead of the invoiced 1404 / 26674 and
# left a fractional-EUR residue that never cleared on the client account.
CAR_152392 = {
    "id": 500, "anexa_id": 4, "line_number": 1, "nr_comanda": "152392",
    "model": "A3 Limuzina S line 40 TFSI quattro", "culoare": "Negru Mythos",
    "list_price_eur": 0, "selling_price_eur": 28078, "qty": 1,
}
ADV_5PCT = {  # 5% advance stored as 1403.89 (a fractional slice), @ 5.0954
    "id": 940, "invoice_number": 9200881, "total_amount_eur": 1403.89,
    "split_mode": "proportional", "kurs_applied": 5.0954,
    "issued_date": "2026-02-25", "line_ids": [500],
}
ADV_95PCT = {  # 95% advance stored as 26674.10, @ 5.2414
    "id": 941, "invoice_number": 9201802, "total_amount_eur": 26674.10,
    "split_mode": "proportional", "kurs_applied": 5.2414,
    "issued_date": "2026-07-17", "line_ids": [500],
}
STORNO_152392 = {
    "id": 400, "invoice_type": "STORNO", "invoice_number": 9201914, "anexa_id": 4,
    "total_amount_eur": -28078, "split_mode": "equal", "kurs_applied": 5.2340982,
    "issued_date": "2026-08-04", "line_ids": [500],
}
FINAL_152392 = {
    "id": 401, "invoice_type": "FINAL", "invoice_number": 9201915, "anexa_id": 4,
    "total_amount_eur": 28078, "split_mode": "proportional", "kurs_applied": 5.2340982,
    "issued_date": "2026-08-04", "line_ids": [500],
}


class DecimalAdvanceFakeRepo(FakeRepo):
    """Audi A3 152392: two advances stored with sub-EUR decimals (1403.89 / 26674.10)."""

    def get_anexa_by_id(self, anexa_id):
        return {"id": 4, "contract_id": 1, "anexa_number": 1}

    def get_lines_by_anexa(self, anexa_id):
        return [dict(CAR_152392)]

    def get_document_number_map(self, invoice_id):
        return {500: 9201914} if invoice_id == 400 else {500: 9201915}

    def query_all(self, sql, params=None):
        if "facturare_invoice_links" in sql:
            return [{"source_invoice_id": 940}, {"source_invoice_id": 941}]
        if "invoice_type = 'STORNO'" in sql:               # matching-storno lookup (final)
            return [{"id": 400, "total_amount_eur": -28078, "line_ids": [500]}]
        if "id IN" in sql:                                 # reversed advances for storno / final
            return [dict(ADV_5PCT), dict(ADV_95PCT)]
        if "invoice_type = 'INVOICE'" in sql:              # fallback: all advances for anexa
            return [dict(ADV_5PCT), dict(ADV_95PCT)]
        return []


# ── Mixed-advance multi-car storno (prod: 9 Polo, contract 142, anexa 23) ──────
# A multi-car storno reverses a whole-anexa proportional advance (covers ALL
# cars) PLUS one single-car 90% advance per car. Each single-car advance must
# reverse ONLY its own car. The export instead looped every reversed advance
# across every storno car, fanning each single-car 90% advance onto all cars
# (prod: 10 advances × 9 cars = 90 rows instead of 18, over-reversing ~8×).
# Simplified here to 2 cars: a 10% whole-anexa advance + one 90% advance each.
MIX_CAR_1 = {
    "id": 601, "anexa_id": 5, "line_number": 1, "nr_comanda": "601",
    "model": "Polo Prime 1.0 TSI DSG", "culoare": "", "list_price_eur": 0,
    "selling_price_eur": 10000, "qty": 1,
}
MIX_CAR_2 = {
    "id": 602, "anexa_id": 5, "line_number": 2, "nr_comanda": "602",
    "model": "Polo Prime 1.0 TSI DSG", "culoare": "", "list_price_eur": 0,
    "selling_price_eur": 10000, "qty": 1,
}
ADV_MIX_BIG = {  # 10% whole-anexa advance covering BOTH cars
    "id": 950, "invoice_type": "INVOICE", "invoice_number": 700, "anexa_id": 5,
    "total_amount_eur": 2000, "split_mode": "proportional", "kurs_applied": 5.0961,
    "issued_date": "2026-02-02", "line_ids": [601, 602],
}
ADV_MIX_CAR_1 = {  # 90% advance covering ONLY car 601
    "id": 951, "invoice_type": "INVOICE", "invoice_number": 701, "anexa_id": 5,
    "total_amount_eur": 9000, "split_mode": "proportional", "kurs_applied": 5.2435,
    "issued_date": "2026-08-14", "line_ids": [601],
}
ADV_MIX_CAR_2 = {  # 90% advance covering ONLY car 602
    "id": 952, "invoice_type": "INVOICE", "invoice_number": 702, "anexa_id": 5,
    "total_amount_eur": 9000, "split_mode": "proportional", "kurs_applied": 5.2435,
    "issued_date": "2026-08-14", "line_ids": [602],
}
STORNO_MIXED = {
    "id": 600, "invoice_type": "STORNO", "invoice_number": 9104148, "anexa_id": 5,
    "total_amount_eur": -20000, "split_mode": "equal", "kurs_applied": 5.2025,
    "issued_date": "2026-08-28", "line_ids": [601, 602],
}


class MixedAdvanceFakeRepo(FakeRepo):
    """2-car storno reversing a whole-anexa advance + one single-car advance each."""

    def get_anexa_by_id(self, anexa_id):
        return {"id": 5, "contract_id": 1, "anexa_number": 23}

    def get_lines_by_anexa(self, anexa_id):
        return [dict(MIX_CAR_1), dict(MIX_CAR_2)]

    def get_invoices_by_anexa(self, anexa_id):
        return [dict(ADV_MIX_BIG), dict(ADV_MIX_CAR_1), dict(ADV_MIX_CAR_2),
                dict(STORNO_MIXED)]

    def get_document_number_map(self, invoice_id):
        return {601: 9104148, 602: 9104148}

    def query_all(self, sql, params=None):
        if "facturare_invoice_links" in sql:
            return [{"source_invoice_id": 950}, {"source_invoice_id": 951},
                    {"source_invoice_id": 952}]
        if "invoice_type = 'STORNO'" in sql:
            return [dict(STORNO_MIXED)]
        if "id IN" in sql:                              # reversed advances for storno
            return [dict(ADV_MIX_BIG), dict(ADV_MIX_CAR_1), dict(ADV_MIX_CAR_2)]
        if "invoice_type = 'INVOICE'" in sql:
            return [dict(ADV_MIX_BIG), dict(ADV_MIX_CAR_1), dict(ADV_MIX_CAR_2)]
        return []


def test_multi_car_storno_reverses_only_covered_cars(monkeypatch):
    """Each single-car advance must reverse ONLY its own car — no cross-car fan-out.

    2 cars, 3 reversed advances (1 whole-anexa + 2 single-car). Correct output is
    4 rows (each car: whole-anexa share + its own advance), not 3 × 2 = 6, and the
    storno must reverse exactly its own total (-20000), not over-reverse.
    """
    monkeypatch.setattr(routes_orders, "_repo", MixedAdvanceFakeRepo())
    _cfg, order_lines = routes_orders._build_eurofib_batch(STORNO_MIXED)

    assert len(order_lines) == 4, (
        f"expected 4 rows (no cross-car fan-out), got {len(order_lines)}"
    )
    assert sum(ol.advance for ol in order_lines) == -20000

    per_car = {}
    for ol in order_lines:
        per_car[ol.comanda] = per_car.get(ol.comanda, 0) + ol.advance
    assert per_car == {601: -10000, 602: -10000}


def _full_betrag(ws, row):
    """Full-precision betrag (Excel evaluates =M*AG with no 2-decimal rounding)."""
    return (ws.cell(row=row, column=COL_FWBETRAG).value
            * ws.cell(row=row, column=COL_KURS).value)


def test_storno_uses_whole_euro_matching_invoice(monkeypatch):
    """Storno EUR (fwbetrag) must be the whole-EUR amount printed on the invoice
    (1404 / 26674), NOT the stored fractional slice (1403.89 / 26674.10)."""
    monkeypatch.setattr(routes_orders, "_repo", DecimalAdvanceFakeRepo())
    ws = _render(STORNO_152392)
    # rows 3/4 = 5% advance, rows 5/6 = 95% advance (debit/credit pairs)
    assert ws.cell(row=3, column=COL_FWBETRAG).value == -1404
    assert ws.cell(row=5, column=COL_FWBETRAG).value == -26674


def test_decimal_storno_fully_reverses_advance_and_final_nets(monkeypatch):
    """The storno RON must reverse the whole-EUR advances (1404@5.0954,
    26674@5.2414) and the final must net the storno to ~0 RON at full precision."""
    monkeypatch.setattr(routes_orders, "_repo", DecimalAdvanceFakeRepo())
    storno_batch = routes_orders._build_eurofib_batch(STORNO_152392)
    final_batch = routes_orders._build_eurofib_batch(FINAL_152392)
    xlsx = EurofibXlsxRenderer.render_multi_to_bytes([storno_batch, final_batch])
    ws = load_workbook(io.BytesIO(xlsx)).active
    # Debit rows: storno 5% (3), storno 95% (5), final (7).
    assert round(_full_betrag(ws, 3), 2) == round(-1404 * 5.0954, 2)    # -7153.94
    assert round(_full_betrag(ws, 5), 2) == round(-26674 * 5.2414, 2)   # -139809.10
    assert ws.cell(row=7, column=COL_FWBETRAG).value == 28078
    net = _full_betrag(ws, 3) + _full_betrag(ws, 5) + _full_betrag(ws, 7)
    assert abs(net) < 0.01, f"storno + final should net to zero, got {net}"
