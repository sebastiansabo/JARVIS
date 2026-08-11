"""Tests for the centralizator XLSX importer (CarPark Dispo, Phase 5 Task 5.2).

Runs against localhost/defaultdb via the probe in conftest.py (same
REAL_DB_AVAILABLE dance as test_dispo_repository_sql.py / test_phase2_e2e.py)
— even the pure dry-run test needs a real DB, since Vânzător/Achizitor name
resolution queries the real `users` table.

Sentinel company_id=990003 — distinct from TEST_COMPANY_ID=990001
(test_dispo_repository_sql.py's dispo_seed) and E2E_COMPANY_ID=990002
(test_phase2_e2e.py), so none of the three can ever collide.

Invocation:
    DATABASE_URL=postgresql://localhost/defaultdb \
        venv/bin/python -m pytest jarvis/tests/carpark/test_import_centralizator.py -v
"""
import io
from datetime import date

import openpyxl
import pytest

from decimal import Decimal

from database import get_db, get_cursor, release_db
from carpark.services.import_service import CentralizatorImporter, _parse_money
from carpark.repositories.vehicle_repository import VehicleRepository
from carpark.repositories.cost_repository import CostRepository
from carpark.repositories.revenue_repository import RevenueRepository

from .conftest import REAL_DB_AVAILABLE

IMPORT_COMPANY_ID = 990003
# Second sentinel for the cross-company guard test — a VIN pre-existing
# under this company must NOT be updated when the import targets
# IMPORT_COMPANY_ID.
OTHER_COMPANY_ID = 990004

_HEADERS = [
    'Marca', 'Model', 'IMPUS', 'SCOS DIN EVIDENTA', 'Serie_sasiu', 'Furnizor',
    'Locatie', 'Data achizitie', 'PV intrare', 'Data plata', 'Pret achizitie',
    'Data promovare', 'Data vanzarii', 'Total costuri', 'Pret vanzare',
    'Bonus leasing', 'Tip vanzare', 'Client', 'Vanzator', 'Achizitor',
    'DOSAR GW', 'PV livrare', 'Chelt cu vanzarea',
]

# VINs used by the fixture — distinctive TESTIMPORT prefix so they can never
# collide with real data. Lengths intentionally vary: row 2's VIN is 16
# chars (dirty-but-tolerated), everything else is a clean 17.
VIN_READY = 'TESTIMPORT0000001'[:17]      # no listing/sale/delivery -> READY_FOR_SALE
VIN_DIRTY = 'TESTIMPORT000002'             # 16 chars -> warning, still imported
VIN_SOLD = 'TESTIMPORT0000004'
VIN_DELIVERED = 'TESTIMPORT0000005'
VIN_COSTS = 'TESTIMPORT0000006'            # Total costuri + Chelt cu vanzarea + Bonus leasing

assert len(VIN_READY) == 17
assert len(VIN_DIRTY) == 16
assert len(VIN_SOLD) == 17
assert len(VIN_DELIVERED) == 17
assert len(VIN_COSTS) == 17


def _row(vin, brand='Audi', model='A6', **overrides):
    values = {h: None for h in _HEADERS}
    values['Marca'] = brand
    values['Model'] = model
    values['Serie_sasiu'] = vin
    values['Furnizor'] = 'TEST SUPPLIER'
    values['Locatie'] = 'TEST LOCATION'
    values['Data achizitie'] = date(2026, 1, 5)
    values.update(overrides)
    return [values[h] for h in _HEADERS]


def _build_fixture_workbook() -> io.BytesIO:
    """6 rows covering every case the task asks for:
      1. clean 17-char VIN, no dates          -> ok, READY_FOR_SALE
      2. dirty 16-char VIN                    -> ok + warning, READY_FOR_SALE
      3. blank VIN                            -> skipped_no_vin
      4. Data vanzarii present                -> ok, SOLD
      5. PV livrare present                   -> ok, DELIVERED
      6. Total costuri + Chelt cu vanzarea +
         Bonus leasing, salesperson resolvable -> ok, 2 cost rows + 1 revenue row
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AUDI'
    ws.append(_HEADERS)

    ws.append(_row(VIN_READY))
    ws.append(_row(VIN_DIRTY))
    ws.append(_row(None, brand='Audi', model='Blank VIN row'))  # no VIN -> skipped
    ws.append(_row(VIN_SOLD, **{'Data vanzarii': date(2026, 2, 1), 'Pret vanzare': 15000}))
    ws.append(_row(VIN_DELIVERED, **{'PV livrare': date(2026, 2, 10)}))
    ws.append(_row(VIN_COSTS, **{
        'Total costuri': 2000, 'Chelt cu vanzarea': 300, 'Bonus leasing': 150,
        'Vanzator': 'Sebastian Sabo',
    }))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────
# _parse_money — pure-function unit tests (no DB), RO + US conventions.
# The real AUDI sheet's money cells are NUMERIC (int/float), so they hit the
# passthrough path; these cover the TEXT-cell path that a hand-edit produces.
# ─────────────────────────────────────────────────────────────────────────

@pytest.mark.parametrize('value, expected', [
    # numeric passthrough (must stay EXACTLY as-is — real-sheet path)
    (1234, Decimal('1234')),
    (1234.56, Decimal('1234.56')),
    (0, Decimal('0')),
    (Decimal('9.99'), Decimal('9.99')),
    (-45273.52, Decimal('-45273.52')),
    # RO decimal-comma text
    ('1234,56', Decimal('1234.56')),
    ('1.234,56', Decimal('1234.56')),
    # US decimal-dot text
    ('1,234.56', Decimal('1234.56')),
    ('1234.56', Decimal('1234.56')),
    # thousands-only / no-separator
    ('1234', Decimal('1234')),
    ('1.234.567', Decimal('1234567')),   # RO thousands (dots)
    ('1,234,567', Decimal('1234567')),   # US thousands (commas)
    # currency symbols / whitespace stripped
    ('1.234,56 RON', Decimal('1234.56')),
    ('€1.234,56', Decimal('1234.56')),
    ('-1234,56', Decimal('-1234.56')),
    # blank / unparseable → None (never raises)
    ('', None),
    (None, None),
    ('   ', None),
    ('garbage', None),
    ('-', None),
])
def test_parse_money_handles_both_conventions(value, expected):
    assert _parse_money(value) == expected


@pytest.fixture
def require_real_db():
    if not REAL_DB_AVAILABLE:
        pytest.skip(
            'Real Postgres not available (DATABASE_URL unreachable or psycopg2 '
            'mocked) — skipping centralizator importer DB-backed test'
        )


@pytest.fixture
def cleanup_import_company():
    """Deletes every carpark_vehicles row for IMPORT_COMPANY_ID before and
    after the test (cost/revenue/status_history cascade via `vehicle_id ...
    ON DELETE CASCADE`), asserting zero remain at teardown."""
    if not REAL_DB_AVAILABLE:
        pytest.skip(
            'Real Postgres not available (DATABASE_URL unreachable or psycopg2 '
            'mocked) — skipping centralizator importer DB-backed test'
        )

    def _delete():
        conn = get_db()
        try:
            cur = get_cursor(conn)
            cur.execute('DELETE FROM carpark_vehicles WHERE company_id = %s', (IMPORT_COMPANY_ID,))
            conn.commit()
        finally:
            release_db(conn)

    _delete()
    try:
        yield
    finally:
        _delete()
        conn = get_db()
        try:
            cur = get_cursor(conn)
            cur.execute('SELECT COUNT(*) AS cnt FROM carpark_vehicles WHERE company_id = %s',
                        (IMPORT_COMPANY_ID,))
            remaining = cur.fetchone()['cnt']
        finally:
            release_db(conn)
        assert remaining == 0, (
            f'teardown left {remaining} orphan carpark_vehicles row(s) for '
            f'company_id={IMPORT_COMPANY_ID}'
        )


@pytest.fixture
def cleanup_both_companies():
    """Cross-company test: cleans BOTH sentinels (IMPORT_COMPANY_ID and
    OTHER_COMPANY_ID) before and after, asserting zero remain for each."""
    if not REAL_DB_AVAILABLE:
        pytest.skip(
            'Real Postgres not available (DATABASE_URL unreachable or psycopg2 '
            'mocked) — skipping centralizator importer DB-backed test'
        )

    def _delete():
        conn = get_db()
        try:
            cur = get_cursor(conn)
            cur.execute('DELETE FROM carpark_vehicles WHERE company_id = ANY(%s)',
                        ([IMPORT_COMPANY_ID, OTHER_COMPANY_ID],))
            conn.commit()
        finally:
            release_db(conn)

    _delete()
    try:
        yield
    finally:
        _delete()
        conn = get_db()
        try:
            cur = get_cursor(conn)
            cur.execute('SELECT COUNT(*) AS cnt FROM carpark_vehicles WHERE company_id = ANY(%s)',
                        ([IMPORT_COMPANY_ID, OTHER_COMPANY_ID],))
            remaining = cur.fetchone()['cnt']
        finally:
            release_db(conn)
        assert remaining == 0, (
            f'teardown left {remaining} orphan carpark_vehicles row(s) for '
            f'company_id in ({IMPORT_COMPANY_ID}, {OTHER_COMPANY_ID})'
        )


# ─────────────────────────────────────────────────────────────────────────
# DRY RUN
# ─────────────────────────────────────────────────────────────────────────

def test_dry_run_counts_statuses_and_writes_nothing(require_real_db, cleanup_import_company):
    importer = CentralizatorImporter()
    wb = _build_fixture_workbook()

    report = importer.run(wb, IMPORT_COMPANY_ID, dry_run=True)

    assert report.dry_run is True
    assert report.total == 6
    assert report.skipped_no_vin == 1
    assert report.rejects == 0
    assert report.ok == 5
    assert report.warnings == 1  # only the dirty 16-char VIN row

    by_vin = {r['vin']: r for r in report.rows if r['vin']}

    assert by_vin[VIN_READY]['status'] == 'ok'
    assert by_vin[VIN_READY]['inferred_status'] == 'READY_FOR_SALE'
    assert by_vin[VIN_READY]['action'] == 'create'

    assert by_vin[VIN_DIRTY]['status'] == 'ok'
    assert by_vin[VIN_DIRTY]['inferred_status'] == 'READY_FOR_SALE'
    assert any('16 chars' in w for w in by_vin[VIN_DIRTY]['warnings'])

    assert by_vin[VIN_SOLD]['inferred_status'] == 'SOLD'
    assert by_vin[VIN_DELIVERED]['inferred_status'] == 'DELIVERED'
    assert by_vin[VIN_COSTS]['inferred_status'] == 'READY_FOR_SALE'

    skipped = [r for r in report.rows if r['status'] == 'skipped_no_vin']
    assert len(skipped) == 1

    # dry-run must not have written ANY of these VINs to the DB
    vehicle_repo = VehicleRepository()
    for vin in (VIN_READY, VIN_DIRTY, VIN_SOLD, VIN_DELIVERED, VIN_COSTS):
        assert vehicle_repo.get_by_vin(vin) is None


def test_dry_run_rejects_vin_over_17_chars(require_real_db, cleanup_import_company):
    """VARCHAR(17) column can't hold an 18-char VIN — parse() must reject
    it (not silently truncate, not warn-and-import)."""
    importer = CentralizatorImporter()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AUDI'
    ws.append(_HEADERS)
    ws.append(_row('TESTIMPORT00000018X'[:18]))  # 18 chars
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    report = importer.run(buf, IMPORT_COMPANY_ID, dry_run=True)

    assert report.total == 1
    assert report.rejects == 1
    assert report.ok == 0
    assert 'exceeds' in report.rows[0]['reject_reason']


# ─────────────────────────────────────────────────────────────────────────
# COMMIT
# ─────────────────────────────────────────────────────────────────────────

def test_commit_upserts_writes_costs_revenues_and_is_idempotent(
        require_real_db, cleanup_import_company):
    importer = CentralizatorImporter()
    vehicle_repo = VehicleRepository()
    cost_repo = CostRepository()
    revenue_repo = RevenueRepository()

    # ── first commit: everything is a CREATE ──
    report1 = importer.run(_build_fixture_workbook(), IMPORT_COMPANY_ID, dry_run=False)

    assert report1.dry_run is False
    assert report1.error is None
    assert report1.committed_vehicles_created == 5
    assert report1.committed_vehicles_updated == 0
    assert report1.committed_cost_rows == 2       # istoric_import + cheltuieli_vanzare (VIN_COSTS row)
    assert report1.committed_revenue_rows == 1     # bonus_leasing (VIN_COSTS row)

    ready = vehicle_repo.get_by_vin(VIN_READY)
    assert ready is not None
    assert ready['status'] == 'READY_FOR_SALE'

    dirty = vehicle_repo.get_by_vin(VIN_DIRTY)
    assert dirty is not None

    sold = vehicle_repo.get_by_vin(VIN_SOLD)
    assert sold['status'] == 'SOLD'

    delivered = vehicle_repo.get_by_vin(VIN_DELIVERED)
    assert delivered['status'] == 'DELIVERED'

    costs_vehicle = vehicle_repo.get_by_id(
        vehicle_repo.get_by_vin(VIN_COSTS)['id'])
    assert costs_vehicle['salesperson_user_id'] == 1  # resolved "Sebastian Sabo" -> user id 1

    vehicle_costs = cost_repo.get_by_vehicle(costs_vehicle['id'])
    cost_types = sorted(c['cost_type'] for c in vehicle_costs)
    assert cost_types == ['cheltuieli_vanzare', 'istoric_import']
    assert {round(float(c['amount']), 2) for c in vehicle_costs} == {2000.0, 300.0}

    vehicle_revenues = revenue_repo.get_by_vehicle(costs_vehicle['id'])
    assert len(vehicle_revenues) == 1
    assert vehicle_revenues[0]['revenue_type'] == 'bonus_leasing'
    assert round(float(vehicle_revenues[0]['amount']), 2) == 150.0

    # ── second commit (re-run the same file): idempotent, not duplicated ──
    report2 = importer.run(_build_fixture_workbook(), IMPORT_COMPANY_ID, dry_run=False)

    assert report2.committed_vehicles_created == 0
    assert report2.committed_vehicles_updated == 5
    assert report2.committed_cost_rows == 2
    assert report2.committed_revenue_rows == 1

    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute('SELECT COUNT(*) AS cnt FROM carpark_vehicles WHERE company_id = %s',
                    (IMPORT_COMPANY_ID,))
        assert cur.fetchone()['cnt'] == 5  # still 5, no duplicate vehicles

        cur.execute('SELECT COUNT(*) AS cnt FROM carpark_vehicle_costs WHERE vehicle_id = %s',
                    (costs_vehicle['id'],))
        assert cur.fetchone()['cnt'] == 2  # still 2, not 4

        cur.execute('SELECT COUNT(*) AS cnt FROM carpark_vehicle_revenues WHERE vehicle_id = %s',
                    (costs_vehicle['id'],))
        assert cur.fetchone()['cnt'] == 1  # still 1, not 2
    finally:
        release_db(conn)


# ─────────────────────────────────────────────────────────────────────────
# CROSS-COMPANY VIN GUARD (protects both the HTTP route and the CLI)
# ─────────────────────────────────────────────────────────────────────────

VIN_CROSS = 'TESTIMPORTX000003'
assert len(VIN_CROSS) == 17


def test_cross_company_vin_is_rejected_and_other_company_vehicle_unchanged(
        require_real_db, cleanup_both_companies):
    """A VIN that already belongs to company A must NOT be updated (nor
    have its company_id reassigned) when the import targets company B — it's
    reported as a cross_company conflict and company A's vehicle is left
    exactly as it was."""
    importer = CentralizatorImporter()
    vehicle_repo = VehicleRepository()

    # ── seed VIN_CROSS under company A (OTHER_COMPANY_ID) ──
    conn = get_db()
    conn.autocommit = False
    cur = get_cursor(conn)
    try:
        cur.execute('''
            INSERT INTO carpark_vehicles (vin, brand, model, status, company_id, category)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id
        ''', (VIN_CROSS, 'CompanyA-Brand', 'CompanyA-Model', 'LISTED',
              OTHER_COMPANY_ID, 'SH'))
        company_a_vehicle_id = cur.fetchone()['id']
        conn.commit()
    finally:
        release_db(conn)

    # ── import file carries the SAME VIN with DIFFERENT data, targeting company B ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AUDI'
    ws.append(_HEADERS)
    ws.append(_row(VIN_CROSS, brand='Import-Brand', model='Import-Model',
                   **{'Data promovare': date(2026, 3, 1)}))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # ── DRY RUN: reported as a cross-company conflict, not ok ──
    dry = importer.run(buf, IMPORT_COMPANY_ID, dry_run=True)
    assert dry.total == 1
    assert dry.ok == 0
    assert dry.cross_company == 1
    assert dry.rejects == 0
    conflict_row = dry.rows[0]
    assert conflict_row['status'] == 'cross_company_conflict'
    assert str(OTHER_COMPANY_ID) in conflict_row['reject_reason']

    # ── COMMIT: nothing written; company A's vehicle is untouched ──
    committed = importer.run(_rebuild_cross_workbook(), IMPORT_COMPANY_ID, dry_run=False)
    assert committed.cross_company == 1
    assert committed.committed_vehicles_created == 0
    assert committed.committed_vehicles_updated == 0

    company_a_vehicle = vehicle_repo.get_by_id(company_a_vehicle_id)
    assert company_a_vehicle is not None
    assert company_a_vehicle['company_id'] == OTHER_COMPANY_ID   # NOT reassigned
    assert company_a_vehicle['brand'] == 'CompanyA-Brand'        # NOT overwritten
    assert company_a_vehicle['model'] == 'CompanyA-Model'
    assert company_a_vehicle['status'] == 'LISTED'               # NOT changed

    # ── and nothing was created under company B ──
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute('SELECT COUNT(*) AS cnt FROM carpark_vehicles WHERE company_id = %s',
                    (IMPORT_COMPANY_ID,))
        assert cur.fetchone()['cnt'] == 0
    finally:
        release_db(conn)


# ─────────────────────────────────────────────────────────────────────────
# IMPUS / SCOS mutual exclusivity (INVARIANT: NOT (is_impus AND stock_removed))
# ─────────────────────────────────────────────────────────────────────────

VIN_BOTH_FLAGS = 'TESTIMPORT0000007'
assert len(VIN_BOTH_FLAGS) == 17


def _both_flags_workbook() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AUDI'
    ws.append(_HEADERS)
    ws.append(_row(VIN_BOTH_FLAGS, **{'IMPUS': 'DA', 'SCOS DIN EVIDENTA': 'DA'}))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_row_with_both_impus_and_scos_keeps_scos_and_warns(
        require_real_db, cleanup_import_company):
    """A centralizator row can carry both IMPUS='DA' and SCOS DIN
    EVIDENTA='DA' (a vehicle flagged IMPUS that was later removed from
    stock). That's a contradiction under the mutual-exclusivity invariant —
    SCOS wins as the later lifecycle state, is_impus is cleared, and the row
    is still imported (`ok`), just carrying a warning so the discrepancy
    stays visible in the report."""
    importer = CentralizatorImporter()
    vehicle_repo = VehicleRepository()

    report = importer.run(_both_flags_workbook(), IMPORT_COMPANY_ID, dry_run=True)

    assert report.total == 1
    assert report.ok == 1
    assert report.rejects == 0
    assert report.warnings == 1
    row = report.rows[0]
    assert row['vin'] == VIN_BOTH_FLAGS
    assert row['status'] == 'ok'
    assert any('IMPUS' in w and 'SCOS' in w for w in row['warnings'])

    # dry-run writes nothing
    assert vehicle_repo.get_by_vin(VIN_BOTH_FLAGS) is None

    committed = importer.run(_both_flags_workbook(), IMPORT_COMPANY_ID, dry_run=False)
    assert committed.committed_vehicles_created == 1

    # get_by_vin() only projects id/vin/brand/model/status — go through
    # get_by_id() for the full row (mirrors the VIN_COSTS lookup above).
    vehicle = vehicle_repo.get_by_id(vehicle_repo.get_by_vin(VIN_BOTH_FLAGS)['id'])
    assert vehicle is not None
    assert vehicle['stock_removed'] is True
    assert vehicle['is_impus'] is False


def _rebuild_cross_workbook() -> io.BytesIO:
    """Fresh stream for the second (commit) importer.run in the
    cross-company test — read_only workbooks consume their stream."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AUDI'
    ws.append(_HEADERS)
    ws.append(_row(VIN_CROSS, brand='Import-Brand', model='Import-Model',
                   **{'Data promovare': date(2026, 3, 1)}))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
