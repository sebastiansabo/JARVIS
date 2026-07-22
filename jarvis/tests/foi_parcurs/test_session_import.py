"""Tests for bulk session import (template + parse/validate/insert + routes).
Mirrors the fixtures/monkeypatching pattern in test_test_drive_submit.py."""
import os

os.environ.setdefault('DATABASE_URL', 'postgresql://localhost/defaultdb')

import io
from datetime import datetime

import pytest
from flask import Flask

from foi_parcurs import foi_parcurs_bp
from foi_parcurs.services import session_import_service as sis
import foi_parcurs.routes.session_import as si_routes


# ── Task 1: parse + validate ──

def test_parse_dt_accepts_romanian_and_iso():
    assert sis.parse_dt('02.07.2026 10:00') == datetime(2026, 7, 2, 10, 0)
    assert sis.parse_dt('02.07.2026') == datetime(2026, 7, 2, 0, 0)
    assert sis.parse_dt('2026-07-02 10:00') == datetime(2026, 7, 2, 10, 0)
    assert sis.parse_dt(datetime(2026, 7, 2, 8, 30)) == datetime(2026, 7, 2, 8, 30)
    assert sis.parse_dt('') is None
    assert sis.parse_dt('nonsense') is None


def _row(**kw):
    base = {c: '' for c in sis.SESSION_COLUMNS}
    base.update(kw)
    return base


def test_row_error_missing_vin():
    assert sis.row_error(_row(VIN=''), None, 1) == 'VIN lipsă'


def test_row_error_wrong_company():
    veh = {'vin': 'V1', 'company_id': 99}
    assert sis.row_error(_row(VIN='V1', **{'KM start': '10', 'KM end': '20', 'Plecare': '02.07.2026'}), veh, 1) \
        == 'VIN aparține altei companii'


def test_row_error_new_vin_needs_make_model():
    r = _row(VIN='NEW', **{'KM start': '10', 'KM end': '20', 'Plecare': '02.07.2026'})
    assert sis.row_error(r, None, 1) == 'Mașină nouă necesită Marcă și Model'


def test_row_error_bad_km():
    veh = {'vin': 'V1', 'company_id': 1}
    r = _row(VIN='V1', **{'KM start': '20', 'KM end': '20', 'Plecare': '02.07.2026'})
    assert sis.row_error(r, veh, 1) == 'KM invalizi (KM end trebuie > KM start)'


def test_row_error_ok():
    veh = {'vin': 'V1', 'company_id': 1}
    r = _row(VIN='V1', **{'KM start': '10', 'KM end': '25', 'Plecare': '02.07.2026 10:00'})
    assert sis.row_error(r, veh, 1) is None


# ── Task 2: template ──

def test_build_template_has_sheets_and_headers(monkeypatch):
    from openpyxl import load_workbook
    monkeypatch.setattr(
        sis._veh_repo, 'get_all',
        lambda active_only=True: [
            {'vin': 'V1', 'mark': 'Audi', 'model': 'A4', 'company_id': 1},
            {'vin': 'V2', 'mark': 'VW', 'model': 'Golf', 'company_id': 2},
        ],
    )
    wb = load_workbook(io.BytesIO(sis.build_template_xlsx(1)))
    assert wb['Sesiuni'][1][0].value == 'VIN'
    assert [c.value for c in wb['Sesiuni'][1]] == sis.SESSION_COLUMNS
    vins = [row[2].value for row in wb['Mașini'].iter_rows(min_row=2) if row[2].value]
    assert vins == ['V1']


# ── Task 3: orchestration ──

def test_import_inserts_skips_and_creates_car(monkeypatch):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Sesiuni'
    for col, name in enumerate(sis.SESSION_COLUMNS, start=1):
        ws.cell(row=1, column=col, value=name)
    rows = [
        ['V1', '', '', '', '', '', '', '02.07.2026 10:00', '', 100, 130, 'Ana'],
        ['V1', '', '', '', '', '', '', '02.07.2026 10:00', '', 100, 130, 'Ana'],
        ['VNEW', 'Audi', 'A4', 'B1', '', '', '', '03.07.2026 09:00', '', 0, 40, 'Ion'],
    ]
    for r, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)
    buf = io.BytesIO(); wb.save(buf); data = buf.getvalue()

    monkeypatch.setattr(sis._veh_repo, 'get_by_vin',
                        lambda vin: {'vin': 'V1', 'company_id': 1, 'fuel_tank_capacity_liters': 55,
                                     'registration_number': 'B-V1'} if vin == 'V1' else None)
    seen = {'inserted': [], 'vehicles': []}

    def fake_query_one(sql, params=()):
        if 'td_km_max' in sql:
            return {'td_km_max': 50}
        if 'FROM foi_de_parcurs' in sql:
            cid = params[0]
            return {'x': 1} if cid in seen['inserted'] else None
        return None

    def fake_execute(sql, params=(), **kw):
        if 'INTO fp_vehicles' in sql:
            seen['vehicles'].append(params[0])
        elif 'INTO foi_de_parcurs' in sql:
            seen['inserted'].append(params[0])
        return 1

    monkeypatch.setattr(sis._fp_repo, 'query_one', fake_query_one)
    monkeypatch.setattr(sis._fp_repo, 'execute', fake_execute)

    res = sis.import_sessions(1, data, 'Tester')
    assert res['inserted'] == 2
    assert res['skipped'] == 1
    assert res['cars_created'] == 1
    assert res['errors'] == []


# ── Task 4: routes ──

@pytest.fixture
def client():
    app = Flask(__name__)
    app.register_blueprint(foi_parcurs_bp)
    app.config['TESTING'] = True
    app.config['LOGIN_DISABLED'] = True
    return app.test_client()


def test_template_route_returns_xlsx(client, monkeypatch):
    monkeypatch.setattr(si_routes, 'build_template_xlsx', lambda cid: b'PK\x03\x04xlsx')
    resp = client.get('/api/foi-parcurs/sessions/import-template?company_id=1')
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp.headers['Content-Type']
    assert resp.data == b'PK\x03\x04xlsx'


def test_import_route_returns_report(client, monkeypatch):
    monkeypatch.setattr(si_routes, 'import_sessions',
                        lambda company_id, file_bytes, user_name: {
                            'inserted': 2, 'skipped': 1, 'cars_created': 1, 'errors': []})
    data = {'company_id': '1', 'file': (io.BytesIO(b'xlsxbytes'), 'sessions.xlsx')}
    resp = client.post('/api/foi-parcurs/sessions/import', data=data,
                       content_type='multipart/form-data')
    assert resp.status_code == 200
    body = resp.get_json()
    assert body['success'] is True and body['inserted'] == 2 and body['cars_created'] == 1


def test_import_route_requires_file_and_company(client):
    resp = client.post('/api/foi-parcurs/sessions/import', data={'company_id': '1'},
                       content_type='multipart/form-data')
    assert resp.status_code == 400
