"""Tests for the Bilant formula engine — parsing, evaluation, metrics."""

import pytest
import pandas as pd
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'jarvis'))

from accounting.bilant.formula_engine import (
    prepare_balanta,
    extract_ct_formula,
    extract_row_formula,
    parse_ct_formula,
    sum_accounts_by_prefix,
    eval_ct_expression,
    eval_row_formula,
    process_bilant_from_template,
    calculate_metrics_from_config,
    eval_metric_formula,
)


# ── Fixtures ──

def make_balanta(rows):
    """Create a Balanta DataFrame from list of (account, sfd, sfc) tuples."""
    return pd.DataFrame(rows, columns=['Cont', 'SFD', 'SFC'])


@pytest.fixture
def sample_balanta():
    return make_balanta([
        ('201', 50000, 2500),
        ('2801', 2500, 0),
        ('203', 30000, 0),
        ('2803', 3000, 0),
        ('2903', 1000, 0),
        ('211', 100000, 0),
        ('212', 80000, 0),
        ('2811', 10000, 0),
        ('2812', 5000, 0),
        ('345', 0, 15000),
        ('346', 20000, 0),
        ('348', 5000, 3000),
        ('4428', 1200, 0),
    ])


# ══════════════════════════════════════════════════════════════
# prepare_balanta
# ══════════════════════════════════════════════════════════════

class TestPrepareBalanta:
    def test_skip_header_row(self):
        df = pd.DataFrame([['Cont', 'SFD', 'SFC'], ['201', 50000, 0]])
        result = prepare_balanta(df)
        assert len(result) == 1
        assert result.iloc[0, 0] == '201'

    def test_no_header_row(self):
        df = make_balanta([('201', 50000, 0)])
        result = prepare_balanta(df)
        assert len(result) == 1

    def test_empty_dataframe(self):
        df = pd.DataFrame(columns=['Cont', 'SFD', 'SFC'])
        result = prepare_balanta(df)
        assert len(result) == 0


# ══════════════════════════════════════════════════════════════
# extract_ct_formula
# ══════════════════════════════════════════════════════════════

class TestExtractCtFormula:
    def test_simple(self):
        assert extract_ct_formula("1.Cheltuieli (ct.201-2801)") == "201-2801"

    def test_complex(self):
        assert extract_ct_formula("Item (ct. 205 + 208 - 2805 - 2808 - 2905 - 2908)") == "205+208-2805-2808-2905-2908"

    def test_dynamic_sign(self):
        assert extract_ct_formula("Item (ct.345+346+/-348)") == "345+346+/-348"

    def test_dinct(self):
        assert extract_ct_formula("Item (ct.345-dinct.4428)") == "345-dinct.4428"

    def test_no_formula(self):
        assert extract_ct_formula("A. ACTIVE IMOBILIZATE") == ""

    def test_none(self):
        assert extract_ct_formula(None) == ""

    def test_nan(self):
        assert extract_ct_formula(float('nan')) == ""

    def test_asterisks_removed(self):
        assert extract_ct_formula("Item (ct.*201*-*2801*)") == "201-2801"

    def test_whitespace_removed(self):
        assert extract_ct_formula("Item (ct. 201 + 202 )") == "201+202"

    def test_no_dot_no_match(self):
        """'ct' without dot should not match (e.g., 'active')."""
        assert extract_ct_formula("Active circulante totale") == ""


# ══════════════════════════════════════════════════════════════
# extract_row_formula
# ══════════════════════════════════════════════════════════════

class TestExtractRowFormula:
    def test_range(self):
        assert extract_row_formula("TOTAL (rd. 01 la 06)") == "01+02+03+04+05+06"

    def test_range_two_digits(self):
        assert extract_row_formula("TOTAL (rd. 08 la 16)") == "08+09+10+11+12+13+14+15+16"

    def test_simple_addition(self):
        result = extract_row_formula("TOTAL (rd. 31+32+33)")
        assert '31' in result and '32' in result and '33' in result

    def test_35a_preserved(self):
        """35a should be preserved as-is (column B numbering)."""
        result = extract_row_formula("TOTAL (rd. 35a)")
        assert '35a' in result

    def test_no_formula(self):
        assert extract_row_formula("Just a description") == ""

    def test_none(self):
        assert extract_row_formula(None) == ""


# ══════════════════════════════════════════════════════════════
# parse_ct_formula
# ══════════════════════════════════════════════════════════════

class TestParseCtFormula:
    def test_simple_plus(self):
        items = parse_ct_formula("201")
        assert items == [('201', 'normal_plus')]

    def test_plus_minus(self):
        items = parse_ct_formula("201-2801")
        assert items == [('201', 'normal_plus'), ('2801', 'normal_minus')]

    def test_multiple(self):
        items = parse_ct_formula("205+208-2805-2808")
        assert len(items) == 4
        assert items[0] == ('205', 'normal_plus')
        assert items[1] == ('208', 'normal_plus')
        assert items[2] == ('2805', 'normal_minus')
        assert items[3] == ('2808', 'normal_minus')

    def test_dynamic(self):
        items = parse_ct_formula("345+346+/-348")
        assert items == [
            ('345', 'normal_plus'),
            ('346', 'normal_plus'),
            ('348', 'dynamic'),
        ]

    def test_dinct(self):
        items = parse_ct_formula("345-dinct.4428")
        assert items == [
            ('345', 'normal_plus'),
            ('4428', 'normal_minus'),
        ]

    def test_empty(self):
        assert parse_ct_formula("") == []
        assert parse_ct_formula(None) == []

    def test_dinct_inherits_plus_sign(self):
        """dinct. should inherit the preceding sign — positive when preceded by +."""
        items = parse_ct_formula("345+dinct.4428")
        assert items == [
            ('345', 'normal_plus'),
            ('4428', 'normal_plus'),
        ]

    def test_dinct_inherits_minus_sign(self):
        """dinct. should inherit the preceding sign — negative when preceded by -."""
        items = parse_ct_formula("345-dinct.4428")
        assert items == [
            ('345', 'normal_plus'),
            ('4428', 'normal_minus'),
        ]

    def test_dinct_default_sign_is_plus(self):
        """dinct. at start of expression defaults to plus."""
        items = parse_ct_formula("dinct.4428")
        assert items == [('4428', 'normal_plus')]

    def test_complex_mixed(self):
        items = parse_ct_formula("345+346-2801+/-348-dinct.4428")
        assert len(items) == 5
        assert ('345', 'normal_plus') in items
        assert ('346', 'normal_plus') in items
        assert ('2801', 'normal_minus') in items
        assert ('348', 'dynamic') in items
        assert ('4428', 'normal_minus') in items


# ══════════════════════════════════════════════════════════════
# sum_accounts_by_prefix
# ══════════════════════════════════════════════════════════════

class TestSumAccountsByPrefix:
    def test_normal_sum(self, sample_balanta):
        total, details = sum_accounts_by_prefix(sample_balanta, '201', use_net=False)
        assert total == 52500  # abs(50000) + abs(2500)
        assert len(details) == 1

    def test_net_sum(self, sample_balanta):
        total, details = sum_accounts_by_prefix(sample_balanta, '348', use_net=True)
        assert total == 2000  # 5000 - 3000
        assert len(details) == 1

    def test_prefix_matching(self, sample_balanta):
        """Prefix '28' should match all accounts starting with 28."""
        total, details = sum_accounts_by_prefix(sample_balanta, '28', use_net=False)
        assert len(details) == 4  # 2801, 2803, 2811, 2812

    def test_no_match(self, sample_balanta):
        total, details = sum_accounts_by_prefix(sample_balanta, '999', use_net=False)
        assert total == 0
        assert len(details) == 0


# ══════════════════════════════════════════════════════════════
# eval_ct_expression
# ══════════════════════════════════════════════════════════════

class TestEvalCtExpression:
    def test_simple_subtraction(self, sample_balanta):
        total, details = eval_ct_expression("201-2801", sample_balanta)
        # 201: abs(50000)+abs(2500)=52500, minus 2801: abs(2500)+abs(0)=2500
        assert total == 50000  # 52500 - 2500

    def test_no_accounts(self, sample_balanta):
        total, details = eval_ct_expression("999", sample_balanta)
        assert total == 0
        assert any(d[1] == 'No Val.' for d in details)

    def test_dynamic_sign(self, sample_balanta):
        total, details = eval_ct_expression("+/-348", sample_balanta)
        assert total == 2000  # SFD(5000) - SFC(3000)


# ══════════════════════════════════════════════════════════════
# eval_row_formula
# ══════════════════════════════════════════════════════════════

class TestEvalRowFormula:
    def test_simple_sum(self):
        values = {'1': 100, '2': 200, '3': 300}
        assert eval_row_formula("1+2+3", values) == 600

    def test_subtraction(self):
        values = {'1': 100, '2': 50}
        assert eval_row_formula("1-2", values) == 50

    def test_leading_zeros(self):
        values = {'1': 100, '2': 200}
        assert eval_row_formula("01+02", values) == 300

    def test_missing_row(self):
        values = {'1': 100}
        assert eval_row_formula("1+2", values) == 100  # row 2 defaults to 0

    def test_35a_token(self):
        """eval_row_formula should handle alphanumeric '35a' tokens."""
        values = {'31': 100, '35': 200, '35a': 50}
        assert eval_row_formula("31+35+35a", values) == 350

    def test_empty(self):
        assert eval_row_formula("", {}) == 0
        assert eval_row_formula(None, {}) == 0


# ══════════════════════════════════════════════════════════════
# process_bilant_from_template
# ══════════════════════════════════════════════════════════════

class TestProcessBilantFromTemplate:
    def test_ct_and_rd_formulas(self, sample_balanta):
        template_rows = [
            {'id': 1, 'description': 'Row 1 (ct.201-2801)', 'nr_rd': '1',
             'formula_ct': '201-2801', 'formula_rd': '', 'row_type': 'data',
             'is_bold': False, 'indent_level': 0, 'sort_order': 0},
            {'id': 2, 'description': 'Row 2 (ct.203-2803-2903)', 'nr_rd': '2',
             'formula_ct': '203-2803-2903', 'formula_rd': '', 'row_type': 'data',
             'is_bold': False, 'indent_level': 0, 'sort_order': 1},
            {'id': 3, 'description': 'TOTAL', 'nr_rd': '3',
             'formula_ct': '', 'formula_rd': '1+2', 'row_type': 'total',
             'is_bold': True, 'indent_level': 0, 'sort_order': 2},
        ]

        bilant_values, results = process_bilant_from_template(sample_balanta, template_rows)

        assert len(results) == 3
        # Row 1: 201(52500) - 2801(2500) = 50000
        assert results[0]['value'] == 50000
        # Row 2: 203(30000) - 2803(3000) - 2903(1000) = 26000
        assert results[1]['value'] == 26000
        # Row 3 (TOTAL): row1 + row2 = 76000
        assert results[2]['value'] == 76000
        assert bilant_values['3'] == 76000

    def test_empty_template(self, sample_balanta):
        bilant_values, results = process_bilant_from_template(sample_balanta, [])
        assert len(results) == 0
        assert len(bilant_values) == 0

    def test_verification_details(self, sample_balanta):
        template_rows = [
            {'id': 1, 'description': 'Test', 'nr_rd': '1',
             'formula_ct': '201', 'formula_rd': '', 'row_type': 'data',
             'is_bold': False, 'indent_level': 0, 'sort_order': 0},
        ]
        _, results = process_bilant_from_template(sample_balanta, template_rows)
        assert '201' in results[0]['verification']
        assert '52500.00' in results[0]['verification']


# ══════════════════════════════════════════════════════════════
# calculate_metrics_from_config
# ══════════════════════════════════════════════════════════════

class TestCalculateMetricsFromConfig:
    def test_basic_metrics(self):
        bilant_values = {
            '25': 60000,   # Active Imobilizate
            '42': 40000,   # Active Circulante
            '30': 15000,   # Stocuri
            '39': 8000,    # Disponibilitati
            '40': 17000,   # Creante
            '54': 25000,   # Datorii < 1 an
            '55': 10000,   # Datorii > 1 an
            '101': 65000,  # Capitaluri Proprii
            '81': 30000,   # Capital Social
        }
        metric_configs = [
            {'metric_key': 'active_imobilizate', 'metric_label': 'Active Imobilizate', 'nr_rd': '25', 'metric_group': 'summary'},
            {'metric_key': 'active_circulante', 'metric_label': 'Active Circulante', 'nr_rd': '42', 'metric_group': 'summary'},
            {'metric_key': 'stocuri', 'metric_label': 'Stocuri', 'nr_rd': '30', 'metric_group': 'ratio_input'},
            {'metric_key': 'disponibilitati', 'metric_label': 'Disponibilitati', 'nr_rd': '39', 'metric_group': 'ratio_input'},
            {'metric_key': 'creante', 'metric_label': 'Creante', 'nr_rd': '40', 'metric_group': 'ratio_input'},
            {'metric_key': 'datorii_termen_scurt', 'metric_label': 'Datorii < 1 an', 'nr_rd': '54', 'metric_group': 'ratio_input'},
            {'metric_key': 'datorii_termen_lung', 'metric_label': 'Datorii > 1 an', 'nr_rd': '55', 'metric_group': 'ratio_input'},
            {'metric_key': 'capitaluri_proprii', 'metric_label': 'Capitaluri Proprii', 'nr_rd': '101', 'metric_group': 'summary'},
            {'metric_key': 'capital_social', 'metric_label': 'Capital Social', 'nr_rd': '81', 'metric_group': 'ratio_input'},
        ]

        metrics = calculate_metrics_from_config(bilant_values, metric_configs)

        # Summary
        assert metrics['summary']['total_active'] == 100000
        assert metrics['summary']['active_imobilizate'] == 60000
        assert metrics['summary']['active_circulante'] == 40000
        assert metrics['summary']['capitaluri_proprii'] == 65000
        assert metrics['summary']['total_datorii'] == 35000

        # Ratios
        assert metrics['ratios']['lichiditate_curenta'] == 1.6   # 40000 / 25000
        assert metrics['ratios']['lichiditate_rapida'] == 1.0    # (40000-15000) / 25000
        assert metrics['ratios']['lichiditate_imediata'] == 0.32  # 8000 / 25000
        assert metrics['ratios']['solvabilitate'] == 65.0         # 65000/100000*100
        assert metrics['ratios']['indatorare'] == 35.0            # 35000/100000*100
        assert metrics['ratios']['autonomie_financiara'] == 65.0  # 65000/(65000+35000)*100

        # Structure
        assert len(metrics['structure']['assets']) == 4
        assert len(metrics['structure']['liabilities']) == 3

    def test_zero_denom(self):
        """Ratios should be None when denominators are 0."""
        bilant_values = {'25': 0, '42': 0}
        metric_configs = [
            {'metric_key': 'active_imobilizate', 'metric_label': 'AI', 'nr_rd': '25', 'metric_group': 'summary'},
            {'metric_key': 'active_circulante', 'metric_label': 'AC', 'nr_rd': '42', 'metric_group': 'summary'},
        ]
        metrics = calculate_metrics_from_config(bilant_values, metric_configs)
        assert metrics['ratios']['lichiditate_curenta'] is None
        assert metrics['ratios']['solvabilitate'] is None

    def test_empty_configs(self):
        metrics = calculate_metrics_from_config({'25': 100}, [])
        # Empty configs → no summary/derived, auto-derive totals from missing keys = 0
        assert metrics['summary'] == {}
        assert metrics['ratios']['lichiditate_curenta'] is None

    def test_dynamic_ratio_configs(self):
        """When ratio configs exist, use formula_expr instead of STANDARD_RATIOS."""
        bilant_values = {'25': 60000, '42': 40000, '54': 25000}
        metric_configs = [
            {'metric_key': 'active_imobilizate', 'metric_label': 'AI', 'nr_rd': '25', 'metric_group': 'summary'},
            {'metric_key': 'active_circulante', 'metric_label': 'AC', 'nr_rd': '42', 'metric_group': 'summary'},
            {'metric_key': 'datorii_termen_scurt', 'metric_label': 'DTS', 'nr_rd': '54', 'metric_group': 'ratio_input'},
            {
                'metric_key': 'my_ratio',
                'metric_label': 'My Custom Ratio',
                'metric_group': 'ratio',
                'formula_expr': 'active_circulante / datorii_termen_scurt',
                'display_format': 'ratio',
                'interpretation': 'Should be > 1',
                'threshold_good': 2.0,
                'threshold_warning': 1.0,
            },
        ]
        metrics = calculate_metrics_from_config(bilant_values, metric_configs)
        # Should use dynamic ratio, not fallback STANDARD_RATIOS
        assert 'my_ratio' in metrics['ratios']
        ratio = metrics['ratios']['my_ratio']
        assert isinstance(ratio, dict)
        assert ratio['value'] == 1.6  # 40000 / 25000
        assert ratio['label'] == 'My Custom Ratio'
        assert ratio['interpretation'] == 'Should be > 1'
        # Fallback ratios should NOT be present
        assert 'lichiditate_curenta' not in metrics['ratios']

    def test_derived_configs(self):
        """Derived metrics use formula_expr over other metric_keys."""
        bilant_values = {'25': 60000, '42': 40000}
        metric_configs = [
            {'metric_key': 'active_imobilizate', 'metric_label': 'AI', 'nr_rd': '25', 'metric_group': 'summary'},
            {'metric_key': 'active_circulante', 'metric_label': 'AC', 'nr_rd': '42', 'metric_group': 'summary'},
            {
                'metric_key': 'total_active',
                'metric_label': 'Total Active',
                'metric_group': 'derived',
                'formula_expr': 'active_imobilizate + active_circulante',
                'display_format': 'currency',
            },
        ]
        metrics = calculate_metrics_from_config(bilant_values, metric_configs)
        assert metrics['summary']['total_active'] == 100000
        assert metrics['summary']['active_imobilizate'] == 60000

    def test_structure_configs(self):
        """Structure configs populate assets/liabilities from config."""
        bilant_values = {'25': 60000, '42': 40000, '101': 70000, '54': 30000}
        metric_configs = [
            {'metric_key': 'active_imobilizate', 'metric_label': 'AI', 'nr_rd': '25', 'metric_group': 'summary'},
            {'metric_key': 'active_circulante', 'metric_label': 'AC', 'nr_rd': '42', 'metric_group': 'summary'},
            {'metric_key': 'capitaluri_proprii', 'metric_label': 'CP', 'nr_rd': '101', 'metric_group': 'summary'},
            {'metric_key': 'datorii_termen_scurt', 'metric_label': 'DTS', 'nr_rd': '54', 'metric_group': 'ratio_input'},
            {'metric_key': 'str_ai', 'metric_label': 'Active Imobilizate', 'nr_rd': '25', 'metric_group': 'structure', 'structure_side': 'assets'},
            {'metric_key': 'str_ac', 'metric_label': 'Active Circulante', 'nr_rd': '42', 'metric_group': 'structure', 'structure_side': 'assets'},
            {'metric_key': 'str_cp', 'metric_label': 'Capitaluri Proprii', 'nr_rd': '101', 'metric_group': 'structure', 'structure_side': 'liabilities'},
        ]
        metrics = calculate_metrics_from_config(bilant_values, metric_configs)
        assert len(metrics['structure']['assets']) == 2
        assert len(metrics['structure']['liabilities']) == 1
        assert metrics['structure']['assets'][0]['value'] == 60000
        assert metrics['structure']['assets'][0]['percent'] == 60.0  # 60000/100000*100


# ══════════════════════════════════════════════════════════════
# eval_metric_formula
# ══════════════════════════════════════════════════════════════

class TestEvalMetricFormula:
    def test_simple_division(self):
        vals = {'a': 100, 'b': 50}
        assert eval_metric_formula('a / b', vals) == 2.0

    def test_compound_expression(self):
        vals = {'x': 10, 'y': 5, 'z': 2}
        assert eval_metric_formula('(x - y) / z', vals) == 2.5

    def test_multiplication(self):
        vals = {'a': 50, 'b': 100}
        assert eval_metric_formula('a / b * 100', vals) == 50.0

    def test_nested_parens(self):
        vals = {'a': 10, 'b': 20, 'c': 30}
        result = eval_metric_formula('a + (b * (c - a))', vals)
        assert result == 410.0  # 10 + (20 * 20)

    def test_missing_variable_returns_none(self):
        vals = {'a': 100}
        assert eval_metric_formula('a / missing', vals) is None

    def test_division_by_zero_returns_none(self):
        vals = {'a': 100, 'b': 0}
        assert eval_metric_formula('a / b', vals) is None

    def test_unary_minus(self):
        vals = {'a': 5}
        assert eval_metric_formula('-a', vals) == -5.0

    def test_numeric_literals(self):
        vals = {'a': 50}
        assert eval_metric_formula('a * 100', vals) == 5000.0

    def test_empty_returns_none(self):
        assert eval_metric_formula('', {}) is None
        assert eval_metric_formula('  ', {}) is None
        assert eval_metric_formula(None, {}) is None

    def test_invalid_chars_raise(self):
        with pytest.raises(ValueError):
            eval_metric_formula('a; b', {'a': 1, 'b': 2})

    def test_unclosed_paren_raises(self):
        with pytest.raises(ValueError):
            eval_metric_formula('(a + b', {'a': 1, 'b': 2})

    def test_real_ratio_formula(self):
        """Test with actual Romanian accounting ratio formula."""
        vals = {
            'active_circulante': 40000,
            'datorii_termen_scurt': 25000,
            'stocuri': 15000,
        }
        # Lichiditate curenta
        assert eval_metric_formula('active_circulante / datorii_termen_scurt', vals) == 1.6
        # Lichiditate rapida
        result = eval_metric_formula('(active_circulante - stocuri) / datorii_termen_scurt', vals)
        assert result == 1.0


# ════════════════════════════════════════════════════════════════
# ANAF Parser Tests
# ════════════════════════════════════════════════════════════════

from accounting.bilant.anaf_parser import (
    _strip_html,
    _is_numeric_rd,
    _extract_rd_formula_strict,
    generate_row_mapping,
)


class TestAnafParserHelpers:
    """Test ANAF parser helper functions."""

    def test_strip_html_basic(self):
        assert _strip_html('<b>Bold</b>') == 'Bold'

    def test_strip_html_entities(self):
        assert _strip_html('A &amp; B') == 'A & B'

    def test_strip_html_nested(self):
        assert _strip_html('<p><span style="color:red">Text</span></p>') == 'Text'

    def test_strip_html_empty(self):
        assert _strip_html('') == ''

    def test_is_numeric_rd_valid(self):
        assert _is_numeric_rd('01') is True
        assert _is_numeric_rd('103') is True
        assert _is_numeric_rd('35a') is True
        assert _is_numeric_rd('7') is True

    def test_is_numeric_rd_invalid(self):
        assert _is_numeric_rd('abc') is False
        assert _is_numeric_rd('1234') is False
        assert _is_numeric_rd('') is False


class TestExtractRdFormulaStrict:
    """Test strict RD formula extraction (avoids false positives)."""

    def test_basic_rd_formula(self):
        desc = 'TOTAL (rd. 01 la 06)'
        result = _extract_rd_formula_strict(desc)
        assert result == '01+02+03+04+05+06'

    def test_rd_with_addition(self):
        desc = 'TOTAL (rd. 23+24+25)'
        result = _extract_rd_formula_strict(desc)
        assert result == '23+24+25'

    def test_rd_with_subtraction(self):
        desc = 'TOTAL (rd. 40-41+42)'
        result = _extract_rd_formula_strict(desc)
        assert result == '40-41+42'

    def test_no_false_positive_acordate(self):
        """'rd' inside 'acordate' should NOT match."""
        desc = 'Creante legate de participatiile acordate (ct.2671+2672-2964)'
        result = _extract_rd_formula_strict(desc)
        assert result == ''

    def test_no_false_positive_pierderea(self):
        """'rd' inside 'pierderea' should NOT match."""
        desc = 'VI. PROFITUL SAU PIERDEREA EXERCITIULUI FINANCIAR'
        result = _extract_rd_formula_strict(desc)
        assert result == ''

    def test_no_false_positive_abordate(self):
        desc = 'Probleme abordate in context (ct.100)'
        result = _extract_rd_formula_strict(desc)
        assert result == ''

    def test_empty_input(self):
        assert _extract_rd_formula_strict('') == ''
        assert _extract_rd_formula_strict(None) == ''

    def test_rd_without_dot(self):
        desc = 'TOTAL (rd 01+02+03)'
        result = _extract_rd_formula_strict(desc)
        assert result == '01+02+03'


class TestGenerateRowMapping:
    """Test ANAF field name mapping generation."""

    def test_basic_mapping(self):
        rows = [
            {'nr_rd': '01', 'description': 'Test'},
            {'nr_rd': '103', 'description': 'Test 2'},
            {'nr_rd': None, 'description': 'Section'},
        ]
        mapping = generate_row_mapping(rows)
        assert 'F10_0011' in mapping
        assert mapping['F10_0011'] == {'row': '01', 'col': 'C1'}
        assert 'F10_0012' in mapping
        assert mapping['F10_0012'] == {'row': '01', 'col': 'C2'}
        assert 'F10_1031' in mapping
        assert mapping['F10_1031'] == {'row': '103', 'col': 'C1'}
        # Section row with no nr_rd should not be in mapping
        assert len(mapping) == 4  # 2 rows * 2 columns

    def test_35a_mapping(self):
        """Row '35a' (B numbering) should produce ANAF field codes F10_3011/F10_3012."""
        rows = [{'nr_rd': '35a', 'description': 'Dividende'}]
        mapping = generate_row_mapping(rows)
        assert 'F10_3011' in mapping
        assert mapping['F10_3011'] == {'row': '35a', 'col': 'C1'}
        assert 'F10_3012' in mapping
        assert mapping['F10_3012'] == {'row': '35a', 'col': 'C2'}

    def test_empty_rows(self):
        assert generate_row_mapping([]) == {}


# ════════════════════════════════════════════════════════════════
# PDF Handler Tests
# ════════════════════════════════════════════════════════════════

from accounting.bilant.pdf_handler import generate_bilant_pdf, _fmt_number, _strip_diacritics


class TestPdfHandler:
    """Test PDF generation functions."""

    def test_fmt_number_basic(self):
        assert _fmt_number(1234567) == '1.234.567'
        assert _fmt_number(0) == ''
        assert _fmt_number(None) == ''
        assert _fmt_number('') == ''

    def test_fmt_number_negative(self):
        assert _fmt_number(-5000) == '-5.000'

    def test_strip_diacritics(self):
        assert _strip_diacritics('ăâîșț') == 'aais' or _strip_diacritics('ăâîșț') in ('aaist', 'aais', 'aaiST', 'aaist')
        # At minimum, should not contain combining characters
        result = _strip_diacritics('ăâî')
        assert all(ord(c) < 768 for c in result)  # No combining chars

    def test_generate_pdf_basic(self):
        gen = {'company_name': 'Test SRL', 'period_label': 'Q4 2025'}
        results = [
            {'nr_rd': '01', 'description': 'Active imobilizate', 'value': 50000,
             'row_type': 'data', 'is_bold': False, 'indent_level': 1},
            {'nr_rd': '06', 'description': 'TOTAL', 'value': 50000,
             'row_type': 'total', 'is_bold': True, 'indent_level': 0},
            {'nr_rd': None, 'description': 'B. ACTIVE CIRCULANTE', 'value': 0,
             'row_type': 'section', 'is_bold': True, 'indent_level': 0},
        ]
        output = generate_bilant_pdf(gen, results)
        assert output is not None
        content = output.getvalue()
        assert len(content) > 0
        assert content[:4] == b'%PDF'  # Valid PDF header

    def test_generate_pdf_with_prior(self):
        gen = {'company_name': 'Test', 'period_label': '2025'}
        results = [
            {'nr_rd': '01', 'description': 'Row 1', 'value': 100, 'row_type': 'data',
             'is_bold': False, 'indent_level': 0},
        ]
        prior = {'01': 80}
        output = generate_bilant_pdf(gen, results, prior_results=prior)
        assert output.getvalue()[:4] == b'%PDF'


# ════════════════════════════════════════════════════════════════
# ANAF Excel Export Tests
# ════════════════════════════════════════════════════════════════

from accounting.bilant.excel_handler import generate_anaf_excel


class TestAnafExcelExport:
    """Test ANAF-format Excel export."""

    def test_basic_export(self):
        gen = {'company_name': 'Test SRL', 'period_label': 'Q4 2025'}
        results = [
            {'nr_rd': '01', 'description': 'Row 1', 'value': 5000, 'is_bold': False,
             'row_type': 'data', 'indent_level': 1, 'sort_order': 0},
            {'nr_rd': '06', 'description': 'TOTAL', 'value': 5000, 'is_bold': True,
             'row_type': 'total', 'indent_level': 0, 'sort_order': 1},
        ]
        output = generate_anaf_excel(gen, results)
        assert output is not None
        # Verify it's a valid xlsx
        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert 'F10' in wb.sheetnames
        assert 'Bilant' in wb.sheetnames
        # Verify F10 sheet has field codes
        ws = wb['F10']
        assert ws.cell(1, 1).value == 'Field Code'
        # Should have rows for F10_0011, F10_0012, F10_0061, F10_0062
        field_codes = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        assert 'F10_0011' in field_codes
        assert 'F10_0062' in field_codes

    def test_export_with_prior_results(self):
        gen = {'company_name': 'Test', 'period_label': '2025'}
        results = [
            {'nr_rd': '01', 'description': 'R1', 'value': 1000, 'is_bold': False,
             'row_type': 'data', 'indent_level': 0, 'sort_order': 0},
        ]
        prior = {'01': 800}
        output = generate_anaf_excel(gen, results, prior_results=prior)
        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb['F10']
        # Find F10_0011 (C1 = prior) and verify value
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == 'F10_0011':
                assert ws.cell(r, 2).value == 800
                break

    def test_export_with_row_mapping(self):
        gen = {'company_name': 'Test', 'period_label': '2025'}
        results = [
            {'nr_rd': '01', 'description': 'R1', 'value': 500, 'is_bold': False,
             'row_type': 'data', 'indent_level': 0, 'sort_order': 0},
        ]
        mapping = {'F10_0011': {'row': '01', 'col': 'C1'}, 'F10_0012': {'row': '01', 'col': 'C2'}}
        output = generate_anaf_excel(gen, results, row_mapping=mapping)
        assert output is not None


class TestFillAnafPdf:
    """Test XFA PDF field filling."""

    TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'jarvis',
                                  'accounting', 'bilant', 'static', 'anaf_f10l_template.pdf')

    @pytest.fixture
    def template_exists(self):
        return os.path.exists(self.TEMPLATE_PATH)

    def test_fill_basic_values(self, template_exists):
        if not template_exists:
            pytest.skip('ANAF template PDF not available')
        from accounting.bilant.anaf_parser import fill_anaf_pdf
        values = {'01': 1234567, '02': 890123, '103': 9999}
        result = fill_anaf_pdf(values)
        assert result is not None
        data = result.getvalue()
        assert len(data) > 100000  # Should be a substantial PDF

    def test_fill_with_prior_values(self, template_exists):
        if not template_exists:
            pytest.skip('ANAF template PDF not available')
        from accounting.bilant.anaf_parser import fill_anaf_pdf
        import pikepdf
        import xml.etree.ElementTree as ET

        values = {'01': 5000, '35': 8000}
        prior = {'01': 3000, '35': 6000}
        result = fill_anaf_pdf(values, prior_values=prior)

        # Verify by re-reading the filled PDF
        result.seek(0)
        pdf = pikepdf.Pdf.open(result)
        xfa = pdf.Root.AcroForm.XFA
        for i in range(0, len(xfa), 2):
            if str(xfa[i]) == 'datasets':
                xml_data = bytes(xfa[i + 1].read_bytes())
                root = ET.fromstring(xml_data)
                ns = '{http://www.xfa.org/schema/xfa-data/1.0/}'
                table = root.find(f'{ns}data').find('form1').find('F10L').find('Table1')
                r01 = table.find('R01')
                assert r01.find('C2').text == '5000'
                assert r01.find('C1').text == '3000'
                r35 = table.find('R35')
                assert r35.find('C2').text == '8000'
                assert r35.find('C1').text == '6000'
                break
        pdf.close()

    def test_fill_preserves_pdf_structure(self, template_exists):
        if not template_exists:
            pytest.skip('ANAF template PDF not available')
        from accounting.bilant.anaf_parser import fill_anaf_pdf
        import pikepdf

        values = {'01': 100}
        result = fill_anaf_pdf(values)
        result.seek(0)
        pdf = pikepdf.Pdf.open(result)
        # Should still have XFA
        assert hasattr(pdf.Root, 'AcroForm')
        xfa = pdf.Root.AcroForm.XFA
        assert len(xfa) == 18  # Same stream count as original
        # Should have pages
        assert len(pdf.pages) > 0
        pdf.close()

    def test_fill_zero_values_skipped(self, template_exists):
        if not template_exists:
            pytest.skip('ANAF template PDF not available')
        from accounting.bilant.anaf_parser import fill_anaf_pdf
        import pikepdf
        import xml.etree.ElementTree as ET

        values = {'01': 0, '02': 500}
        result = fill_anaf_pdf(values)
        result.seek(0)
        pdf = pikepdf.Pdf.open(result)
        xfa = pdf.Root.AcroForm.XFA
        for i in range(0, len(xfa), 2):
            if str(xfa[i]) == 'datasets':
                xml_data = bytes(xfa[i + 1].read_bytes())
                root = ET.fromstring(xml_data)
                ns = '{http://www.xfa.org/schema/xfa-data/1.0/}'
                table = root.find(f'{ns}data').find('form1').find('F10L').find('Table1')
                r01 = table.find('R01')
                # C2 for R01 should remain empty (zero skipped)
                assert r01.find('C2').text is None or r01.find('C2').text == ''
                r02 = table.find('R02')
                assert r02.find('C2').text == '500'
                break
        pdf.close()

    def test_fill_special_rows(self, template_exists):
        """Test filling R301 via B-numbered '35a' key."""
        if not template_exists:
            pytest.skip('ANAF template PDF not available')
        from accounting.bilant.anaf_parser import fill_anaf_pdf
        import pikepdf
        import xml.etree.ElementTree as ET

        values = {'35a': 77777}  # B numbering — maps to XFA element R301
        result = fill_anaf_pdf(values)
        result.seek(0)
        pdf = pikepdf.Pdf.open(result)
        xfa = pdf.Root.AcroForm.XFA
        for i in range(0, len(xfa), 2):
            if str(xfa[i]) == 'datasets':
                xml_data = bytes(xfa[i + 1].read_bytes())
                root = ET.fromstring(xml_data)
                ns = '{http://www.xfa.org/schema/xfa-data/1.0/}'
                table = root.find(f'{ns}data').find('form1').find('F10L').find('Table1')
                r301 = table.find('R301')
                assert r301.find('C2').text == '77777'
                break
        pdf.close()

    def test_fill_missing_template_raises(self):
        from accounting.bilant.anaf_parser import fill_anaf_pdf
        with pytest.raises(ValueError, match='template not found'):
            fill_anaf_pdf({'01': 100}, template_path='/nonexistent/path.pdf')

    def test_fill_rounds_floats(self, template_exists):
        """Values should be rounded to integers (ANAF format)."""
        if not template_exists:
            pytest.skip('ANAF template PDF not available')
        from accounting.bilant.anaf_parser import fill_anaf_pdf
        import pikepdf
        import xml.etree.ElementTree as ET

        values = {'01': 1234.56}
        result = fill_anaf_pdf(values)
        result.seek(0)
        pdf = pikepdf.Pdf.open(result)
        xfa = pdf.Root.AcroForm.XFA
        for i in range(0, len(xfa), 2):
            if str(xfa[i]) == 'datasets':
                xml_data = bytes(xfa[i + 1].read_bytes())
                root = ET.fromstring(xml_data)
                ns = '{http://www.xfa.org/schema/xfa-data/1.0/}'
                table = root.find(f'{ns}data').find('form1').find('F10L').find('Table1')
                r01 = table.find('R01')
                assert r01.find('C2').text == '1235'  # Rounded
                break
        pdf.close()


class TestAnafExportFormats:
    """Tests for ANAF XML and TXT export formats."""

    def test_generate_anaf_txt_basic(self):
        """Generate balanta.txt with C2 values."""
        from accounting.bilant.anaf_parser import generate_anaf_txt
        values = {'01': 1000, '02': 500, '10': 2000}
        txt = generate_anaf_txt(values, company_name='TEST SRL', cif='12345678')
        lines = txt.strip().split('\r\n')
        assert lines[0].startswith('BL,12345678,TEST SRL,')
        # 27 fields in identification line (since 12/2022)
        assert len(lines[0].split(',')) == 27
        assert ',F10L,R01,C2,,1000' in txt
        assert ',F10L,R02,C2,,500' in txt
        assert ',F10L,R10,C2,,2000' in txt

    def test_generate_anaf_txt_with_prior(self):
        """Generate balanta.txt with C1 and C2 values."""
        from accounting.bilant.anaf_parser import generate_anaf_txt
        values = {'01': 1000}
        prior = {'01': 800}
        txt = generate_anaf_txt(values, prior_values=prior)
        assert ',F10L,R01,C2,,1000' in txt
        assert ',F10L,R01,C1,,800' in txt

    def test_generate_anaf_txt_negative_values(self):
        """Negative values get sign field '-'."""
        from accounting.bilant.anaf_parser import generate_anaf_txt
        values = {'01': -500}
        txt = generate_anaf_txt(values)
        assert ',F10L,R01,C2,-,500' in txt

    def test_generate_anaf_txt_crlf(self):
        """Lines end with CR+LF."""
        from accounting.bilant.anaf_parser import generate_anaf_txt
        txt = generate_anaf_txt({'01': 100})
        assert '\r\n' in txt

    def test_generate_anaf_txt_zero_skipped(self):
        """Zero values are not included."""
        from accounting.bilant.anaf_parser import generate_anaf_txt
        values = {'01': 0, '02': 100}
        txt = generate_anaf_txt(values)
        assert 'R01' not in txt
        assert ',F10L,R02,C2,,100' in txt

    def test_generate_anaf_txt_35a(self):
        """Row '35a' (B numbering) should export as R301 in ANAF TXT."""
        from accounting.bilant.anaf_parser import generate_anaf_txt
        values = {'35a': 5000}
        txt = generate_anaf_txt(values)
        assert ',F10L,R301,C2,,5000' in txt
        assert 'R35a' not in txt  # Should use ANAF code, not B numbering

    def test_generate_anaf_xml_basic(self):
        """Generate ANAF XML with Bilant1002 root and F10 attributes."""
        from accounting.bilant.anaf_parser import generate_anaf_xml
        values = {'01': 1000, '02': 500}
        xml_bytes = generate_anaf_xml(values, company_name='TEST SRL', cif='12345678')
        text = xml_bytes.decode('utf-8')
        # Root tag is Bilant1002 for F10L
        assert '<Bilant1002' in text
        assert 'cui="12345678"' in text
        assert 'den="TEST SRL"' in text
        # F10 element with field attributes: F10_XXXC
        assert 'F10_0012="1000"' in text  # row 01, col 2 (C2)
        assert 'F10_0022="500"' in text   # row 02, col 2 (C2)
        assert '</Bilant1002>' in text

    def test_generate_anaf_xml_with_prior(self):
        """XML includes C1 values as F10_XXX1 attributes."""
        from accounting.bilant.anaf_parser import generate_anaf_xml
        values = {'01': 1000}
        prior = {'01': 800}
        xml_bytes = generate_anaf_xml(values, prior_values=prior)
        text = xml_bytes.decode('utf-8')
        assert 'F10_0011="800"' in text   # row 01, col 1 (C1 prior)
        assert 'F10_0012="1000"' in text  # row 01, col 2 (C2 current)

    def test_generate_anaf_xml_period_date(self):
        """Period date populates luna and an attributes."""
        from accounting.bilant.anaf_parser import generate_anaf_xml
        xml_bytes = generate_anaf_xml({'01': 100}, period_date='2024-12-31')
        text = xml_bytes.decode('utf-8')
        assert 'an="2024"' in text
        assert 'luna="12"' in text

    def test_generate_anaf_xml_has_declaration(self):
        """XML output starts with <?xml version="1.0"?>."""
        from accounting.bilant.anaf_parser import generate_anaf_xml
        xml_bytes = generate_anaf_xml({'01': 100})
        assert xml_bytes.startswith(b'<?xml version="1.0"?>')

    def test_generate_anaf_xml_bilant1002_marker(self):
        """ANAF import function searches for 'Bilant1002' string."""
        from accounting.bilant.anaf_parser import generate_anaf_xml
        xml_bytes = generate_anaf_xml({'01': 100}, form='F10L')
        assert b'Bilant1002' in xml_bytes
        # Small entities use Bilant1003
        xml_bytes_s = generate_anaf_xml({'01': 100}, form='F10S')
        assert b'Bilant1003' in xml_bytes_s

    def test_generate_anaf_xml_namespace(self):
        """XML includes required ANAF namespace/schema."""
        from accounting.bilant.anaf_parser import generate_anaf_xml
        xml_bytes = generate_anaf_xml({'01': 100})
        text = xml_bytes.decode('utf-8')
        assert 'xmlns:xsi=' in text
        assert 'mfp:anaf:dgti:s1002' in text

    def test_generate_anaf_xml_field_padding(self):
        """Row numbers are 3-digit zero-padded in field names."""
        from accounting.bilant.anaf_parser import generate_anaf_xml
        values = {'1': 100, '10': 200, '103': 300}
        xml_bytes = generate_anaf_xml(values)
        text = xml_bytes.decode('utf-8')
        assert 'F10_0012="100"' in text   # row '1' padded to '001'
        assert 'F10_0102="200"' in text   # row '10' padded to '010'
        assert 'F10_1032="300"' in text   # row '103' stays '103'

    def test_generate_anaf_xml_35a(self):
        """Row '35a' (B numbering) should export as F10_301X in ANAF XML."""
        from accounting.bilant.anaf_parser import generate_anaf_xml
        values = {'35': 1000, '35a': 2000, '36': 3000}
        xml_bytes = generate_anaf_xml(values)
        text = xml_bytes.decode('utf-8')
        assert 'F10_0352="1000"' in text   # row '35' → '035'
        assert 'F10_3012="2000"' in text   # row '35a' → '301'
        assert 'F10_0362="3000"' in text   # row '36' → '036'
